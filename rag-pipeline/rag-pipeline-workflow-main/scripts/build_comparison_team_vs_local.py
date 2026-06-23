#!/usr/bin/env python3
"""So sánh file team (OpenAI pipeline thật) vs file local test (BM25) -> workbook."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEAM=os.path.join(ROOT,"reports","goldns_emni_rag_vs_gold_comparison.xlsx")
MINE=os.path.join(ROOT,"reports","local_rag_chunk_vs_excel_test_20260619.xlsx")
OUT=os.path.join(ROOT,"reports","SO_SANH_team_vs_local_20260619.xlsx")

def load(f,sheet,idcol):
    wb=openpyxl.load_workbook(f,read_only=True,data_only=True); ws=wb[sheet]
    rows=list(ws.iter_rows(values_only=True)); hdr=list(rows[0]); idx={h:i for i,h in enumerate(hdr)}
    return [{h:r[idx[h]] for h in hdr} for r in rows[1:] if r and r[idx[idcol]]]

team={}
for sh in ["goldns_compare","emni_compare"]:
    for r in load(TEAM,sh,"question_id"): team[str(r['question_id']).lower()]=r
mine=load(MINE,"detail","id")
norm=lambda x:str(x).lower().replace("-q","-")

wb=openpyxl.Workbook()
BOLD=Font(bold=True); HEAD=Font(bold=True,color="FFFFFF"); FILL=PatternFill("solid",fgColor="305496")
def hrow(ws,cells):
    ws.append(cells)
    for c in ws[ws.max_row]: c.font=HEAD; c.fill=FILL

# 1) Tong quan metric
ws=wb.active; ws.title="1_doi_chieu_metric"
ws.append(["ĐỐI CHIẾU: TEAM (OpenAI thật) vs LOCAL (BM25 sandbox)"]); ws["A1"].font=Font(bold=True,size=14)
ws.append([])
hrow(ws,["Tiêu chí","TEAM — goldns_emni_rag_vs_gold","LOCAL — local_rag_chunk_vs_excel"])
data=[
 ["Pipeline","OpenAI text-embedding-3-small + Qdrant + FlashRank rerank + gpt-4o-mini (generative)","BM25 lexical + value substring match (KHÔNG generation, KHÔNG rerank)"],
 ["Corpus (văn bản gốc)","Local JSON/DART thật: 2025_empSttus.json, 2025_재무_OFS.json, 제재이력*.json ...","7 file .md TalkFile (đã curate từ nguồn)"],
 ["Số câu","530 (goldns 251 + emni 279) — TOÀN BỘ","200 (sample: 54 'answerable' + 146 abstain)"],
 ["answer_correct","530/530 = 100%","(không sinh câu trả lời — chỉ đo match)"],
 ["retrieval_hit_top1","526/530 = 99.2% (4 miss)","value-match 51/54 = 94.4% (proxy cho source_match)"],
 ["source_match_top1","526/530 = 99.2%","— (chỉ match chuỗi giá trị, không match cấp document)"],
 ["abstain_correct","463/463 = 100%","146/146 = 100%"],
 ["overall_score (v5)","0.9702","không tính (thiếu generation + 5-metric)"],
 ["Lỗi còn lại","4 semantic_ambiguity + 2 coverage_gap (FTC blocked)","3 MISS = proxy data (không có trong .md gốc)"],
 ["Chẩn đoán/diagnostics","Đầy đủ: fail_type, review_owner (SME/Dataset/RAG), semantic_audit, coverage_gap","Tối giản: HIT/MISS/ABSTAIN-OK"],
]
for row in data:
    ws.append(row)
    ws[ws.max_row][0].font=BOLD
for col,w in {"A":24,"B":52,"C":52}.items(): ws.column_dimensions[col].width=w
for r in ws.iter_rows(min_row=3):
    for c in r: c.alignment=Alignment(wrap_text=True,vertical="top")

# 2) Join theo question_id
ws2=wb.create_sheet("2_join_theo_question_id")
hrow(ws2,["question_id","cat","sub","gold_2025","LOCAL_verdict","TEAM_partition",
          "TEAM_retrieval_hit","TEAM_source_match","TEAM_status","TEAM_top1_doc","Ghi chú"])
green=PatternFill("solid",fgColor="C6EFCE"); red=PatternFill("solid",fgColor="FFC7CE"); yel=PatternFill("solid",fgColor="FFF2CC")
n_join=n_ans=n_agree=n_falsepos=n_misclass=0
for m in mine:
    qid=norm(m['id']); t=team.get(qid)
    if not t: continue
    n_join+=1
    part=str(t['partition'])
    note=""
    local_ans = m['verdict'] in ('HIT','MISS')
    team_ans = part=='answerable_gold'
    if local_ans and not team_ans:
        note="LOCAL phân loại answerable nhưng TEAM = abstain (heuristic local thừa)"; n_misclass+=1
    if m['verdict']=='HIT' and str(t['source_match_top1'])=='False':
        note="LOCAL HIT giả (substring '1' trùng ngẫu nhiên) — TEAM bắt đúng là miss"; n_falsepos+=1
    if team_ans:
        n_ans+=1
        if (m['verdict']=='HIT')==(str(t['source_match_top1'])=='True'): n_agree+=1
    if note or part=='answerable_gold':
        ws2.append([qid,m['cat'],m['sub'],str(m['gold_2025']),m['verdict'],part,
                    str(t['retrieval_hit_top1']),str(t['source_match_top1']),
                    str(t['comparison_status']),str(t['rag_top1_doc'])[:30],note])
        cell=ws2.cell(ws2.max_row,5)
        cell.fill=green if m['verdict']=='HIT' else red if m['verdict']=='MISS' else yel
for col,w in {"A":13,"B":18,"C":22,"D":12,"E":13,"F":15,"L":40,"K":48}.items(): ws2.column_dimensions[col].width=w

# 3) Ket luan
ws3=wb.create_sheet("3_ket_luan")
ws3.append(["KẾT LUẬN SO SÁNH"]); ws3["A1"].font=Font(bold=True,size=14)
for line in [
 "",
 f"• Join được {n_join}/200 câu local với file team.",
 f"• Trong sample, TEAM coi {n_ans} câu là answerable_gold; LOCAL gắn nhãn answerable cho 54 câu",
 f"  -> {n_misclass} câu local nhận nhầm 'answerable' (thực ra team xếp abstain). Heuristic local (chỉ",
 "   dựa cột '2025 value') yếu hơn partition đã reconcile của team.",
 f"• {n_falsepos} câu LOCAL=HIT giả (goldns-0237, emni-0237): BM25 + substring khớp số '1' ngẫu nhiên;",
 "   TEAM xác định đúng là retrieval miss vì nguồn thật (제재이력/sanction JSON) không được truy xuất.",
 "",
 "ĐÁNH GIÁ:",
 "• TEAM (OpenAI pipeline thật) là chuẩn: full 530 câu, corpus JSON/DART gốc, có generation + abstain,",
 "  source_match cấp document, diagnostics đầy đủ -> answer 100%, retrieval 99.2%, overall 0.9702.",
 "• LOCAL chỉ là smoke-test feasibility trong sandbox (không gọi được OpenAI/Qdrant): xác nhận",
 "  'chunk văn bản gốc + đối chiếu Excel' chạy được, nhưng yếu ở 3 điểm:",
 "   (1) phân loại answerable/abstain thô; (2) substring match gây HIT giả; (3) corpus .md nhỏ hơn,",
 "   không có generation nên không đo được chất lượng câu trả lời.",
 "",
 "KHUYẾN NGHỊ: dùng LOCAL để kiểm tra nhanh logic chunk/retrieve; số liệu chính thức lấy từ pipeline",
 "team (OpenAI + Qdrant). Khi cần chạy lại OpenAI thật, dùng scripts/openai_chunk_vs_excel_test.py trên máy.",
]:
    ws3.append([line])
ws3.column_dimensions["A"].width=100
wb.save(OUT)
print("joined",n_join,"team_answerable_in_sample",n_ans,"agree",n_agree,
      "misclass",n_misclass,"falsepos",n_falsepos)
print("Saved:",OUT)
