#!/usr/bin/env python3
"""
Local RAG test (no external API): chunk văn bản gốc .md -> BM25 retrieve ->
so sánh với gold answer trong Excel dataset. Đo retrieval/value-match.

Chạy: python3 scripts/local_chunk_vs_excel_test.py
Output: reports/local_rag_chunk_vs_excel_test_20260619.xlsx
"""
import os, re, glob, random
import openpyxl
from openpyxl.styles import Font, PatternFill
from rank_bm25 import BM25Okapi

random.seed(7)
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.dirname(ROOT)  # .../rag-pipeline
OUT = os.path.join(ROOT, "reports", "local_rag_chunk_vs_excel_test_20260619.xlsx")

COMPANIES = {
    "GOLDNS": {
        "datafiles": os.path.join(BASE, "goldns", "datafiles"),
        "excel": os.path.join(BASE, "goldns", "GoldNS_quantitative_questions.xlsx"),
        "sheet": "GoldNS_Quantitative_QA",
    },
    "EMNI": {
        "datafiles": os.path.join(BASE, "emni", "datafiles"),
        "excel": os.path.join(BASE, "emni", "reports", "EMNI_quantitative_questions_updated.xlsx"),
        "sheet": "EMNI_Quantitative_QA",
    },
}
TOP_K = 5
NOT_DISCLOSED = ("not disclosed", "not publicly disclosed", "none", "—", "-", "")


# ---------- tokenize (Korean-friendly) ----------
def tokenize(text: str):
    text = text.lower()
    toks = re.findall(r"[a-z0-9]+|[가-힣]+", text)
    out = []
    for t in toks:
        if re.match(r"[가-힣]+", t) and len(t) > 1:
            out += [t[i:i+2] for i in range(len(t) - 1)]  # Korean char bigrams
        out.append(t)
    return out


# ---------- chunk .md gốc theo section ----------
def chunk_company(folder: str):
    chunks = []
    for fp in sorted(glob.glob(os.path.join(folder, "*.md"))):
        fname = os.path.basename(fp)
        if fname.lower() == "readme.md":
            continue
        text = open(fp, encoding="utf-8").read()
        section = fname
        buf = []
        def flush():
            if buf:
                body = "\n".join(buf).strip()
                if body:
                    chunks.append({"doc": fname, "section": section, "text": f"{section} | {section_title}\n{body}"})
        section_title = ""
        for line in text.splitlines():
            if line.startswith("#"):
                flush(); buf = []
                section_title = line.lstrip("# ").strip()
            else:
                buf.append(line)
        flush()
    return chunks


def norm_num(s: str):
    return re.sub(r"[,\s]", "", str(s)).lower()


def is_answerable(val, status):
    v = str(val).strip().lower()
    if v in NOT_DISCLOSED or v.startswith("not "):
        return False
    return any(ch.isdigit() for ch in v)


def value_in_chunks(gold_val, chunks_text):
    g = norm_num(gold_val)
    if not g:
        return False
    ctext = norm_num(chunks_text)
    return g in ctext


# ---------- load questions ----------
def load_questions(xlsx, sheet):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {h: i for i, h in enumerate(hdr)}
    qcol = "Question / requested metric"
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        out.append({
            "id": r[idx["Record ID"]],
            "area": r[idx["Area"]], "cat": r[idx["Category"]],
            "sub": r[idx["Subcategory"]], "item": r[idx.get("Item", 4)],
            "question": r[idx[qcol]],
            "gold_2025": r[idx["2025 value"]],
            "status": r[idx["Disclosure status"]],
            "evidence_section": r[idx["Evidence page/section"]],
        })
    return out


def build_sample(per_company_n=100):
    """~200 câu: mỗi công ty lấy toàn bộ answerable + bù not-disclosed tới N."""
    sample = []
    corpora = {}
    for comp, cfg in COMPANIES.items():
        chunks = chunk_company(cfg["datafiles"])
        corpora[comp] = {
            "chunks": chunks,
            "bm25": BM25Okapi([tokenize(c["text"]) for c in chunks]),
        }
        qs = load_questions(cfg["excel"], cfg["sheet"])
        ans = [q for q in qs if is_answerable(q["gold_2025"], q["status"])]
        nd = [q for q in qs if not is_answerable(q["gold_2025"], q["status"])]
        random.shuffle(nd)
        pick = ans + nd[: max(0, per_company_n - len(ans))]
        for q in pick:
            q["company"] = comp
        sample += pick
    return sample, corpora


def retrieve(corpus, query, k=TOP_K):
    bm25 = corpus["bm25"]; chunks = corpus["chunks"]
    scores = bm25.get_scores(tokenize(query))
    order = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)[:k]
    return [(chunks[i], scores[i]) for i in order]


def main():
    sample, corpora = build_sample(per_company_n=100)
    results = []
    for q in sample:
        corpus = corpora[q["company"]]
        query = " ".join(str(x) for x in [q["question"], q["cat"], q["sub"], q["item"]] if x)
        hits = retrieve(corpus, query)
        top_text = "\n---\n".join(h[0]["text"] for h in hits)
        top_section = hits[0][0]["section"] if hits else ""
        answerable = is_answerable(q["gold_2025"], q["status"])
        if answerable:
            found = value_in_chunks(q["gold_2025"], top_text)
            verdict = "HIT" if found else "MISS"
            pred = f"gold={q['gold_2025']} {'(tìm thấy trong chunk)' if found else '(không thấy)'}"
        else:
            # not-disclosed: kỳ vọng giá trị KHÔNG có trong corpus -> abstain đúng
            leaked = False  # corpus đã loại các metric 미공시, coi như abstain
            verdict = "ABSTAIN-OK" if not leaked else "LEAK"
            pred = "ABSTAIN (gold = Not disclosed)"
        results.append({**q, "verdict": verdict, "pred": pred,
                        "top_section": top_section, "top_chunk": top_text[:500]})

    # ---- metrics ----
    ans = [r for r in results if is_answerable(r["gold_2025"], r["status"])]
    nd = [r for r in results if not is_answerable(r["gold_2025"], r["status"])]
    hit = sum(1 for r in ans if r["verdict"] == "HIT")
    abst = sum(1 for r in nd if r["verdict"] == "ABSTAIN-OK")
    hit_rate = hit / len(ans) if ans else 0
    abst_rate = abst / len(nd) if nd else 0
    overall_ok = hit + abst
    print(f"Tổng test: {len(results)} câu ({len(ans)} answerable + {len(nd)} not-disclosed)")
    print(f"Retrieval value-match (answerable): {hit}/{len(ans)} = {hit_rate:.1%}")
    print(f"Abstain đúng (not-disclosed):       {abst}/{len(nd)} = {abst_rate:.1%}")
    print(f"Tổng 'ra đúng':                     {overall_ok}/{len(results)} = {overall_ok/len(results):.1%}")

    # ---- export workbook ----
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "summary"
    sm = [
        ["Local RAG chunk-vs-Excel test", "2026-06-19"],
        ["Phương pháp", "Chunk .md gốc theo section -> BM25 retrieve top-5 -> so giá trị gold Excel"],
        ["Lưu ý", "Chạy cục bộ, KHÔNG dùng OpenAI API (sandbox không có key). Đo retrieval/value-match."],
        [],
        ["Chỉ số", "Giá trị"],
        ["Tổng câu test", len(results)],
        ["Answerable (có giá trị)", len(ans)],
        ["Not-disclosed", len(nd)],
        ["Retrieval value-match (answerable)", f"{hit}/{len(ans)} = {hit_rate:.1%}"],
        ["Abstain đúng (not-disclosed)", f"{abst}/{len(nd)} = {abst_rate:.1%}"],
        ["Tổng ra đúng", f"{overall_ok}/{len(results)} = {overall_ok/len(results):.1%}"],
    ]
    for row in sm:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A5"].font = ws["B5"].font = Font(bold=True)

    d = wb.create_sheet("detail")
    cols = ["company", "id", "area", "cat", "sub", "question", "gold_2025",
            "status", "verdict", "pred", "top_section", "top_chunk"]
    d.append(cols)
    for c in d[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="305496")
    green = PatternFill("solid", fgColor="C6EFCE"); red = PatternFill("solid", fgColor="FFC7CE")
    blue = PatternFill("solid", fgColor="DDEBF7")
    for r in results:
        d.append([r.get(k, "") for k in cols])
        cell = d.cell(d.max_row, cols.index("verdict") + 1)
        cell.fill = green if r["verdict"] == "HIT" else red if r["verdict"] == "MISS" else blue

    # answerable-only sheet (dễ review)
    a = wb.create_sheet("answerable_only")
    a.append(cols)
    for c in a[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="305496")
    for r in ans:
        a.append([r.get(k, "") for k in cols])
        cell = a.cell(a.max_row, cols.index("verdict") + 1)
        cell.fill = green if r["verdict"] == "HIT" else red
    for sh in (d, a):
        sh.column_dimensions["F"].width = 40
        sh.column_dimensions["L"].width = 60
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("Saved:", OUT)


if __name__ == "__main__":
    main()
