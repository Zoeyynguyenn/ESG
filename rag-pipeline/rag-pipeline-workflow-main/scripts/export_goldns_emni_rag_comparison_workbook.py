#!/usr/bin/env python3
"""Export Excel workbook comparing Dataset gold vs frozen RAG eval results."""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from ingest_esg_excel_workbook import EXPECTED_HEADERS, WorkbookSpec, _load_rows  # noqa: E402

BASE_RAG_COLUMNS = [
    "question_id",
    "partition",
    "question_family",
    "gold_answer_canonical",
    "rag_predicted_answer",
    "rag_predicted_abstain",
    "answer_correct",
    "abstain_correct",
    "retrieval_hit_top1",
    "source_match_top1",
    "top_score",
    "rag_top1_doc",
    "rag_topk_docs",
    "predict_reason",
    "fail_type",
    "diagnostic_tags",
    "semantic_ambiguity",
    "semantic_audit_note",
    "comparison_status",
    "comparison_note",
]

REVIEW_COLUMNS = [
    "retrieval_review_status",
    "review_priority",
    "review_bucket",
    "needs_sme_check",
    "needs_dataset_check",
    "needs_rag_check",
    "review_owner_hint",
    "top1_vs_gold_doc_match",
    "source_gap_reason",
    "business_meaning_note",
]

EXTENSION_COLUMNS = BASE_RAG_COLUMNS + REVIEW_COLUMNS

# Cot hien thi som de reviewer khong phai scroll den cot 40+
CONTEXT_FRONT_HEADERS = ["영역", "카테고리", "서브카테고리", "항목", "Year", "Value", "단위", "Disclosure status"]
REVIEW_PINNED_HEADERS = [
    "question_id",
    "comparison_status",
    "review_priority",
    "review_owner_hint",
    "retrieval_review_status",
    "business_meaning_note",
    "rag_predicted_answer",
    "gold_answer_canonical",
    "comparison_note",
    "review_bucket",
    "needs_sme_check",
    "needs_dataset_check",
    "needs_rag_check",
    "top1_vs_gold_doc_match",
    "source_gap_reason",
]
COMPACT_REVIEW_TAIL = [
    "question_family",
    "partition",
    "retrieval_hit_top1",
    "source_match_top1",
    "rag_top1_doc",
    "Source document/page",
    "Source URL",
    "fail_type",
    "diagnostic_tags",
    "semantic_audit_note",
]

# Chi freeze 4 cot dinh danh — tranh loi Excel "cua so qua nho"
FREEZE_PANES = "E2"
SHEET_ZOOM = 85
FREEZE_PANES_COMPACT = "I2"  # freeze den Value (cot Year/Value/단위/Disclosure trong vung freeze compact)

BOLD_HEADERS = {
    "Value",
    "rag_predicted_answer",
    "comparison_status",
    "comparison_note",
    "review_priority",
    "review_owner_hint",
    "retrieval_review_status",
}
WRAP_HEADERS = {
    "comparison_note",
    "business_meaning_note",
    "Source detail / calculation note",
    "Evidence",
    "semantic_audit_note",
}

GREEN_STATUSES = {"MATCH", "ABSTAIN_OK"}

STATUS_FILLS = {
    "MATCH": PatternFill(fill_type="solid", fgColor="A9D18E"),
    "ABSTAIN_OK": PatternFill(fill_type="solid", fgColor="C6E0B4"),
    "RETRIEVAL_MISS_BUT_ANSWER_OK": PatternFill(fill_type="solid", fgColor="FFD966"),
    "ANSWER_MISMATCH": PatternFill(fill_type="solid", fgColor="F8696B"),
    "COVERAGE_GAP": PatternFill(fill_type="solid", fgColor="F4B084"),
    "SEMANTIC_AMBIGUITY": PatternFill(fill_type="solid", fgColor="B4A7D6"),
    "REVIEW_NEEDED": PatternFill(fill_type="solid", fgColor="BFBFBF"),
}

RETRIEVAL_STATUS_FILL = PatternFill(fill_type="solid", fgColor="FFE699")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="2F5597")
HEADER_FONT = Font(color="FFFFFF", bold=True)
REVIEW_ROW_BORDER = Border(
    left=Side(style="medium", color="7F7F7F"),
    right=Side(style="medium", color="7F7F7F"),
    top=Side(style="medium", color="7F7F7F"),
    bottom=Side(style="medium", color="7F7F7F"),
)

REVIEW_GUIDE_LINES = [
    ("review_guide", "Hướng dẫn review workbook RAG vs Gold"),
    ("", ""),
    ("Status / nhóm", "Ý nghĩa", "Reviewer nên làm gì"),
    ("MATCH", "RAG trả đúng đáp án gold và retrieval top-1 khớp nguồn", "Không cần review thêm — có thể bỏ qua"),
    ("ABSTAIN_OK", "Gold không có provenance; RAG abstain đúng", "Không cần review thêm — có thể bỏ qua"),
    ("SEMANTIC_AMBIGUITY", "Đáp án có thể đúng số nhưng label workbook chưa map rõ với account/source", "SME kiểm tra ý nghĩa chỉ tiêu và mapping gold"),
    ("COVERAGE_GAP", "Thiếu raw source thật (ví dụ FTC blocked)", "Dataset team bổ sung nguồn; không vá extractor"),
    ("ANSWER_MISMATCH", "RAG trả sai đáp án hoặc abstain sai", "RAG team kiểm tra retrieval/extractor"),
    ("TOP1_MISS_BUT_ANSWER_OK", "Đáp án đúng nhưng top-1 retrieval trỏ sai tài liệu/lane", "RAG xem rerank/routing; SME xem nếu kèm semantic tag"),
    ("", ""),
    ("Sheet review", "Mục đích"),
    ("review_all_non_green", "Tất cả dòng cần xem — không phải MATCH/ABSTAIN_OK"),
    ("review_semantic_ambiguity", "Chỉ các dòng cần SME"),
    ("review_coverage_gap", "Chỉ các dòng thiếu source"),
    ("review_retrieval_miss_answer_ok", "Answer đúng nhưng retrieval top-1 sai — tách riêng để không bị chìm"),
    ("review_answer_mismatch", "Chỉ các dòng RAG sai đáp án (nếu có)"),
    ("", ""),
    ("Cột needs_*", "Y = cần team tương ứng xử lý"),
    ("review_owner_hint", "Gợi ý owner chính: SME / Dataset / RAG / None"),
    ("", ""),
    ("Luu y Excel", "Sheet review_* ~25 cot, zoom 90%. Compare: cot review o cot I+; freeze chi 4 cot dau."),
]


@dataclass
class CompanySpec:
    company_id: str
    company_name: str
    workbook_path: Path
    partition_csv: Path
    questions_jsonl: Path | None = None


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def _norm(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _as_bool(value: Any) -> bool | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return value
    if str(value).strip().lower() in {"1", "true", "y", "yes"}:
        return True
    if str(value).strip().lower() in {"0", "false", "n", "no"}:
        return False
    return None


def _business_key(
    company_id: str,
    row: dict[str, Any],
    partition: dict[str, Any] | None = None,
) -> tuple[str, ...]:
    source = partition or row
    return (
        company_id,
        _norm(source.get("year") or row.get("Year")),
        _norm(source.get("source_url") or row.get("Source URL")),
        _norm(source.get("file_url") or row.get("File URL")),
        _norm(source.get("gold_answer_raw") or row.get("Value")),
        _norm(source.get("disclosure_status") or row.get("Disclosure status")),
        _norm(source.get("doc_title") or row.get("Source document/page")),
        _norm(source.get("evidence_text") or row.get("Evidence")),
    )


def _load_questions_maps(path: Path) -> tuple[dict[int, str], dict[str, dict[str, Any]]]:
    by_row_index: dict[int, str] = {}
    by_question_id: dict[str, dict[str, Any]] = {}
    for row in _read_jsonl(path):
        qid = row["question_id"]
        by_question_id[qid] = row
        row_index = row.get("source_row_index")
        if row_index is not None:
            by_row_index[int(row_index)] = qid
    return by_row_index, by_question_id


def _load_results(path: Path) -> dict[str, dict[str, Any]]:
    return {row["question_id"]: row for row in _read_jsonl(path)}


def _load_partitions(path: Path) -> dict[str, dict[str, Any]]:
    return {row["question_id"]: row for row in _read_csv(path)}


def _format_top_docs(top_doc_titles: Any) -> tuple[str, str]:
    if not top_doc_titles:
        return "", ""
    docs = [str(d) for d in top_doc_titles if d]
    if not docs:
        return "", ""
    return docs[0], " | ".join(docs)


def _is_answerable(partition: dict[str, Any] | None, result: dict[str, Any] | None = None) -> bool:
    scoring = (partition or {}).get("scoring_rule") or (result or {}).get("scoring_rule") or ""
    return scoring != "abstain_expected"


def _comparison_status_and_note(
    partition: dict[str, Any] | None,
    result: dict[str, Any] | None,
) -> tuple[str, str]:
    if result is None:
        return "REVIEW_NEEDED", "Không tìm thấy kết quả RAG cho dòng này"

    tags = list(result.get("diagnostic_tags") or [])
    is_abstain = not _is_answerable(partition, result)

    if result.get("coverage_gap") or "coverage_gap" in tags:
        gap = result.get("coverage_gap") or "coverage_gap"
        if "ftc" in str(gap).lower():
            return "COVERAGE_GAP", "Sai do thiếu nguồn FTC raw HTML"
        return "COVERAGE_GAP", "Sai do thiếu raw source / coverage gap"

    if (
        result.get("semantic_ambiguity")
        or result.get("semantic_audit_note")
        or "semantic_ambiguity" in tags
        or result.get("fail_type") == "semantic_ambiguity"
    ):
        note = result.get("semantic_audit_note") or result.get("semantic_ambiguity")
        if note and "OFS" in str(note):
            return "SEMANTIC_AMBIGUITY", "Có ambiguity giữa label workbook và account OFS"
        if note and "FTC" in str(note):
            return "SEMANTIC_AMBIGUITY", "FTC blocked — giả định gold=0 khi không có evidence"
        return "SEMANTIC_AMBIGUITY", "Có semantic ambiguity cần SME audit"

    if is_abstain:
        if result.get("abstain_correct"):
            return "ABSTAIN_OK", "RAG abstain đúng theo gold"
        return "ANSWER_MISMATCH", "RAG không abstain đúng khi gold không có provenance"

    if result.get("answer_correct"):
        if not result.get("retrieval_hit_top1"):
            return "RETRIEVAL_MISS_BUT_ANSWER_OK", "Đáp án đúng nhưng top-1 retrieval sai tài liệu"
        return "MATCH", "RAG khớp đáp án gold"

    if not result.get("retrieval_hit_top1") and result.get("predicted_answer") in (None, "", "Not disclosed"):
        return "ANSWER_MISMATCH", "Retrieval miss và không trích xuất được đáp án"

    if result.get("fail_type") == "answer_fail" or not result.get("answer_correct"):
        return "ANSWER_MISMATCH", "RAG trả đáp án không khớp gold"

    return "REVIEW_NEEDED", "Không đủ thông tin để phân loại — cần review thủ công"


def _retrieval_review_status(partition: dict[str, Any] | None, result: dict[str, Any] | None) -> str:
    if not result or not _is_answerable(partition, result):
        return ""
    if result.get("retrieval_hit_top1"):
        return "TOP1_OK"
    if result.get("answer_correct"):
        return "TOP1_MISS_BUT_ANSWER_OK"
    return "TOP1_MISS_AND_ANSWER_WRONG"


def _top1_vs_gold_doc_match(
    merged: dict[str, Any],
    partition: dict[str, Any] | None,
    result: dict[str, Any] | None,
) -> str:
    if not _is_answerable(partition, result):
        return ""
    if result and result.get("source_match_top1"):
        return "MATCH"
    gold_doc = _norm((partition or {}).get("doc_title") or merged.get("Source document/page"))
    top1 = _norm(merged.get("rag_top1_doc"))
    if not gold_doc and not top1:
        return ""
    if gold_doc and top1 and (gold_doc == top1 or gold_doc in top1 or top1 in gold_doc):
        return "MATCH"
    return "MISMATCH"


def _source_gap_reason(result: dict[str, Any] | None, comparison_status: str) -> str:
    if comparison_status != "COVERAGE_GAP" and not (result or {}).get("coverage_gap"):
        tags = (result or {}).get("diagnostic_tags") or []
        if "coverage_gap" not in tags:
            return ""
    gap = str((result or {}).get("coverage_gap") or "")
    if "ftc" in gap.lower():
        return "FTC blocked"
    if gap:
        return gap
    return "missing raw source"


def _needs_dataset_check(
    partition: dict[str, Any] | None,
    result: dict[str, Any] | None,
    comparison_status: str,
    source_gap_reason: str,
) -> str:
    if comparison_status == "COVERAGE_GAP" or source_gap_reason:
        return "Y"
    tags = (result or {}).get("diagnostic_tags") or []
    if "coverage_gap" in tags:
        return "Y"
    disclosure = _norm((partition or {}).get("disclosure_status") or "")
    if disclosure == "matched" and not _norm((partition or {}).get("source_url")) and not _norm((partition or {}).get("file_url")):
        return "Y"
    return "N"


def _needs_sme_check(result: dict[str, Any] | None, comparison_status: str) -> str:
    if comparison_status == "SEMANTIC_AMBIGUITY":
        return "Y"
    if result and (result.get("semantic_ambiguity") or result.get("semantic_audit_note")):
        return "Y"
    tags = (result or {}).get("diagnostic_tags") or []
    return "Y" if "semantic_ambiguity" in tags else "N"


def _needs_rag_check(result: dict[str, Any] | None, partition: dict[str, Any] | None) -> str:
    if not result:
        return "Y"
    if not _is_answerable(partition, result):
        return "N" if result.get("abstain_correct") else "Y"
    if not result.get("answer_correct"):
        return "Y"
    if not result.get("retrieval_hit_top1"):
        return "Y"
    return "N"


def _review_priority(comparison_status: str, retrieval_review_status: str) -> str:
    if comparison_status in {"ANSWER_MISMATCH", "COVERAGE_GAP"}:
        return "HIGH"
    if comparison_status in {"SEMANTIC_AMBIGUITY", "RETRIEVAL_MISS_BUT_ANSWER_OK"}:
        return "MEDIUM"
    if retrieval_review_status == "TOP1_MISS_BUT_ANSWER_OK" and comparison_status in GREEN_STATUSES.union(
        {"SEMANTIC_AMBIGUITY", "COVERAGE_GAP"}
    ):
        return "MEDIUM"
    if comparison_status in GREEN_STATUSES:
        return "LOW"
    return "MEDIUM"


def _review_bucket(comparison_status: str, retrieval_review_status: str) -> str:
    if comparison_status in GREEN_STATUSES:
        if retrieval_review_status == "TOP1_MISS_BUT_ANSWER_OK":
            return "RAG_RETRIEVAL_NEEDS_REVIEW"
        return "RAG_OUTPUT_OK"
    if comparison_status == "COVERAGE_GAP":
        return "SOURCE_COVERAGE_GAP"
    if comparison_status == "SEMANTIC_AMBIGUITY":
        return "SME_SEMANTIC_CHECK"
    if comparison_status == "ANSWER_MISMATCH":
        return "ANSWER_ERROR"
    if comparison_status == "RETRIEVAL_MISS_BUT_ANSWER_OK":
        return "RAG_RETRIEVAL_NEEDS_REVIEW"
    return "RAG_OUTPUT_OK"


def _review_owner_hint(needs_sme: str, needs_dataset: str, needs_rag: str) -> str:
    if needs_sme == "Y":
        return "SME"
    if needs_dataset == "Y":
        return "Dataset"
    if needs_rag == "Y":
        return "RAG"
    return "None"


def _business_meaning_note(
    comparison_status: str,
    comparison_note: str,
    retrieval_review_status: str,
    source_gap_reason: str,
) -> str:
    if comparison_status == "MATCH" and retrieval_review_status == "TOP1_OK":
        return "Đáp án đúng, không cần review thêm"
    if comparison_status == "ABSTAIN_OK":
        return "Không có disclosure trong gold — RAG abstain đúng"
    if comparison_status == "COVERAGE_GAP":
        if source_gap_reason == "FTC blocked":
            return "Thiếu nguồn raw FTC nên chưa xác nhận retrieval đúng lane"
        return "Thiếu raw source — Dataset cần bổ sung nguồn trước khi kết luận retrieval"
    if comparison_status == "SEMANTIC_AMBIGUITY":
        if retrieval_review_status == "TOP1_MISS_BUT_ANSWER_OK":
            return "Đáp án đúng nhưng cần SME xác nhận ý nghĩa chỉ tiêu; đồng thời retrieval top-1 chưa khớp doc gold"
        return "Đáp án đúng nhưng cần SME xác nhận ý nghĩa chỉ tiêu"
    if comparison_status == "RETRIEVAL_MISS_BUT_ANSWER_OK":
        return "Đáp án đúng nhưng top-1 retrieval trỏ sai tài liệu — RAG cần xem routing/rerank"
    if comparison_status == "ANSWER_MISMATCH":
        return "RAG trả sai đáp án — cần kiểm tra extractor/retrieval"
    if retrieval_review_status == "TOP1_MISS_BUT_ANSWER_OK":
        return "Đáp án đúng nhưng top-1 retrieval chưa khớp doc gold"
    return comparison_note or "Cần review thủ công"


def _enrich_review_fields(
    merged: dict[str, Any],
    partition: dict[str, Any] | None,
    result: dict[str, Any] | None,
) -> None:
    comparison_status = merged.get("comparison_status") or "REVIEW_NEEDED"
    retrieval_status = _retrieval_review_status(partition, result)
    gap_reason = _source_gap_reason(result, comparison_status)
    needs_sme = _needs_sme_check(result, comparison_status)
    needs_dataset = _needs_dataset_check(partition, result, comparison_status, gap_reason)
    needs_rag = _needs_rag_check(result, partition)

    merged.update(
        {
            "retrieval_review_status": retrieval_status,
            "review_priority": _review_priority(comparison_status, retrieval_status),
            "review_bucket": _review_bucket(comparison_status, retrieval_status),
            "needs_sme_check": needs_sme,
            "needs_dataset_check": needs_dataset,
            "needs_rag_check": needs_rag,
            "review_owner_hint": _review_owner_hint(needs_sme, needs_dataset, needs_rag),
            "top1_vs_gold_doc_match": _top1_vs_gold_doc_match(merged, partition, result),
            "source_gap_reason": gap_reason,
            "business_meaning_note": _business_meaning_note(
                comparison_status,
                merged.get("comparison_note") or "",
                retrieval_status,
                gap_reason,
            ),
        }
    )


def _assign_question_ids(
    company_id: str,
    excel_rows: list[dict[str, Any]],
    questions_by_row: dict[int, str],
) -> list[tuple[dict[str, Any], str, str]]:
    assigned: list[tuple[dict[str, Any], str, str]] = []
    for ordinal, row in enumerate(excel_rows, start=1):
        ordinal_qid = f"{company_id}-{ordinal:04d}"
        row_index = int(row.get("_row_index") or 0)
        ingest_qid = questions_by_row.get(row_index)
        if ingest_qid and ingest_qid != ordinal_qid:
            question_id = ingest_qid
            match_method = "row_index_mismatch"
        elif ingest_qid:
            question_id = ingest_qid
            match_method = "source_row_index"
        else:
            question_id = ordinal_qid
            match_method = "ordinal_fallback"
        assigned.append((row, question_id, match_method))
    return assigned


def _build_company_rows(
    spec: CompanySpec,
    partitions: dict[str, dict[str, Any]],
    results: dict[str, dict[str, Any]],
    partition_by_key: dict[tuple[str, ...], str],
    result_by_key: dict[tuple[str, ...], str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    excel_rows = _load_rows(WorkbookSpec(spec.workbook_path, spec.company_id, spec.company_name))
    questions_by_row: dict[int, str] = {}
    if spec.questions_jsonl and spec.questions_jsonl.exists():
        questions_by_row, _ = _load_questions_maps(spec.questions_jsonl)

    assigned = _assign_question_ids(spec.company_id, excel_rows, questions_by_row)
    output_rows: list[dict[str, Any]] = []
    match_methods = Counter()
    missing_partition = 0
    missing_result = 0
    key_fallback = 0

    for excel_row, question_id, match_method in assigned:
        match_methods[match_method] += 1
        partition = partitions.get(question_id)
        result = results.get(question_id)

        if partition is None:
            key = _business_key(spec.company_id, excel_row)
            alt_qid = partition_by_key.get(key)
            if alt_qid:
                partition = partitions.get(alt_qid)
                if result is None:
                    result = results.get(alt_qid)
                question_id = alt_qid
                match_method = "business_key"
                key_fallback += 1
            else:
                missing_partition += 1

        if result is None and partition is not None:
            key = _business_key(spec.company_id, excel_row, partition)
            alt_qid = result_by_key.get(key)
            if alt_qid:
                result = results.get(alt_qid)
                question_id = alt_qid
                match_method = "business_key"
                key_fallback += 1
            else:
                missing_result += 1

        status, note = _comparison_status_and_note(partition, result)
        top1_doc, topk_docs = _format_top_docs((result or {}).get("top_doc_titles"))

        merged: dict[str, Any] = {"company_id": spec.company_id}
        for col in EXPECTED_HEADERS:
            merged[col] = excel_row.get(col)

        merged.update(
            {
                "question_id": question_id,
                "partition": (partition or {}).get("partition"),
                "question_family": (result or {}).get("question_family"),
                "gold_answer_canonical": (partition or {}).get("gold_answer_raw"),
                "rag_predicted_answer": (result or {}).get("predicted_answer"),
                "rag_predicted_abstain": (result or {}).get("predicted_abstain"),
                "answer_correct": (result or {}).get("answer_correct"),
                "abstain_correct": (result or {}).get("abstain_correct"),
                "retrieval_hit_top1": (result or {}).get("retrieval_hit_top1"),
                "source_match_top1": (result or {}).get("source_match_top1"),
                "top_score": (result or {}).get("top_score"),
                "rag_top1_doc": top1_doc,
                "rag_topk_docs": topk_docs,
                "predict_reason": (result or {}).get("predict_reason"),
                "fail_type": (result or {}).get("fail_type"),
                "diagnostic_tags": ", ".join((result or {}).get("diagnostic_tags") or []),
                "semantic_ambiguity": (result or {}).get("semantic_ambiguity"),
                "semantic_audit_note": (result or {}).get("semantic_audit_note"),
                "comparison_status": status,
                "comparison_note": note,
            }
        )
        _enrich_review_fields(merged, partition, result)
        output_rows.append(merged)

    meta = {
        "company_id": spec.company_id,
        "excel_rows": len(excel_rows),
        "output_rows": len(output_rows),
        "match_methods": dict(match_methods),
        "missing_partition": missing_partition,
        "missing_result": missing_result,
        "key_fallback": key_fallback,
        "mapped_pct": round(
            100.0 * sum(1 for r in output_rows if results.get(r["question_id"])) / max(1, len(output_rows)),
            2,
        ),
    }
    return output_rows, meta


def _sheet_headers(include_company_id: bool = False, *, layout: str = "pinned") -> list[str]:
    """pinned = cot review gan dau; compact = sheet review; legacy = tat ca cot goc + extension."""
    if layout == "legacy":
        prefix = ["company_id"] if include_company_id else []
        return prefix + EXPECTED_HEADERS + EXTENSION_COLUMNS

    if layout == "compact":
        prefix = ["company_id"] if include_company_id else []
        return prefix + CONTEXT_FRONT_HEADERS + REVIEW_PINNED_HEADERS + COMPACT_REVIEW_TAIL

    # pinned (compare sheets)
    prefix = ["company_id"] if include_company_id else []
    remaining_expected = [h for h in EXPECTED_HEADERS if h not in CONTEXT_FRONT_HEADERS]
    pinned_set = set(REVIEW_PINNED_HEADERS)
    remaining_ext = [h for h in EXTENSION_COLUMNS if h not in pinned_set]
    return prefix + CONTEXT_FRONT_HEADERS + REVIEW_PINNED_HEADERS + remaining_expected + remaining_ext


def _column_width(header: str) -> float:
    if header in {"영역", "카테고리", "서브카테고리", "항목", "Year", "단위", "Disclosure status"}:
        return 14
    if header in {"Value", "rag_predicted_answer", "gold_answer_canonical"}:
        return 12
    if header in {
        "comparison_status",
        "review_priority",
        "review_owner_hint",
        "review_bucket",
        "retrieval_review_status",
        "top1_vs_gold_doc_match",
        "question_id",
        "company_id",
        "partition",
        "question_family",
    }:
        return 16
    if header in {"comparison_note", "business_meaning_note", "semantic_audit_note"}:
        return 28
    if header in {"Source URL", "File URL", "rag_topk_docs"}:
        return 22
    if header in {"Evidence", "Source detail / calculation note", "기준 및 설명"}:
        return 24
    if header in {"Source document/page", "rag_top1_doc"}:
        return 20
    return 14


def _apply_sheet_view(ws, *, zoom: int = SHEET_ZOOM) -> None:
    ws.sheet_view.zoomScale = zoom
    ws.sheet_view.showGridLines = True


def _row_fill(row: dict[str, Any]) -> PatternFill:
    status = row.get("comparison_status") or "REVIEW_NEEDED"
    if status in GREEN_STATUSES:
        return STATUS_FILLS[status]
    return STATUS_FILLS.get(status, STATUS_FILLS["REVIEW_NEEDED"])


def _needs_review_border(row: dict[str, Any]) -> bool:
    status = row.get("comparison_status") or ""
    return status not in GREEN_STATUSES


def _set_column_widths(ws, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = _column_width(header)


def _freeze_pane_for_layout(layout: str) -> str:
    if layout == "compact":
        return FREEZE_PANES_COMPACT
    return FREEZE_PANES


def _write_data_sheet(
    ws,
    rows: list[dict[str, Any]],
    sheet_title: str,
    *,
    include_company_id: bool = False,
    emphasize_review: bool = False,
    layout: str = "pinned",
) -> Counter[str]:
    headers = _sheet_headers(include_company_id=include_company_id, layout=layout)
    ws.title = sheet_title[:31]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    status_counts: Counter[str] = Counter()

    for row_idx, row in enumerate(rows, start=2):
        status = row.get("comparison_status") or "REVIEW_NEEDED"
        status_counts[status] += 1
        fill = _row_fill(row)
        border = REVIEW_ROW_BORDER if emphasize_review and _needs_review_border(row) else None

        for col_idx, header in enumerate(headers, start=1):
            value = row.get(header)
            if header in ("rag_predicted_abstain", "answer_correct", "abstain_correct", "retrieval_hit_top1", "source_match_top1"):
                display = value if value not in (None, "") else ""
            else:
                display = value
            cell = ws.cell(row=row_idx, column=col_idx, value=display)
            cell.fill = fill
            if border:
                cell.border = border
            if header in WRAP_HEADERS:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if header in BOLD_HEADERS:
                cell.font = Font(bold=True)
            if header == "retrieval_review_status" and value == "TOP1_MISS_BUT_ANSWER_OK":
                cell.fill = RETRIEVAL_STATUS_FILL
                cell.font = Font(bold=True)

    ws.freeze_panes = _freeze_pane_for_layout(layout)
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
    _set_column_widths(ws, headers)
    _apply_sheet_view(ws, zoom=90 if layout == "compact" else SHEET_ZOOM)
    return status_counts


def _write_summary_sheet(
    ws,
    all_rows: list[dict[str, Any]],
    company_stats: dict[str, dict[str, Any]],
    eval_summary: dict[str, Any],
) -> None:
    ws.title = "summary"
    ws.append(["Goldns/Emni RAG vs Gold — Summary (review-friendly)"])
    ws.append([])

    safe_ignore = sum(1 for r in all_rows if r.get("comparison_status") in GREEN_STATUSES)
    needs_review = len(all_rows) - safe_ignore
    ws.append(["Rows safe to ignore (MATCH + ABSTAIN_OK)", safe_ignore])
    ws.append(["Rows needing review", needs_review])
    ws.append([])

    ws.append(["By company — comparison_status"])
    ws.append(["company_id", "total_rows", "MATCH", "ABSTAIN_OK", "SEMANTIC_AMBIGUITY", "COVERAGE_GAP", "ANSWER_MISMATCH", "RETRIEVAL_MISS_BUT_ANSWER_OK", "REVIEW_NEEDED"])
    for company_id, stats in company_stats.items():
        sc = stats.get("status_counts", {})
        ws.append(
            [
                company_id,
                stats.get("total_rows"),
                sc.get("MATCH", 0),
                sc.get("ABSTAIN_OK", 0),
                sc.get("SEMANTIC_AMBIGUITY", 0),
                sc.get("COVERAGE_GAP", 0),
                sc.get("ANSWER_MISMATCH", 0),
                sc.get("RETRIEVAL_MISS_BUT_ANSWER_OK", 0),
                sc.get("REVIEW_NEEDED", 0),
            ]
        )
    ws.append([])

    for section_title, key in (
        ("By company — review_bucket", "review_bucket"),
        ("By company — review_owner_hint", "review_owner_hint"),
        ("By company — review_priority", "review_priority"),
    ):
        ws.append([section_title])
        ws.append(["company_id", key, "count"])
        grouped: dict[tuple[str, str], int] = defaultdict(int)
        for row in all_rows:
            grouped[(row.get("company_id") or "", row.get(key) or "")] += 1
        for (company_id, bucket), count in sorted(grouped.items()):
            if bucket:
                ws.append([company_id, bucket, count])
        ws.append([])

    ws.append(["Retrieval review (answerable only)"])
    ws.append(["company_id", "retrieval_review_status", "count"])
    grouped_ret: dict[tuple[str, str], int] = defaultdict(int)
    for row in all_rows:
        rs = row.get("retrieval_review_status") or ""
        if rs:
            grouped_ret[(row.get("company_id") or "", rs)] += 1
    for (company_id, rs), count in sorted(grouped_ret.items()):
        ws.append([company_id, rs, count])
    ws.append([])

    ws.append(["Global metrics (frozen v5 eval)"])
    ws.append(["metric", "value"])
    for key in ("retrieval_hit_top1", "answer_accuracy", "abstain_accuracy", "overall_score"):
        ws.append([key, eval_summary.get(key)])
    ws.append([])

    ws.append(["Per-company eval metrics"])
    ws.append(["company_id", "retrieval_hit_top1", "answer_accuracy", "abstain_accuracy"])
    for company_id in sorted((eval_summary.get("by_company") or {})):
        item = eval_summary["by_company"][company_id]
        ws.append([company_id, item.get("retrieval_hit_top1"), item.get("answer_accuracy"), item.get("abstain_accuracy")])

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    _apply_sheet_view(ws, zoom=100)


def _write_review_guide_sheet(ws) -> None:
    ws.title = "review_guide"
    for line in REVIEW_GUIDE_LINES:
        ws.append(list(line))
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 58
    ws.column_dimensions["C"].width = 48
    _apply_sheet_view(ws, zoom=100)
    for row_idx in range(1, 4):
        for col_idx in range(1, 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")


def _filter_review_rows(all_rows: list[dict[str, Any]], predicate) -> list[dict[str, Any]]:
    return [row for row in all_rows if predicate(row)]


def export_workbook(
    companies: list[CompanySpec],
    results_path: Path,
    eval_summary_path: Path,
    output_path: Path,
) -> dict[str, Any]:
    results = _load_results(results_path)
    eval_summary = json.loads(eval_summary_path.read_text(encoding="utf-8"))

    all_partitions: dict[str, dict[str, Any]] = {}
    partition_by_key: dict[tuple[str, ...], str] = {}
    result_by_key: dict[tuple[str, ...], str] = {}

    for spec in companies:
        parts = _load_partitions(spec.partition_csv)
        all_partitions.update(parts)
        for qid, row in parts.items():
            partition_by_key[_business_key(spec.company_id, row, row)] = qid
    for qid, row in results.items():
        result_by_key[_business_key(row["company_id"], row, row)] = qid

    wb = Workbook()
    wb.remove(wb.active)

    company_stats: dict[str, dict[str, Any]] = {}
    rows_by_company: dict[str, list[dict[str, Any]]] = {}
    sheet_map = {"goldns": "goldns_compare", "emni": "emni_compare"}

    for spec in companies:
        rows, meta = _build_company_rows(spec, all_partitions, results, partition_by_key, result_by_key)
        rows_by_company[spec.company_id] = rows
        status_counts = Counter(r.get("comparison_status") or "REVIEW_NEEDED" for r in rows)
        answerable = sum(
            1 for r in rows if (all_partitions.get(r["question_id"]) or {}).get("scoring_rule") != "abstain_expected"
        )
        abstain = sum(
            1 for r in rows if (all_partitions.get(r["question_id"]) or {}).get("scoring_rule") == "abstain_expected"
        )
        bucket_counts = Counter(r.get("review_bucket") or "" for r in rows)
        company_stats[spec.company_id] = {
            **meta,
            "total_rows": len(rows),
            "answerable_rows": answerable,
            "abstain_rows": abstain,
            "status_counts": dict(status_counts),
            "review_bucket_counts": dict(bucket_counts),
        }
        ws = wb.create_sheet(sheet_map.get(spec.company_id, f"{spec.company_id}_compare"))
        _write_data_sheet(ws, rows, ws.title, include_company_id=False, emphasize_review=False, layout="pinned")

    all_rows = [row for spec in companies for row in rows_by_company[spec.company_id]]

    review_specs = [
        (
            "review_all_non_green",
            lambda r: r.get("comparison_status") not in GREEN_STATUSES,
        ),
        (
            "review_semantic_ambiguity",
            lambda r: r.get("comparison_status") == "SEMANTIC_AMBIGUITY",
        ),
        (
            "review_coverage_gap",
            lambda r: r.get("comparison_status") == "COVERAGE_GAP",
        ),
        (
            "review_retrieval_miss_answer_ok",
            lambda r: r.get("retrieval_review_status") == "TOP1_MISS_BUT_ANSWER_OK",
        ),
    ]
    review_sheet_counts: dict[str, int] = {}
    for sheet_name, predicate in review_specs:
        filtered = _filter_review_rows(all_rows, predicate)
        review_sheet_counts[sheet_name] = len(filtered)
        ws = wb.create_sheet(sheet_name)
        _write_data_sheet(
            ws, filtered, sheet_name, include_company_id=True, emphasize_review=True, layout="compact"
        )

    mismatch_rows = _filter_review_rows(all_rows, lambda r: r.get("comparison_status") == "ANSWER_MISMATCH")
    review_sheet_counts["review_answer_mismatch"] = len(mismatch_rows)
    if mismatch_rows:
        ws = wb.create_sheet("review_answer_mismatch")
        _write_data_sheet(
            ws, mismatch_rows, "review_answer_mismatch", include_company_id=True, emphasize_review=True, layout="compact"
        )

    summary_ws = wb.create_sheet("summary", 0)
    _write_summary_sheet(summary_ws, all_rows, company_stats, eval_summary)

    guide_ws = wb.create_sheet("review_guide", 1)
    _write_review_guide_sheet(guide_ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    global_review_buckets = Counter(r.get("review_bucket") or "" for r in all_rows)
    retrieval_miss_count = sum(1 for r in all_rows if r.get("retrieval_review_status") == "TOP1_MISS_BUT_ANSWER_OK")

    return {
        "output_path": str(output_path),
        "company_stats": company_stats,
        "review_sheet_counts": review_sheet_counts,
        "retrieval_miss_answer_ok_total": retrieval_miss_count,
        "review_bucket_totals": dict(global_review_buckets),
        "safe_to_ignore_rows": sum(1 for r in all_rows if r.get("comparison_status") in GREEN_STATUSES),
        "needs_review_rows": sum(1 for r in all_rows if r.get("comparison_status") not in GREEN_STATUSES),
        "eval_summary_metrics": {
            k: eval_summary.get(k)
            for k in ("retrieval_hit_top1", "answer_accuracy", "abstain_accuracy", "overall_score")
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Export RAG vs gold comparison workbook")
    parser.add_argument("--output", default="reports/goldns_emni_rag_vs_gold_comparison.xlsx")
    parser.add_argument("--results", default="reports/goldns_emni_rag_eval_20260618-100003/results.jsonl")
    parser.add_argument("--eval-summary", default="reports/goldns_emni_rag_eval_20260618-100003/summary.json")
    parser.add_argument("--eval-root", default="data/dataset_excel_eval_ready/20260617_goldns_emni")
    parser.add_argument("--intake-root", default="data/dataset_excel_intake/20260617_goldns_emni")
    parser.add_argument(
        "--goldns-workbook",
        default=r"C:\Users\nguye\Downloads\data-company\dataset-excel\골드앤에스_Final_ESG_Data.xlsx",
    )
    parser.add_argument(
        "--emni-workbook",
        default=r"C:\Users\nguye\Downloads\data-company\dataset-excel\이엠앤아이_Final_ESG_Data.xlsx",
    )
    args = parser.parse_args()

    eval_root = ROOT / args.eval_root
    intake_root = ROOT / args.intake_root
    companies = [
        CompanySpec(
            company_id="goldns",
            company_name="골드앤에스",
            workbook_path=Path(args.goldns_workbook),
            partition_csv=eval_root / "goldns" / "partition_all.csv",
            questions_jsonl=intake_root / "goldns" / "questions.jsonl",
        ),
        CompanySpec(
            company_id="emni",
            company_name="이엠앤아이",
            workbook_path=Path(args.emni_workbook),
            partition_csv=eval_root / "emni" / "partition_all.csv",
            questions_jsonl=intake_root / "emni" / "questions.jsonl",
        ),
    ]

    for spec in companies:
        if not spec.workbook_path.exists():
            raise SystemExit(f"Workbook not found: {spec.workbook_path}")
        if not spec.partition_csv.exists():
            raise SystemExit(f"Partition CSV not found: {spec.partition_csv}")

    report = export_workbook(
        companies,
        ROOT / args.results,
        ROOT / args.eval_summary,
        ROOT / args.output,
    )

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
