from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE_CSV = ROOT / "data" / "golden_set" / "golden_answer_fill_preliminary_ko_20260609.csv"
OUT_CSV = ROOT / "data" / "golden_set" / "working_set_v1_ko_20260609.csv"
OUT_XLSX = ROOT / "data" / "golden_set" / "working_set_v1_ko_20260609.xlsx"
OUT_REPORT = ROOT / "reports" / "working-set-v1-20-cau-thuc-dung-20260609.md"


def load_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def map_working_status(fill_status: str) -> str:
    if fill_status == "filled_from_dataset":
        return "grounded"
    if fill_status == "partial_from_dataset":
        return "partial"
    if fill_status == "not_found_in_current_dataset":
        return "needs_review"
    return "exclude"


def select_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    selected: list[dict[str, str]] = []
    for row in rows:
        if row["company"] != "한샘":
            continue
        working_status = map_working_status(row["fill_status"])
        if working_status == "exclude":
            continue

        selected.append(
            {
                "working_version": "v1",
                "working_status": working_status,
                "company": row["company"],
                "package_name": row["package_name"],
                "question_id": row["question_id"],
                "question_type": row["question_type"],
                "area": row["area"],
                "category": row["category"],
                "subcategory": row["subcategory"],
                "question_text_ko": row["question_text_ko"],
                "answer_type": row["answer_type"],
                "unit": row["unit"],
                "gold_answer_ko": row["gold_answer_ko"],
                "evidence_record_id": row["evidence_record_id"],
                "evidence_excerpt_ko": row["evidence_excerpt_ko"],
                "forbidden_rule": row["forbidden_rule"],
                "source_fill_status": row["fill_status"],
                "review_priority": "high" if working_status in {"grounded", "partial"} else "medium",
                "review_action": (
                    "use_for_eval_now"
                    if working_status == "grounded"
                    else "review_before_scoring"
                ),
                "answer_note": row["answer_note"],
            }
        )
    return selected


def write_csv(rows: list[dict[str, str]]) -> None:
    fieldnames = list(rows[0].keys())
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Info"
    info_rows = [
        ("working_version", "v1"),
        ("scope", "Hansem only, 20-question working set"),
        ("selection_rule", "Keep Hansem rows with filled_from_dataset / partial_from_dataset / not_found_in_current_dataset"),
        ("status_meaning_grounded", "Có thể dùng chấm ngay"),
        ("status_meaning_partial", "Có thể dùng tạm nhưng nên reviewer xác nhận"),
        ("status_meaning_needs_review", "Giữ lại để đủ phạm vi làm việc, nhưng cần điền/soát thêm trước khi coi là gold"),
    ]
    for idx, (k, v) in enumerate(info_rows, start=1):
        ws_info.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws_info.cell(row=idx, column=2, value=v)
    ws_info.column_dimensions["A"].width = 28
    ws_info.column_dimensions["B"].width = 100

    ws = wb.create_sheet("WorkingSetV1")
    headers = list(rows[0].keys())
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[header])

    widths = {
        1: 14, 2: 16, 3: 12, 4: 36, 5: 12, 6: 14, 7: 10, 8: 18, 9: 24, 10: 42,
        11: 14, 12: 10, 13: 60, 14: 24, 15: 60, 16: 30, 17: 22, 18: 16, 19: 22, 20: 28,
    }
    for idx, width in widths.items():
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_report(rows: list[dict[str, str]]) -> None:
    counts: dict[str, int] = {}
    lines = []
    for row in rows:
        counts[row["working_status"]] = counts.get(row["working_status"], 0) + 1
        lines.append(
            f"| {row['question_id']} | {row['question_type']} | {row['area']} | {row['working_status']} | {row['review_action']} |"
        )

    report = f"""# Working set v1 - 20 cau thuc dung - 2026-06-09

## Muc tieu

Tao mot `working set` lon hon `golden_set_v1` de workbook cham pipeline dung duoc thuc te hon, nhung van ghi ro muc do tin cay cua tung cau.

## Nguyen tac chon

- Khong dung `dataset_issue` cua `레이시온` va `무신사`.
- Lay toan bo `20` cau shortlist cua `한샘`.
- Mapping trang thai:
  - `filled_from_dataset` -> `grounded`
  - `partial_from_dataset` -> `partial`
  - `not_found_in_current_dataset` -> `needs_review`

## Ket qua

- Tong so cau: {len(rows)}
- Cong ty: `한샘`
- Phan bo trang thai: {", ".join(f"{k} ({v})" for k, v in sorted(counts.items()))}

## Danh sach

| Question ID | Type | Area | Working Status | Review Action |
|---|---|---|---|---|
{chr(10).join(lines)}

## Cach dung

- `grounded`: co the dung de cham ngay.
- `partial`: co the dua vao eval sheet, nhung reviewer nen xac nhan truoc khi ket luan.
- `needs_review`: giu lai de du 20 cau cho vong lam viec; chua nen coi la gold answer hoan chinh.
"""
    OUT_REPORT.write_text(report, encoding="utf-8")


def main() -> None:
    rows = select_rows(load_rows())
    write_csv(rows)
    write_xlsx(rows)
    write_report(rows)
    print(f"Wrote {len(rows)} rows to {OUT_CSV}")


if __name__ == "__main__":
    main()
