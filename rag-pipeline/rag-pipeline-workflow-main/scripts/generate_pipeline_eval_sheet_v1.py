from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
GOLD_CSV = ROOT / "data" / "golden_set" / "golden_set_v1_ko_20260609.csv"
OUT_XLSX = ROOT / "data" / "golden_set" / "pipeline_eval_sheet_v1_20260609.xlsx"
OUT_CSV = ROOT / "data" / "golden_set" / "pipeline_eval_input_v1_20260609.csv"
OUT_REPORT = ROOT / "reports" / "pipeline-eval-sheet-v1-20260609.md"


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
AUTO_FILL = PatternFill("solid", fgColor="E2F0D9")


def load_gold_rows() -> list[dict[str, str]]:
    with GOLD_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_input_csv(rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "run_id",
        "model_name",
        "question_id",
        "company",
        "question_text_ko",
        "pipeline_answer_ko",
        "pipeline_evidence_record_id",
        "pipeline_evidence_excerpt_ko",
        "latency_ms",
        "pipeline_notes",
    ]
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "run_id": "",
                    "model_name": "",
                    "question_id": row["question_id"],
                    "company": row["company"],
                    "question_text_ko": row["question_text_ko"],
                    "pipeline_answer_ko": "",
                    "pipeline_evidence_record_id": "",
                    "pipeline_evidence_excerpt_ko": "",
                    "latency_ms": "",
                    "pipeline_notes": "",
                }
            )


def style_headers(ws, headers: list[str], fill: PatternFill = HEADER_FILL) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill


def set_widths(ws, widths: dict[int, int]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[chr(64 + idx)].width = width


def write_info_sheet(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "Info"
    info_rows = [
        ("file_name", "pipeline_eval_sheet_v1_20260609.xlsx"),
        ("gold_source", "golden_set_v1_ko_20260609.csv"),
        ("question_count", str(len(rows))),
        ("usage_step_1", "Mở sheet PipelineInput và dán output pipeline vào các cột màu vàng."),
        ("usage_step_2", "Giữ nguyên question_id; không đổi thứ tự nếu muốn xem trực tiếp theo hàng."),
        ("usage_step_3", "Xem sheet Score để đối chiếu gold_answer với pipeline_answer."),
        ("usage_step_4", "Nếu auto check chưa đủ, điền manual_answer_score / manual_evidence_score / reviewer_note."),
        ("note_language", "Question / gold answer / pipeline answer giữ Korean-only để tránh lệch nghĩa."),
    ]
    for idx, (k, v) in enumerate(info_rows, start=1):
        ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=v)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def write_gold_sheet(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("GoldSet")
    headers = list(rows[0].keys())
    style_headers(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[header])
    set_widths(
        ws,
        {
            1: 12, 2: 12, 3: 36, 4: 12, 5: 14, 6: 10, 7: 18, 8: 24, 9: 42, 10: 14,
            11: 10, 12: 60, 13: 24, 14: 60, 15: 30, 16: 14, 17: 22, 18: 18, 19: 22, 20: 28,
        },
    )
    ws.freeze_panes = "A2"


def write_pipeline_input_sheet(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("PipelineInput")
    headers = [
        "run_id",
        "model_name",
        "question_id",
        "company",
        "question_text_ko",
        "pipeline_answer_ko",
        "pipeline_evidence_record_id",
        "pipeline_evidence_excerpt_ko",
        "latency_ms",
        "pipeline_notes",
    ]
    style_headers(ws, headers)
    input_cols = {1, 2, 6, 7, 8, 9, 10}

    for row_idx, row in enumerate(rows, start=2):
        values = [
            "",
            "",
            row["question_id"],
            row["company"],
            row["question_text_ko"],
            "",
            "",
            "",
            "",
            "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in input_cols:
                cell.fill = INPUT_FILL

    set_widths(
        ws,
        {1: 18, 2: 24, 3: 12, 4: 12, 5: 42, 6: 60, 7: 24, 8: 60, 9: 12, 10: 28},
    )
    ws.freeze_panes = "A2"


def write_score_sheet(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("Score")
    headers = [
        "question_id",
        "company",
        "question_text_ko",
        "gold_answer_ko",
        "pipeline_answer_ko",
        "gold_evidence_record_id",
        "pipeline_evidence_record_id",
        "auto_answer_exact",
        "auto_evidence_record_match",
        "manual_answer_score",
        "manual_evidence_score",
        "manual_forbidden_violation",
        "final_score",
        "reviewer_note",
    ]
    style_headers(ws, headers)

    for row_idx, row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=f"=GoldSet!D{row_idx}")
        ws.cell(row=row_idx, column=2, value=f"=GoldSet!B{row_idx}")
        ws.cell(row=row_idx, column=3, value=f"=GoldSet!I{row_idx}")
        ws.cell(row=row_idx, column=4, value=f"=GoldSet!L{row_idx}")
        ws.cell(row=row_idx, column=5, value=f"=PipelineInput!F{row_idx}")
        ws.cell(row=row_idx, column=6, value=f"=GoldSet!M{row_idx}")
        ws.cell(row=row_idx, column=7, value=f"=PipelineInput!G{row_idx}")
        ws.cell(row=row_idx, column=8, value=f'=IF(E{row_idx}="","",IF(EXACT(TRIM(E{row_idx}),TRIM(D{row_idx})),1,0))')
        ws.cell(row=row_idx, column=9, value=f'=IF(G{row_idx}="","",IF(EXACT(TRIM(G{row_idx}),TRIM(F{row_idx})),1,0))')
        ws.cell(row=row_idx, column=10, value="")
        ws.cell(row=row_idx, column=11, value="")
        ws.cell(row=row_idx, column=12, value="")
        ws.cell(row=row_idx, column=13, value=(
            f'=IF(OR(J{row_idx}<>"",K{row_idx}<>"",L{row_idx}<>""),'
            f'IF(L{row_idx}=1,0,ROUND(((IF(J{row_idx}="",0,J{row_idx})+IF(K{row_idx}="",0,K{row_idx}))/2),4)),'
            f'IF(OR(H{row_idx}="",I{row_idx}=""),"",ROUND((H{row_idx}+I{row_idx})/2,4)))'
        ))
        ws.cell(row=row_idx, column=14, value="")

    for col_idx in [8, 9, 13]:
        for row_idx in range(2, len(rows) + 2):
            ws.cell(row=row_idx, column=col_idx).fill = AUTO_FILL

    for col_idx in [10, 11, 12, 14]:
        for row_idx in range(2, len(rows) + 2):
            ws.cell(row=row_idx, column=col_idx).fill = INPUT_FILL

    set_widths(
        ws,
        {1: 12, 2: 12, 3: 42, 4: 60, 5: 60, 6: 24, 7: 24, 8: 16, 9: 20, 10: 18, 11: 20, 12: 22, 13: 14, 14: 32},
    )
    ws.freeze_panes = "A2"


def write_summary_sheet(wb: Workbook, row_count: int) -> None:
    ws = wb.create_sheet("Summary")
    info = [
        ("total_questions", f"=COUNTA(Score!A2:A{row_count + 1})"),
        ("answered_by_pipeline", f'=COUNTIF(Score!E2:E{row_count + 1},"<>")'),
        ("auto_answer_exact_count", f'=COUNTIF(Score!H2:H{row_count + 1},1)'),
        ("auto_evidence_match_count", f'=COUNTIF(Score!I2:I{row_count + 1},1)'),
        ("final_score_avg", f'=IFERROR(AVERAGE(Score!M2:M{row_count + 1}),"")'),
        ("manual_forbidden_violation_count", f'=COUNTIF(Score!L2:L{row_count + 1},1)'),
    ]
    for idx, (k, v) in enumerate(info, start=1):
        ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=v)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18


def write_report(rows: list[dict[str, str]]) -> None:
    report = f"""# Pipeline eval sheet v1 - 2026-06-09

## Muc tieu

Tao workbook de team co the do output pipeline vao va doi chieu truc tiep voi `gold_answer` cua `golden_set_v1`.

## Artifact

- `data/golden_set/pipeline_eval_sheet_v1_20260609.xlsx`
- `data/golden_set/pipeline_eval_input_v1_20260609.csv`

## Cau truc workbook

- `Info`: huong dan cach dung nhanh.
- `GoldSet`: ban sao `golden_set_v1` de doi chieu.
- `PipelineInput`: sheet de dan output pipeline vao cac cot mau vang.
- `Score`: sheet doi chieu theo tung `question_id`, co auto check exact match va evidence record match.
- `Summary`: tong hop so cau da tra loi, exact match va diem trung binh.

## Ghi chu

- Auto check trong `Score` chi la check ky thuat muc co ban (`EXACT` text va `evidence_record_id`).
- Cac cau tra loi dien dat khac nhung van dung can reviewer dien `manual_answer_score` / `manual_evidence_score`.
- `manual_forbidden_violation = 1` se keo `final_score` ve `0`.
- Workbook nay phu hop de bat dau phase cham pipeline voi tap `golden_set_v1` hien tai ({len(rows)} cau).
"""
    OUT_REPORT.write_text(report, encoding="utf-8")


def main() -> None:
    rows = load_gold_rows()
    write_input_csv(rows)

    wb = Workbook()
    write_info_sheet(wb, rows)
    write_gold_sheet(wb, rows)
    write_pipeline_input_sheet(wb, rows)
    write_score_sheet(wb, rows)
    write_summary_sheet(wb, len(rows))
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    write_report(rows)
    print(f"Wrote workbook to {OUT_XLSX}")


if __name__ == "__main__":
    main()
