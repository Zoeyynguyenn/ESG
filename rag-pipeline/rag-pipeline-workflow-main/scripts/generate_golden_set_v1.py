from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
SOURCE_CSV = ROOT / "data" / "golden_set" / "golden_answer_fill_preliminary_ko_20260609.csv"
OUT_CSV = ROOT / "data" / "golden_set" / "golden_set_v1_ko_20260609.csv"
OUT_XLSX = ROOT / "data" / "golden_set" / "golden_set_v1_ko_20260609.xlsx"
OUT_REPORT = ROOT / "reports" / "golden-set-v1-loc-theo-dataset-thuc-te-20260609.md"


def load_rows() -> list[dict[str, str]]:
    with SOURCE_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def filter_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    kept: list[dict[str, str]] = []
    for row in rows:
        if row["fill_status"] != "filled_from_dataset":
            continue
        kept.append(
            {
                "golden_version": "v1",
                "company": row["company"],
                "package_name": row["package_name"],
                "question_id": row["question_id"],
                "question_type": row["question_type"],
                "area": row["area"],
                "category": row["category"],
                "subcategory": row["subcategory"],
                "question_text_ko": row["question_text_ko"],
                "answer_type": row["answer_type"],
                "unit": row["unit"],
                "gold_answer_ko": row["gold_answer_ko"],
                "evidence_record_id": row["evidence_record_id"],
                "evidence_excerpt_ko": row["evidence_excerpt_ko"],
                "forbidden_rule": row["forbidden_rule"],
                "confidence_level": "high",
                "inclusion_rule": "filled_from_dataset_only",
                "exclusion_note": "",
                "review_status": "ready_for_pipeline_eval",
                "answer_note": row["answer_note"],
            }
        )
    return kept


def write_csv(rows: list[dict[str, str]]) -> None:
    fieldnames = list(rows[0].keys()) if rows else [
        "golden_version",
        "company",
        "package_name",
        "question_id",
        "question_type",
        "area",
        "category",
        "subcategory",
        "question_text_ko",
        "answer_type",
        "unit",
        "gold_answer_ko",
        "evidence_record_id",
        "evidence_excerpt_ko",
        "forbidden_rule",
        "confidence_level",
        "inclusion_rule",
        "exclusion_note",
        "review_status",
        "answer_note",
    ]
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "Info"
    info_rows = [
        ("golden_version", "v1"),
        ("selection_rule", "Chỉ giữ các dòng có fill_status = filled_from_dataset"),
        ("excluded_status", "partial_from_dataset, not_found_in_current_dataset, dataset_issue"),
        ("scope_note", "Tập v1 phản ánh đúng phần dataset hiện có thể chứng minh chắc chắn, chưa đại diện đủ 3 công ty"),
        ("ready_for", "Pipeline accuracy evaluation với câu hỏi định tính Korean-only"),
    ]
    for idx, (k, v) in enumerate(info_rows, start=1):
        ws_info.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws_info.cell(row=idx, column=2, value=v)

    ws = wb.create_sheet("GoldenSetV1")
    headers = list(rows[0].keys()) if rows else []
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[header])

    widths = {
        1: 12, 2: 12, 3: 36, 4: 12, 5: 14, 6: 10, 7: 18, 8: 24, 9: 42, 10: 14,
        11: 10, 12: 60, 13: 24, 14: 60, 15: 28, 16: 14, 17: 22, 18: 18, 19: 22, 20: 28,
    }
    for idx, width in widths.items():
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = "A2"
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


def write_report(rows: list[dict[str, str]]) -> None:
    by_company: dict[str, int] = {}
    by_area: dict[str, int] = {}
    lines = []
    for row in rows:
        by_company[row["company"]] = by_company.get(row["company"], 0) + 1
        by_area[row["area"]] = by_area.get(row["area"], 0) + 1
        lines.append(
            f"| {row['company']} | {row['question_id']} | {row['area']} | {row['category']} | {row['question_text_ko']} | {row['evidence_record_id']} |"
        )

    report = f"""# Golden set V1 loc theo dataset thuc te - 2026-06-09

## Muc tieu

Chot mot `golden set v1` nho, chi gom cac cau hoi co the doi chieu chac chan voi dataset `05_company_export_json` hien tai.

## Tieu chi giu lai

- Chi giu cac dong co `fill_status = filled_from_dataset`.
- Loai bo toan bo `partial_from_dataset`, `not_found_in_current_dataset`, `dataset_issue`.
- Khong co dich thuat; `question`, `gold_answer`, `evidence_excerpt` giu Korean-only.
- Moi dong phai tro duoc ve `company_evidence` bang `evidence_record_id`.

## Ket qua loc

- Tong so dong trong `golden_set_v1`: {len(rows)}
- Pham vi cong ty: {", ".join(f"{k} ({v})" for k, v in sorted(by_company.items()))}
- Phan bo theo vung: {", ".join(f"{k} ({v})" for k, v in sorted(by_area.items()))}
- Tat ca cau hoi trong v1 hien la `qualitative`.

## Danh sach cau giu lai

| Company | Question ID | Area | Category | Question (KO) | Evidence Record |
|---|---|---|---|---|---|
{chr(10).join(lines)}

## Ghi chu van hanh

- Tap v1 nay dung de chay thu pipeline accuracy tren cac cau hoi dinh tinh co bang chung ro.
- Khong nen dung tap nay de ket luan nang luc dinh luong, do hien chua co cau dinh luong nao duoc chung minh chac chan.
- Khong nen xem tap nay la representative cho du 3 cong ty; hien tai no phan anh phan dataset co chat luong dung duoc ngay.
"""
    OUT_REPORT.write_text(report, encoding="utf-8")


def main() -> None:
    rows = filter_rows(load_rows())
    write_csv(rows)
    write_xlsx(rows)
    write_report(rows)
    print(f"Wrote {len(rows)} rows to {OUT_CSV}")


if __name__ == "__main__":
    main()
