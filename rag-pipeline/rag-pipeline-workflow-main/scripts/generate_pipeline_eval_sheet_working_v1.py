from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKING_CSV = ROOT / "data" / "golden_set" / "working_set_v1_ko_20260609.csv"
OUT_XLSX = ROOT / "data" / "golden_set" / "pipeline_eval_sheet_working_v1_20260609.xlsx"
OUT_CSV = ROOT / "data" / "golden_set" / "pipeline_eval_input_working_v1_20260609.csv"
OUT_REPORT = ROOT / "reports" / "pipeline-eval-sheet-working-v1-20260609.md"

HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
AUTO_FILL = PatternFill("solid", fgColor="E2F0D9")


def load_rows() -> list[dict[str, str]]:
    with WORKING_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_input_csv(rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "run_id",
        "model_name",
        "question_id",
        "company",
        "question_text_ko",
        "pipeline_answer_ko",
        "pipeline_evidence_record_id",
        "pipeline_evidence_excerpt_ko",
        "latency_ms",
        "pipeline_notes",
    ]
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "run_id": "",
                    "model_name": "",
                    "question_id": row["question_id"],
                    "company": row["company"],
                    "question_text_ko": row["question_text_ko"],
                    "pipeline_answer_ko": "",
                    "pipeline_evidence_record_id": "",
                    "pipeline_evidence_excerpt_ko": "",
                    "latency_ms": "",
                    "pipeline_notes": "",
                }
            )


def style_headers(ws, headers: list[str], fill: PatternFill = HEADER_FILL) -> None:
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = fill


def set_widths(ws, widths: dict[int, int]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[chr(64 + idx)].width = width


def write_info(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.active
    ws.title = "Info"
    info_rows = [
        ("file_name", "pipeline_eval_sheet_working_v1_20260609.xlsx"),
        ("working_source", "working_set_v1_ko_20260609.csv"),
        ("question_count", str(len(rows))),
        ("note", "Workbook này dùng cho working set 20 câu, không chỉ cho gold sạch."),
        ("status_rule", "Xem cột working_status trong sheet WorkingSet."),
        ("review_rule", "Câu grounded dùng chấm ngay; partial và needs_review cần reviewer chú ý hơn."),
    ]
    for idx, (k, v) in enumerate(info_rows, start=1):
        ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=v)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100


def write_working_set(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("WorkingSet")
    headers = list(rows[0].keys())
    style_headers(ws, headers)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[header])
    set_widths(
        ws,
        {1: 14, 2: 16, 3: 12, 4: 36, 5: 12, 6: 14, 7: 10, 8: 18, 9: 24, 10: 42, 11: 14, 12: 10, 13: 60, 14: 24, 15: 60, 16: 30, 17: 22, 18: 16, 19: 22, 20: 28},
    )
    ws.freeze_panes = "A2"


def write_pipeline_input(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("PipelineInput")
    headers = [
        "run_id", "model_name", "question_id", "company", "question_text_ko",
        "pipeline_answer_ko", "pipeline_evidence_record_id", "pipeline_evidence_excerpt_ko",
        "latency_ms", "pipeline_notes",
    ]
    style_headers(ws, headers)
    input_cols = {1, 2, 6, 7, 8, 9, 10}
    for row_idx, row in enumerate(rows, start=2):
        values = ["", "", row["question_id"], row["company"], row["question_text_ko"], "", "", "", "", ""]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx in input_cols:
                cell.fill = INPUT_FILL
    set_widths(ws, {1: 18, 2: 24, 3: 12, 4: 12, 5: 42, 6: 60, 7: 24, 8: 60, 9: 12, 10: 28})
    ws.freeze_panes = "A2"


def write_score(wb: Workbook, rows: list[dict[str, str]]) -> None:
    ws = wb.create_sheet("Score")
    headers = [
        "question_id",
        "working_status",
        "review_action",
        "question_text_ko",
        "gold_answer_ko",
        "pipeline_answer_ko",
        "gold_evidence_record_id",
        "pipeline_evidence_record_id",
        "auto_answer_exact",
        "auto_evidence_record_match",
        "manual_answer_score",
        "manual_evidence_score",
        "manual_forbidden_violation",
        "final_score",
        "reviewer_note",
    ]
    style_headers(ws, headers)
    for row_idx, _row in enumerate(rows, start=2):
        ws.cell(row=row_idx, column=1, value=f"=WorkingSet!E{row_idx}")
        ws.cell(row=row_idx, column=2, value=f"=WorkingSet!B{row_idx}")
        ws.cell(row=row_idx, column=3, value=f"=WorkingSet!S{row_idx}")
        ws.cell(row=row_idx, column=4, value=f"=WorkingSet!J{row_idx}")
        ws.cell(row=row_idx, column=5, value=f"=WorkingSet!M{row_idx}")
        ws.cell(row=row_idx, column=6, value=f"=PipelineInput!F{row_idx}")
        ws.cell(row=row_idx, column=7, value=f"=WorkingSet!N{row_idx}")
        ws.cell(row=row_idx, column=8, value=f"=PipelineInput!G{row_idx}")
        ws.cell(row=row_idx, column=9, value=f'=IF(F{row_idx}="","",IF(EXACT(TRIM(F{row_idx}),TRIM(E{row_idx})),1,0))')
        ws.cell(row=row_idx, column=10, value=f'=IF(H{row_idx}="","",IF(EXACT(TRIM(H{row_idx}),TRIM(G{row_idx})),1,0))')
        ws.cell(row=row_idx, column=11, value="")
        ws.cell(row=row_idx, column=12, value="")
        ws.cell(row=row_idx, column=13, value="")
        ws.cell(row=row_idx, column=14, value=(
            f'=IF(OR(K{row_idx}<>"",L{row_idx}<>"",M{row_idx}<>""),'
            f'IF(M{row_idx}=1,0,ROUND(((IF(K{row_idx}="",0,K{row_idx})+IF(L{row_idx}="",0,L{row_idx}))/2),4)),'
            f'IF(OR(I{row_idx}="",J{row_idx}=""),"",ROUND((I{row_idx}+J{row_idx})/2,4)))'
        ))
        ws.cell(row=row_idx, column=15, value="")

    for col_idx in [9, 10, 14]:
        for row_idx in range(2, len(rows) + 2):
            ws.cell(row=row_idx, column=col_idx).fill = AUTO_FILL
    for col_idx in [11, 12, 13, 15]:
        for row_idx in range(2, len(rows) + 2):
            ws.cell(row=row_idx, column=col_idx).fill = INPUT_FILL

    set_widths(ws, {1: 12, 2: 16, 3: 22, 4: 42, 5: 60, 6: 60, 7: 24, 8: 24, 9: 16, 10: 20, 11: 18, 12: 20, 13: 22, 14: 14, 15: 32})
    ws.freeze_panes = "A2"


def write_summary(wb: Workbook, row_count: int) -> None:
    ws = wb.create_sheet("Summary")
    info = [
        ("total_questions", f"=COUNTA(Score!A2:A{row_count + 1})"),
        ("grounded_count", f'=COUNTIF(Score!B2:B{row_count + 1},"grounded")'),
        ("partial_count", f'=COUNTIF(Score!B2:B{row_count + 1},"partial")'),
        ("needs_review_count", f'=COUNTIF(Score!B2:B{row_count + 1},"needs_review")'),
        ("answered_by_pipeline", f'=COUNTIF(Score!F2:F{row_count + 1},"<>")'),
        ("final_score_avg", f'=IFERROR(AVERAGE(Score!N2:N{row_count + 1}),"")'),
    ]
    for idx, (k, v) in enumerate(info, start=1):
        ws.cell(row=idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=v)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def write_report(rows: list[dict[str, str]]) -> None:
    report = f"""# Pipeline eval sheet working v1 - 2026-06-09

## Muc tieu

Mo rong workbook cham tu `golden_set_v1` sang `working_set_v1` de co it nhat 20 cau phuc vu vong lam viec thuc te hon.

## Artifact

- `data/golden_set/pipeline_eval_sheet_working_v1_20260609.xlsx`
- `data/golden_set/pipeline_eval_input_working_v1_20260609.csv`

## Cau truc

- `Info`
- `WorkingSet`
- `PipelineInput`
- `Score`
- `Summary`

## Khac biet so voi ban gold-only

- `Score` co them `working_status` va `review_action`.
- Khong phai moi dong deu la gold chac chan.
- Reviewer can uu tien tin tuong:
  - `grounded`
  - sau do `partial`
  - cuoi cung `needs_review`

## Quy mo

- Tong so cau trong working set: {len(rows)}
"""
    OUT_REPORT.write_text(report, encoding="utf-8")


def main() -> None:
    rows = load_rows()
    write_input_csv(rows)
    wb = Workbook()
    write_info(wb, rows)
    write_working_set(wb, rows)
    write_pipeline_input(wb, rows)
    write_score(wb, rows)
    write_summary(wb, len(rows))
    wb.save(OUT_XLSX)
    write_report(rows)
    print(f"Wrote workbook to {OUT_XLSX}")


if __name__ == "__main__":
    main()
