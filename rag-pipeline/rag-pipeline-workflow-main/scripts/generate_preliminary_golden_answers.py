from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATASET_ROOT = ROOT / "data" / "rag_dataset" / "05_company_export_json"
QUESTION_CSV = ROOT / "data" / "golden_set" / "pilot_question_bank_shortlist_ko_20260609.csv"
OUT_CSV = ROOT / "data" / "golden_set" / "golden_answer_fill_preliminary_ko_20260609.csv"
OUT_XLSX = ROOT / "data" / "golden_set" / "golden_answer_fill_preliminary_ko_20260609.xlsx"


def load_questions() -> list[dict[str, str]]:
    with QUESTION_CSV.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def detect_company_dirs() -> list[Path]:
    return sorted([p for p in DATASET_ROOT.iterdir() if p.is_dir()])


def load_manifest(pkg: Path) -> dict:
    return json.loads((pkg / "manifest.json").read_text(encoding="utf-8"))


def build_hanssem_answers() -> dict[str, dict[str, str]]:
    record_2022 = "rec_66100907c00656ec"
    record_2025 = "rec_86c98b945fc03e6d"
    return {
        "QL-001": {
            "gold_answer_ko": (
                "한샘은 ESG 경영 성과와 전략 과제를 지속가능경영보고서로 공개하고 있으며, "
                "환경·사회·거버넌스 전 영역에서 지속 가능한 성장 체계를 구축한 "
                "‘가장 존경받는 환경기업’을 지향한다고 밝혔다."
            ),
            "evidence_record_id": record_2022,
            "evidence_excerpt_ko": (
                "한샘은 지난 2020년부터 3년 연속으로 지속가능경영보고서를 발간하면서 ESG 경영 활동을 이어가고 있다. "
                "… ‘가장 존경받는 환경기업’이 되겠다는 비전을 가지고 있다."
            ),
            "fill_status": "filled_from_dataset",
            "answer_note": "지속가능경영보고서 발간 배경 + 비전 문구 기반 요약",
        },
        "QL-003": {
            "gold_answer_ko": (
                "한샘은 이중 중대성 평가를 통해 기후변화 완화, 책임있는 조달, 지속 가능한 제품 설계, "
                "사업장 근무 조건, 인권경영, 협력사 동반성장, 제품 안전 및 품질, 공정거래 등 "
                "8개 중대 이슈를 선정하고 이를 중심으로 관리한다고 설명했다."
            ),
            "evidence_record_id": record_2025,
            "evidence_excerpt_ko": (
                "이중 중대성 평가(Double Materiality Assessment)… 총 8개 중대 이슈를 선정했다."
            ),
            "fill_status": "filled_from_dataset",
            "answer_note": "ESG 리스크 식별/우선순위화 방식 확인 가능",
        },
        "QL-004": {
            "gold_answer_ko": (
                "한샘은 사회 분야에서 인적자원 개발, 고용 평등, 안전보건 강화, 동반성장에 집중하고 있으며, "
                "하반기 내 안전보건경영시스템(ISO45001) 인증을 추진하는 등 안전 관리 수준을 높일 계획이라고 밝혔다."
            ),
            "evidence_record_id": record_2022,
            "evidence_excerpt_ko": (
                "사회 분야에서는 ▲인적자원 개발 ▲고용 평등 ▲안전보건 강화 ▲동반성장에 집중한다. "
                "… ISO45001 인증을 받는 등 안전 관리 수준을 한층 더 높일 방침이다."
            ),
            "fill_status": "filled_from_dataset",
            "answer_note": "안전보건 목표/방향은 확인되나 세부 정책 문서 자체는 아님",
        },
        "QL-008": {
            "gold_answer_ko": (
                "한샘은 2024년 최초로 인권실태조사를 실시해 임직원의 인권 감수성을 진단했고, "
                "이를 바탕으로 인권영향평가를 수행해 인권경영 체계를 강화했다고 설명했다."
            ),
            "evidence_record_id": record_2025,
            "evidence_excerpt_ko": (
                "2024년에는 최초로 인권실태조사를 실시해 임직원의 인권 감수성을 진단하고, "
                "이를 바탕으로 인권영향평가를 수행해 인권경영 체계를 강화했다."
            ),
            "fill_status": "filled_from_dataset",
            "answer_note": "노동·인권 관리 체계에 대한 직접 진술 확보",
        },
        "QL-011": {
            "gold_answer_ko": (
                "한샘은 2025년까지 환경경영 시스템을 완성하고 온실가스 배출 관리에 집중하겠다고 밝혔다. "
                "또한 탄소배출 관리 범위를 Scope 3까지 확대하고 CDP에도 가입해 자발적으로 평가를 받고 있다고 설명했다."
            ),
            "evidence_record_id": record_2022,
            "evidence_excerpt_ko": (
                "환경 분야에서는 오는 2025년까지 환경경영 시스템을 완성한다. "
                "특히 온실가스 배출 관리에 집중한다. … ‘스코프(scope) 3’으로 확대한 바 있다."
            ),
            "fill_status": "filled_from_dataset",
            "answer_note": "환경 정책/목표에 대한 직접 진술 확보",
        },
        "QL-012": {
            "gold_answer_ko": (
                "한샘은 이중 중대성 평가를 통해 기후변화 완화, 책임있는 조달, 지속 가능한 제품 설계 등 "
                "환경 관련 이슈를 중대 이슈로 선정하고, 온실가스 배출 관리를 Scope 3까지 확대하며 "
                "CDP 가입을 통해 관리 수준을 높이고 있다고 설명했다."
            ),
            "evidence_record_id": record_2025,
            "evidence_excerpt_ko": (
                "이중 중대성 평가… 기후변화 완화… 총 8개 중대 이슈를 선정했다. "
                "또한 정부의 2050 탄소중립 목표에 적극 동참하고 있다."
            ),
            "fill_status": "filled_from_dataset",
            "answer_note": "환경 리스크 관리 답변으로 사용 가능하나 일부는 환경정책 서술과 결합됨",
        },
        "QL-013": {
            "gold_answer_ko": (
                "한샘은 2021년 이사회 중심 경영 체제를 구축하고 이사회 산하 ESG 위원회를 설치했으며, "
                "감사·ESG·보상 등 6개 소위원회를 운영해 지배구조의 투명성과 책임성을 강화했다고 밝혔다."
            ),
            "evidence_record_id": record_2022,
            "evidence_excerpt_ko": (
                "지난 2021년 이사회 중심 경영 체제를 구축하고, 이사회 산하 ESG 위원회를 설치… "
                "이사회 산하에 감사와 ESG, 보상 등 6개 소위원회를 설치"
            ),
            "fill_status": "partial_from_dataset",
            "answer_note": "윤리경영 전담 조직 그 자체보다 ESG/이사회 거버넌스 중심 근거",
        },
        "QL-014": {
            "gold_answer_ko": (
                "한샘은 준법윤리지수 평가를 자체 시행하고 윤리 프로그램 운영 강화, 정보공개 확대, "
                "협력사 준법 점검과 ESG 평가 항목 반영 등을 통해 준법·윤리 리스크를 관리하겠다고 설명했다."
            ),
            "evidence_record_id": record_2022,
            "evidence_excerpt_ko": (
                "‘준법윤리지수 평가’ 자체 시행… 협력사 준법 점검… ESG 평가 항목을 넣어 공급망의 지속가능성을 강화"
            ),
            "fill_status": "partial_from_dataset",
            "answer_note": "준법 리스크 관리 방향은 보이나 통제 절차 세부도는 낮음",
        },
    }


def package_default_status(pkg_name: str) -> tuple[str, str]:
    if pkg_name.startswith("레이시온_"):
        return (
            "dataset_issue",
            "패키지 내 주요 sustainability_report 텍스트가 레이시온 고유 ESG 본문이 아니라 공공기관/타사 페이지와 혼재되어 있어 gold answer로 사용하기 어려움",
        )
    if pkg_name.startswith("무신사_"):
        return (
            "dataset_issue",
            "sustainability_report 레코드가 실제 본문보다 리포트 게시/목록 페이지 성격이 강해 질문별 gold answer 근거로 쓰기 어려움",
        )
    return (
        "not_found_in_current_dataset",
        "현재 package의 company_evidence에서 질문에 직접 대응하는 근거를 아직 확인하지 못함",
    )


def build_rows() -> list[dict[str, str]]:
    questions = load_questions()
    hanssem_answers = build_hanssem_answers()
    rows: list[dict[str, str]] = []

    for pkg in detect_company_dirs():
        manifest = load_manifest(pkg)
        company = manifest.get("dataset_name", pkg.name)
        default_status, default_note = package_default_status(pkg.name)
        for q in questions:
            row = {
                "package_name": pkg.name,
                "company": company,
                "question_id": q["question_id"],
                "question_type": q["질문유형"],
                "area": q["영역"],
                "category": q["카테고리"],
                "subcategory": q["세부항목"],
                "question_text_ko": q["질문"],
                "answer_type": q["답변형식"],
                "unit": q["단위"],
                "forbidden_rule": q["금지규칙"],
                "gold_answer_ko": "",
                "evidence_record_id": "",
                "evidence_excerpt_ko": "",
                "fill_status": default_status,
                "answer_note": default_note,
            }

            if company == "한샘" and q["question_id"] in hanssem_answers:
                row.update(hanssem_answers[q["question_id"]])
            rows.append(row)
    return rows


def write_csv(rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "package_name",
        "company",
        "question_id",
        "question_type",
        "area",
        "category",
        "subcategory",
        "question_text_ko",
        "answer_type",
        "unit",
        "forbidden_rule",
        "gold_answer_ko",
        "evidence_record_id",
        "evidence_excerpt_ko",
        "fill_status",
        "answer_note",
    ]
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "PreliminaryFill"
    headers = [
        "package_name",
        "company",
        "question_id",
        "question_type",
        "area",
        "category",
        "subcategory",
        "question_text_ko",
        "answer_type",
        "unit",
        "forbidden_rule",
        "gold_answer_ko",
        "evidence_record_id",
        "evidence_excerpt_ko",
        "fill_status",
        "answer_note",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[header])

    widths = {
        1: 42,
        2: 14,
        3: 12,
        4: 14,
        5: 10,
        6: 18,
        7: 22,
        8: 38,
        9: 14,
        10: 10,
        11: 42,
        12: 48,
        13: 24,
        14: 54,
        15: 24,
        16: 54,
    }
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def main() -> None:
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    print(OUT_CSV)
    print(OUT_XLSX)


if __name__ == "__main__":
    main()
