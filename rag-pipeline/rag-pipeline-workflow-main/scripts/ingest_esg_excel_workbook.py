#!/usr/bin/env python3
"""Convert ESG result workbook(s) into canonical questions/gold_answers/sources artifacts."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


EXPECTED_HEADERS = [
    "영역",
    "카테고리",
    "서브카테고리",
    "항목",
    "기준 및 설명",
    "단위",
    "단위_설명",
    "GRI",
    "GRI_설명",
    "SASB",
    "KBIZ",
    "K-ESG",
    "Value",
    "Year",
    "Answer unit",
    "Disclosure status",
    "Reporting boundary",
    "Source document/page",
    "Source URL",
    "Link check",
    "Source detail / calculation note",
    "Confidence",
    "File URL",
    "Evidence",
]

CONTEXT_FIELDS = EXPECTED_HEADERS[:12]


@dataclass
class WorkbookSpec:
    path: Path
    company_id: str
    company_name: str


def _slugify_metric(parts: Iterable[str]) -> str:
    text = "_".join(p.strip().lower() for p in parts if p and str(p).strip())
    text = re.sub(r"[^0-9a-zA-Z가-힣_]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text or "metric"


def _sanitize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_answer(value: Any, status: str) -> str | int | float | None:
    if status == "Not disclosed":
        return "NOT_DISCLOSED"
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        if "." in text:
            return float(text)
        return int(text)
    except ValueError:
        return text


def _question_type(unit: str | None, value: Any) -> str:
    numeric_units = {"명", "%", "건", "원", "시간", "회", "톤", "tCO2e", "MWh", "kWh", "m3"}
    if unit and unit in numeric_units:
        return "quantitative"
    if isinstance(value, (int, float)):
        return "quantitative"
    text = _sanitize_text(value)
    if text in {"Y", "N", "Yes", "No", "예", "아니오"}:
        return "yes_no"
    return "qualitative"


def _build_question_text(record: dict[str, Any]) -> str:
    year = record.get("Year")
    prefix = f"{year}년 " if year else ""
    path_parts = [
        record.get("영역"),
        record.get("카테고리"),
        record.get("서브카테고리"),
        record.get("항목"),
    ]
    path_text = " / ".join(p for p in path_parts if p)
    unit = record.get("단위") or record.get("Answer unit")
    if unit and unit != "None":
        return f"{prefix}{path_text} 값은 무엇인가? (단위: {unit})"
    return f"{prefix}{path_text} 값은 무엇인가?"


def _load_rows(spec: WorkbookSpec) -> list[dict[str, Any]]:
    wb = load_workbook(spec.path, read_only=True, data_only=True)
    ws = wb["ESG Results"]
    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    if header != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected header in {spec.path.name}")

    rows: list[dict[str, Any]] = []
    context: dict[str, Any] = {field: None for field in CONTEXT_FIELDS}
    for idx, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = dict(zip(EXPECTED_HEADERS, raw))
        if all(record.get(k) is None for k in EXPECTED_HEADERS):
            continue
        for field in CONTEXT_FIELDS:
            if record.get(field) is None:
                record[field] = context[field]
            else:
                context[field] = record[field]
        record["_row_index"] = idx
        rows.append(record)
    return rows


def _build_artifacts(spec: WorkbookSpec, rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    questions: list[dict[str, Any]] = []
    gold_answers: list[dict[str, Any]] = []
    sources: list[dict[str, Any]] = []
    source_registry: dict[str, dict[str, Any]] = {}

    for ordinal, row in enumerate(rows, start=1):
        metric_name = _slugify_metric(
            [row.get("영역"), row.get("카테고리"), row.get("서브카테고리"), row.get("항목"), str(row.get("Year") or "")]
        )
        question_id = f"{spec.company_id}-{ordinal:04d}"
        disclosure_status = _sanitize_text(row.get("Disclosure status")) or "UNKNOWN"
        question_type = _question_type(_sanitize_text(row.get("단위")), row.get("Value"))

        questions.append(
            {
                "question_id": question_id,
                "company_id": spec.company_id,
                "company_name": spec.company_name,
                "question_text": _build_question_text(row),
                "question_type": question_type,
                "metric_name": metric_name,
                "expected_answer_language": "ko",
                "requires_abstain_when_missing": disclosure_status == "Not disclosed",
                "dimension": _sanitize_text(row.get("영역")),
                "category": _sanitize_text(row.get("카테고리")),
                "subcategory": _sanitize_text(row.get("서브카테고리")),
                "item": _sanitize_text(row.get("항목")),
                "year": row.get("Year"),
                "unit": _sanitize_text(row.get("단위")),
                "source_row_index": row["_row_index"],
            }
        )

        gold_answers.append(
            {
                "question_id": question_id,
                "company_id": spec.company_id,
                "gold_answer_raw": row.get("Value"),
                "gold_answer_normalized": _normalize_answer(row.get("Value"), disclosure_status),
                "answer_unit": _sanitize_text(row.get("Answer unit")) or _sanitize_text(row.get("단위")),
                "disclosure_status": disclosure_status,
                "expert_verified": True,
                "normalization_rule": "not_disclosed_to_canonical_token" if disclosure_status == "Not disclosed" else "identity_numeric_if_possible",
                "scoring_rule": "abstain_expected" if disclosure_status == "Not disclosed" else "value_match_with_unit",
                "reporting_boundary": _sanitize_text(row.get("Reporting boundary")),
                "confidence": _sanitize_text(row.get("Confidence")),
                "evidence_text": _sanitize_text(row.get("Evidence")),
            }
        )

        source_url = _sanitize_text(row.get("Source URL"))
        file_url = _sanitize_text(row.get("File URL"))
        sources.append(
            {
                "question_id": question_id,
                "company_id": spec.company_id,
                "source_url": source_url,
                "doc_title": _sanitize_text(row.get("Source document/page")),
                "source_type": "dataset_excel_reference",
                "page_or_section": _sanitize_text(row.get("Source document/page")),
                "source_priority": 1 if disclosure_status == "matched" else None,
                "crawl_allowed": bool(source_url),
                "link_check": _sanitize_text(row.get("Link check")),
                "source_detail_note": _sanitize_text(row.get("Source detail / calculation note")),
                "file_url": file_url,
                "evidence_text": _sanitize_text(row.get("Evidence")),
                "confidence": _sanitize_text(row.get("Confidence")),
            }
        )

        source_key = source_url or file_url
        if source_key and source_key not in source_registry:
            source_registry[source_key] = {
                "company_id": spec.company_id,
                "company_name": spec.company_name,
                "source_url": source_url,
                "file_url": file_url,
                "doc_title": _sanitize_text(row.get("Source document/page")),
                "source_type": "dataset_excel_reference",
                "link_check": _sanitize_text(row.get("Link check")),
                "confidence": _sanitize_text(row.get("Confidence")),
            }

    return questions, gold_answers, sources, list(source_registry.values())


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    fieldnames = list(rows[0].keys())
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _write_summary(path: Path, all_summaries: list[dict[str, Any]]) -> None:
    lines = [
        "# ESG Excel Intake Summary",
        "",
        "| Company | Rows | Matched | Not disclosed | Unique sources | Crawlable URLs |",
        "|---|---:|---:|---:|---:|---:|",
    ]
    for item in all_summaries:
        lines.append(
            f"| {item['company_name']} | {item['rows']} | {item['matched']} | {item['not_disclosed']} | "
            f"{item['unique_sources']} | {item['crawlable_urls']} |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_manifest(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-root", default="data/dataset_excel_intake/20260617_goldns_emni")
    parser.add_argument(
        "--input",
        action="append",
        nargs=3,
        metavar=("XLSX_PATH", "COMPANY_ID", "COMPANY_NAME"),
        required=True,
        help="One workbook spec: path, company_id, company_name",
    )
    args = parser.parse_args()

    output_root = Path(args.output_root)
    summaries: list[dict[str, Any]] = []

    for xlsx_path, company_id, company_name in args.input:
        spec = WorkbookSpec(path=Path(xlsx_path), company_id=company_id, company_name=company_name)
        rows = _load_rows(spec)
        questions, gold_answers, sources, source_registry = _build_artifacts(spec, rows)

        company_dir = output_root / company_id
        _write_jsonl(company_dir / "questions.jsonl", questions)
        _write_jsonl(company_dir / "gold_answers.jsonl", gold_answers)
        _write_jsonl(company_dir / "sources.jsonl", sources)
        _write_jsonl(company_dir / "source_registry.jsonl", source_registry)

        _write_csv(company_dir / "questions.csv", questions)
        _write_csv(company_dir / "gold_answers.csv", gold_answers)
        _write_csv(company_dir / "sources.csv", sources)
        _write_csv(company_dir / "source_registry.csv", source_registry)

        status_counter = Counter(item["disclosure_status"] for item in gold_answers)
        manifest = {
            "company_id": company_id,
            "company_name": company_name,
            "input_workbook": str(spec.path),
            "row_count": len(rows),
            "question_count": len(questions),
            "gold_answer_count": len(gold_answers),
            "source_row_count": len(sources),
            "source_registry_count": len(source_registry),
            "matched_count": status_counter.get("matched", 0),
            "not_disclosed_count": status_counter.get("Not disclosed", 0),
        }
        _write_manifest(company_dir / "manifest.json", manifest)
        summaries.append(
            {
                "company_id": company_id,
                "company_name": company_name,
                "rows": len(rows),
                "matched": status_counter.get("matched", 0),
                "not_disclosed": status_counter.get("Not disclosed", 0),
                "unique_sources": len(source_registry),
                "crawlable_urls": sum(1 for item in source_registry if item.get("source_url")),
            }
        )

    _write_summary(output_root / "README.md", summaries)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
