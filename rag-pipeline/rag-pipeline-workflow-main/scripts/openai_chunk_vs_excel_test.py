#!/usr/bin/env python3
"""
RAG test với OpenAI THẬT (chạy trên máy có OPENAI_API_KEY + mạng OpenAI).
Chunk văn bản gốc .md -> embed text-embedding-3-small -> retrieve top-k (cosine, in-memory)
-> gpt-4o-mini sinh câu trả lời bám evidence -> so sánh với gold Excel.

KHÔNG cần Qdrant (corpus nhỏ ~50 chunk/công ty -> cosine in-memory là đủ).

Cách chạy (Windows / Git Bash):
  # đảm bảo OPENAI_API_KEY có trong env hoặc trong .env của app gemma4
  cd D:/esg-te/ai-workflows/rag-pipeline/rag-pipeline-workflow-main
  python scripts/openai_chunk_vs_excel_test.py --env "D:/ESG/Testing data/ai-gemma4-main/.env" --per-company 100

Output: reports/openai_rag_chunk_vs_excel_test_20260619.xlsx
"""
import os, re, glob, json, argparse, random, math, time
import openpyxl
from openpyxl.styles import Font, PatternFill

random.seed(7)
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.dirname(ROOT)
EMBED_MODEL = "text-embedding-3-small"
GEN_MODEL = "gpt-4o-mini"
TOP_K = 5
NOT_DISCLOSED = ("not disclosed", "not publicly disclosed", "none", "—", "-", "")

COMPANIES = {
    "GOLDNS": dict(datafiles=os.path.join(BASE, "goldns", "datafiles"),
                   excel=os.path.join(BASE, "goldns", "GoldNS_quantitative_questions.xlsx"),
                   sheet="GoldNS_Quantitative_QA"),
    "EMNI": dict(datafiles=os.path.join(BASE, "emni", "datafiles"),
                 excel=os.path.join(BASE, "emni", "reports", "EMNI_quantitative_questions_updated.xlsx"),
                 sheet="EMNI_Quantitative_QA"),
}


def load_env(path):
    if path and os.path.exists(path):
        for line in open(path, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def chunk_company(folder):
    chunks = []
    for fp in sorted(glob.glob(os.path.join(folder, "*.md"))):
        fn = os.path.basename(fp)
        if fn.lower() == "readme.md":
            continue
        title = fn
        buf, body_title = [], ""
        text = open(fp, encoding="utf-8").read()
        def flush():
            if buf:
                body = "\n".join(buf).strip()
                if body:
                    chunks.append(dict(doc=fn, section=f"{fn} > {body_title}",
                                       text=f"[{fn} | {body_title}]\n{body}"))
        for line in text.splitlines():
            if line.startswith("#"):
                flush(); buf = []; body_title = line.lstrip("# ").strip()
            else:
                buf.append(line)
        flush()
    return chunks


def norm_num(s):
    return re.sub(r"[,\s]", "", str(s)).lower()


def is_answerable(val, status):
    v = str(val).strip().lower()
    if v in NOT_DISCLOSED or v.startswith("not "):
        return False
    return any(c.isdigit() for c in v)


def load_questions(xlsx, sheet):
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet]; rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0]); idx = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        out.append(dict(id=r[idx["Record ID"]], area=r[idx["Area"]], cat=r[idx["Category"]],
                        sub=r[idx["Subcategory"]], item=r[idx.get("Item", 4)],
                        question=r[idx["Question / requested metric"]],
                        gold_2025=r[idx["2025 value"]], status=r[idx["Disclosure status"]],
                        evidence_section=r[idx["Evidence page/section"]]))
    return out


def embed_batch(client, texts):
    out = []
    for i in range(0, len(texts), 64):
        r = client.embeddings.create(model=EMBED_MODEL, input=texts[i:i+64])
        out += [d.embedding for d in r.data]
    return out


def cosine(a, b):
    s = sum(x*y for x, y in zip(a, b))
    na = math.sqrt(sum(x*x for x in a)); nb = math.sqrt(sum(y*y for y in b))
    return s/(na*nb+1e-9)


GEN_SYS = (
    "Bạn là trợ lý trích xuất dữ liệu ESG. Chỉ dùng EVIDENCE được cung cấp. "
    "Nếu evidence không có thông tin, trả lời đúng chuỗi 'NOT_DISCLOSED'. "
    "Nếu có, trả lời JSON: {\"value\": <giá trị/số>, \"unit\": <đơn vị hoặc null>}."
)


def generate(client, question, evidence):
    msg = [{"role": "system", "content": GEN_SYS},
           {"role": "user", "content": f"CÂU HỎI:\n{question}\n\nEVIDENCE:\n{evidence}"}]
    r = client.chat.completions.create(model=GEN_MODEL, messages=msg, temperature=0, max_tokens=120)
    return r.choices[0].message.content.strip()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--env", default=os.path.join(BASE, "..", "ESG", "Testing data", "ai-gemma4-main", ".env"))
    ap.add_argument("--per-company", type=int, default=100)
    args = ap.parse_args()
    load_env(args.env)
    from openai import OpenAI
    client = OpenAI()

    out_path = os.path.join(ROOT, "reports", "openai_rag_chunk_vs_excel_test_20260619.xlsx")
    sample, corpora, results = [], {}, []

    for comp, cfg in COMPANIES.items():
        chunks = chunk_company(cfg["datafiles"])
        vecs = embed_batch(client, [c["text"] for c in chunks])
        corpora[comp] = (chunks, vecs)
        qs = load_questions(cfg["excel"], cfg["sheet"])
        ans = [q for q in qs if is_answerable(q["gold_2025"], q["status"])]
        nd = [q for q in qs if not is_answerable(q["gold_2025"], q["status"])]
        random.shuffle(nd)
        for q in ans + nd[:max(0, args.per_company - len(ans))]:
            q["company"] = comp; sample.append(q)

    for n, q in enumerate(sample, 1):
        chunks, vecs = corpora[q["company"]]
        query = " ".join(str(x) for x in [q["question"], q["cat"], q["sub"], q["item"]] if x)
        qv = embed_batch(client, [query])[0]
        ranked = sorted(range(len(chunks)), key=lambda i: cosine(qv, vecs[i]), reverse=True)[:TOP_K]
        evidence = "\n---\n".join(chunks[i]["text"] for i in ranked)
        top_section = chunks[ranked[0]]["section"]
        raw = generate(client, query, evidence)
        answerable = is_answerable(q["gold_2025"], q["status"])
        abstained = "NOT_DISCLOSED" in raw.upper()
        pred_val = ""
        if not abstained:
            m = re.search(r'"value"\s*:\s*"?([^",}]+)', raw)
            pred_val = m.group(1).strip() if m else raw
        if answerable:
            ok = (not abstained) and norm_num(q["gold_2025"]) in norm_num(pred_val + " " + evidence)
            verdict = "HIT" if ok else "MISS"
        else:
            verdict = "ABSTAIN-OK" if abstained else "OVER-ANSWER"
        results.append({**q, "verdict": verdict, "rag_raw": raw[:300],
                        "pred_val": pred_val, "top_section": top_section})
        if n % 20 == 0:
            print(f"  {n}/{len(sample)} ...")

    ans = [r for r in results if is_answerable(r["gold_2025"], r["status"])]
    nd = [r for r in results if not is_answerable(r["gold_2025"], r["status"])]
    hit = sum(r["verdict"] == "HIT" for r in ans)
    abst = sum(r["verdict"] == "ABSTAIN-OK" for r in nd)
    print(f"\nTổng {len(results)} câu | answer_match {hit}/{len(ans)}={hit/max(len(ans),1):.1%} | "
          f"abstain {abst}/{len(nd)}={abst/max(len(nd),1):.1%} | "
          f"overall {(hit+abst)}/{len(results)}={(hit+abst)/len(results):.1%}")

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "summary"
    for row in [["OpenAI RAG chunk-vs-Excel test", "2026-06-19"],
                ["Stack", f"{EMBED_MODEL} + cosine in-memory top-{TOP_K} + {GEN_MODEL}"],
                [], ["Chỉ số", "Giá trị"],
                ["Tổng câu", len(results)],
                ["answer_match (answerable)", f"{hit}/{len(ans)} = {hit/max(len(ans),1):.1%}"],
                ["abstain đúng (not-disclosed)", f"{abst}/{len(nd)} = {abst/max(len(nd),1):.1%}"],
                ["overall", f"{hit+abst}/{len(results)} = {(hit+abst)/len(results):.1%}"]]:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    d = wb.create_sheet("detail")
    cols = ["company", "id", "cat", "sub", "question", "gold_2025", "status",
            "verdict", "pred_val", "rag_raw", "top_section"]
    d.append(cols)
    g = PatternFill("solid", fgColor="C6EFCE"); rd = PatternFill("solid", fgColor="FFC7CE")
    bl = PatternFill("solid", fgColor="DDEBF7")
    for r in results:
        d.append([r.get(k, "") for k in cols])
        c = d.cell(d.max_row, cols.index("verdict") + 1)
        c.fill = g if r["verdict"] == "HIT" else rd if r["verdict"] in ("MISS", "OVER-ANSWER") else bl
    d.column_dimensions["E"].width = 40; d.column_dimensions["J"].width = 50
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print("Saved:", out_path)


if __name__ == "__main__":
    main()
