"""Unified ESG answer resolution — merge dataset RAG + internal-doc structured output."""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterator

from unified_esg_resolution_policy import (
    DEFAULT_POLICY,
    map_dataset_family,
    normalize_metric_name,
    normalize_year,
    resolve_pair,
)

ROOT = Path(__file__).resolve().parents[1]
SCHEMA_PATH = ROOT / "data/unified_esg/unified_answer_schema.json"

DEFAULT_DATASET_RESULTS = ROOT / "reports/goldns_emni_rag_eval_20260618-100003/results.jsonl"
DEFAULT_EVAL_ROOT = ROOT / "data/dataset_excel_eval_ready/20260617_goldns_emni"
DEFAULT_INTERNAL_RECORDS = (
    ROOT / "reports/enterprise_docs_structured_esg_hardening_20260619-090700/structured_esg_records.jsonl"
)


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    rows: list[dict[str, Any]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.strip():
            rows.append(json.loads(line))
    return rows


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for row in rows:
            f.write(json.dumps(row, ensure_ascii=False) + "\n")


def load_partition_enrichment(eval_root: Path) -> dict[str, dict[str, Any]]:
    """Load metric_name/year from partition CSVs keyed by question_id."""
    out: dict[str, dict[str, Any]] = {}
    for company_dir in eval_root.iterdir():
        if not company_dir.is_dir():
            continue
        for csv_name in ("answerable_gold.csv", "abstain_gold.csv"):
            csv_path = company_dir / csv_name
            if not csv_path.exists():
                continue
            import csv

            with csv_path.open(encoding="utf-8-sig", newline="") as f:
                for row in csv.DictReader(f):
                    qid = row.get("question_id")
                    if qid:
                        out[qid] = row
    return out


def business_key(
    *,
    company_id: str,
    family_id: str,
    metric_name: str,
    year: Any,
) -> str:
    return "::".join(
        [
            str(company_id or "").strip().lower(),
            str(family_id or "unknown").strip().lower(),
            normalize_metric_name(metric_name),
            normalize_year(year),
        ]
    )


def identity_key(
    row: dict[str, Any],
    *,
    source: str,
    partition_enrichment: dict[str, dict[str, Any]] | None = None,
) -> tuple[str | None, str]:
    """Return (question_id or None, business_key)."""
    qid = row.get("question_id")
    company_id = str(row.get("company_id") or "")

    if source == "dataset":
        enrich = (partition_enrichment or {}).get(qid or "") or {}
        family_id = map_dataset_family(row.get("question_family"))
        metric = enrich.get("metric_name") or row.get("metric_name") or _metric_from_question(row.get("question_text"))
        year = enrich.get("year") or _year_from_question(row.get("question_text"))
    else:
        family_id = str(row.get("family_id") or "unknown")
        metric = row.get("metric_name") or ""
        year = row.get("year")

    bk = business_key(company_id=company_id, family_id=family_id, metric_name=str(metric), year=year)
    return (str(qid) if qid else None, bk)


def _year_from_question(text: str | None) -> str:
    if not text:
        return ""
    m = re.search(r"(20\d{2})", str(text))
    return m.group(1) if m else ""


def _metric_from_question(text: str | None) -> str:
    if not text:
        return ""
    # e.g. "2025년 사회 / 구성원 현황 / 총 구성원 수"
    parts = [p.strip() for p in str(text).split("/") if p.strip()]
    if parts:
        return parts[-1].split("(")[0].strip()
    return ""


def index_dataset_rows(
    rows: list[dict[str, Any]],
    *,
    partition_enrichment: dict[str, dict[str, Any]] | None = None,
) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    by_qid: dict[str, dict[str, Any]] = {}
    by_bk: dict[str, dict[str, Any]] = {}
    for row in rows:
        qid, bk = identity_key(row, source="dataset", partition_enrichment=partition_enrichment)
        if qid:
            by_qid[qid] = row
        by_bk[bk] = row
    return by_qid, by_bk


def index_internal_rows(rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    by_qid: dict[str, dict[str, Any]] = {}
    by_bk: dict[str, dict[str, Any]] = {}
    for row in rows:
        qid, bk = identity_key(row, source="internal_doc")
        if qid:
            by_qid[qid] = row
        by_bk[bk] = row
    return by_qid, by_bk


def _merge_keys(
    ds_by_qid: dict[str, dict[str, Any]],
    ds_by_bk: dict[str, dict[str, Any]],
    in_by_qid: dict[str, dict[str, Any]],
    in_by_bk: dict[str, dict[str, Any]],
) -> list[tuple[str, str | None, str]]:
    """Union of join keys: (join_type, question_id, business_key)."""
    seen: set[str] = set()
    keys: list[tuple[str, str | None, str]] = []

    for qid in set(ds_by_qid) | set(in_by_qid):
        token = f"qid::{qid}"
        if token in seen:
            continue
        seen.add(token)
        ds = ds_by_qid.get(qid)
        in_row = in_by_qid.get(qid)
        _, bk = identity_key(ds or in_row or {}, source="dataset" if ds else "internal_doc")
        keys.append(("question_id", qid, bk))

    for bk in set(ds_by_bk) | set(in_by_bk):
        token = f"bk::{bk}"
        if token in seen:
            continue
        # skip if already covered by qid join
        ds = ds_by_bk.get(bk)
        in_row = in_by_bk.get(bk)
        qid = (ds or {}).get("question_id") or (in_row or {}).get("question_id")
        if qid and qid in ds_by_qid | in_by_qid:
            continue
        seen.add(token)
        keys.append(("business_key", str(qid) if qid else None, bk))

    return keys


def build_unified_record(
    *,
    join_type: str,
    question_id: str | None,
    business_key_str: str,
    dataset_row: dict[str, Any] | None,
    internal_row: dict[str, Any] | None,
    policy: dict[str, Any] | None = None,
) -> dict[str, Any]:
    policy = policy or DEFAULT_POLICY
    resolution = resolve_pair(dataset_row, internal_row, policy=policy)

    company_id = (dataset_row or internal_row or {}).get("company_id")
    family_id = None
    metric_name = None
    year = None
    if dataset_row:
        family_id = map_dataset_family(dataset_row.get("question_family"))
        metric_name = dataset_row.get("metric_name")
        year = _year_from_question(dataset_row.get("question_text"))
    if internal_row:
        family_id = family_id or internal_row.get("family_id")
        metric_name = metric_name or internal_row.get("metric_name")
        year = year if year else internal_row.get("year")

    evidence = []
    if internal_row and internal_row.get("primary_evidence"):
        evidence.append({"source": "internal_doc", **(internal_row.get("primary_evidence") or {})})
    if dataset_row and dataset_row.get("doc_title"):
        evidence.append(
            {
                "source": "dataset",
                "document_id": dataset_row.get("doc_title"),
                "snippet": dataset_row.get("evidence_text") or dataset_row.get("question_text"),
            }
        )

    readiness = (internal_row or {}).get("readiness_state")
    conflict = (internal_row or {}).get("conflict_status")
    if resolution["resolution_status"] == "CONFLICT_REVIEW_REQUIRED":
        conflict = "cross_source_conflict"

    confidence = 0.0
    if resolution["best_answer_origin"] == "both_confirmed":
        confidence = max(float((internal_row or {}).get("confidence") or 0), 0.9)
    elif resolution["best_answer_origin"] == "dataset":
        confidence = 0.9 if (dataset_row or {}).get("answer_correct") else 0.7
    elif resolution["best_answer_origin"] == "internal_doc":
        confidence = float((internal_row or {}).get("confidence") or 0)

    return {
        "schema_version": "1.0.0",
        "identity": {
            "question_id": question_id,
            "business_key": business_key_str,
            "join_type": join_type,
            "company_id": company_id,
            "family_id": family_id,
            "metric_name": metric_name,
            "year": year,
        },
        "best_answer": resolution["best_answer"],
        "best_answer_origin": resolution["best_answer_origin"],
        "resolution_status": resolution["resolution_status"],
        "auto_confirm": resolution["auto_confirm"],
        "review_required": resolution["review_required"],
        "review_owner": resolution["review_owner"],
        "resolution_note": resolution["resolution_note"],
        "confidence": round(confidence, 4),
        "readiness_state": readiness,
        "conflict_status": conflict,
        "sources": {
            "dataset": {
                "present": resolution["dataset_present"],
                "answer": resolution["dataset_answer"],
                "partition": (dataset_row or {}).get("partition"),
                "answer_correct": (dataset_row or {}).get("answer_correct"),
                "predicted_abstain": (dataset_row or {}).get("predicted_abstain"),
                "retrieval_hit_top1": (dataset_row or {}).get("retrieval_hit_top1"),
            },
            "internal_doc": {
                "present": resolution["internal_present"],
                "answer": resolution["internal_answer"],
                "extraction_success": (internal_row or {}).get("extraction_success"),
                "multi_source_confirmed": (internal_row or {}).get("multi_source_confirmed"),
                "field_id": (internal_row or {}).get("field_id"),
            },
        },
        "supporting_evidence": evidence,
    }


def run_unified_resolution(
    *,
    dataset_results_path: Path | None = None,
    internal_records_path: Path | None = None,
    eval_root: Path | None = None,
    policy: dict[str, Any] | None = None,
) -> dict[str, Any]:
    policy = policy or DEFAULT_POLICY
    ds_path = dataset_results_path or DEFAULT_DATASET_RESULTS
    in_path = internal_records_path or DEFAULT_INTERNAL_RECORDS
    eval_root = eval_root or DEFAULT_EVAL_ROOT

    partition_enrichment = load_partition_enrichment(eval_root)
    dataset_rows = _read_jsonl(ds_path)
    internal_rows = _read_jsonl(in_path)

    ds_by_qid, ds_by_bk = index_dataset_rows(dataset_rows, partition_enrichment=partition_enrichment)
    in_by_qid, in_by_bk = index_internal_rows(internal_rows)

    merge_keys = _merge_keys(ds_by_qid, ds_by_bk, in_by_qid, in_by_bk)
    unified: list[dict[str, Any]] = []

    for join_type, qid, bk in merge_keys:
        ds_row = ds_by_qid.get(qid) if qid else ds_by_bk.get(bk)
        if ds_row is None and qid:
            ds_row = ds_by_bk.get(bk)
        in_row = in_by_qid.get(qid) if qid else in_by_bk.get(bk)
        if in_row is None and qid:
            in_row = in_by_bk.get(bk)

        unified.append(
            build_unified_record(
                join_type=join_type,
                question_id=qid,
                business_key_str=bk,
                dataset_row=ds_row,
                internal_row=in_row,
                policy=policy,
            )
        )

    status_counts = Counter(r["resolution_status"] for r in unified)
    return {
        "unified_records": unified,
        "inputs": {
            "dataset_results_path": str(ds_path.relative_to(ROOT)).replace("\\", "/") if ds_path.exists() else str(ds_path),
            "internal_records_path": str(in_path.relative_to(ROOT)).replace("\\", "/") if in_path.exists() else str(in_path),
            "dataset_row_count": len(dataset_rows),
            "internal_row_count": len(internal_rows),
            "unified_row_count": len(unified),
            "join_by_question_id": sum(1 for j, _, _ in merge_keys if j == "question_id"),
            "join_by_business_key_only": sum(1 for j, _, _ in merge_keys if j == "business_key"),
        },
        "status_breakdown": dict(status_counts),
        "policy_version": policy.get("version"),
    }


def mandatory_resolution_answers(result: dict[str, Any]) -> dict[str, Any]:
    sb = result.get("status_breakdown") or {}
    return {
        "1_business_key_mapping": {
            "answer": (
                "Ưu tiên join theo question_id khi cả hai nguồn có; "
                "fallback business_key = company_id::family_id::metric_name_norm::year. "
                "Dataset family map qua FAMILY_ALIASES (employee_status→employee_headcount)."
            ),
            "primary_key": "question_id",
            "fallback_key": ["company_id", "family_id", "metric_name", "year"],
        },
        "2_best_answer_rules": {
            "answer": (
                "MATCH_CONFIRMED khi giá trị khớp; BACKFILL_FROM_DATASET khi chỉ RAG có; "
                "BACKFILL_FROM_INTERNAL khi chỉ internal-doc có; CONFLICT_REVIEW_REQUIRED khi cả hai khác; "
                "NO_ANSWER_FOUND / INSUFFICIENT_EVIDENCE khi không đủ signal."
            ),
            "policy_file": "resolution_policy.json",
        },
        "3_auto_confirm_vs_review": {
            "auto_confirm": ["MATCH_CONFIRMED", "BACKFILL_FROM_DATASET (answer_correct)", "BACKFILL_FROM_INTERNAL (conf>=0.85+sufficiency)"],
            "review_required": ["CONFLICT_REVIEW_REQUIRED", "INSUFFICIENT_EVIDENCE", "BACKFILL_FROM_INTERNAL candidate"],
            "status_counts": sb,
        },
        "4_unified_output_fields": {
            "answer": "identity, best_answer, best_answer_origin, resolution_status, confidence, readiness_state, conflict_status, sources{dataset,internal_doc}, supporting_evidence, auto_confirm, review_required",
            "schema": "data/unified_esg/unified_answer_schema.json",
        },
        "5_review_artifact_without_touching_excel": {
            "answer": "Có — workbook/JSONL riêng (unified_answers.jsonl, unified_esg_review.xlsx); không ghi đè Excel gốc hay results.jsonl frozen.",
            "artifacts": ["unified_answers.jsonl", "unified_esg_review.xlsx", "review_workbook_plan.md"],
        },
        "6_next_step": {
            "answer": (
                "Khi có công ty overlap cả dataset + internal-doc: chạy resolution trên cùng company_id; "
                "SME review sheet CONFLICT; publish unified layer cho báo cáo ESG; không mở LangGraph/synthesis."
            ),
        },
    }


def export_review_workbook(
    unified_records: list[dict[str, Any]],
    output_path: Path,
) -> dict[str, Any]:
    """Export unified review workbook — separate artifact, does not modify source Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return {"exported": False, "reason": "openpyxl not available"}

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)

    sheets: dict[str, list[dict[str, Any]]] = {
        "all_unified": unified_records,
        "MATCH_CONFIRMED": [r for r in unified_records if r["resolution_status"] == "MATCH_CONFIRMED"],
        "BACKFILL_INTERNAL": [r for r in unified_records if r["resolution_status"] == "BACKFILL_FROM_INTERNAL"],
        "BACKFILL_DATASET": [r for r in unified_records if r["resolution_status"] == "BACKFILL_FROM_DATASET"],
        "CONFLICT_REVIEW": [r for r in unified_records if r["resolution_status"] == "CONFLICT_REVIEW_REQUIRED"],
        "NO_ANSWER": [r for r in unified_records if r["resolution_status"] in ("NO_ANSWER_FOUND", "INSUFFICIENT_EVIDENCE")],
    }

    columns = [
        "question_id",
        "company_id",
        "family_id",
        "metric_name",
        "year",
        "resolution_status",
        "best_answer",
        "best_answer_origin",
        "auto_confirm",
        "review_required",
        "dataset_answer",
        "internal_answer",
        "confidence",
        "resolution_note",
    ]

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
        for rec in rows:
            ident = rec.get("identity") or {}
            src = rec.get("sources") or {}
            ws.append(
                [
                    ident.get("question_id"),
                    ident.get("company_id"),
                    ident.get("family_id"),
                    ident.get("metric_name"),
                    ident.get("year"),
                    rec.get("resolution_status"),
                    rec.get("best_answer"),
                    rec.get("best_answer_origin"),
                    rec.get("auto_confirm"),
                    rec.get("review_required"),
                    (src.get("dataset") or {}).get("answer"),
                    (src.get("internal_doc") or {}).get("answer"),
                    rec.get("confidence"),
                    rec.get("resolution_note"),
                ]
            )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"exported": True, "path": str(output_path), "sheet_counts": {k: len(v) for k, v in sheets.items()}}
