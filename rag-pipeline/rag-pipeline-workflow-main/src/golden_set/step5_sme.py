"""Step 5: Export SME review workbook (human-in-the-loop)."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Dict, List

from golden_set.io_utils import read_jsonl


SME_FIELDS = [
    "silver_id",
    "company",
    "package_name",
    "question_type",
    "difficulty",
    "gri_code",
    "question",
    "ground_truth_answer",
    "ground_truth_record_id",
    "context_excerpt",
    "qc_reason",
    "sme_decision",
    "sme_revised_question",
    "sme_revised_answer",
    "sme_notes",
    "forbidden_rule",
]


def run_step5(*, input_path: Path, csv_path: Path, xlsx_path: Path | None = None) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    out_rows: List[Dict[str, str]] = []
    for r in rows:
        out_rows.append(
            {
                "silver_id": r.get("silver_id", ""),
                "company": r.get("company", ""),
                "package_name": r.get("package_name", ""),
                "question_type": r.get("question_type", ""),
                "difficulty": r.get("difficulty", ""),
                "gri_code": r.get("gri_code", ""),
                "question": r.get("question", ""),
                "ground_truth_answer": r.get("ground_truth_answer", ""),
                "ground_truth_record_id": r.get("ground_truth_record_id", ""),
                "context_excerpt": (r.get("context_excerpt") or "")[:500],
                "qc_reason": r.get("qc_reason", ""),
                "sme_decision": "",
                "sme_revised_question": "",
                "sme_revised_answer": "",
                "sme_notes": "",
                "forbidden_rule": "Không thêm thông tin ngoài context; không suy đoán nếu thiếu số liệu",
            }
        )

    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SME_FIELDS)
        w.writeheader()
        w.writerows(out_rows)

    if xlsx_path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "SME Review"
            header_fill = PatternFill("solid", fgColor="1D4ED8")
            header_font = Font(color="FFFFFF", bold=True)
            ws.append(SME_FIELDS)
            for c in range(1, len(SME_FIELDS) + 1):
                cell = ws.cell(row=1, column=c)
                cell.fill = header_fill
                cell.font = header_font
            for row in out_rows:
                ws.append([row.get(k, "") for k in SME_FIELDS])
            wb.save(xlsx_path)
        except ImportError:
            pass

    return {"rows": len(out_rows), "csv": str(csv_path), "xlsx": str(xlsx_path) if xlsx_path else ""}
