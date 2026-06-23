"""Canonical Rewrite + Fact-Cluster Dedupe Round 2 for reference seed workbook."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.io_utils import read_jsonl, write_jsonl

CANONICAL_VERSION = "ref_canonical_r2"

KEEP_DECISIONS = {"canonical_keep", "canonical_keep_after_rewrite"}

# (cluster_id, question_patterns, fact_line_patterns)
FACT_CLUSTER_RULES: List[Tuple[str, List[re.Pattern], List[re.Pattern]]] = [
    (
        "FC_NET_ZERO_2050",
        [re.compile(p) for p in [r"2050", r"기후 목표", r"탄소중립", r"넷제로", r"net zero"]],
        [re.compile(p) for p in [r"2050", r"탄소중립", r"넷제로", r"net zero", r"환경경영 시스템"]],
    ),
    (
        "FC_BOARD_2022",
        [re.compile(p) for p in [r"이사회", r"14회", r"44건", r"안건"]],
        [re.compile(p) for p in [r"14회", r"44건", r"이사회를 개최"]],
    ),
    (
        "FC_ESG_GOVERNANCE",
        [re.compile(p) for p in [r"거버넌스", r"ESG 위원회", r"ESG위원회", r"경영 체계", r"소위원회"]],
        [re.compile(p) for p in [r"ESG 위원회", r"ESG위원회", r"이사회 중심", r"소위원회", r"경영 체계"]],
    ),
    (
        "FC_TCFD",
        [re.compile(p) for p in [r"TCFD", r"기후변화 대응 공시", r"기후.*공시"]],
        [re.compile(p) for p in [r"TCFD"]],
    ),
    (
        "FC_MATERIAL_8",
        [re.compile(p) for p in [r"중대 이슈", r"몇 개의 중대"]],
        [re.compile(p) for p in [r"8개 중대", r"중대 이슈를 선정"]],
    ),
    (
        "FC_KGCS_A",
        [re.compile(p) for p in [r"KGCS", r"ESG 평가", r"등급"]],
        [re.compile(p) for p in [r"KGCS", r"A.?등급", r"등급을 획득"]],
    ),
    (
        "FC_HUMAN_RIGHTS",
        [re.compile(p) for p in [r"인권"]],
        [re.compile(p) for p in [r"인권실태", r"인권영향", r"인권경영"]],
    ),
]

TCFD_DEFINITION_ONLY = re.compile(
    r"TCFD는\s*2015년.*협의체", re.IGNORECASE
)
TRUNCATED_ENDINGS = ("BIS)과", "금융안정위원회(FSB)가", "인증 제도인", "국제 친환경")


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _fact_lines(seed: Dict[str, Any]) -> List[str]:
    raw = seed.get("facts_tuple") or ""
    lines: List[str] = []
    for line in str(raw).splitlines():
        parts = line.split("|")
        if len(parts) >= 3:
            lines.append(_norm_ws(parts[2]))
        elif line.strip():
            lines.append(_norm_ws(line))
    return [x for x in lines if len(x) >= 10]


def _match_rules(text: str, patterns: Sequence[re.Pattern]) -> bool:
    return any(p.search(text) for p in patterns)


def infer_fact_cluster(text: str) -> str:
    for cluster_id, q_patterns, _ in FACT_CLUSTER_RULES:
        if _match_rules(text, q_patterns):
            return cluster_id
    return "FC_UNKNOWN"


def infer_fact_target_id(question: str, passage: str = "") -> str:
    q = question or ""
    blob = f"{q} {passage}"
    for cluster_id, q_patterns, f_patterns in FACT_CLUSTER_RULES:
        if _match_rules(q, q_patterns):
            return cluster_id
    if _match_rules(blob, [re.compile(r"8개")]):
        return "FC_MATERIAL_8"
    return "FC_UNKNOWN"


def _intent_key(question: str, fact_target: str) -> str:
    q = _norm_ws(question).lower()
    q = re.sub(r"\s+", " ", q)
    return f"{fact_target}::{q}"


def _pick_fact_line(lines: List[str], fact_target: str) -> Optional[str]:
    _, _, f_patterns = next(
        ((a, b, c) for a, b, c in FACT_CLUSTER_RULES if a == fact_target),
        ("FC_UNKNOWN", [], []),
    )
    if f_patterns:
        for line in lines:
            if _match_rules(line, f_patterns) and not _is_noise_fact(line, fact_target):
                return line
    for line in lines:
        if not _is_noise_fact(line, fact_target):
            return line
    return lines[0] if lines else None


def _is_noise_fact(line: str, fact_target: str) -> bool:
    if TCFD_DEFINITION_ONLY.search(line) and fact_target != "FC_TCFD":
        return True
    if fact_target == "FC_TCFD" and TCFD_DEFINITION_ONLY.search(line):
        return False  # definition-only for TCFD intent is still weak company fact
    if "발간했다고" in line and fact_target not in ("FC_MATERIAL_8",):
        if "8개" not in line:
            return True
    return False


def _is_truncated(text: str) -> bool:
    t = _norm_ws(text)
    if len(t) < 50:
        return True
    if t.startswith(("완성하고", "또한")) and "한샘" not in t[:50]:
        return True
    return any(t.endswith(end) for end in TRUNCATED_ENDINGS)


def _disclosure_answers_target(disclosure: str, fact_target: str, fact_line: Optional[str]) -> bool:
    if not disclosure or not fact_line:
        return False
    d = disclosure.lower()
    fl = fact_line.lower()
    if fact_target == "FC_NET_ZERO_2050":
        return "2050" in d or "탄소중립" in d or "넷제로" in d
    if fact_target == "FC_BOARD_2022":
        return "14회" in d and "44건" in d
    if fact_target == "FC_ESG_GOVERNANCE":
        return ("esg 위원회" in d or "esg위원회" in d or "경영 체계" in d) and "tcfd는" not in d[:80]
    if fact_target == "FC_TCFD":
        return "tcfd" in d and "한샘" in d
    if fact_target == "FC_MATERIAL_8":
        return "8개" in d and "중대" in d
    if fact_target == "FC_KGCS_A":
        return "kgcs" in d or "a" in d and "등급" in d
    return fl[:40] in d or d[:40] in fl


def _r1_rewrite_drifted(seed: Dict[str, Any], original_target: str) -> Tuple[bool, str]:
    rq = seed.get("rewritten_question_draft") or ""
    rd = seed.get("rewritten_acceptable_disclosure") or ""
    rewritten_target = infer_fact_target_id(rq, rd)
    if rewritten_target != original_target and rewritten_target != "FC_UNKNOWN":
        return True, f"rewrite_target_{rewritten_target}_vs_original_{original_target}"
    if not _disclosure_answers_target(rd, original_target, _pick_fact_line(_fact_lines(seed), original_target)):
        return True, "answer_target_mismatch_in_r1_rewrite"
    return False, ""


def _canonical_question(seed: Dict[str, Any], fact_target: str) -> str:
    company = seed.get("company") or "한샘"
    templates = {
        "FC_NET_ZERO_2050": f"{company}는 2050년까지 어떤 탄소중립 목표를 공개했는가?",
        "FC_BOARD_2022": f"{company}는 2022년 이사회를 몇 회 개최하고 몇 건의 안건을 심의했는가?",
        "FC_ESG_GOVERNANCE": f"{company}의 ESG 거버넌스 체계는 어떻게 운영되는가?",
        "FC_TCFD": f"{company}는 기후변화 대응 공시에 어떤 프레임워크를 따르는가?",
        "FC_MATERIAL_8": f"{company}는 이중 중대성 평가를 통해 몇 개의 중대 이슈를 선정했는가?",
        "FC_KGCS_A": f"{company}는 KGCS ESG경영 평가에서 어떤 등급을 획득했는가?",
        "FC_HUMAN_RIGHTS": f"{company}는 인권경영 체계를 어떻게 강화했는가?",
    }
    q0 = seed.get("question_draft") or ""
    if infer_fact_target_id(q0) == fact_target:
        return q0
    return templates.get(fact_target, q0)


def _canonical_disclosure(fact_line: str, fact_target: str) -> str:
    line = _norm_ws(fact_line)
    if fact_target == "FC_KGCS_A" and "KGCS" in line:
        m = re.search(r"한샘은[^.]*KGCS[^.]*등급[^.]*\.", line)
        if m:
            return m.group(0)
        m2 = re.search(r"한샘은[^.]*등급을 획득[^.]*\.", line)
        if m2:
            return m2.group(0)
    if fact_target == "FC_MATERIAL_8":
        m = re.search(r"[^.]*8개 중대 이슈를 선정[^.]*\.", line)
        if m:
            return m.group(0)
        if "8개 중대" in line:
            return line[:280]
    return line[:420]


def _passage_too_noisy_for_target(seed: Dict[str, Any], fact_target: str, fact_line: Optional[str]) -> bool:
    if not fact_line:
        return True
    if fact_target == "FC_TCFD":
        lines = _fact_lines(seed)
        company_tcfd = any("한샘" in ln and "tcfd" in ln.lower() for ln in lines)
        if not company_tcfd and all(TCFD_DEFINITION_ONLY.search(ln) or "tcfd" in ln.lower() for ln in lines if "tcfd" in ln.lower()):
            return True
    if _is_truncated(fact_line):
        return True
    return False


def _cleanliness_score(seed: Dict[str, Any], fact_line: str) -> float:
    score = float(seed.get("rank") or 0)
    if seed.get("curation_decision") == "keep_strong":
        score += 5
    passage = seed.get("passage_text") or ""
    if "발간했다고" in passage or "(009240)" in passage:
        score -= 3
    if _is_truncated(passage):
        score -= 5
    if len(fact_line) > 80:
        score += 2
    return score


def canonicalize_row(seed: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(seed)
    question = seed.get("question_draft") or ""
    original_target = infer_fact_target_id(question, seed.get("passage_text") or "")
    fact_line = _pick_fact_line(_fact_lines(seed), original_target)
    cluster_id = original_target

    out.update(
        {
            "canonical_version": CANONICAL_VERSION,
            "fact_target": original_target,
            "fact_cluster_id": cluster_id,
            "original_fact_target": original_target,
            "rewritten_fact_target": infer_fact_target_id(
                seed.get("rewritten_question_draft") or "",
                seed.get("rewritten_acceptable_disclosure") or "",
            ),
        }
    )

    r1_drift, r1_drift_reason = _r1_rewrite_drifted(seed, original_target)
    out["r1_rewrite_drift"] = r1_drift
    out["r1_drift_reason"] = r1_drift_reason if r1_drift else ""

    # Hard drops before canonical rewrite (ignore bad R1 rewrite — rebuild from original target)
    if original_target == "FC_UNKNOWN":
        out.update(
            _drop_fields("drop_still_too_noisy", "unknown_fact_target", "Không xác định được fact target từ câu hỏi.")
        )
        return out

    if _passage_too_noisy_for_target(seed, original_target, fact_line):
        if _is_truncated(fact_line or ""):
            out.update(
                _drop_fields(
                    "drop_truncated_unsalvageable",
                    "truncated_fact_line",
                    "Passage/fact bị cắt, không đủ grounded disclosure.",
                )
            )
            return out
        out.update(
            _drop_fields(
                "drop_still_too_noisy",
                "no_company_fact_for_target",
                "Passage không có fact công ty đủ sạch cho target.",
            )
        )
        return out

    canonical_q = _canonical_question(seed, original_target)
    canonical_disc = _canonical_disclosure(fact_line or "", original_target)
    canonical_prohib = seed.get("prohibited_claims") or seed.get("rewritten_prohibited_claims") or ""

    out["rewritten_fact_target"] = infer_fact_target_id(canonical_q, canonical_disc)
    preserved = out["rewritten_fact_target"] == original_target and _disclosure_answers_target(
        canonical_disc, original_target, fact_line
    )
    out["fact_target_preserved"] = "yes" if preserved else "no"

    if not preserved:
        drop_decision = "drop_semantic_drift" if r1_drift else "drop_answer_target_mismatch"
        out.update(
            _drop_fields(
                drop_decision,
                "canonical_disclosure_mismatch" if not r1_drift else r1_drift_reason,
                "Canonical disclosure không trả lời đúng fact target ban đầu."
                if not r1_drift
                else "Không thể cứu row mà vẫn giữ fact target gốc.",
            )
        )
        return out

    needs_rewrite = (
        r1_drift
        or seed.get("curation_decision") == "keep_but_needs_rewrite"
        or "발간했다고" in (seed.get("passage_text") or "")
        or len(canonical_disc) < len(seed.get("acceptable_disclosure") or "") * 0.4
    )
    if seed.get("curation_decision") == "keep_strong" and not r1_drift and not needs_rewrite:
        decision = "canonical_keep"
        reason = "clean_canonical_fact"
        notes = "Seed sạch, fact target rõ, không cần rewrite thêm."
        out.update(
            {
                "canonical_decision": decision,
                "canonical_reason": reason,
                "canonical_notes": notes,
                "canonical_question": canonical_q,
                "canonical_acceptable_disclosure": canonical_disc,
                "canonical_prohibited_claims": canonical_prohib,
                "review_priority": seed.get("review_priority") or "medium",
                "intent_key": _intent_key(canonical_q, original_target),
                "_clean_score": _cleanliness_score(seed, fact_line or ""),
            }
        )
        return out

    decision = "canonical_keep_after_rewrite" if needs_rewrite else "canonical_keep"
    reason = "fact_target_preserved_rewrite" if needs_rewrite else "clean_canonical_fact"
    notes = (
        f"Fact target giữ nguyên; disclosure thu gọn theo fact line."
        + (f" (R1 drift đã sửa: {r1_drift_reason})" if r1_drift else "")
        if needs_rewrite
        else "Seed sạch, fact target rõ, không cần rewrite thêm."
    )

    out.update(
        {
            "canonical_decision": decision,
            "canonical_reason": reason,
            "canonical_notes": notes,
            "canonical_question": canonical_q,
            "canonical_acceptable_disclosure": canonical_disc,
            "canonical_prohibited_claims": canonical_prohib,
            "review_priority": seed.get("review_priority") or "medium",
            "intent_key": _intent_key(canonical_q, original_target),
            "_clean_score": _cleanliness_score(seed, fact_line or ""),
        }
    )
    return out


def _drop_fields(decision: str, reason: str, notes: str) -> Dict[str, Any]:
    return {
        "canonical_decision": decision,
        "canonical_reason": reason,
        "canonical_notes": notes,
        "canonical_question": None,
        "canonical_acceptable_disclosure": None,
        "canonical_prohibited_claims": None,
        "review_priority": "n/a",
        "intent_key": None,
        "_clean_score": -999,
    }


def _dedupe_canonical_candidates(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Keep at most one canonical row per company + fact_cluster_id + intent."""
    keep_candidates = [r for r in rows if r.get("canonical_decision") in KEEP_DECISIONS]
    dropped = [r for r in rows if r.get("canonical_decision") not in KEEP_DECISIONS]

    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in keep_candidates:
        key = f"{row.get('company')}::{row.get('fact_cluster_id')}::{row.get('intent_key')}"
        groups[key].append(row)

    final_keep: List[Dict[str, Any]] = []
    for _key, group in groups.items():
        group.sort(key=lambda x: (-x.get("_clean_score", 0), x.get("seed_id", "")))
        winner = group[0]
        final_keep.append(winner)
        for loser in group[1:]:
            dup = dict(loser)
            dup.update(
                _drop_fields(
                    "drop_duplicate_fact_cluster",
                    f"duplicate_of_{winner.get('seed_id')}",
                    f"Trùng fact cluster/intent với {winner.get('seed_id')}; giữ bản sạch hơn.",
                )
            )
            dropped.append(dup)

    for row in rows:
        if row.get("canonical_decision") not in KEEP_DECISIONS and row not in dropped:
            dropped.append(row)

    # Re-add non-keep that weren't in dropped yet (already in dropped from canonicalize)
    seen_ids = {r["seed_id"] for r in final_keep + dropped}
    for row in rows:
        if row["seed_id"] not in seen_ids:
            dropped.append(row)

    final_keep.sort(key=lambda x: (-x.get("_clean_score", 0), x.get("seed_id", "")))
    return final_keep, dropped


def _write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws_info = wb.active
    ws_info.title = "안내"
    ws_info.append(["Reference Seed Workbook — Canonical R2"])
    ws_info.append(["Gate", "Fact-target preservation + cluster dedupe"])
    ws_info.append(["Scope", "Hansem-only partial workbook"])

    header = [
        "seed_id",
        "company",
        "question_type",
        "question_draft",
        "canonical_question",
        "source_record_id",
        "fact_cluster_id",
        "original_fact_target",
        "canonical_acceptable_disclosure",
        "canonical_prohibited_claims",
        "canonical_decision",
        "canonical_reason",
        "review_priority",
        "notes",
    ]
    ws = wb.create_sheet("작성")
    ws.append(header)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    for row in rows:
        ws.append([row.get(h, "") for h in header])

    ws_sum = wb.create_sheet("요약")
    ws_sum.append(["metric", "value"])
    ws_sum.append(["canonical_total", len(rows)])
    for k, v in Counter(r["canonical_decision"] for r in rows).items():
        ws_sum.append([k, v])
    for cid, n in sorted(Counter(r["fact_cluster_id"] for r in rows).items()):
        ws_sum.append([f"cluster:{cid}", n])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_canonical_report(
    summary: Dict[str, Any],
    examples: Dict[str, List[Dict]],
    path: Path,
) -> None:
    lines = [
        "# Golden Set — Reference Workbook Canonical R2",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Canonical hóa curated R1 theo **fact target**: giữ đúng fact mà seed đại diện, loại semantic drift, dedupe cluster — tạo workbook chặt hơn trước review nội dung.",
        "",
        "## Vấn đề còn lại của R1",
        "",
        "1. Curated set **Hansem-only** (14/14) — không còn 무신사/레이시온.",
        "2. 11 row `keep_but_needs_rewrite` trùng fact cluster (Net Zero, board 14/44, KGCS, material 8).",
        "3. R1 rewrite **semantic drift**: ví dụ HS-G-Q01 hỏi 2050 goal nhưng rewrite sang governance.",
        "4. **Answer target mismatch**: HS-G-T03 disclosure rewrite rơi vào TCFD/BIS thay vì Net Zero.",
        "5. `keep_strong` Q06/L06 trùng KGCS — cần dedupe.",
        "",
        "## Quy tắc canonicalization R2",
        "",
        "| Decision | Điều kiện |",
        "|----------|-----------|",
        "| `canonical_keep` | Fact target rõ, passage sạch, không drift |",
        "| `canonical_keep_after_rewrite` | Fact salvageable, rewrite giữ nguyên target |",
        "| `drop_semantic_drift` | Rewrite R1 đổi fact target |",
        "| `drop_answer_target_mismatch` | Disclosure không trả lời fact câu hỏi |",
        "| `drop_duplicate_fact_cluster` | Trùng cluster+intent, giữ 1 bản tốt nhất |",
        "| `drop_truncated_unsalvageable` | Fact line cắt cụt |",
        "| `drop_still_too_noisy` | Không có company fact cho target |",
        "",
        "## Tổng số row đầu vào",
        "",
        f"**{summary.get('input_total', 0)}** row từ `reference_seed_candidates_curated_r1.jsonl`.",
        "",
        "## Kết quả canonical",
        "",
        f"| Chỉ số | Giá trị |",
        f"|--------|--------:|",
        f"| canonical_keep | {summary.get('canonical_keep', 0)} |",
        f"| canonical_keep_after_rewrite | {summary.get('canonical_keep_after_rewrite', 0)} |",
        f"| canonical usable total | {summary.get('canonical_usable_total', 0)} |",
        f"| independent fact clusters | {summary.get('independent_fact_clusters', 0)} |",
        "",
        "### Drop theo nhóm",
        "",
    ]
    for k, v in sorted(summary.get("dropped_by_decision", {}).items()):
        lines.append(f"- `{k}`: **{v}**")

    lines.extend(["", "## Fact cluster collapse", ""])
    for cid, detail in summary.get("cluster_collapse", {}).items():
        lines.append(f"- **{cid}**: {detail}")

    lines.extend(["", "## Ví dụ cụ thể", ""])
    for group in [
        "drop_semantic_drift",
        "drop_answer_target_mismatch",
        "drop_duplicate_fact_cluster",
        "canonical_keep",
        "canonical_keep_after_rewrite",
    ]:
        items = examples.get(group, [])
        if items:
            lines.append(f"### `{group}`")
            for ex in items[:2]:
                lines.append(
                    f"- **{ex.get('seed_id')}**: {ex.get('canonical_notes', ex.get('canonical_reason', ''))[:100]}"
                )
            lines.append("")

    lines.extend(
        [
            "## Đánh giá cuối",
            "",
            f"- **Seed canonical usable:** {summary.get('canonical_usable_total', 0)}",
            f"- **Fact cluster độc lập:** {summary.get('independent_fact_clusters', 0)}",
            f"- **Coverage:** {summary.get('coverage_note', '')}",
            f"- **Đủ review nội dung?** {summary.get('content_review_ready', '')}",
            "",
            "## Kết luận (3 câu)",
            "",
            f"1. **Sau R2 còn bao nhiêu seed canonical usable?** **{summary.get('canonical_usable_total', 0)}** seed.",
            f"2. **Có bao nhiêu fact cluster độc lập?** **{summary.get('independent_fact_clusters', 0)}** cluster.",
            f"3. **Bước tiếp theo?** {summary.get('next_step', '')}",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_canonical_r2(
    *,
    input_jsonl: Path,
    canonical_jsonl: Path,
    dropped_jsonl: Path,
    canonical_xlsx: Path,
) -> Tuple[Dict[str, Any], Dict[str, List[Dict]]]:
    seeds = read_jsonl(input_jsonl)
    processed = [canonicalize_row(s) for s in seeds]
    canonical, dropped = _dedupe_canonical_candidates(processed)

    # strip internal fields
    for row in canonical + dropped:
        row.pop("_clean_score", None)

    write_jsonl(canonical_jsonl, canonical)
    write_jsonl(dropped_jsonl, dropped)
    _write_workbook(canonical, canonical_xlsx)

    all_rows = canonical + dropped
    examples: Dict[str, List[Dict]] = defaultdict(list)
    for row in all_rows:
        examples[row.get("canonical_decision", "unknown")].append(row)

    dec_counts = Counter(r["canonical_decision"] for r in all_rows)
    cluster_in = Counter(infer_fact_target_id(r.get("question_draft") or "") for r in seeds)
    cluster_out = Counter(r.get("fact_cluster_id") for r in canonical)

    cluster_collapse = {}
    all_clusters = {c for c in set(cluster_in) | set(cluster_out) if c and c != "FC_UNKNOWN"}
    for cid in sorted(all_clusters):
        cluster_collapse[cid] = f"input {cluster_in.get(cid, 0)} → canonical {cluster_out.get(cid, 0)}"

    usable = len(canonical)
    companies = set(r.get("company") for r in canonical)
    content_ready = usable >= 4 and cluster_out.get("FC_UNKNOWN", 0) == 0

    if len(companies) == 1 and "무신사" not in companies:
        coverage = "**Hansem-only canonical workbook** — chưa có multi-company coverage."
        next_step = (
            "**Chưa** đủ cho review nội dung đầy đủ 3 công ty. "
            "Ưu tiên: (a) review nội dung round 3 trên Hansem canonical nếu cần validate format; "
            "(b) **rebuild seed workbook v2** cho 무신사/레이시온 từ corpus sạch."
        )
    else:
        coverage = "Multi-company partial."
        next_step = "Review nội dung round 3."

    summary = {
        "canonical_version": CANONICAL_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_total": len(seeds),
        "canonical_keep": dec_counts.get("canonical_keep", 0),
        "canonical_keep_after_rewrite": dec_counts.get("canonical_keep_after_rewrite", 0),
        "canonical_usable_total": usable,
        "independent_fact_clusters": len(cluster_out),
        "dropped_total": len(dropped),
        "dropped_by_decision": {
            k: v
            for k, v in dec_counts.items()
            if k not in KEEP_DECISIONS
        },
        "cluster_collapse": cluster_collapse,
        "canonical_by_company": dict(Counter(r.get("company") for r in canonical)),
        "canonical_seed_ids": [r["seed_id"] for r in canonical],
        "semantic_drift_dropped": [
            r["seed_id"]
            for r in dropped
            if r.get("canonical_decision") == "drop_semantic_drift"
        ],
        "answer_target_mismatch_dropped": [
            r["seed_id"]
            for r in dropped
            if r.get("canonical_decision") == "drop_answer_target_mismatch"
        ],
        "r1_drift_detected_in_input": [
            r["seed_id"] for r in processed if r.get("r1_rewrite_drift")
        ],
        "r1_drift_fixed_in_canonical": [
            r["seed_id"] for r in canonical if r.get("r1_rewrite_drift")
        ],
        "coverage_note": coverage,
        "content_review_ready": (
            "Chưa — Hansem-only skeleton (4 cluster); đủ pilot format, chưa đủ review nội dung 3 công ty"
            if usable < 8
            else "Có thể bắt đầu review nội dung"
        ),
        "next_step": next_step,
        "canonical_jsonl": str(canonical_jsonl),
        "dropped_jsonl": str(dropped_jsonl),
        "canonical_xlsx": str(canonical_xlsx),
    }
    return summary, dict(examples)


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Canonicalize reference seed workbook R2")
    root = Path(__file__).resolve().parents[2]
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_curated_r1.jsonl",
    )
    parser.add_argument(
        "--canonical-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_canonical_r2.jsonl",
    )
    parser.add_argument(
        "--dropped-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_dropped_r2.jsonl",
    )
    parser.add_argument(
        "--canonical-xlsx",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_canonical_r2.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_reference_workbook_canonical_r2.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_reference_workbook_canonical_r2_summary.json",
    )
    args = parser.parse_args(argv)

    summary, examples = run_canonical_r2(
        input_jsonl=root / args.input,
        canonical_jsonl=root / args.canonical_jsonl,
        dropped_jsonl=root / args.dropped_jsonl,
        canonical_xlsx=root / args.canonical_xlsx,
    )
    write_canonical_report(summary, examples, root / args.report)
    (root / args.summary_json).write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "input": summary["input_total"],
                "canonical_usable": summary["canonical_usable_total"],
                "clusters": summary["independent_fact_clusters"],
                "drift_dropped": len(summary["semantic_drift_dropped"]),
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
