"""Gold Decision Round 4 — core 한샘 + 무신사 only."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import METRIC_HINT_RE
from golden_set.canonicalize_reference_seed_workbook_r2 import (
    TCFD_DEFINITION_ONLY,
    infer_fact_target_id,
)
from golden_set.io_utils import read_jsonl, write_jsonl
from golden_set.prepare_manual_review_round2 import GENERIC_Q, _norm_ws

GOLD_VERSION = "ref_gold_decision_r4"
CORE_COMPANIES = frozenset({"한샘", "무신사"})
KNOWN_CLUSTERS = frozenset(
    {
        "FC_NET_ZERO_2050",
        "FC_BOARD_2022",
        "FC_ESG_GOVERNANCE",
        "FC_TCFD",
        "FC_MATERIAL_8",
        "FC_KGCS_A",
        "FC_HUMAN_RIGHTS",
    }
)

LISTING_NOISE = ["목록 글쓰기", "조회수", "lg헬로비전", "코스맥스", "69180kb", "share.google"]
NEWS_CHROME = ["기사와 관련", "요기요", "앱 ui", "당신이 좋아할", "혼용률", "기자", "발행일"]
TRUNCATED_END = ("보다", "BIS)과", "2023년보다")

WORKBOOK_COLS = [
    "seed_id",
    "company",
    "question_type",
    "cluster_id",
    "canonical_decision",
    "final_question",
    "final_disclosure",
    "canonical_prohibited_claims",
    "gold_decision",
    "gold_reason",
    "gold_notes",
    "gold_ready_flag",
    "needs_cluster_rename",
    "needs_prohibited_claims_refine",
]


def _q(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("final_question") or "")


def _d(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("final_disclosure") or "")


def _prohibited(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("canonical_prohibited_claims") or row.get("prohibited_claims") or "")


def _fact_target(row: Dict[str, Any]) -> str:
    return row.get("canonical_fact_target") or infer_fact_target_id(_q(row), _d(row))


def _is_generic(q: str) -> bool:
    return bool(GENERIC_Q.search(q))


def _listing_noise(d: str) -> bool:
    blob = d.lower()
    return sum(1 for m in LISTING_NOISE if m in blob) >= 1


def _news_chrome(d: str) -> bool:
    blob = d.lower()
    return sum(1 for m in NEWS_CHROME if m in blob) >= 1


def _is_truncated(d: str) -> bool:
    if len(d) < 55:
        return True
    return d.rstrip().endswith(TRUNCATED_END)


def _company_in_text(company: str, text: str) -> bool:
    if company in text:
        return True
    aliases = {"한샘": ["㈜한샘", "한샘이"], "무신사": ["musinsa"]}
    return any(a in text for a in aliases.get(company, []))


def _resolve_unknown_cluster(q: str, d: str) -> Optional[str]:
    """If FC_UNKNOWN but content maps to a known cluster, return resolved name."""
    inferred = infer_fact_target_id(q, d)
    if inferred != "FC_UNKNOWN":
        return inferred
    if "임팩트" in d or "impact report" in d.lower():
        return "FC_REPORT_IMPACT"
    if "스탠다드" in d and "매장" in d:
        return "FC_OFFLINE_RETAIL"
    if "사외이사" in d and "이사회" in d:
        return "FC_ESG_GOVERNANCE"
    return None


def _fact_clarity_score(company: str, q: str, d: str) -> int:
    score = 0
    if _company_in_text(company, d):
        score += 2
    if METRIC_HINT_RE.search(d) or re.search(r"\d", d):
        score += 2
    if not _is_generic(q):
        score += 2
    if len(d) >= 80:
        score += 1
    if any(k in d for k in ("중대", "탄소", "온실", "이사회", "esg", "지속가능", "인권")):
        score += 1
    return score


def _prohibited_ok(text: str) -> bool:
    if not text:
        return False
    required = ("금지", "미공시")
    return any(r in text for r in required)


def decide_gold(row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row)
    company = row.get("company") or ""
    q = _q(row)
    d = _d(row)
    prohib = _prohibited(row)
    target = _fact_target(row)
    canon_dec = row.get("canonical_decision") or ""

    needs_cluster_rename = target == "FC_UNKNOWN"
    resolved = _resolve_unknown_cluster(q, d) if needs_cluster_rename else target
    clarity = _fact_clarity_score(company, q, d)

    gold_decision = "gold_revise_hold"
    gold_reason = "default_hold_review"
    gold_notes = ""
    gold_ready = False
    needs_prohib_refine = not _prohibited_ok(prohib)

    # Hard reject signals
    if _listing_noise(d) and clarity < 6:
        gold_decision = "gold_reject"
        gold_reason = "listing_or_cross_report_meta"
        gold_notes = "Disclosure còn listing/meta — không đủ gold."
    elif _is_truncated(d) and clarity < 5:
        gold_decision = "gold_reject"
        gold_reason = "truncated_disclosure"
        gold_notes = "Disclosure cắt cụt, không đủ grounded gold."
    elif TCFD_DEFINITION_ONLY.search(d) and target == "FC_TCFD":
        gold_decision = "gold_reject"
        gold_reason = "tcfd_definition_only"
        gold_notes = "Chỉ có TCFD definition, không company fact."
    elif not _company_in_text(company, d) and clarity < 4:
        gold_decision = "gold_reject"
        gold_reason = "weak_company_anchor"
        gold_notes = "Disclosure không anchor company đủ mạnh."

    # FC_UNKNOWN special path (no auto approve unless very clear)
    elif target == "FC_UNKNOWN":
        if resolved and clarity >= 6 and not _listing_noise(d) and not _news_chrome(d):
            if _news_chrome(d) or len(d) > 350:
                gold_decision = "gold_revise_hold"
                gold_reason = "unknown_cluster_salvageable"
                gold_notes = f"Fact có thể map → {resolved}; cần rename cluster + trim disclosure."
            else:
                gold_decision = "gold_approve"
                gold_reason = "unknown_resolved_clear_fact"
                gold_notes = f"FC_UNKNOWN nhưng fact rõ — đề xuất rename → {resolved}."
                gold_ready = True
                needs_cluster_rename = True
        elif clarity >= 4 and not _listing_noise(d):
            gold_decision = "gold_revise_hold"
            gold_reason = "unknown_cluster_needs_review"
            gold_notes = "FC_UNKNOWN — cần human gán cluster + chỉnh disclosure."
        else:
            gold_decision = "gold_reject"
            gold_reason = "unknown_cluster_too_vague"
            gold_notes = "FC_UNKNOWN và fact không đủ rõ cho gold."

    # Known cluster path
    elif gold_decision == "gold_revise_hold":  # not rejected yet
        if target in KNOWN_CLUSTERS:
            if (
                canon_dec == "canonical_keep"
                and clarity >= 6
                and not _news_chrome(d)
                and not _is_generic(q)
                and not needs_prohib_refine
            ):
                gold_decision = "gold_approve"
                gold_reason = "known_cluster_clean_anchor"
                gold_notes = "Cluster rõ, Q/disclosure sạch — gold approve."
                gold_ready = True
            elif _news_chrome(d) or len(d) > 400:
                gold_decision = "gold_revise_hold"
                gold_reason = "known_cluster_news_chrome"
                gold_notes = "Fact đúng cluster nhưng disclosure còn press chrome — trim trước gold."
            elif canon_dec == "canonical_keep_after_merge" or needs_prohib_refine:
                gold_decision = "gold_revise_hold"
                gold_reason = "needs_wording_or_prohibited_refine"
                gold_notes = "Cần refine wording/prohibited claims trước promote."
            elif clarity >= 5 and not _is_generic(q):
                gold_decision = "gold_approve"
                gold_reason = "known_cluster_usable"
                gold_notes = "Cluster known, fact usable."
                gold_ready = True
            else:
                gold_decision = "gold_revise_hold"
                gold_reason = "known_cluster_borderline"
                gold_notes = "Borderline — reviewer spot-check."
        else:
            gold_decision = "gold_revise_hold"
            gold_reason = "nonstandard_cluster"
            gold_notes = f"Cluster {target} — hold để đặt tên chuẩn."

    out.update(
        {
            "gold_version": GOLD_VERSION,
            "gold_decision": gold_decision,
            "gold_reason": gold_reason,
            "gold_notes": gold_notes,
            "gold_ready_flag": gold_ready,
            "needs_cluster_rename": needs_cluster_rename and gold_decision != "gold_reject",
            "needs_prohibited_claims_refine": needs_prohib_refine and gold_decision == "gold_revise_hold",
            "resolved_cluster_hint": resolved or target,
        }
    )
    return out


def write_workbook(
    approved: Sequence[Dict[str, Any]],
    hold: Sequence[Dict[str, Any]],
    rejected: Sequence[Dict[str, Any]],
    path: Path,
) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide.append(["Gold Decision Round 4 — Core HS+MS"])
    guide.append(["RX", "Backlog only — không trong approved core"])

    sheets = [
        ("Approved_Core", approved, "166534"),
        ("Revise_Hold", hold, "1D4ED8"),
        ("Rejected", rejected, "991B1B"),
    ]
    for name, rows, color in sheets:
        ws = wb.create_sheet(name)
        ws.append(WORKBOOK_COLS)
        fill = PatternFill("solid", fgColor=color)
        font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(WORKBOOK_COLS) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = fill
            c.font = font
        for r in sorted(rows, key=lambda x: (x.get("company", ""), x.get("seed_id", ""))):
            ws.append([r.get(c, "") for c in WORKBOOK_COLS])

    ws_rx = wb.create_sheet("RX_Backlog_Status")
    ws_rx.append(["status", "detail"])
    ws_rx.append(["rx_backlog", "source-acquisition dependent — không trong gold core round 4"])
    ws_rx.append(["rx_survivor", "1 row từ manual review — giữ riêng"])
    ws_rx.append(["lane_c", "22 row backlog — chưa xử lý"])
    ws_rx.append(["action", "Acquire SR/Impact PDF trước khi mở RX gold lane"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    ex: Dict[str, List[Dict[str, str]]] = {"approve": [], "hold": [], "reject": []}
    for dec in ("gold_approve", "gold_revise_hold", "gold_reject"):
        key = dec.replace("gold_", "").replace("_", "")
        if dec == "gold_revise_hold":
            key = "hold"
        for r in rows:
            if r.get("gold_decision") != dec:
                continue
            bucket = {"gold_approve": "approve", "gold_revise_hold": "hold", "gold_reject": "reject"}[dec]
            if len(ex[bucket]) >= 3:
                break
            ex[bucket].append(
                {
                    "seed_id": r.get("seed_id", ""),
                    "company": r.get("company", ""),
                    "reason": r.get("gold_reason", ""),
                    "target": r.get("canonical_fact_target", ""),
                }
            )
    return ex


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Gold Decision Core Round 4",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Gold decision round cho core **한샘 + 무신사**: approve / revise_hold / reject — không promote hàng loạt.",
        "",
        "## Vì sao mở gold decision round lúc này",
        "",
        "- Core canonical 45 row (HS 36, MS 9) đủ chặt sau merge.",
        "- Cần tách row gold-ready vs hold vs reject trước promote.",
        "- `FC_UNKNOWN=15` yêu cầu siết riêng — không auto-approve.",
        "",
        "## Rule approve / hold / reject",
        "",
        "| Decision | Điều kiện |",
        "|----------|-----------|",
        "| `gold_approve` | Cluster rõ (hoặc unknown đã resolve), Q/disclosure sạch |",
        "| `gold_revise_hold` | Fact đúng nhưng wording/cluster/prohibited cần sửa |",
        "| `gold_reject` | Listing/meta, truncated, unknown quá mơ hồ |",
        "",
        "## Rule riêng cho `FC_UNKNOWN`",
        "",
        "- Không auto-approve.",
        "- Approve chỉ khi `final_question + final_disclosure` xác định fact rõ (clarity≥6, không listing).",
        "- Còn lại: hold hoặc reject.",
        "",
        "## Kết quả tổng quan",
        "",
        f"- Input: **{summary.get('input_core_count', 0)}**",
        f"- approve: **{summary.get('gold_approve_count', 0)}**",
        f"- revise_hold: **{summary.get('gold_revise_hold_count', 0)}**",
        f"- reject: **{summary.get('gold_reject_count', 0)}**",
        f"- **Core gold ready (approve):** **{summary.get('core_gold_ready_count', 0)}**",
        "",
        "### Breakdown theo công ty",
        "",
    ]
    for co, stats in summary.get("by_company", {}).items():
        lines.append(
            f"- **{co}**: approve={stats.get('gold_approve', 0)}, "
            f"hold={stats.get('gold_revise_hold', 0)}, reject={stats.get('gold_reject', 0)}"
        )

    lines.extend(["", "### Breakdown theo cluster", ""])
    for cid, n in summary.get("by_cluster", {}).items():
        lines.append(f"- `{cid}`: {n}")

    fc = summary.get("fc_unknown_breakdown", {})
    lines.extend(
        [
            "",
            "### FC_UNKNOWN breakdown",
            "",
            f"- approve: **{fc.get('gold_approve', 0)}**",
            f"- hold: **{fc.get('gold_revise_hold', 0)}**",
            f"- reject: **{fc.get('gold_reject', 0)}**",
            f"- **Chưa chốt (hold+reject):** **{fc.get('unresolved', 0)}**",
            "",
            "## Ví dụ",
            "",
        ]
    )
    for title, key in (("Approve mạnh", "approve"), ("Hold wording/cluster", "hold"), ("Reject", "reject")):
        lines.append(f"### {title}")
        for e in examples.get(key, []):
            lines.append(f"- `{e['seed_id']}` ({e['company']}, {e['target']}): `{e['reason']}`")
        lines.append("")

    lines.extend(
        [
            "## Kết luận",
            "",
            f"- Core gold ready: **{summary.get('core_gold_ready_count', 0)}** (HS {summary.get('hs_gold_ready', 0)}, MS {summary.get('ms_gold_ready', 0)})",
            f"- RX status: **{summary.get('rx_backlog_status', '')}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_gold_decision(
    *,
    input_path: Path,
    approved_path: Path,
    hold_path: Path,
    rejected_path: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    core = [r for r in rows if r.get("company") in CORE_COMPANIES]

    decided = [decide_gold(r) for r in core]
    approved = [r for r in decided if r.get("gold_decision") == "gold_approve"]
    hold = [r for r in decided if r.get("gold_decision") == "gold_revise_hold"]
    rejected = [r for r in decided if r.get("gold_decision") == "gold_reject"]

    write_jsonl(approved_path, approved)
    write_jsonl(hold_path, hold)
    write_jsonl(rejected_path, rejected)
    write_workbook(approved, hold, rejected, workbook_path)

    fc_rows = [r for r in decided if r.get("canonical_fact_target") == "FC_UNKNOWN"]
    fc_bd = dict(Counter(r.get("gold_decision") for r in fc_rows))
    fc_bd["unresolved"] = fc_bd.get("gold_revise_hold", 0) + fc_bd.get("gold_reject", 0)

    by_co: Dict[str, Dict[str, int]] = {}
    for co in sorted(CORE_COMPANIES):
        sub = [r for r in decided if r.get("company") == co]
        by_co[co] = dict(Counter(r.get("gold_decision") for r in sub))

    by_cluster = dict(Counter(r.get("canonical_fact_target") for r in approved))

    hs_gold = sum(1 for r in approved if r.get("company") == "한샘")
    ms_gold = sum(1 for r in approved if r.get("company") == "무신사")

    summary = {
        "gold_version": GOLD_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_core_count": len(core),
        "gold_approve_count": len(approved),
        "gold_revise_hold_count": len(hold),
        "gold_reject_count": len(rejected),
        "core_gold_ready_count": len(approved),
        "hs_gold_ready": hs_gold,
        "ms_gold_ready": ms_gold,
        "by_company": by_co,
        "by_cluster": by_cluster,
        "fc_unknown_breakdown": fc_bd,
        "rx_backlog_status": "source-acquisition dependent — 1 survivor + 22 lane C + reject rows; không trong gold core",
        "output_approved": str(approved_path),
        "output_hold": str(hold_path),
        "output_rejected": str(rejected_path),
        "output_workbook": str(workbook_path),
    }

    examples = _pick_examples(decided)
    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Gold decision round 4 — core HS+MS")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_core_canonical_round3.jsonl",
    )
    parser.add_argument(
        "--approved",
        default="data/golden_set/v2/step6_gold/golden_set_core_round4_approved.jsonl",
    )
    parser.add_argument(
        "--hold",
        default="data/golden_set/v2/step6_gold/golden_set_core_round4_hold.jsonl",
    )
    parser.add_argument(
        "--rejected",
        default="data/golden_set/v2/step6_gold/golden_set_core_round4_rejected.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/step6_gold/golden_set_core_round4_decision.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_gold_decision_core_round4.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_gold_decision_core_round4_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_gold_decision(
        input_path=root / args.input,
        approved_path=root / args.approved,
        hold_path=root / args.hold,
        rejected_path=root / args.rejected,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "input_core_count",
                    "gold_approve_count",
                    "gold_revise_hold_count",
                    "gold_reject_count",
                    "core_gold_ready_count",
                    "fc_unknown_breakdown",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
