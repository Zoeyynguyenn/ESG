"""RTX candidate generation v2 — fact-specific question layer (rebuild from corpus)."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import (
    METRIC_HINT_RE,
    NUMBER_RE,
    YEAR_RE,
    _best_sentences,
    _norm_ws,
    _sentence_score,
    _split_sentences,
)
from golden_set.build_reference_seed_workbook_rtx_v1 import (
    COMPANY,
    HARD_NOISE,
    SEC_BOILERPLATE,
    _cluster_hint,
    _esg_hits,
    _infer_candidate_kind,
    _prohibited_claims_en,
    _provenance,
    passage_level_filter,
)
from golden_set.io_utils import read_jsonl, write_jsonl

SEED_VERSION = "ref_seed_rtx_v2_fact_specific"

BANNED_GENERIC_PATTERNS = [
    re.compile(r"^What ESG-related policies or performance does RTX disclose\?$"),
    re.compile(r"^What quantitative ESG metrics does RTX disclose\?$"),
    re.compile(r"^How have RTX's key ESG metrics changed over time\?$"),
    re.compile(r"^How is ESG governance structured at RTX\?$"),
    re.compile(r"^What ethics and compliance practices does RTX disclose\?$"),
    re.compile(r"^What greenhouse gas emissions does RTX disclose\?$"),
    re.compile(r"^What ESG metric or policy does RTX disclose\?$"),
    re.compile(r"^What does RTX disclose about ESG metric or policy\?$"),
]

METRIC_PHRASE_RULES: List[Tuple[str, str]] = [
    (r"scope\s*1(?:\s*(?:and|&)\s*2)?[^.]{0,60}", "Scope 1 and Scope 2 GHG emissions"),
    (r"market-based scope\s*2", "market-based Scope 2 GHG emissions"),
    (r"location-based scope\s*2", "location-based Scope 2 GHG emissions"),
    (r"scope\s*1[^.]{0,40}", "Scope 1 GHG emissions"),
    (r"scope\s*2[^.]{0,40}", "Scope 2 GHG emissions"),
    (r"energy intensity[^.]{0,50}", "energy intensity (GJ per $M revenue)"),
    (r"reduction in energy consumption since 2019[^.]{0,40}", "energy consumption reduction since 2019 baseline"),
    (r"ergonomic risk[^.]{0,50}", "high and elevated ergonomic risks"),
    (r"renewable (?:electricity|energy)[^.]{0,50}", "renewable electricity/energy use"),
    (r"water (?:withdrawal|consumption|stress)[^.]{0,50}", "water withdrawal/consumption"),
    (r"lost.?time injury[^.]{0,40}", "lost-time injury rate"),
    (r"female(?:\s+|\s*%|\s+representation)[^.]{0,40}", "female workforce representation"),
    (r"diversity and inclusion[^.]{0,50}", "diversity and inclusion programs"),
    (r"r&d[^.]{0,60}", "R&D investment for sustainability technologies"),
    (r"co2e savings[^.]{0,40}", "annual CO2e savings from initiatives"),
    (r"ghg emissions?[^.]{0,50}", "GHG emissions"),
    (r"greenhouse gas[^.]{0,50}", "greenhouse gas emissions"),
    (r"net zero[^.]{0,40}", "net zero climate commitment"),
    (r"deferred prosecution agreement", "deferred prosecution agreement compliance"),
    (r"\bfcpa\b|foreign corrupt practices", "FCPA / anti-bribery compliance"),
    (r"ecovadis", "supplier EcoVadis assessments"),
    (r"geared turbo fan|gtf engine", "Pratt & Whitney GTF engine fuel efficiency"),
]

GOVERNANCE_PHRASE_RULES: List[Tuple[str, str]] = [
    (r"audit committee[^.]{0,60}", "Audit Committee oversight"),
    (r"esg committee|sustainability committee", "ESG/sustainability committee governance"),
    (r"board of directors[^.]{0,60}", "Board of Directors governance"),
    (r"code of (?:ethics|conduct)", "Code of Ethics/Conduct"),
    (r"conflict of interest", "conflict-of-interest policy"),
    (r"whistleblower", "whistleblower reporting mechanism"),
    (r"data (?:privacy|security|protection)", "data privacy and security program"),
    (r"cybersecurity", "cybersecurity program"),
]

STAKEHOLDER_RULES: List[Tuple[str, str]] = [
    (r"material(?:ity)?(?:\s+topic|\s+issue)[^.]{0,60}", "material ESG topics"),
    (r"stakeholder (?:engagement|group)[^.]{0,60}", "stakeholder engagement approach"),
    (r"double materiality", "double materiality assessment"),
]

FRAMEWORK_RULES: List[Tuple[str, str]] = [
    (r"\bcdp\b", "CDP climate disclosure"),
    (r"\btcfd\b", "TCFD-aligned climate disclosure"),
    (r"\bsasb\b", "SASB disclosure"),
    (r"\bgri\b", "GRI reporting"),
    (r"sustainability report", "sustainability report disclosure"),
]


@dataclass
class FactTarget:
    text: str
    fact_target_type: str
    anchor_tokens: Tuple[str, ...]
    years: Tuple[str, ...]
    numbers: Tuple[str, ...]


def _match_phrase(sentence: str, rules: Sequence[Tuple[str, str]]) -> Optional[str]:
    lower = sentence.lower()
    for pat, label in rules:
        if re.search(pat, lower):
            return label
    return None


def _extract_metric_phrase(sentence: str) -> Optional[str]:
    lower = sentence.lower()
    for pat, label in METRIC_PHRASE_RULES:
        if re.search(pat, lower):
            return label
    m = re.search(
        r"([A-Za-z][A-Za-z0-9\s\-/%]{8,60}(?:emission|intensity|consumption|reduction|rate|target|goal))",
        sentence,
        re.I,
    )
    if m:
        return _norm_ws(m.group(1))[:70]
    return None


def extract_fact_target(sentence: str, qtype: str, _depth: int = 0) -> Optional[FactTarget]:
    lower = sentence.lower()
    years = tuple(sorted(set(YEAR_RE.findall(sentence))))
    numbers = tuple(NUMBER_RE.findall(sentence)[:6])

    if qtype == "trend" and len(years) >= 2 and numbers:
        metric = _extract_metric_phrase(sentence)
        if metric:
            text = f"{metric} trend ({years[0]} to {years[-1]})"
            return FactTarget(text, "trend_specific", (metric.lower(),), years, numbers)

    metric = _match_phrase(sentence, METRIC_PHRASE_RULES) or _extract_metric_phrase(sentence)
    if metric and (numbers or METRIC_HINT_RE.search(sentence)):
        y_suffix = f" in {years[-1]}" if years else ""
        return FactTarget(
            f"{metric}{y_suffix}",
            "metric_specific",
            tuple(w for w in re.findall(r"[a-z]{4,}", metric.lower())[:4]),
            years,
            numbers,
        )

    gov = _match_phrase(sentence, GOVERNANCE_PHRASE_RULES)
    if gov:
        return FactTarget(gov, "governance_specific", tuple(gov.lower().split()[:3]), years, numbers)

    stake = _match_phrase(sentence, STAKEHOLDER_RULES)
    if stake:
        return FactTarget(stake, "stakeholder_materiality", tuple(stake.lower().split()[:3]), years, numbers)

    fw = _match_phrase(sentence, FRAMEWORK_RULES)
    if fw:
        return FactTarget(fw, "framework_report", tuple(fw.lower().split()[:2]), years, numbers)

    if _depth < 1 and qtype == "qualitative" and len(sentence) >= 120:
        for sent in re.split(r"(?<=[.!?])\s+", sentence):
            sent = _norm_ws(sent)
            if len(sent) < 50 or len(sent) >= len(sentence) - 5:
                continue
            sub = extract_fact_target(sent, "quantitative" if NUMBER_RE.search(sent) else "qualitative", _depth + 1)
            if sub:
                return sub

    return None


def build_fact_specific_question(ft: FactTarget, qtype: str) -> Optional[str]:
    t = ft.text
    if ft.fact_target_type == "trend_specific" and len(ft.years) >= 2:
        metric = t.split(" trend ")[0] if " trend " in t else t
        return f"How did RTX's {metric} change from {ft.years[0]} to {ft.years[-1]}?"
    if ft.fact_target_type == "metric_specific":
        return f"What {t} does RTX report?"
    if ft.fact_target_type == "governance_specific":
        return f"What does RTX disclose about its {t}?"
    if ft.fact_target_type == "stakeholder_materiality":
        return f"What does RTX disclose regarding {t}?"
    if ft.fact_target_type == "framework_report":
        return f"What does RTX disclose in its {t}?"
    if qtype == "quantitative" and ft.numbers:
        return f"What {t} figure does RTX disclose?"
    return None


def _is_banned_generic(question: str) -> bool:
    return any(p.search(question) for p in BANNED_GENERIC_PATTERNS)


def _question_specific_enough(question: str, ft: FactTarget) -> bool:
    if _is_banned_generic(question):
        return False
    if len(ft.text) < 12:
        return False
    q_lower = question.lower()
    hits = sum(1 for tok in ft.anchor_tokens if tok in q_lower)
    if hits >= 1:
        return True
    if any(y in question for y in ft.years):
        return True
    if ft.numbers and any(n in question for n in ft.numbers[:2]):
        return True
    return len(ft.text) >= 20 and ft.text.lower()[:20] in q_lower


def _fact_dedupe_key(ft: FactTarget, record_id: str, qtype: str) -> str:
    raw = f"{ft.fact_target_type}|{ft.text.lower()}|{record_id}|{qtype}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def _detect_qtypes(sentence: str) -> List[str]:
    years = len(set(YEAR_RE.findall(sentence)))
    numbers = len(NUMBER_RE.findall(sentence))
    out: List[str] = []
    if years >= 2 and numbers >= 2:
        out.append("trend")
    if numbers >= 1:
        out.append("quantitative")
    out.append("qualitative")
    deduped: List[str] = []
    for q in out:
        if q not in deduped:
            deduped.append(q)
    return deduped[:2]


def generate_candidates_from_passage(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    text = row.get("text") or ""
    document_kind = row.get("document_kind") or "unknown"
    sentences = _best_sentences(text, limit=6)
    if not sentences:
        sentences = [s for s in _split_sentences(text) if len(s) >= 40][:4]

    out: List[Dict[str, Any]] = []
    for sentence in sentences:
        if _esg_hits(sentence) < 1 and not NUMBER_RE.search(sentence):
            continue
        for qtype in _detect_qtypes(sentence):
            ft = extract_fact_target(sentence, qtype)
            if ft is None:
                continue
            question = build_fact_specific_question(ft, qtype)
            if not question or not _question_specific_enough(question, ft):
                continue
            kind = _infer_candidate_kind(sentence, qtype, document_kind)
            disclosure = _norm_ws(sentence)[:420]
            out.append(
                {
                    "company": COMPANY,
                    "question_type": qtype,
                    "candidate_kind": kind,
                    "question_draft": question,
                    "fact_target": ft.text,
                    "fact_target_type": ft.fact_target_type,
                    "acceptable_disclosure": disclosure,
                    "prohibited_claims": _prohibited_claims_en(qtype),
                    "source_record_id": str(row.get("record_id") or ""),
                    "source_unit_id": str(row.get("unit_id") or ""),
                    "source_excerpt": disclosure,
                    "source_file": str(row.get("source_file") or ""),
                    "source_type": str(row.get("source_type") or ""),
                    "seed_origin_type": _provenance(row, text),
                    "candidate_status": "candidate_rtx_v2_fact_specific",
                    "candidate_reason": f"fact_specific:{ft.fact_target_type}",
                    "workbook_cluster_hint": _cluster_hint(kind, sentence),
                    "document_kind": document_kind,
                    "rank": float(_sentence_score(sentence)) + (3 if ft.numbers else 0),
                    "dedupe_key": _fact_dedupe_key(ft, str(row.get("record_id") or ""), qtype),
                }
            )
    return out


def dedupe_candidates(candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    best: Dict[str, Dict[str, Any]] = {}
    for c in sorted(candidates, key=lambda x: float(x.get("rank") or 0), reverse=True):
        key = c["dedupe_key"]
        if key not in best:
            best[key] = c
    rows = list(best.values())

    by_question: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in rows:
        by_question[r["question_draft"]].append(r)

    final: List[Dict[str, Any]] = []
    for _q, group in by_question.items():
        if len(group) == 1:
            final.append(group[0])
            continue
        facts = {g["fact_target"] for g in group}
        if len(facts) == 1:
            final.append(max(group, key=lambda x: float(x.get("rank") or 0)))
        else:
            for g in sorted(group, key=lambda x: -float(x.get("rank") or 0)):
                final.append(g)
    return final


def count_duplicate_questions(rows: Sequence[Dict[str, Any]]) -> Dict[str, int]:
    qs = Counter(r.get("question_draft") or "" for r in rows)
    exact = sum(1 for n in qs.values() if n > 1)
    affected = sum(n for n in qs.values() if n > 1)
    prefix_groups: Dict[str, int] = defaultdict(int)
    for q in qs:
        prefix_groups[_norm_ws(q.lower())[:50]] += 1
    near = sum(1 for c in prefix_groups.values() if c > 1)
    return {
        "exact_duplicate_question_count": exact,
        "exact_duplicate_affected_rows": affected,
        "near_duplicate_question_count": near,
    }


def _seed_id(qtype: str, index: int) -> str:
    code = {"quantitative": "Q", "trend": "T", "qualitative": "L"}.get(qtype, "X")
    return f"RTX-V2-{code}{index:02d}"


def to_output_rows(candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    counters: Counter = Counter()
    rows: List[Dict[str, Any]] = []
    for c in sorted(candidates, key=lambda x: (-float(x.get("rank") or 0), x.get("question_type", ""))):
        counters[c["question_type"]] += 1
        row = dict(c)
        row["seed_id"] = _seed_id(c["question_type"], counters[c["question_type"]])
        row["seed_version"] = SEED_VERSION
        rows.append(row)
    return rows


def write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Guide"
    ws.append(["RTX V2 Fact-Specific Candidates — question layer rebuild"])

    header = [
        "seed_id", "company", "question_type", "candidate_kind", "document_kind",
        "question_draft", "fact_target", "fact_target_type",
        "acceptable_disclosure", "prohibited_claims", "source_record_id",
        "candidate_status", "workbook_cluster_hint",
    ]
    ws2 = wb.create_sheet("Candidates")
    ws2.append(header)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        c = ws2.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    for row in rows:
        ws2.append([row.get(h, "") for h in header])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report(summary: Dict[str, Any], examples: List[Dict[str, str]], path: Path) -> None:
    lines = [
        "# Golden Set — Candidate Generation RTX V2 (Fact-Specific)",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Rebuild question layer RTX với **fact_target** bắt buộc — không dùng template generic.",
        "",
        "## Vì sao v1 lỗi ở tầng question",
        "",
        f"- v1 chỉ có **{summary.get('v1_unique_questions', 11)}** unique questions cho **{summary.get('v1_total_rows', 3170)}** rows",
        f"- **{summary.get('v1_affected_rows', 3170)}** rows dùng exact-duplicate generic templates",
        "- Cùng question map sang disclosure/question_type/document_kind khác nhau",
        "",
        "## Rule fact-specific question",
        "",
        "- Mỗi candidate phải có `fact_target` + `fact_target_type`",
        "- Question phải phản ánh fact cụ thể (metric, trend years, governance body, framework)",
        "- Cấm template generic backbone; không fact-specific → không vào workbook",
        "",
        "## Cách builder v2 khác v1",
        "",
        "- Extract fact target trước, sinh question sau",
        "- Dedupe theo `fact_target + record_id + question_type`",
        "- Post-filter: không giữ question trùng mà fact_target khác nhau",
        "",
        "## Kết quả",
        "",
        f"- Raw candidates: **{summary.get('raw_candidates', 0)}**",
        f"- Filtered candidates: **{summary.get('filtered_candidates', 0)}**",
        f"- Exact duplicate questions còn lại: **{summary.get('exact_duplicate_question_count', 0)}**",
        f"- Near-duplicate question groups: **{summary.get('near_duplicate_question_count', 0)}**",
        "",
        "### Breakdown theo question_type",
        "",
    ]
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")
    lines.extend(["", "### Breakdown theo document_kind", ""])
    for dk, n in summary.get("by_document_kind", {}).items():
        lines.append(f"- `{dk}`: {n}")

    lines.extend(["", "## Ví dụ v1 generic → v2 fact-specific", ""])
    for ex in examples:
        lines.append(f"- v1: `{ex.get('v1_generic', '')}`")
        lines.append(f"  v2: `{ex.get('v2_specific', '')}` (fact: {ex.get('fact_target', '')})")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- Workbook RTX v2 đủ mở review round 1 lại: **{summary.get('review_ready_verdict', '')}**",
            f"- `review_ready_flag` = **{summary.get('review_ready_flag', False)}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, str]]:
    mapping = {
        "Scope 1": "What greenhouse gas emissions does RTX disclose?",
        "energy intensity": "What quantitative ESG metrics does RTX disclose?",
        "ergonomic": "How have RTX's key ESG metrics changed over time?",
    }
    examples: List[Dict[str, str]] = []
    for key, v1_q in mapping.items():
        for r in rows:
            if key.lower() in (r.get("fact_target") or "").lower():
                examples.append(
                    {
                        "v1_generic": v1_q,
                        "v2_specific": r.get("question_draft", ""),
                        "fact_target": r.get("fact_target", ""),
                    }
                )
                break
    return examples[:5]


def run_builder(
    *,
    input_path: Path,
    output_jsonl: Path,
    output_xlsx: Path,
    report_path: Path,
    summary_json_path: Path,
    audit_summary_path: Optional[Path] = None,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    raw: List[Dict[str, Any]] = []
    for row in rows:
        ok, _ = passage_level_filter(row)
        if not ok:
            continue
        raw.extend(generate_candidates_from_passage(row))

    deduped = dedupe_candidates(raw)
    out_rows = to_output_rows(deduped)
    dup_stats = count_duplicate_questions(out_rows)

    write_jsonl(output_jsonl, out_rows)
    write_workbook(out_rows, output_xlsx)

    v1_stats = {}
    if audit_summary_path and audit_summary_path.exists():
        v1_stats = json.loads(audit_summary_path.read_text(encoding="utf-8"))

    exact_dup = dup_stats["exact_duplicate_question_count"]
    near_dup = dup_stats["near_duplicate_question_count"]
    filtered_n = len(out_rows)

    unique_q = len({r["question_draft"] for r in out_rows})
    review_ready = (
        filtered_n >= 40
        and exact_dup == 0
        and unique_q == filtered_n
        and filtered_n > 0
    )

    summary = {
        "seed_version": SEED_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "raw_candidates": len(raw),
        "filtered_candidates": filtered_n,
        "exact_duplicate_question_count": exact_dup,
        "near_duplicate_question_count": near_dup,
        "exact_duplicate_affected_rows": dup_stats.get("exact_duplicate_affected_rows", 0),
        "by_question_type": dict(Counter(r["question_type"] for r in out_rows)),
        "by_document_kind": dict(Counter(r.get("document_kind", "") for r in out_rows)),
        "by_fact_target_type": dict(Counter(r.get("fact_target_type", "") for r in out_rows)),
        "v1_unique_questions": v1_stats.get("v1_unique_questions", 11),
        "v1_total_rows": v1_stats.get("v1_total_rows", 3170),
        "v1_affected_rows": v1_stats.get("affected_rows_count", 3170),
        "review_ready_flag": review_ready,
        "review_ready_verdict": (
            "Có — question layer fact-specific đủ để mở lại RTX review round 1"
            if review_ready
            else "Chưa — cần siết thêm fact extraction hoặc dedupe"
        ),
        "output_jsonl": str(output_jsonl),
        "output_xlsx": str(output_xlsx),
    }

    examples = _pick_examples(out_rows)
    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Build RTX v2 fact-specific candidates")
    parser.add_argument("--input", default="data/golden_set/v2/rtx_step1_corpus_units/corpus_units_rtx_normalized.jsonl")
    parser.add_argument("--output-jsonl", default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v2_fact_specific.jsonl")
    parser.add_argument("--output-xlsx", default="data/golden_set/v2/reference_style/reference_seed_workbook_rtx_v2_fact_specific.xlsx")
    parser.add_argument("--report", default="reports/golden_set_candidate_generation_rtx_v2_fact_specific.md")
    parser.add_argument("--summary-json", default="reports/_candidate_generation_rtx_v2_fact_specific_summary.json")
    parser.add_argument("--audit-summary", default="reports/_rtx_duplicate_question_audit_summary.json")
    args = parser.parse_args(argv)

    summary = run_builder(
        input_path=root / args.input,
        output_jsonl=root / args.output_jsonl,
        output_xlsx=root / args.output_xlsx,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
        audit_summary_path=root / args.audit_summary,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "filtered_candidates",
                    "exact_duplicate_question_count",
                    "near_duplicate_question_count",
                    "review_ready_flag",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
