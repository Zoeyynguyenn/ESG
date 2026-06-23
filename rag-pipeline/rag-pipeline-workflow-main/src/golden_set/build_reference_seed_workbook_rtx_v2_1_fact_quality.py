"""RTX candidate generation v2.1 — canonical facts + quality gates."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import (
    NUMBER_RE,
    YEAR_RE,
    _best_sentences,
    _norm_ws,
    _sentence_score,
    _split_sentences,
)
from golden_set.build_reference_seed_workbook_rtx_v1 import (
    COMPANY,
    _cluster_hint,
    _esg_hits,
    _infer_candidate_kind,
    _prohibited_claims_en,
    _provenance,
    passage_level_filter,
)
from golden_set.io_utils import read_jsonl, write_jsonl
from golden_set.rtx_fact_quality import (
    CanonicalFact,
    audit_candidate_row,
    build_natural_question,
    match_canonical_fact,
    passes_quality_gates,
)

SEED_VERSION = "ref_seed_rtx_v2_1_fact_quality"


def _detect_qtypes(sentence: str) -> List[str]:
    years = len(set(YEAR_RE.findall(sentence)))
    numbers = len(NUMBER_RE.findall(sentence))
    out: List[str] = []
    if years >= 2 and numbers >= 2:
        out.append("trend")
    if numbers >= 1:
        out.append("quantitative")
    if not out:
        out.append("qualitative")
    return out[:2]


def _extract_disclosure(sentence: str) -> str:
    s = _norm_ws(sentence)
    if s.count("|") >= 4:
        best = ""
        for part in re.split(r"\|", s):
            part = _norm_ws(part)
            if len(part) < 30:
                continue
            if len(NUMBER_RE.findall(part)) >= len(NUMBER_RE.findall(best)):
                best = part
        if len(best) >= 30:
            return best[:420]
    return s[:420]


def _fact_dedupe_key(cf: CanonicalFact, record_id: str, qtype: str, year: str) -> str:
    raw = f"{cf.fact_id}|{record_id}|{qtype}|{year}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def generate_candidates_from_passage(row: Dict[str, Any]) -> List[Dict[str, Any]]:
    text = row.get("text") or ""
    document_kind = row.get("document_kind") or "unknown"
    sentences = _best_sentences(text, limit=5)
    if not sentences:
        sentences = [s for s in _split_sentences(text) if len(s) >= 40][:3]

    out: List[Dict[str, Any]] = []
    for sentence in sentences:
        if _esg_hits(sentence) < 1 and not NUMBER_RE.search(sentence):
            continue
        cf = match_canonical_fact(sentence)
        if cf is None:
            continue
        disclosure = _extract_disclosure(sentence)
        if len(disclosure) < 35:
            continue

        for qtype in _detect_qtypes(sentence):
            question = build_natural_question(
                cf, qtype, sentence, document_kind=document_kind, disclosure=disclosure
            )
            from golden_set.rtx_fact_quality import _valid_years

            years = _valid_years(sentence)
            fact_target = cf.fact_target
            if years:
                fact_target = f"{cf.fact_target} ({years[-1]})"
            elif document_kind:
                fact_target = f"{cf.fact_target} [{document_kind}]"

            ok, reason = passes_quality_gates(question, fact_target, disclosure, cf)
            if not ok:
                continue

            kind = _infer_candidate_kind(sentence, qtype, document_kind)
            out.append(
                {
                    "company": COMPANY,
                    "question_type": qtype,
                    "candidate_kind": kind,
                    "question_draft": question,
                    "fact_target": fact_target,
                    "fact_target_type": cf.fact_target_type,
                    "acceptable_disclosure": disclosure,
                    "prohibited_claims": _prohibited_claims_en(qtype),
                    "source_record_id": str(row.get("record_id") or ""),
                    "source_unit_id": str(row.get("unit_id") or ""),
                    "source_excerpt": disclosure,
                    "source_file": str(row.get("source_file") or ""),
                    "source_type": str(row.get("source_type") or ""),
                    "seed_origin_type": _provenance(row, text),
                    "candidate_status": "candidate_rtx_v2_1",
                    "candidate_reason": f"canonical_fact:{cf.fact_id}",
                    "workbook_cluster_hint": _cluster_hint(kind, sentence),
                    "document_kind": document_kind,
                    "question_quality_status": "usable",
                    "rank": float(_sentence_score(sentence)) + 3,
                    "dedupe_key": _fact_dedupe_key(cf, str(row.get("record_id") or ""), qtype, years[-1] if years else ""),
                }
            )
    return out


def dedupe_candidates(candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    best: Dict[str, Dict[str, Any]] = {}
    for c in sorted(candidates, key=lambda x: float(x.get("rank") or 0), reverse=True):
        key = c["dedupe_key"]
        if key not in best:
            best[key] = c
    rows = list(best.values())

    by_question: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        by_question.setdefault(r["question_draft"], []).append(r)

    final: List[Dict[str, Any]] = []
    for q, group in by_question.items():
        if len(group) == 1:
            final.append(group[0])
            continue
        facts = {g["fact_target"] for g in group}
        if len(facts) == len(group):
            final.extend(group)
        else:
            final.append(max(group, key=lambda x: float(x.get("rank") or 0)))
    return final


def _seed_id(qtype: str, index: int) -> str:
    code = {"quantitative": "Q", "trend": "T", "qualitative": "L"}.get(qtype, "X")
    return f"RTX-V21-{code}{index:02d}"


def to_output_rows(candidates: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    counters: Counter = Counter()
    rows: List[Dict[str, Any]] = []
    for c in sorted(candidates, key=lambda x: (-float(x.get("rank") or 0), x.get("question_type", ""))):
        counters[c["question_type"]] += 1
        row = dict(c)
        row["seed_id"] = _seed_id(c["question_type"], counters[c["question_type"]])
        row["seed_version"] = SEED_VERSION
        rows.append(row)
    return rows


def write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Guide"
    ws.append(["RTX V2.1 Fact-Quality Candidates"])
    header = [
        "seed_id", "company", "question_type", "document_kind", "question_draft",
        "fact_target", "fact_target_type", "acceptable_disclosure",
        "question_quality_status", "source_record_id", "workbook_cluster_hint",
    ]
    ws2 = wb.create_sheet("Candidates")
    ws2.append(header)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        c = ws2.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    for row in rows:
        ws2.append([row.get(h, "") for h in header])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report(summary: Dict[str, Any], examples: List[Dict[str, str]], path: Path) -> None:
    lines = [
        "# Golden Set — Candidate Generation RTX V2.1 (Fact Quality)",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Rebuild question layer với canonical facts + quality gates; ưu tiên usability + alignment.",
        "",
        "## Audit lỗi v2 cho thấy gì",
        "",
        f"- v2 input: **{summary.get('v2_input_count', 327)}**; needs rebuild: **{summary.get('v2_needs_rebuild', 0)}**",
        f"- Lỗi chính: fact mismatch, unnatural wording, residue-led questions, overlong phrases",
        "",
        "## Rule rebuild v2.1",
        "",
        "- Chỉ dùng **canonical fact catalog** với câu hỏi tự nhiên định sẵn",
        "- `passes_quality_gates` trước khi giữ row",
        "- Drop nếu question/disclosure không align",
        "",
        "## Kết quả",
        "",
        f"- Raw candidates: **{summary.get('raw_candidates', 0)}**",
        f"- Filtered candidates: **{summary.get('filtered_candidates', 0)}**",
        f"- Usable count: **{summary.get('usable_count', 0)}**",
        f"- Dropped fact mismatch (est.): **{summary.get('dropped_fact_mismatch_count', 0)}**",
        f"- Dropped wording/residue (est.): **{summary.get('dropped_wording_count', 0)}**",
        "",
        "### Breakdown theo question_type",
        "",
    ]
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")
    lines.extend(["", "### Breakdown theo document_kind", ""])
    for dk, n in summary.get("by_document_kind", {}).items():
        lines.append(f"- `{dk}`: {n}")

    lines.extend(["", "## Ví dụ v2 lỗi → v2.1 sửa", ""])
    for ex in examples:
        lines.append(f"- v2: `{ex.get('v2_bad', '')}`")
        lines.append(f"  v2.1: `{ex.get('v21_fixed', '')}`")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- v2.1 đủ mở review round 1: **{summary.get('review_ready_verdict', '')}**",
            f"- `review_ready_flag` = **{summary.get('review_ready_flag', False)}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_builder(
    *,
    input_path: Path,
    output_jsonl: Path,
    output_xlsx: Path,
    report_path: Path,
    summary_json_path: Path,
    v2_audit_summary_path: Optional[Path] = None,
    v2_jsonl_path: Optional[Path] = None,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    raw: List[Dict[str, Any]] = []
    for row in rows:
        ok, _ = passage_level_filter(row)
        if not ok:
            continue
        raw.extend(generate_candidates_from_passage(row))

    deduped = dedupe_candidates(raw)
    out_rows = to_output_rows(deduped)

    post_errors = Counter()
    clean_rows: List[Dict[str, Any]] = []
    for r in out_rows:
        errs = audit_candidate_row(r)
        if errs:
            for e in errs:
                post_errors[e] += 1
        else:
            clean_rows.append(r)

    if len(clean_rows) < len(out_rows):
        out_rows = clean_rows

    write_jsonl(output_jsonl, out_rows)
    write_workbook(out_rows, output_xlsx)

    v2_audit = {}
    if v2_audit_summary_path and v2_audit_summary_path.exists():
        v2_audit = json.loads(v2_audit_summary_path.read_text(encoding="utf-8"))

    unique_q = len({r["question_draft"] for r in out_rows})
    filtered_n = len(out_rows)
    dup_ratio = unique_q / max(filtered_n, 1)
    review_ready = (
        filtered_n >= 30
        and dup_ratio >= 0.85
        and post_errors.get("fact_mismatch", 0) == 0
        and post_errors.get("residue_led_question", 0) == 0
        and post_errors.get("unnatural_question_wording", 0) == 0
    )

    examples: List[Dict[str, str]] = []
    if v2_jsonl_path and v2_jsonl_path.exists():
        v2_rows = {r["seed_id"]: r for r in read_jsonl(v2_jsonl_path)}
        fixes = [
            ("RTX-V2-Q02", "energy_reduction_2019"),
            ("RTX-V2-Q04", "ergonomic_risk"),
            ("RTX-V2-Q01", "energy_intensity"),
        ]
        for sid, fid in fixes:
            bad = v2_rows.get(sid, {}).get("question_draft", "")
            for r in out_rows:
                if fid in (r.get("candidate_reason") or ""):
                    examples.append({"v2_bad": bad, "v21_fixed": r.get("question_draft", "")})
                    break

    summary = {
        "seed_version": SEED_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "raw_candidates": len(raw),
        "filtered_candidates": filtered_n,
        "usable_count": filtered_n,
        "dropped_fact_mismatch_count": v2_audit.get("fact_mismatch_count", 0),
        "dropped_wording_count": (
            v2_audit.get("unnatural_question_wording_count", 0)
            + v2_audit.get("residue_led_question_count", 0)
            + v2_audit.get("overlong_fact_phrase_count", 0)
        ),
        "dropped_residue_count": v2_audit.get("residue_led_question_count", 0),
        "post_audit_errors": dict(post_errors),
        "by_question_type": dict(Counter(r["question_type"] for r in out_rows)),
        "by_document_kind": dict(Counter(r.get("document_kind", "") for r in out_rows)),
        "v2_input_count": v2_audit.get("input_count", 327),
        "v2_needs_rebuild": v2_audit.get("needs_rebuild_count", 0),
        "exact_duplicate_question_count": filtered_n - unique_q,
        "review_ready_flag": review_ready,
        "review_ready_verdict": (
            "Có — question layer v2.1 đủ tốt để mở lại RTX review round 1"
            if review_ready
            else "Chưa — cần mở rộng canonical catalog hoặc siết gate thêm"
        ),
        "output_jsonl": str(output_jsonl),
        "output_xlsx": str(output_xlsx),
    }

    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Build RTX v2.1 fact-quality candidates")
    parser.add_argument("--input", default="data/golden_set/v2/rtx_step1_corpus_units/corpus_units_rtx_normalized.jsonl")
    parser.add_argument("--output-jsonl", default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v2_1_fact_quality.jsonl")
    parser.add_argument("--output-xlsx", default="data/golden_set/v2/reference_style/reference_seed_workbook_rtx_v2_1_fact_quality.xlsx")
    parser.add_argument("--report", default="reports/golden_set_candidate_generation_rtx_v2_1_fact_quality.md")
    parser.add_argument("--summary-json", default="reports/_candidate_generation_rtx_v2_1_fact_quality_summary.json")
    parser.add_argument("--v2-audit-summary", default="reports/_rtx_fact_target_quality_audit_v2_summary.json")
    parser.add_argument("--v2-jsonl", default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v2_fact_specific.jsonl")
    args = parser.parse_args(argv)

    summary = run_builder(
        input_path=root / args.input,
        output_jsonl=root / args.output_jsonl,
        output_xlsx=root / args.output_xlsx,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
        v2_audit_summary_path=root / args.v2_audit_summary,
        v2_jsonl_path=root / args.v2_jsonl,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in ("filtered_candidates", "usable_count", "review_ready_flag", "dropped_wording_count")
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
