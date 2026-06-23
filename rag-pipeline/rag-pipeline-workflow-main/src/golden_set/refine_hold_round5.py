"""Refine Hold Round 5 — two-lane expansion prep for gold_core_v1_1."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.canonicalize_reference_seed_workbook_r2 import infer_fact_target_id
from golden_set.gold_decision_core_round4 import (
    KNOWN_CLUSTERS,
    NEWS_CHROME,
    _company_in_text,
    _fact_clarity_score,
    _is_generic,
    _is_truncated,
    _listing_noise,
    _news_chrome,
    _resolve_unknown_cluster,
)
from golden_set.io_utils import read_jsonl, write_jsonl
from golden_set.prepare_manual_review_round2 import _norm_ws

REFINE_VERSION = "ref_hold_refine_r5"
CORE_COMPANIES = frozenset({"한샘", "무신사"})
PROMOTABLE_CLUSTERS = frozenset(
    set(KNOWN_CLUSTERS)
    | {
        "FC_REPORT_FRAMEWORK",
        "FC_CLIMATE_GHG",
        "FC_QUAL_POLICY",
    }
)

WORKBOOK_COLS = [
    "seed_id",
    "company",
    "question_type",
    "cluster_id",
    "final_question",
    "final_disclosure",
    "canonical_prohibited_claims",
    "hold_refine_lane",
    "hold_refine_decision",
    "hold_refine_reason",
    "refined_question",
    "refined_disclosure",
    "refined_prohibited_claims",
    "refined_fact_cluster_id",
    "cluster_resolution_status",
    "promote_candidate_v1_1",
    "notes",
]

META_CHROME = [
    "e-mail",
    "tel.",
    "font ",
    "닫기",
    "기념촬영",
    "이미지./제공",
    "사진=",
    "국문 영문",
    "당신이 좋아할",
    "요기요",
    "앱 ui",
    "이 기사와 관련",
    "혼용률",
]
TOC_PATTERN = re.compile(r"^\d{3}\s+[\w가-힣·]+(\s+\d{3}\s+[\w가-힣·]+){2,}")
PAGE_NUM_TAIL = re.compile(r"\s+0\d{2}\s+0\d{2}\s+0\d{2}")


def _q(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("final_question") or "")


def _d(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("final_disclosure") or "")


def _cluster(row: Dict[str, Any]) -> str:
    return (
        row.get("resolved_cluster_hint")
        or row.get("canonical_fact_target")
        or "FC_UNKNOWN"
    )


def _normalize_prohibited(text: str, qtype: str) -> str:
    t = _norm_ws(text)
    if t and "금지" in t:
        lines = [ln.strip() for ln in re.split(r"[\n;]+", t) if ln.strip()]
        return "\n".join(lines)
    if qtype == "quantitative":
        return "원문에 없는 수치 추가 금지\n단위 변경 추정 금지\n미공시 항목 단정 금지"
    if qtype == "trend":
        return "원문에 없는 원인 추정 금지\n향후 개선 보장 금지\n미공시 수치 보완 금지"
    return "완전 달성/전면 준수 단정 금지\n원문 밖 정책 확장 해석 금지\n미공시 성과 추정 금지"


def _has_meta_chrome(d: str) -> bool:
    blob = d.lower()
    return any(m in blob for m in META_CHROME)


def _is_toc_disclosure(d: str) -> bool:
    if TOC_PATTERN.search(d):
        return True
    if PAGE_NUM_TAIL.search(d):
        return True
    if re.search(r"\b0\d{2}\b.*\b0\d{2}\b.*\b0\d{2}\b", d):
        return True
    if "GOVERNANCE" in d and "·" in d and len(d) < 200:
        return True
    return False


def _extract_governance_sentence(d: str, company: str) -> Optional[str]:
    if "사외이사" in d and "이사회" in d:
        idx = d.rfind("사외이사")
        start = max(d.rfind(company, 0, idx), 0)
        if start == 0 and company not in d[: idx + 1]:
            start = max(d.rfind("무신사", 0, idx), 0)
        snippet = _norm_ws(d[start:])
        if len(snippet) >= 30:
            return snippet[:320]
    patterns = [
        r"[^.!?…]*사외이사[^.!?…]*이사회[^.!?…]*",
        r"[^.!?…]*이사회[^.!?…]*독립성[^.!?…]*",
        r"[^.!?…]*거버넌스[^.!?…]*",
    ]
    for pat in patterns:
        m = re.search(pat, d)
        if m:
            s = _norm_ws(m.group(0))
            if company in s or company == "무신사":
                return s[:320]
    return None


def _trim_news_chrome(d: str, company: str) -> str:
    if not _news_chrome(d) and not _has_meta_chrome(d):
        return d
    gov = _extract_governance_sentence(d, company)
    if gov:
        return gov
    parts = re.split(r"(?<=[.!?다])\s+|…+", d)
    kept: List[str] = []
    for p in parts:
        p = _norm_ws(p)
        if len(p) < 25:
            continue
        if any(c in p.lower() for c in NEWS_CHROME + META_CHROME):
            continue
        if company in p:
            kept.append(p)
    if kept:
        return kept[-1][:320]
    return d


def _trim_report_portal(d: str) -> str:
    years = sorted(set(re.findall(r"(20\d{2})년", d)), reverse=True)
    if years and "지속가능경영보고서" in d:
        if len(years) >= 2:
            return f"한샘은 {years[-1]}년~{years[0]}년 지속가능경영보고서를 발간·공개하고 있다."
        return f"한샘은 {years[0]}년 지속가능경영보고서를 발간했다."
    m = re.search(r"['\"]?20\d{2}[^'\"]*지속가능경영보고서['\"]?[^.]*발간[^.]*", d)
    if m:
        return _norm_ws(m.group(0))[:320]
    return d


def _trim_verification_meta(d: str) -> str:
    if "보고검증" in d or "한국표준협회" in d:
        m = re.search(
            r"보고주기[^.]*보고검증[^.]*한국표준협회[^.]*",
            d,
        )
        if m:
            return _norm_ws(m.group(0))[:280]
    return d


def _strip_leading_noise(d: str) -> str:
    d = re.sub(r"^\[.*?\]\s*", "", d)
    d = re.sub(r"^[\d\s]+한샘\s+", "한샘 ", d)
    d = re.sub(r"font\s*\d+\s*닫기\s*", "", d, flags=re.IGNORECASE)
    return _norm_ws(d)


def _assign_lane(row: Dict[str, Any]) -> str:
    cluster = _cluster(row)
    if cluster == "FC_UNKNOWN" or row.get("needs_cluster_rename"):
        return "lane_2_fc_unknown_resolution"
    return "lane_1_known_cluster_cleanup"


def _resolve_cluster_lane2(q: str, d: str, company: str) -> Tuple[str, str]:
    """Return (cluster_id, resolution_status)."""
    inferred = infer_fact_target_id(q, d)
    if inferred != "FC_UNKNOWN":
        return inferred, "resolved"

    resolved = _resolve_unknown_cluster(q, d)
    if resolved:
        return resolved, "resolved"

    blob = f"{q} {d}"
    if "지속가능경영보고서" in d and ("발간" in d or "보고서" in d):
        return "FC_REPORT_FRAMEWORK", "resolved"
    if re.search(r"20\d{2}년.*지속가능경영보고서", d):
        return "FC_REPORT_FRAMEWORK", "resolved"
    if "보고주기" in d and "한국표준협회" in d:
        return "FC_REPORT_FRAMEWORK", "resolved"
    if "이해관계자" in q and "이해관계자" in d:
        return "FC_MATERIAL_8", "partial"
    if "온실가스" in q and "온실가스" in d:
        return "FC_CLIMATE_GHG", "partial"
    if "공급망" in q and ("공급망" in d or "동반성장" in d):
        if _is_toc_disclosure(d):
            return "FC_UNKNOWN", "unresolved"
        return "FC_ESG_GOVERNANCE", "partial"
    if "SDGs" in d or "지속가능개발목표" in d:
        if _is_truncated(d):
            return "FC_UNKNOWN", "unresolved"
        return "FC_QUAL_POLICY", "partial"
    if "GRP" in d or "우수기업" in d:
        return "FC_KGCS_A", "partial"

    return "FC_UNKNOWN", "unresolved"


def _align_question(q: str, d: str, cluster: str, company: str) -> str:
    if cluster == "FC_REPORT_FRAMEWORK":
        if "발간" in d:
            return f"{company}는 지속가능경영보고서 발간 현황을 어떻게 공시하는가?"
    if cluster == "FC_HUMAN_RIGHTS" and "인권" not in q:
        return f"{company}의 인권경영 정책 또는 실사 현황은 무엇인가?"
    if cluster == "FC_ESG_GOVERNANCE" and "거버넌스" not in q:
        return f"{company}의 ESG 거버넌스 체계는 어떻게 운영되는가?"
    if cluster == "FC_CLIMATE_GHG":
        return f"{company}의 온실가스 배출 관리 또는 감축 활동은 무엇인가?"
    return q


def _disclosure_usable(
    d: str,
    company: str,
    cluster: str,
    *,
    q: str = "",
    chrome_cleaned: bool = False,
) -> bool:
    if len(d) < 40:
        return False
    if _is_truncated(d) and cluster == "FC_UNKNOWN":
        return False
    if _listing_noise(d):
        return False
    if _is_toc_disclosure(d):
        return False
    if not chrome_cleaned and (_news_chrome(d) or _has_meta_chrome(d)):
        return False
    anchored = _company_in_text(company, d) or _company_in_text(company, q)
    if not anchored and cluster not in PROMOTABLE_CLUSTERS:
        return False
    return True


def refine_lane1(row: Dict[str, Any]) -> Dict[str, Any]:
    company = row.get("company") or ""
    qtype = row.get("question_type") or "qualitative"
    q = _q(row)
    d = _d(row)
    cluster = _cluster(row)
    prohib = _normalize_prohibited(
        row.get("canonical_prohibited_claims") or row.get("prohibited_claims") or "",
        qtype,
    )

    refined_d = _strip_leading_noise(d)
    chrome_cleaned = False
    if _news_chrome(refined_d) or _has_meta_chrome(refined_d):
        before = refined_d
        refined_d = _trim_news_chrome(refined_d, company)
        chrome_cleaned = refined_d != before and len(refined_d) >= 30

    refined_q = q
    reason = "known_cluster_cleanup"
    clarity = _fact_clarity_score(company, refined_q, refined_d)

    if cluster == "FC_KGCS_A" and "KGCS" in q and "GRP" in refined_d and "KGCS" not in refined_d:
        decision = "keep_hold"
        reason = "question_cluster_mismatch_grp_vs_kgcs"
        resolution = "partial"
    elif _is_toc_disclosure(refined_d):
        decision = "drop_after_refine"
        reason = "toc_or_page_index_only"
        resolution = "unresolved"
    elif cluster in PROMOTABLE_CLUSTERS and clarity >= 6 and _disclosure_usable(
        refined_d, company, cluster, q=refined_q, chrome_cleaned=chrome_cleaned
    ):
        if _is_generic(q):
            refined_q = _align_question(q, refined_d, cluster, company)
        decision = "promote_candidate"
        reason = "known_cluster_cleaned"
        resolution = "resolved"
    elif not _disclosure_usable(
        refined_d, company, cluster, q=refined_q, chrome_cleaned=chrome_cleaned
    ):
        if clarity >= 5 and cluster == "FC_HUMAN_RIGHTS":
            decision = "keep_hold"
            reason = "disclosure_still_meta_heavy"
            resolution = "partial"
        else:
            decision = "drop_after_refine"
            reason = "disclosure_not_clean_after_trim"
            resolution = "unresolved"
    elif clarity >= 5:
        decision = "keep_hold"
        reason = "salvageable_needs_spot_check"
        resolution = "partial"
    else:
        decision = "drop_after_refine"
        reason = "insufficient_fact_clarity"
        resolution = "unresolved"

    promote = decision == "promote_candidate"
    return {
        "hold_refine_lane": "lane_1_known_cluster_cleanup",
        "hold_refine_decision": decision,
        "hold_refine_reason": reason,
        "refined_question": refined_q,
        "refined_disclosure": refined_d,
        "refined_prohibited_claims": prohib,
        "refined_fact_cluster_id": cluster,
        "cluster_resolution_status": resolution,
        "promote_candidate_v1_1": promote,
    }


def refine_lane2(row: Dict[str, Any]) -> Dict[str, Any]:
    company = row.get("company") or ""
    qtype = row.get("question_type") or "qualitative"
    q = _q(row)
    d = _d(row)
    prohib = _normalize_prohibited(
        row.get("canonical_prohibited_claims") or row.get("prohibited_claims") or "",
        qtype,
    )

    refined_d = _strip_leading_noise(d)
    cluster, resolution = _resolve_cluster_lane2(q, refined_d, company)

    if cluster == "FC_REPORT_FRAMEWORK":
        refined_d = _trim_report_portal(refined_d)
        if "보고주기" in d:
            refined_d = _trim_verification_meta(d)
    elif _news_chrome(refined_d) or _has_meta_chrome(refined_d):
        refined_d = _trim_news_chrome(refined_d, company)
    elif _is_toc_disclosure(refined_d):
        refined_d = refined_d  # keep for drop decision

    refined_q = _align_question(q, refined_d, cluster, company)

    reason = "fc_unknown_resolution"
    if cluster == "FC_UNKNOWN":
        if _is_toc_disclosure(d):
            decision = "drop_after_refine"
            reason = "toc_listing_unresolvable"
        elif _is_truncated(d):
            decision = "drop_after_refine"
            reason = "truncated_unresolvable"
        elif "기념촬영" in d or "font" in d.lower():
            decision = "drop_after_refine"
            reason = "press_photo_caption_only"
        elif _fact_clarity_score(company, q, refined_d) >= 4:
            decision = "keep_hold"
            reason = "cluster_unresolved_salvageable"
        else:
            decision = "drop_after_refine"
            reason = "cluster_and_fact_unresolved"
    elif resolution == "partial":
        if _disclosure_usable(refined_d, company, cluster, q=refined_q) and _fact_clarity_score(
            company, refined_q, refined_d
        ) >= 5:
            decision = "keep_hold"
            reason = "cluster_partial_needs_review"
        else:
            decision = "drop_after_refine"
            reason = "partial_cluster_weak_disclosure"
    elif (
        cluster in PROMOTABLE_CLUSTERS
        and _disclosure_usable(refined_d, company, cluster, q=refined_q)
        and _fact_clarity_score(company, refined_q, refined_d) >= 6
    ):
        decision = "promote_candidate"
        reason = "fc_unknown_resolved_clean"
    elif _fact_clarity_score(company, refined_q, refined_d) >= 5:
        decision = "keep_hold"
        reason = "cluster_resolved_disclosure_needs_trim"
    else:
        decision = "drop_after_refine"
        reason = "resolved_but_disclosure_weak"

    if cluster == "FC_UNKNOWN" and decision == "promote_candidate":
        decision = "keep_hold"
        reason = "fc_unknown_cannot_promote"

    promote = decision == "promote_candidate"
    return {
        "hold_refine_lane": "lane_2_fc_unknown_resolution",
        "hold_refine_decision": decision,
        "hold_refine_reason": reason,
        "refined_question": refined_q,
        "refined_disclosure": refined_d,
        "refined_prohibited_claims": prohib,
        "refined_fact_cluster_id": cluster,
        "cluster_resolution_status": resolution,
        "promote_candidate_v1_1": promote,
    }


def refine_row(row: Dict[str, Any]) -> Dict[str, Any]:
    if row.get("company") not in CORE_COMPANIES:
        return {}
    lane = _assign_lane(row)
    base = dict(row)
    if lane == "lane_1_known_cluster_cleanup":
        refined = refine_lane1(row)
    else:
        refined = refine_lane2(row)
    base.update(refined)
    base["refine_version"] = REFINE_VERSION
    base["notes"] = f"gold_reason={row.get('gold_reason', '')}; lane={lane}"
    return base


def write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide.append(["Hold Refine Round 5 — expansion lane for gold_core_v1_1"])
    guide.append(["gold_core_v1", "Frozen — không chỉnh trong round này"])

    groups = {
        "Promote_Candidates_V1_1": [r for r in rows if r.get("hold_refine_decision") == "promote_candidate"],
        "Keep_Hold": [r for r in rows if r.get("hold_refine_decision") == "keep_hold"],
        "Drop_After_Refine": [r for r in rows if r.get("hold_refine_decision") == "drop_after_refine"],
    }
    for title, subset in groups.items():
        ws = wb.create_sheet(title)
        ws.append(WORKBOOK_COLS)
        fill = PatternFill("solid", fgColor="1D4ED8")
        font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(WORKBOOK_COLS) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = fill
            c.font = font
        for r in subset:
            ws.append([r.get(c, "") for c in WORKBOOK_COLS])

    ws_rx = wb.create_sheet("RX_Status")
    ws_rx.append(["item", "status"])
    ws_rx.append(["rx_in_refine_round5", "excluded"])
    ws_rx.append(["rx_status", "source-acquisition dependent"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report(summary: Dict[str, Any], path: Path) -> None:
    lines = [
        "# Golden Set — Refine Hold Round 5",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Refine 17 hold rows thành lane mở rộng có kiểm soát cho **gold_core_v1_1**, không đụng **gold_core_v1** đã freeze.",
        "",
        "## Vì sao refine hold sau khi freeze v1",
        "",
        "- `gold_core_v1` (26 row) đã đủ làm artifact chính thức đầu tiên.",
        "- Hold backlog là lane giá trị cao nhất để mở rộng core.",
        "- Không benchmark — tập vẫn đang mở rộng.",
        "",
        "## Chia lane hold như thế nào",
        "",
        "- **lane_1_known_cluster_cleanup:** cluster đã rõ; trim news chrome / meta / prohibited.",
        "- **lane_2_fc_unknown_resolution:** FC_UNKNOWN — resolve cluster + rewrite nhẹ Q/disclosure.",
        "",
        "## Kết quả tổng quan",
        "",
        f"- **input_hold_count:** {summary.get('input_hold_count', 0)}",
        f"- **promote_candidate:** {summary.get('promote_candidate_count', 0)}",
        f"- **keep_hold:** {summary.get('keep_hold_count', 0)}",
        f"- **drop_after_refine:** {summary.get('drop_after_refine_count', 0)}",
        "",
        "### Breakdown theo lane",
        "",
    ]
    for lane, counts in summary.get("lane_breakdown", {}).items():
        lines.append(f"- `{lane}`: {counts}")

    lines.extend(["", "### Breakdown theo công ty", ""])
    for co, n in summary.get("by_company", {}).items():
        lines.append(f"- **{co}**: {n}")

    lines.extend(["", "### Breakdown theo refined cluster", ""])
    for cid, n in summary.get("by_cluster", {}).items():
        lines.append(f"- `{cid}`: {n}")

    lines.extend(
        [
            "",
            "### FC_UNKNOWN",
            "",
            f"- **resolved:** {summary.get('fc_unknown_resolved_count', 0)}",
            f"- **unresolved:** {summary.get('fc_unknown_unresolved_count', 0)}",
            "",
            "## Ví dụ",
            "",
        ]
    )
    for label, sid in [
        ("Known cluster cleanup thành công", summary.get("example_lane1_promote")),
        ("FC_UNKNOWN resolve thành công", summary.get("example_lane2_promote")),
        ("Drop sau refine", summary.get("example_drop")),
    ]:
        if sid:
            lines.append(f"- **{label}:** `{sid}`")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- Có thể mở **gold_core_v1_1** với **{summary.get('potential_gold_core_v1_1_additions', 0)}** row promote_candidate (chưa promote trong task).",
            f"- Còn **{summary.get('keep_hold_count', 0)}** row giữ hold cho round sau.",
            f"- **RX:** {summary.get('rx_status', '')}",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_refine(
    *,
    hold_path: Path,
    refined_path: Path,
    unresolved_path: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    hold_rows = [r for r in read_jsonl(hold_path) if r.get("company") in CORE_COMPANIES]
    refined_rows = [refine_row(r) for r in hold_rows]

    unresolved_rows = [
        r
        for r in refined_rows
        if r.get("hold_refine_decision") in ("keep_hold", "drop_after_refine")
    ]
    write_jsonl(refined_path, refined_rows)
    write_jsonl(unresolved_path, unresolved_rows)
    write_workbook(refined_rows, workbook_path)

    promote = [r for r in refined_rows if r.get("hold_refine_decision") == "promote_candidate"]
    keep = [r for r in refined_rows if r.get("hold_refine_decision") == "keep_hold"]
    drop = [r for r in refined_rows if r.get("hold_refine_decision") == "drop_after_refine"]

    lane_breakdown: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in refined_rows:
        lane_breakdown[r.get("hold_refine_lane", "")][r.get("hold_refine_decision", "")] += 1

    fc_input = [r for r in hold_rows if _cluster(r) == "FC_UNKNOWN"]
    fc_resolved = [
        r
        for r in refined_rows
        if r.get("hold_refine_lane") == "lane_2_fc_unknown_resolution"
        and r.get("refined_fact_cluster_id") != "FC_UNKNOWN"
    ]
    fc_unresolved = [
        r
        for r in refined_rows
        if r.get("hold_refine_lane") == "lane_2_fc_unknown_resolution"
        and r.get("refined_fact_cluster_id") == "FC_UNKNOWN"
    ]

    summary = {
        "refine_version": REFINE_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_hold_count": len(hold_rows),
        "promote_candidate_count": len(promote),
        "keep_hold_count": len(keep),
        "drop_after_refine_count": len(drop),
        "lane_breakdown": {k: dict(v) for k, v in lane_breakdown.items()},
        "by_company": dict(Counter(r["company"] for r in refined_rows)),
        "by_cluster": dict(Counter(r.get("refined_fact_cluster_id", "") for r in refined_rows)),
        "fc_unknown_input_count": len(fc_input),
        "fc_unknown_resolved_count": len(fc_resolved),
        "fc_unknown_unresolved_count": len(fc_unresolved),
        "potential_gold_core_v1_1_additions": len(promote),
        "rx_status": "source-acquisition dependent — excluded from refine round 5",
        "example_lane1_promote": next(
            (r["seed_id"] for r in promote if r.get("hold_refine_lane") == "lane_1_known_cluster_cleanup"),
            None,
        ),
        "example_lane2_promote": next(
            (r["seed_id"] for r in promote if r.get("hold_refine_lane") == "lane_2_fc_unknown_resolution"),
            None,
        ),
        "example_drop": drop[0]["seed_id"] if drop else None,
        "output_refined": str(refined_path),
        "output_unresolved": str(unresolved_path),
        "output_workbook": str(workbook_path),
    }

    write_report(summary, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Refine hold backlog round 5")
    parser.add_argument(
        "--hold",
        default="data/golden_set/v2/step6_gold/golden_set_core_round4_hold.jsonl",
    )
    parser.add_argument(
        "--output-refined",
        default="data/golden_set/v2/step6_gold/golden_set_core_round5_refined.jsonl",
    )
    parser.add_argument(
        "--output-unresolved",
        default="data/golden_set/v2/step6_gold/golden_set_core_round5_unresolved.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/step6_gold/golden_set_core_round5_refine.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_refine_hold_round5.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_refine_hold_round5_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_refine(
        hold_path=root / args.hold,
        refined_path=root / args.output_refined,
        unresolved_path=root / args.output_unresolved,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "promote_candidate_count",
                    "keep_hold_count",
                    "drop_after_refine_count",
                    "fc_unknown_resolved_count",
                    "potential_gold_core_v1_1_additions",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
