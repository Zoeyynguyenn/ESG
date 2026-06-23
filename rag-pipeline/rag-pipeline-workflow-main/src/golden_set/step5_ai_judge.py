"""Step 5b: LLM-as-judge — auto SME review when no human expert is available."""

from __future__ import annotations

import csv
import json
import os
import re
from pathlib import Path
from typing import Any, Dict, List

from golden_set.step5_sme import SME_FIELDS, run_step5

AI_JUDGE_PROMPT = """You are an ESG disclosure QA reviewer for a Korean RAG golden set.
Judge whether this Q&A pair is safe to use as ground truth for regression testing.

RULES:
- APPROVE only if the answer is fully supported by CONTEXT (no hallucination).
- REJECT if the answer adds facts not in context, confuses report years/editions, or is too vague.
- REVISE if the idea is correct but wording should be tighter; provide fixed question/answer from CONTEXT only.
- Question and answer should match the language of CONTEXT (Korean for Korean ESG reports).

QUESTION: {question}
ANSWER: {answer}
COMPANY: {company}
GRI: {gri}
CONTEXT:
\"\"\"
{context}
\"\"\"

Return JSON only (no markdown):
{{
  "decision": "approve|reject|revise",
  "confidence": 0.0,
  "notes": "short reason in English or Korean",
  "revised_question": "",
  "revised_answer": "",
  "forbidden_rule": "what the RAG must NOT claim beyond context"
}}"""


def _openai_chat(prompt: str, model: str) -> str:
    from openai import OpenAI

    client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY", ""))
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.0,
    )
    return (resp.choices[0].message.content or "").strip()


def _parse_json(text: str) -> Dict[str, Any]:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```\w*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
    return json.loads(text)


def _judge_row(row: Dict[str, str], *, model: str, min_confidence: float) -> Dict[str, str]:
    prompt = AI_JUDGE_PROMPT.format(
        question=row.get("question", ""),
        answer=row.get("ground_truth_answer", ""),
        company=row.get("company", ""),
        gri=row.get("gri_code", "") or "(none)",
        context=(row.get("context_excerpt") or "")[:800],
    )
    try:
        body = _parse_json(_openai_chat(prompt, model))
    except Exception as exc:
        row["sme_decision"] = "reject"
        row["sme_notes"] = f"ai_judge_error:{exc}"
        return row

    decision = str(body.get("decision", "reject")).strip().lower()
    confidence = float(body.get("confidence") or 0.0)
    notes = str(body.get("notes", "")).strip()
    revised_q = str(body.get("revised_question", "")).strip()
    revised_a = str(body.get("revised_answer", "")).strip()
    forbidden = str(body.get("forbidden_rule", "")).strip()

    if decision == "revise" and revised_q and revised_a:
        row["sme_revised_question"] = revised_q
        row["sme_revised_answer"] = revised_a
        row["sme_decision"] = "approve" if confidence >= min_confidence else "reject"
        row["sme_notes"] = f"ai_judge:revise→approve conf={confidence:.2f}; {notes}"
    elif decision == "approve" and confidence >= min_confidence:
        row["sme_decision"] = "approve"
        row["sme_notes"] = f"ai_judge:approve conf={confidence:.2f}; {notes}"
    else:
        row["sme_decision"] = "reject"
        row["sme_notes"] = f"ai_judge:{decision} conf={confidence:.2f}; {notes}"

    if forbidden:
        row["forbidden_rule"] = forbidden
    return row


def _write_sme_files(rows: List[Dict[str, str]], csv_path: Path, xlsx_path: Path | None) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=SME_FIELDS)
        w.writeheader()
        w.writerows(rows)

    if not xlsx_path:
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "SME Review"
        header_fill = PatternFill("solid", fgColor="1D4ED8")
        header_font = Font(color="FFFFFF", bold=True)
        ws.append(SME_FIELDS)
        for c in range(1, len(SME_FIELDS) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row.get(k, "") for k in SME_FIELDS])
        wb.save(xlsx_path)
    except ImportError:
        pass


def _load_sme_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def run_step5_ai_judge(
    *,
    input_path: Path,
    csv_path: Path,
    xlsx_path: Path | None = None,
    model: str = "gpt-4o-mini",
    min_confidence: float = 0.75,
    limit: int = 0,
    reviewer_tag: str = "ai_judge",
) -> Dict[str, Any]:
    if not os.environ.get("OPENAI_API_KEY", "").strip():
        raise SystemExit("OPENAI_API_KEY missing for AI SME judge")

    if csv_path.exists():
        rows = _load_sme_csv(csv_path)
    else:
        run_step5(input_path=input_path, csv_path=csv_path, xlsx_path=None)
        rows = _load_sme_csv(csv_path)

    if limit > 0:
        rows = rows[:limit]

    approved = rejected = 0
    for i, row in enumerate(rows):
        rows[i] = _judge_row(row, model=model, min_confidence=min_confidence)
        if rows[i].get("sme_decision") == "approve":
            approved += 1
        else:
            rejected += 1
        if reviewer_tag and not rows[i].get("sme_notes", "").startswith("ai_judge"):
            rows[i]["sme_notes"] = f"{reviewer_tag}; {rows[i].get('sme_notes', '')}"

    _write_sme_files(rows, csv_path, xlsx_path)

    return {
        "reviewer": reviewer_tag,
        "model": model,
        "min_confidence": min_confidence,
        "total": len(rows),
        "approved": approved,
        "rejected": rejected,
        "csv": str(csv_path),
        "xlsx": str(xlsx_path) if xlsx_path else "",
    }
