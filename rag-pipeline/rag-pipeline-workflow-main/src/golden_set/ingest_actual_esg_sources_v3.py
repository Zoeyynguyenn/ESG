"""Ingest actual / company-primary ESG sources and rebuild reference workbook v3."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader

from golden_set.build_reference_seed_workbook import (
    ESG_SENTENCE_KEYWORDS,
    _make_question,
    _prohibited_claims,
    _split_sentences,
)
from golden_set.canonicalize_reference_seed_workbook_r2 import FACT_CLUSTER_RULES
from golden_set.io_utils import read_jsonl, write_jsonl

INGEST_VERSION = "ref_workbook_v3"
PACKAGE_ROOT = Path("data/rag_dataset/05_company_export_json")

COMPANY_PACKAGES = {
    "무신사": "무신사_dataset_package_20260608T092823",
    "레이시온": "레이시온_dataset_package_20260608T055801",
}

COMPANY_NEWSROOM_PATTERNS = {
    "무신사": ("newsroom.musinsa.com",),
    "레이시온": (),
}

REPORT_BODY_MARKERS = (
    "about this report",
    "보고서 개요",
    "ceo 인사말",
    "ceo message",
    "중대성",
    "materiality",
    "gri standard",
    "목차",
    "table of contents",
)

PORTAL_MARKERS = (
    "민원서비스",
    "정보공개제도",
    "만족도 평가",
    "게시판 목록",
    "print 상태",
)

PRESS_MARKERS = ("기자", "연합뉴스", "발행일", "송고", "구독")

CROSS_COMPANY = (
    "현대트랜시스",
    "삼성전기",
    "삼성전자",
    "여수광양항만공사",
    "코스맥스",
    "lg헬로비전",
)


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _match_rules(text: str, patterns: Sequence[re.Pattern]) -> bool:
    return any(p.search(text) for p in patterns)


def infer_cluster(question: str, passage: str) -> str:
    blob = f"{question} {passage}"
    for cluster_id, q_patterns, _ in FACT_CLUSTER_RULES:
        if _match_rules(question or "", q_patterns) or _match_rules(blob, q_patterns):
            return cluster_id
    extra = [
        ("FC_IMPACT_REPORT", [r"임팩트 리포트", r"impact report"]),
        ("FC_CLIMATE_GHG", [r"온실가스", r"스코프", r"scope", r"기후"]),
        ("FC_STAKEHOLDER_DISCLOSURE", [r"이해관계자", r"소통"]),
        ("FC_GRI_FRAMEWORK", [r"GRI", r"글로벌 리포팅"]),
        ("FC_ISO_GOVERNANCE", [r"ISO 37001", r"ISO 37301", r"윤리·준법"]),
    ]
    for cid, pats in extra:
        if any(re.search(p, blob, re.I) for p in pats):
            return cid
    return "FC_UNKNOWN"


def _pdf_report_signals(text: str, company: str) -> Tuple[int, int]:
    lower = text.lower()
    company_hits = text.count(company) + lower.count("musinsa") + lower.count("무신사")
    report_hits = sum(1 for m in REPORT_BODY_MARKERS if m in lower)
    wrong_hits = sum(1 for m in ("중소기업 연차보고서", "www.mss.go.kr", "mss.go.kr") if m in text)
    return company_hits + report_hits * 3, wrong_hits


def classify_pdf(pdf_path: Path, company: str, max_pages: int = 40) -> Dict[str, Any]:
    try:
        reader = PdfReader(str(pdf_path))
    except Exception as exc:  # noqa: BLE001
        return {
            "usable": False,
            "reason": f"pdf_read_error:{exc}",
            "pages": 0,
        }
    sample = []
    for i, page in enumerate(reader.pages[:max_pages]):
        sample.append(page.extract_text() or "")
    blob = "\n".join(sample)
    pos, neg = _pdf_report_signals(blob, company)
    usable = pos >= 5 and neg == 0 and len(blob) > 2000
    return {
        "usable": usable,
        "reason": "validated_report_body" if usable else "wrong_or_irrelevant_pdf",
        "pages": len(reader.pages),
        "company_signal": pos,
        "wrong_document_signal": neg,
        "preview": _norm_ws(blob)[:400],
    }


def _noise_score(text: str) -> int:
    return sum(1 for m in PORTAL_MARKERS + PRESS_MARKERS if m in text)


def _esg_score(text: str) -> int:
    lower = text.lower()
    return sum(1 for k in ESG_SENTENCE_KEYWORDS if k in lower) + (2 if "esg" in lower else 0)


def _strip_newsroom_chrome(text: str) -> str:
    text = re.sub(
        r"^보도자료\s+무신사,?\s*2024\s*임팩트\s*리포트[^‍]*링크\s*바로가기\s*",
        "",
        text,
        flags=re.I,
    )
    return text.strip()


def _paragraphs(text: str, *, company: str = "") -> List[str]:
    text = _strip_newsroom_chrome(text)
    parts = re.split(r"‍+", text)
    if len(parts) <= 1:
        parts = re.split(r"\n+", text)
    out: List[str] = []
    for p in parts:
        s = _norm_ws(p)
        if len(s) < 60:
            continue
        if re.match(r"^보도자료\s+\d{4}-\d{4}", s):
            continue
        if "링크 바로가기" in s and len(s) < 120:
            continue
        out.append(s)
    if not out and company:
        sentences = _split_sentences(text)
        blob = " ".join(sentences)
        if len(blob) >= 80:
            out.append(blob)
    return out


def _theme_for_passage(passage: str) -> str:
    lower = passage.lower()
    if any(k in lower for k in ("온실가스", "스코프", "scope", "기후", "탄소")):
        return "climate_emissions"
    if any(k in lower for k in ("지배구조", "이사회", "윤리", "준법", "iso 37")):
        return "governance"
    if any(k in lower for k in ("사회", "이해관계자", "지역사회", "기부", "동반성장", "파트너")):
        return "social_supply_chain"
    if any(k in lower for k in ("gri", "임팩트 리포트", "보고서")):
        return "strategy_disclosure"
    if any(k in lower for k in ("중대", "materiality")):
        return "materiality"
    return "general_esg"


def discover_sources(company: str, package_dir: Path) -> List[Dict[str, Any]]:
    discovered: List[Dict[str, Any]] = []
    sources_dir = package_dir / "_sources"
    if sources_dir.exists():
        for pdf in sorted(sources_dir.glob("*.pdf")):
            meta = classify_pdf(pdf, company)
            discovered.append(
                {
                    "company": company,
                    "record_id_or_doc_id": pdf.stem,
                    "source_path": str(pdf.relative_to(package_dir)).replace("\\", "/"),
                    "source_origin": "package_sources_pdf",
                    "source_kind": "actual_report_pdf" if meta["usable"] else "rejected_pdf",
                    "section_title": "",
                    "passage_text": meta.get("preview", ""),
                    "source_decision": "keep_for_ingest" if meta["usable"] else "drop_wrong_document",
                    "source_reason": meta["reason"],
                    "esg_theme": "report_body" if meta["usable"] else "none",
                    "seed_candidate_priority": "high" if meta["usable"] else "drop",
                    "pdf_meta": meta,
                }
            )

    evidence_path = package_dir / "records" / "company_evidence.jsonl"
    if not evidence_path.exists():
        return discovered

    newsroom_best: Dict[str, Dict[str, Any]] = {}
    portal_candidates: List[Dict[str, Any]] = []

    with evidence_path.open(encoding="utf-8") as f:
        for line in f:
            row = json.loads(line)
            if row.get("company") != company:
                continue
            meta = row.get("metadata") or {}
            url = str(meta.get("source_url") or "")
            path = str(meta.get("source_path") or "")
            text = row.get("text") or ""
            rid = str(row.get("record_id") or "")

            if any(name in text for name in CROSS_COMPANY):
                continue

            # Company newsroom (Musinsa primary)
            if any(p in url for p in COMPANY_NEWSROOM_PATTERNS.get(company, ())):
                esg = _esg_score(text)
                noise = _noise_score(text)
                listing_penalty = 500 if re.search(r"보도자료\s+2026-\d{4}", text) else 0
                score = esg * 120 + len(text) - noise * 200 - listing_penalty
                prev = newsroom_best.get(url)
                if prev is None or score > prev["_score"]:
                    newsroom_best[url] = {
                        "row": row,
                        "url": url,
                        "path": path,
                        "_score": score,
                    }
                continue

            # YGPA portal embedded fact (Raysolution)
            if company == "레이시온" and "ygpa.or.kr" in url:
                m = re.search(
                    r"공사의 ESG경영 활동에 대한 이해관계자[^.]*투명하게 공개[^.]*\.",
                    _norm_ws(text),
                )
                if m:
                    portal_candidates.append(
                        {
                            "company": company,
                            "record_id_or_doc_id": rid,
                            "source_path": url or path,
                            "source_origin": "ygpa_portal_page",
                            "source_kind": "portal_embedded_disclosure",
                            "section_title": "ESG stakeholder disclosure",
                            "passage_text": _norm_ws(m.group(0)),
                            "source_decision": "salvage_only",
                            "source_reason": "no_report_body_only_portal_sentence",
                            "esg_theme": "stakeholder",
                            "seed_candidate_priority": "low",
                        }
                    )

    for url, item in newsroom_best.items():
        row = item["row"]
        text = row.get("text") or ""
        if _noise_score(text) > 2 or _esg_score(text) < 3:
            continue
        if "임팩트 리포트" not in text and "impact" not in text.lower():
            continue
        discovered.append(
            {
                "company": company,
                "record_id_or_doc_id": row.get("record_id"),
                "source_path": url or item["path"],
                "source_origin": "company_newsroom",
                "source_kind": "company_primary_report_summary",
                "section_title": "Musinsa newsroom impact report summary",
                "passage_text": text,
                "source_decision": "keep_for_ingest",
                "source_reason": "company_issued_newsroom_narrative_not_full_pdf",
                "esg_theme": "multi_theme",
                "seed_candidate_priority": "high",
            }
        )

    discovered.extend(portal_candidates)
    return discovered


def expand_pool_passages(pool: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    expanded: List[Dict[str, Any]] = []
    for item in pool:
        if item.get("source_decision") != "keep_for_ingest":
            expanded.append(item)
            continue
        if item.get("source_kind") == "company_primary_report_summary":
            company = item.get("company") or ""
            for para in _paragraphs(item.get("passage_text") or "", company=company):
                if _esg_score(para) < 2:
                    continue
                expanded.append(
                    {
                        **item,
                        "section_title": _theme_for_passage(para),
                        "passage_text": para,
                        "esg_theme": _theme_for_passage(para),
                        "source_kind": "company_primary_report_summary",
                        "seed_candidate_priority": "high",
                    }
                )
        elif item.get("source_kind") == "actual_report_pdf":
            expanded.append(item)
    return expanded


def _seed_id(company: str, qtype: str, index: int) -> str:
    prefix = {"무신사": "MS", "레이시온": "RX", "한샘": "HS"}.get(company, "XX")
    code = {"quantitative": "Q", "trend": "T", "qualitative": "L"}.get(qtype, "X")
    return f"{prefix}-G-{code}{index:02d}"


def build_seeds_from_pool(
    pool: Sequence[Dict[str, Any]],
    company: str,
    *,
    start_index: int = 1,
) -> Tuple[List[Dict[str, Any]], Counter]:
    seeds: List[Dict[str, Any]] = []
    origin_counts: Counter = Counter()
    used_clusters: set[str] = set()
    idx = start_index

    eligible = [
        p
        for p in pool
        if p.get("source_decision") == "keep_for_ingest"
        or p.get("source_decision") == "salvage_only"
    ]

    for item in sorted(
        eligible,
        key=lambda x: (
            0 if x.get("source_kind") == "actual_report_pdf" else 1,
            0 if x.get("source_kind") == "company_primary_report_summary" else 1,
            -len(x.get("passage_text") or ""),
        ),
    ):
        passage = item.get("passage_text") or ""
        if len(passage) < 50:
            continue

        kind = item.get("source_kind") or ""
        if kind == "portal_embedded_disclosure":
            origin = "portal_salvage_seed"
        elif kind == "actual_report_pdf":
            origin = "actual_report_body_seed"
        elif kind == "company_primary_report_summary":
            origin = "company_primary_narrative_seed"
        else:
            origin = "unknown"

        theme = item.get("esg_theme") or "general"
        cluster = infer_cluster("", passage)
        dedupe_key = cluster if cluster != "FC_UNKNOWN" else f"{theme}::{passage[:80]}"
        if dedupe_key in used_clusters:
            continue

        qtype = "qualitative"
        if re.search(r"\d", passage):
            qtype = "quantitative"
        question = _make_question(company, passage, "G", qtype) or (
            f"{company}의 ESG 활동 성과는 무엇인가?"
        )

        seeds.append(
            {
                "seed_id": _seed_id(company, qtype, idx),
                "company": company,
                "question_type": qtype,
                "question_draft": question,
                "acceptable_disclosure": passage[:500],
                "prohibited_claims": _prohibited_claims(qtype),
                "source_record_id": item.get("record_id_or_doc_id"),
                "source_path": item.get("source_path"),
                "fact_target": cluster,
                "fact_cluster_id": cluster,
                "seed_status": "candidate_v3",
                "seed_notes": f"{origin}; theme={theme}; kind={kind}",
                "seed_origin_type": origin,
                "seed_version": INGEST_VERSION,
                "anchor_type": "rebuilt_v3",
            }
        )
        origin_counts[origin] += 1
        used_clusters.add(dedupe_key)
        idx += 1

    return seeds, origin_counts


def _frozen_hansem(canonical_path: Path) -> List[Dict[str, Any]]:
    rows = read_jsonl(canonical_path)
    frozen: List[Dict[str, Any]] = []
    for row in rows:
        frozen.append(
            {
                "seed_id": row["seed_id"],
                "company": row["company"],
                "question_type": row.get("question_type"),
                "question_draft": row.get("canonical_question") or row.get("question_draft"),
                "acceptable_disclosure": row.get("canonical_acceptable_disclosure")
                or row.get("acceptable_disclosure"),
                "prohibited_claims": row.get("canonical_prohibited_claims")
                or row.get("prohibited_claims"),
                "source_record_id": row.get("source_record_id"),
                "source_path": "",
                "fact_target": row.get("fact_target"),
                "fact_cluster_id": row.get("fact_cluster_id"),
                "seed_status": "frozen_canonical_r2",
                "seed_notes": "frozen Hansem canonical anchor — do not edit",
                "seed_origin_type": "frozen_canonical_anchor",
                "seed_version": INGEST_VERSION,
                "anchor_type": "frozen",
            }
        )
    return frozen


def _write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "안내"
    ws.append(["Reference Seed Workbook V3 — actual / company-primary sources"])
    header = [
        "seed_id",
        "company",
        "anchor_type",
        "seed_origin_type",
        "question_type",
        "question_draft",
        "acceptable_disclosure",
        "prohibited_claims",
        "source_record_id",
        "source_path",
        "fact_cluster_id",
        "seed_status",
        "seed_notes",
    ]
    ws2 = wb.create_sheet("작성")
    ws2.append(header)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(header) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
    for row in rows:
        ws2.append([row.get(h, "") for h in header])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_audit_report(audit: Dict[str, Any], path: Path) -> None:
    lines = [
        "# Golden Set — Actual Source Audit V3",
        "",
        "## Mục tiêu",
        "",
        "Xác định có **actual ESG report body** trong repo hay không, trước khi rebuild workbook v3.",
        "",
        "## Đã tìm ở đâu",
        "",
        "- `data/rag_dataset/05_company_export_json/*/_sources/*.pdf`",
        "- `records/company_evidence.jsonl` metadata (`source_url`, `source_path`)",
        "- `splits/full.jsonl` (scan cùng package)",
        "",
        "## Musinsa",
        "",
        f"- **Có report PDF usable?** {audit['musinsa']['has_usable_pdf']}",
        f"- **PDF paths:** {', '.join(audit['musinsa']['pdf_paths']) or '—'}",
        f"- **Company newsroom primary:** {audit['musinsa']['has_company_newsroom']}",
        f"- **Newsroom URLs:** {', '.join(audit['musinsa']['newsroom_urls']) or '—'}",
        f"- **Usable for ingest:** {audit['musinsa']['ingest_usable']}",
        "",
        "## Raysolution",
        "",
        f"- **Có report PDF usable?** {audit['raysolution']['has_usable_pdf']}",
        f"- **PDF paths:** {', '.join(audit['raysolution']['pdf_paths']) or '— (không có _sources/)'}",
        f"- **Portal salvage only:** {audit['raysolution']['portal_salvage_only']}",
        f"- **Usable for ingest:** {audit['raysolution']['ingest_usable']}",
        "",
        "## Kết luận",
        "",
        f"- **Musinsa:** {audit['musinsa']['conclusion']}",
        f"- **Raysolution:** {audit['raysolution']['conclusion']}",
        "",
        "### Source acquisition gap",
        "",
    ]
    for gap in audit.get("gaps", []):
        lines.append(f"- {gap}")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_workbook_report(summary: Dict[str, Any], path: Path) -> None:
    lines = [
        "# Golden Set — Reference Workbook V3 (Actual Sources)",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Rebuild workbook từ **actual report body** hoặc **company-primary narrative**, không hardcode salvage record id.",
        "",
        "## Vì sao v2 chưa đủ",
        "",
        "- V2 salvage từ Yonhap/headline/portal với record id cố định.",
        "- Không có ingest từ PDF/report body thật.",
        "",
        "## Source thật đã ingest được",
        "",
    ]
    for src in summary.get("ingested_sources", []):
        lines.append(f"- {src}")

    lines.extend(
        [
            "",
            "## Kết quả seed",
            "",
            f"- Musinsa actual report body: **{summary.get('musinsa_actual_report_body_seeds', 0)}**",
            f"- Musinsa company-primary narrative: **{summary.get('musinsa_company_primary_narrative_seeds', 0)}**",
            f"- Raysolution actual report body: **{summary.get('raysolution_actual_report_body_seeds', 0)}**",
            f"- Seeds salvage (portal/press): **{summary.get('salvage_seeds', 0)}**",
            f"- Hansem frozen: **{summary.get('hansem_frozen', 0)}**",
            f"- **Tổng workbook v3:** **{summary.get('workbook_total', 0)}**",
            "",
            "## Coverage theo công ty",
            "",
        ]
    )
    for co, n in summary.get("seeds_by_company", {}).items():
        lines.append(f"- **{co}**: {n}")
    lines.extend(["", "## Coverage theo fact cluster", ""])
    for cid, n in summary.get("clusters", {}).items():
        lines.append(f"- `{cid}`: {n}")

    lines.extend(
        [
            "",
            "## Đánh giá",
            "",
            f"- Bớt phụ thuộc press/headline so với v2? **{summary.get('less_press_dependent', '')}**",
            f"- Gần 3-company review-ready hơn v2? **{summary.get('review_ready_verdict', '')}**",
            "",
            "## Kết luận",
            "",
            summary.get("conclusion", ""),
            "",
            f"**Bước tiếp:** {summary.get('next_step', '')}",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_ingest_v3(
    *,
    root: Path,
    canonical_path: Path,
    musinsa_pool_path: Path,
    raysolution_pool_path: Path,
    candidates_path: Path,
    workbook_path: Path,
    audit_path: Path,
) -> Dict[str, Any]:
    audit: Dict[str, Any] = {
        "musinsa": {
            "pdf_paths": [],
            "has_usable_pdf": False,
            "has_company_newsroom": False,
            "newsroom_urls": [],
            "ingest_usable": False,
            "conclusion": "",
        },
        "raysolution": {
            "pdf_paths": [],
            "has_usable_pdf": False,
            "portal_salvage_only": True,
            "ingest_usable": False,
            "conclusion": "",
        },
        "gaps": [],
    }

    all_pools: Dict[str, List[Dict[str, Any]]] = {}
    ingested_sources: List[str] = []

    for company, pkg_name in COMPANY_PACKAGES.items():
        pkg_dir = root / PACKAGE_ROOT / pkg_name
        raw = discover_sources(company, pkg_dir)
        expanded = expand_pool_passages(raw)

        key = "musinsa" if company == "무신사" else "raysolution"
        pdfs = [r for r in raw if r.get("source_origin") == "package_sources_pdf"]
        audit[key]["pdf_paths"] = [r["source_path"] for r in pdfs]
        usable_pdf = [r for r in pdfs if r.get("source_decision") == "keep_for_ingest"]
        audit[key]["has_usable_pdf"] = bool(usable_pdf)

        if company == "무신사":
            newsroom = [r for r in raw if r.get("source_origin") == "company_newsroom"]
            audit["musinsa"]["has_company_newsroom"] = bool(newsroom)
            audit["musinsa"]["newsroom_urls"] = [r.get("source_path", "") for r in newsroom]
            audit["musinsa"]["ingest_usable"] = bool(usable_pdf or newsroom)
            if usable_pdf:
                ingested_sources.append(f"Musinsa PDF: {usable_pdf[0]['source_path']}")
            elif newsroom:
                for row in newsroom:
                    ingested_sources.append(
                        f"Musinsa company newsroom (not full PDF): {row.get('source_path')}"
                    )
            if not usable_pdf:
                audit["gaps"].append(
                    "Musinsa: thiếu Impact Report PDF usable trong package _sources/"
                )

        if company == "레이시온":
            portal = [r for r in raw if r.get("source_kind") == "portal_embedded_disclosure"]
            audit["raysolution"]["portal_salvage_only"] = not usable_pdf
            audit["raysolution"]["ingest_usable"] = bool(usable_pdf or portal)
            if not usable_pdf:
                audit["gaps"].append(
                    "레이시온: không có _sources/ PDF; corpus = YGPA portal + cross-company noise"
                )

        all_pools[company] = expanded

    # Audit conclusions
    if audit["musinsa"]["has_usable_pdf"]:
        audit["musinsa"]["conclusion"] = "Có PDF report body — ingest được"
    elif audit["musinsa"]["has_company_newsroom"]:
        audit["musinsa"]["conclusion"] = (
            "Không có PDF report body usable; có company newsroom summary (newsroom.musinsa.com)"
        )
    else:
        audit["musinsa"]["conclusion"] = "Thiếu source thật — cần source acquisition"

    if audit["raysolution"]["has_usable_pdf"]:
        audit["raysolution"]["conclusion"] = "Có PDF report body — ingest được"
    else:
        audit["raysolution"]["conclusion"] = (
            "Không có report body; chỉ portal salvage hoặc thiếu hoàn toàn"
        )

    write_jsonl(musinsa_pool_path, all_pools["무신사"])
    write_jsonl(raysolution_pool_path, all_pools["레이시온"])

    frozen = _frozen_hansem(canonical_path)
    ms_seeds, ms_origins = build_seeds_from_pool(all_pools["무신사"], "무신사")
    rx_seeds, rx_origins = build_seeds_from_pool(all_pools["레이시온"], "레이시온")
    all_seeds = frozen + ms_seeds + rx_seeds
    write_jsonl(candidates_path, all_seeds)
    _write_workbook(all_seeds, workbook_path)

    ms_report_body = sum(
        1 for s in ms_seeds if s.get("seed_origin_type") == "actual_report_body_seed"
    )
    rx_report_body = sum(
        1 for s in rx_seeds if s.get("seed_origin_type") == "actual_report_body_seed"
    )
    ms_primary = sum(
        1 for s in ms_seeds if s.get("seed_origin_type") == "company_primary_narrative_seed"
    )
    rx_primary = sum(
        1 for s in rx_seeds if s.get("seed_origin_type") == "company_primary_narrative_seed"
    )
    salvage = sum(
        1
        for s in ms_seeds + rx_seeds
        if s.get("seed_origin_type") in {"portal_salvage_seed", "press_salvage_seed"}
    )

    by_co = Counter(s["company"] for s in all_seeds)
    clusters = Counter(s.get("fact_cluster_id") for s in all_seeds)

    less_press = (
        "Có — Musinsa dùng newsroom.musinsa.com thay Yonhap/headline v2"
        if ms_primary + ms_report_body > 0
        else "Chưa — vẫn thiếu source"
    )
    review_ready = (
        ms_report_body >= 3
        and rx_report_body >= 2
        and by_co.get("레이시온", 0) >= 3
        and len(all_seeds) >= 10
    )

    summary = {
        "ingest_version": INGEST_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "hansem_frozen": len(frozen),
        "musinsa_actual_report_body_seeds": ms_report_body,
        "raysolution_actual_report_body_seeds": rx_report_body,
        "musinsa_company_primary_narrative_seeds": ms_primary,
        "raysolution_company_primary_narrative_seeds": rx_primary,
        "musinsa_body_seeds": ms_report_body,
        "raysolution_body_seeds": rx_report_body,
        "salvage_seeds": salvage,
        "musinsa_new_seeds": len(ms_seeds),
        "raysolution_new_seeds": len(rx_seeds),
        "workbook_total": len(all_seeds),
        "unique_clusters": len(clusters),
        "seeds_by_company": dict(by_co),
        "clusters": dict(clusters),
        "seed_origin_breakdown": dict(ms_origins + rx_origins),
        "ingested_sources": ingested_sources,
        "audit": audit,
        "less_press_dependent": less_press,
        "review_ready_verdict": (
            "Gần hơn v2 về provenance, nhưng chưa đủ review-ready"
            if not review_ready
            else "Gần review-ready pilot"
        ),
        "conclusion": (
            "V3 cải thiện provenance (company newsroom) nhưng chưa có full report PDF cho cả hai công ty."
        ),
        "next_step": "Source acquisition: Musinsa Impact Report PDF + 레이시온 2024 SR PDF vào package _sources/",
        "seed_ids": [s["seed_id"] for s in all_seeds],
    }

    write_audit_report(audit, audit_path)
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Ingest actual ESG sources and build workbook v3")
    parser.add_argument(
        "--canonical",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_canonical_r2.jsonl",
    )
    parser.add_argument(
        "--musinsa-pool",
        default="data/golden_set/v2/reference_style/source_pool_musinsa_v3.jsonl",
    )
    parser.add_argument(
        "--raysolution-pool",
        default="data/golden_set/v2/reference_style/source_pool_raytheon_v3.jsonl",
    )
    parser.add_argument(
        "--candidates",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v3.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_v3.xlsx",
    )
    parser.add_argument("--audit", default="reports/golden_set_actual_source_audit_v3.md")
    parser.add_argument(
        "--report", default="reports/golden_set_reference_workbook_v3_actual_sources.md"
    )
    parser.add_argument("--summary-json", default="reports/_reference_workbook_v3_actual_sources_summary.json")
    args = parser.parse_args(argv)

    summary = run_ingest_v3(
        root=root,
        canonical_path=root / args.canonical,
        musinsa_pool_path=root / args.musinsa_pool,
        raysolution_pool_path=root / args.raysolution_pool,
        candidates_path=root / args.candidates,
        workbook_path=root / args.workbook,
        audit_path=root / args.audit,
    )
    write_workbook_report(summary, root / args.report)
    (root / args.summary_json).write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "ms_report_body": summary["musinsa_actual_report_body_seeds"],
                "ms_primary": summary["musinsa_company_primary_narrative_seeds"],
                "rx_report_body": summary["raysolution_actual_report_body_seeds"],
                "salvage": summary["salvage_seeds"],
                "total": summary["workbook_total"],
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
