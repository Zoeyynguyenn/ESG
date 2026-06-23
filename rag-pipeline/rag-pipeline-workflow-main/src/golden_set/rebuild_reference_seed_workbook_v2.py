"""Rebuild reference seed workbook v2 — Musinsa + Raysolution from cleaner sources."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import (
    ESG_SENTENCE_KEYWORDS,
    FINANCIAL_NOISE_PATTERNS,
    NOISE_KEYWORDS,
    OTHER_COMPANY_MARKERS,
    _best_sentences,
    _make_question,
    _prohibited_claims,
    _split_sentences,
)
from golden_set.canonicalize_reference_seed_workbook_r2 import FACT_CLUSTER_RULES
from golden_set.io_utils import read_jsonl, write_jsonl

REBUILD_VERSION = "ref_workbook_v2"
TARGET_COMPANIES = ("무신사", "레이시온")

# Extra clusters for Musinsa / Raysolution rebuild (not in R2 Hansem set)
REBUILD_CLUSTER_RULES: List[Tuple[str, List[re.Pattern], List[re.Pattern]]] = [
    (
        "FC_IMPACT_REPORT",
        [re.compile(p) for p in [r"임팩트 리포트", r"ESG 보고서", r"리포트.*발간"]],
        [re.compile(p) for p in [r"임팩트 리포트", r"ESG 보고서", r"발간"]],
    ),
    (
        "FC_CLIMATE_GHG",
        [re.compile(p) for p in [r"온실가스", r"기후 변화", r"기후변화"]],
        [re.compile(p) for p in [r"온실가스", r"기후 변화", r"기후변화"]],
    ),
    (
        "FC_STAKEHOLDER_DISCLOSURE",
        [re.compile(p) for p in [r"이해관계자", r"투명하게 공개"]],
        [re.compile(p) for p in [r"이해관계자", r"투명하게 공개"]],
    ),
    (
        "FC_COMMUNITY_DONATION",
        [re.compile(p) for p in [r"기부", r"지역사회", r"이재민"]],
        [re.compile(p) for p in [r"기부", r"이재민"]],
    ),
    (
        "FC_EXTERNAL_DIRECTOR",
        [re.compile(p) for p in [r"사외이사", r"이사회 독립"]],
        [re.compile(p) for p in [r"사외이사", r"이사회 독립"]],
    ),
]

ALL_CLUSTER_RULES = REBUILD_CLUSTER_RULES + FACT_CLUSTER_RULES

MUSINSA_SALVAGE_RECORDS = {"rec_753b38dd437e46d0", "rec_609ec5bf54523097"}
RAYSOLUTION_SALVAGE_RECORDS = {"rec_40b94d178a66babb"}
MUSINSA_HEADLINE_RECORD = "rec_0f148d12e3f024bc"

PORTAL_MARKERS = [
    "esg 소개 esg 기업정보",
    "게시판 목록",
    "조회 전체 건",
    "민원서비스",
    "정보공개제도",
    "항만운영과 개발",
    "여수항 일반현황",
    "faq english",
    "esg 강의실",
    "만족도 평가",
    "글 작성시 입력하신 비밀번호",
]

LISTING_MARKERS = ["table of contents", "목차", "appendix", "about this report", "보고서 개요"]

NEWS_MARKERS = ["기자", "발행일", "구독", "송고", "연합뉴스", "주소복사", "네이버 채널"]

CROSS_COMPANY_NAMES = [
    "현대트랜시스",
    "삼성전기",
    "삼성전자",
    "에이피알",
    "여수광양항만공사",
    "lg헬로비전",
    "코스맥스",
]


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _match_rules(text: str, patterns: Sequence[re.Pattern]) -> bool:
    return any(p.search(text) for p in patterns)


def infer_rebuild_fact_target(question: str, passage: str = "") -> str:
    blob = f"{question} {passage}"
    for cluster_id, q_patterns, _ in ALL_CLUSTER_RULES:
        if _match_rules(question or "", q_patterns) or _match_rules(blob, q_patterns):
            return cluster_id
    return "FC_UNKNOWN"


def _has_embedded_esg_fact(company: str, text: str) -> bool:
    for sent in _extract_company_sentences(company, text):
        if _sentence_esg_score(sent) >= 3:
            return True
    return False


def _classify_source(row: Dict[str, Any]) -> Tuple[str, str, str]:
    """Return (cleaning_decision, cleaning_reason, substance_class)."""
    text = row.get("text") or ""
    blob = _norm_ws(text).lower()
    company = row.get("company") or ""

    if any(name in text for name in CROSS_COMPANY_NAMES if name != company):
        if company == "레이시온" and "여수광양항만공사" in text and "공사의 esg" not in blob:
            return "drop_before_seed_generation", "cross_company_contamination", "cross_company_contamination"
    if any(m in blob for m in OTHER_COMPANY_MARKERS):
        return "drop_before_seed_generation", "cross_company_contamination", "cross_company_contamination"

    portal_hits = sum(1 for m in PORTAL_MARKERS if m in blob)
    if portal_hits >= 2 and not _has_embedded_esg_fact(company, text):
        return "drop_before_seed_generation", "portal_nav_contact", "contact_navigation_site_text"
    if portal_hits >= 2 and _has_embedded_esg_fact(company, text):
        return "conditional", "portal_with_embedded_esg_fact", "governance_or_policy_grounded"

    if any(m in blob for m in LISTING_MARKERS) and "중대" not in blob:
        return "drop_before_seed_generation", "listing_archive_index", "listing_archive_index"

    fin_hits = sum(1 for p in FINANCIAL_NOISE_PATTERNS if p in blob)
    esg_hits = sum(1 for k in ESG_SENTENCE_KEYWORDS if k in blob)
    has_esg_headline = any(
        k in blob for k in ("사외이사", "기부", "이재민", "온실가스", "esg", "임팩트 리포트")
    )
    if fin_hits >= 2 and esg_hits < 2 and "esg" not in blob and not has_esg_headline:
        return "drop_before_seed_generation", "financial_irrelevant", "financial_or_irrelevant_non_esg"
    if fin_hits >= 2 and has_esg_headline and _has_embedded_esg_fact(company, text):
        return "conditional", "listing_with_esg_headlines", "press_release_mixed"

    if "analyst" in blob and "@kbfg.com" in blob:
        return "drop_before_seed_generation", "analyst_report", "financial_or_irrelevant_non_esg"

    if row.get("source_type") == "annual_report" and "dart disclosure snapshot" in blob:
        return "drop_before_seed_generation", "dart_metadata_only", "too_generic_no_company_fact"

    sentences = _extract_company_sentences(company, text)
    if not sentences:
        if sum(1 for m in NEWS_MARKERS if m in blob) >= 1 and esg_hits >= 2:
            return "conditional", "press_release_mixed", "press_release_mixed"
        return "drop_before_seed_generation", "no_company_esg_fact", "too_generic_no_company_fact"

    if sum(1 for m in NEWS_MARKERS if m in blob) >= 1:
        return "conditional", "press_release_with_esg_fact", "press_release_mixed"

    if esg_hits >= 3 and portal_hits == 0:
        return "keep_for_seed_generation", "report_body_or_esg_narrative", "report_body_narrative"

    if esg_hits >= 2:
        return "conditional", "esg_fact_in_mixed_page", "governance_or_policy_grounded"

    return "drop_before_seed_generation", "insufficient_esg_substance", "too_generic_no_company_fact"


def _extract_company_sentences(company: str, text: str) -> List[str]:
    """Pull sentences that mention company and carry ESG substance."""
    out: List[str] = []
    seen = set()
    for raw in _split_sentences(text):
        s = _norm_ws(raw)
        if len(s) < 30:
            continue
        lower = s.lower()
        if company not in s and company not in lower:
            # Raysolution uses "공사" in some ESG lines
            if company == "레이시온" and "공사" in s and "esg" in lower:
                pass
            else:
                continue
        if any(name in s for name in CROSS_COMPANY_NAMES):
            continue
        if sum(1 for k in ESG_SENTENCE_KEYWORDS if k in lower) < 1 and "esg" not in lower:
            if not re.search(r"\d", s):
                continue
        key = s[:80]
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
    scored = sorted(out, key=lambda x: (_sentence_esg_score(x), len(x)), reverse=True)
    return scored[:6]


def _sentence_esg_score(sentence: str) -> int:
    lower = sentence.lower()
    score = sum(1 for k in ESG_SENTENCE_KEYWORDS if k in lower)
    if "esg" in lower:
        score += 2
    if re.search(r"\d", sentence):
        score += 1
    return score


def _build_clean_passage(sentences: List[str], limit: int = 3) -> str:
    return " ".join(sentences[:limit])


def _priority(decision: str, score: int) -> str:
    if decision == "keep_for_seed_generation":
        return "high"
    if decision == "conditional" and score >= 4:
        return "medium"
    return "low"


def audit_source_pool(rows: Sequence[Dict[str, Any]], company: str) -> List[Dict[str, Any]]:
    pool: List[Dict[str, Any]] = []
    for row in rows:
        if row.get("company") != company:
            continue
        decision, reason, substance = _classify_source(row)
        sentences = _extract_company_sentences(company, row.get("text") or "")
        score = max((_sentence_esg_score(s) for s in sentences), default=0)
        clean = _build_clean_passage(sentences) if sentences else ""
        pool.append(
            {
                "company": company,
                "record_id": row.get("record_id"),
                "source_file": row.get("source_file"),
                "source_type": row.get("source_type"),
                "section_path": row.get("section_path"),
                "cleaning_decision": decision,
                "cleaning_reason": reason,
                "substance_class": substance,
                "clean_passage": clean,
                "noise_notes": f"portal/news classification; salvageable_sentences={len(sentences)}",
                "seed_candidate_priority": _priority(decision, score),
                "esg_sentence_score": score,
            }
        )
    pool.sort(
        key=lambda x: (
            0 if x["cleaning_decision"] == "keep_for_seed_generation" else 1,
            -x.get("esg_sentence_score", 0),
        )
    )
    return pool


def _seed_id(company: str, qtype: str, index: int) -> str:
    prefix = {"무신사": "MS", "레이시온": "RX", "한샘": "HS"}.get(company, "XX")
    code = {"quantitative": "Q", "trend": "T", "qualitative": "L"}.get(qtype, "X")
    return f"{prefix}-G-{code}{index:02d}"


def _headline_facts_musinsa(text: str) -> List[Tuple[str, str, str]]:
    """Return (fact_text, question, cluster_hint) from listing headlines."""
    facts: List[Tuple[str, str, str]] = []
    if "사외이사" in text and "이사회" in text:
        facts.append(
            (
                "무신사, 첫 사외이사 선임…이사회 독립성·투명성 강화",
                "무신사는 이사회 독립성과 투명성을 위해 어떤 조치를 취했는가?",
                "FC_EXTERNAL_DIRECTOR",
            )
        )
    if "기부" in text and "이재민" in text:
        facts.append(
            (
                "무신사, 산불 피해 이재민에 2억5000만원 상당 의류 기부",
                "무신사는 산불 피해 이재민 지원을 위해 어떤 활동을 했는가?",
                "FC_COMMUNITY_DONATION",
            )
        )
    return facts


def _musinsa_press_facts(text: str) -> List[Tuple[str, str, str]]:
    """Extract structured facts from Yonhap Impact Report press article."""
    facts: List[Tuple[str, str, str]] = []
    patterns: List[Tuple[str, str, str, str]] = [
        (
            r"무신사는 지난해 ESG[^']*'[^']+임팩트 리포트'[^.]*발간[^.]*\.",
            "무신사는 2024년 ESG 성과를 어떤 보고서로 공개했는가?",
            "FC_IMPACT_REPORT",
        ),
        (
            r"환경 부문은 온실가스 관리[^.]*기후 변화 대응[^.]*\.",
            "무신사는 환경 부문에서 어떤 기후·온실가스 대응 활동을 추진했는가?",
            "FC_CLIMATE_GHG",
        ),
        (
            r"지배구조 부문은 이사회의 독립적인 운영[^.]*\.",
            "무신사는 지배구조 부문에서 어떤 ESG 거버넌스 성과를 보고했는가?",
            "FC_ESG_GOVERNANCE",
        ),
    ]
    blob = _norm_ws(text)
    for pat, question, cluster in patterns:
        m = re.search(pat, blob)
        if m:
            facts.append((_norm_ws(m.group(0)), question, cluster))
    return facts


def _raysolution_esg_facts(text: str) -> List[Tuple[str, str, str]]:
    facts: List[Tuple[str, str, str]] = []
    m = re.search(
        r"공사의 ESG경영 활동에 대한 이해관계자[^.]*투명하게 공개[^.]*\.",
        _norm_ws(text),
    )
    if m:
        facts.append(
            (
                _norm_ws(m.group(0)),
                "레이시온(공사)은 ESG 경영 활동과 성과를 이해관계자에게 어떻게 공개하는가?",
                "FC_STAKEHOLDER_DISCLOSURE",
            )
        )
    return facts


def _corpus_text_by_record(corpus: Sequence[Dict[str, Any]], company: str) -> Dict[str, str]:
    out: Dict[str, str] = {}
    for row in corpus:
        if row.get("company") == company and row.get("record_id"):
            out[str(row["record_id"])] = row.get("text") or ""
    return out


def generate_seeds_from_pool(
    pool: Sequence[Dict[str, Any]],
    company: str,
    *,
    corpus_text: Optional[Dict[str, str]] = None,
    start_index: int = 1,
) -> List[Dict[str, Any]]:
    seeds: List[Dict[str, Any]] = []
    used_clusters: set[str] = set()
    idx = start_index

    def _add_fact_seed(
        fact: str,
        question: str,
        cluster: str,
        record_id: str,
        notes: str,
        qtype: str = "qualitative",
    ) -> None:
        nonlocal idx
        if cluster in used_clusters or cluster == "FC_UNKNOWN":
            return
        seeds.append(
            _seed_row(
                company=company,
                seed_id=_seed_id(company, qtype, idx),
                qtype=qtype,
                question=question,
                disclosure=fact[:420],
                record_id=record_id,
                fact_target=cluster,
                notes=notes,
            )
        )
        used_clusters.add(cluster)
        idx += 1

    text_lookup = corpus_text or {}

    if company == "무신사":
        # Prefer single best press article (rec_753) over duplicate Yonhap captures
        for p in pool:
            rid = p.get("record_id") or ""
            if rid not in MUSINSA_SALVAGE_RECORDS:
                continue
            raw_text = text_lookup.get(rid) or p.get("clean_passage") or ""
            for fact, question, cluster in _musinsa_press_facts(raw_text):
                _add_fact_seed(
                    fact,
                    question,
                    cluster,
                    rid,
                    "press_release_fact_salvage",
                )
        for p in pool:
            if p.get("record_id") != MUSINSA_HEADLINE_RECORD:
                continue
            headline_text = text_lookup.get(MUSINSA_HEADLINE_RECORD) or p.get("clean_passage") or ""
            for fact, question, cluster in _headline_facts_musinsa(headline_text):
                _add_fact_seed(
                    fact,
                    question,
                    cluster,
                    p["record_id"],
                    "headline_fact_salvage_from_listing_page",
                    "quantitative" if re.search(r"\d", fact) else "qualitative",
                )

    if company == "레이시온":
        for p in pool:
            rid = p.get("record_id") or ""
            if rid not in RAYSOLUTION_SALVAGE_RECORDS:
                continue
            passage = text_lookup.get(rid) or p.get("clean_passage") or ""
            for fact, question, cluster in _raysolution_esg_facts(passage):
                _add_fact_seed(
                    fact,
                    question,
                    cluster,
                    rid,
                    "portal_embedded_esg_fact_salvage",
                )

    return seeds


def _seed_row(
    *,
    company: str,
    seed_id: str,
    qtype: str,
    question: str,
    disclosure: str,
    record_id: str,
    fact_target: str,
    notes: str,
) -> Dict[str, Any]:
    return {
        "seed_id": seed_id,
        "company": company,
        "question_type": qtype,
        "question_draft": question,
        "acceptable_disclosure": disclosure,
        "prohibited_claims": _prohibited_claims(qtype),
        "source_record_id": record_id,
        "fact_target": fact_target,
        "fact_cluster_id": fact_target,
        "seed_status": "candidate_v2",
        "seed_notes": notes,
        "seed_version": REBUILD_VERSION,
        "anchor_type": "rebuilt",
    }


def _frozen_hansem_anchors(canonical_path: Path) -> List[Dict[str, Any]]:
    rows = read_jsonl(canonical_path)
    frozen: List[Dict[str, Any]] = []
    for row in rows:
        frozen.append(
            {
                "seed_id": row["seed_id"],
                "company": row["company"],
                "question_type": row.get("question_type"),
                "question_draft": row.get("canonical_question") or row.get("question_draft"),
                "acceptable_disclosure": row.get("canonical_acceptable_disclosure")
                or row.get("acceptable_disclosure"),
                "prohibited_claims": row.get("canonical_prohibited_claims")
                or row.get("prohibited_claims"),
                "source_record_id": row.get("source_record_id"),
                "fact_target": row.get("fact_target"),
                "fact_cluster_id": row.get("fact_cluster_id"),
                "seed_status": "frozen_canonical_r2",
                "seed_notes": "frozen Hansem canonical anchor — do not edit",
                "seed_version": REBUILD_VERSION,
                "anchor_type": "frozen",
            }
        )
    return frozen


def _write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "안내"
    ws.append(["Reference Seed Workbook V2 — 3-company rebuild"])
    ws.append(["Hansem", "4 frozen canonical anchors"])
    ws.append(["Musinsa / Raysolution", "rebuilt from cleaner source pool"])

    header = [
        "seed_id",
        "company",
        "anchor_type",
        "question_type",
        "question_draft",
        "acceptable_disclosure",
        "prohibited_claims",
        "source_record_id",
        "fact_cluster_id",
        "fact_target",
        "seed_status",
        "seed_notes",
    ]
    ws2 = wb.create_sheet("작성")
    ws2.append(header)
    fill = PatternFill("solid", fgColor="14532D")
    font = Font(color="FFFFFF", bold=True)
    for c in range(1, len(header) + 1):
        cell = ws2.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
    for row in rows:
        ws2.append([row.get(h, "") for h in header])

    ws3 = wb.create_sheet("요약")
    ws3.append(["metric", "value"])
    ws3.append(["total", len(rows)])
    for co, n in sorted(Counter(r["company"] for r in rows).items()):
        ws3.append([f"company:{co}", n])
    for cid, n in sorted(Counter(r.get("fact_cluster_id") for r in rows).items()):
        ws3.append([f"cluster:{cid}", n])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_rebuild_report(summary: Dict[str, Any], path: Path) -> None:
    lines = [
        "# Golden Set — Reference Workbook Rebuild V2",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Khôi phục hướng **workbook-first 3 công ty**: giữ 4 Hansem canonical frozen, rebuild seed pool cho **무신사** và **레이시온** từ source sạch hơn.",
        "",
        "## Vì sao R2 chưa đủ",
        "",
        "- R2 chỉ còn **4 seed Hansem-only** — không phải workbook 3 công ty.",
        "- 무신사 biến mất vì source corpus chủ yếu là **portal/nav** và **news/financial**.",
        "- 레이시온 corpus gần như **항만 portal + 정보공개**, chỉ vài câu ESG disclosure salvageable.",
        "",
        "## Audit source — 무신사",
        "",
        f"- Tổng unit corpus: **{summary.get('musinsa_units_total', 0)}**",
        f"- Pool keep/conditional: **{summary.get('musinsa_pool_eligible', 0)}**",
        f"- Drop chính: portal/nav, news/financial, listing — xem `source_pool_musinsa_r2.jsonl`",
        "",
        "## Audit source — 레이시온",
        "",
        f"- Tổng unit corpus: **{summary.get('raysolution_units_total', 0)}**",
        f"- Pool keep/conditional: **{summary.get('raysolution_pool_eligible', 0)}**",
        f"- Drop chính: portal/nav, 민원/정보공개, cross-company contamination",
        "",
        "## Quy tắc chọn source sạch",
        "",
        "- **keep**: ESG narrative/fact có company name, substance ≥ 2 signals",
        "- **conditional**: press release mixed nhưng trích được câu fact company",
        "- **drop**: portal, listing, financial/analyst, cross-company, DART metadata",
        "",
        "## Kết quả rebuild",
        "",
        f"| Thành phần | Số seed |",
        f"|-----------|--------:|",
        f"| Hansem frozen anchors | {summary.get('hansem_frozen', 0)} |",
        f"| Musinsa new seeds | {summary.get('musinsa_new_seeds', 0)} |",
        f"| Raysolution new seeds | {summary.get('raysolution_new_seeds', 0)} |",
        f"| **Tổng workbook v2** | **{summary.get('workbook_total', 0)}** |",
        "",
        "## Coverage theo công ty",
        "",
    ]
    for co, n in summary.get("seeds_by_company", {}).items():
        lines.append(f"- **{co}**: {n} seed")
    lines.extend(["", "## Coverage theo fact cluster", ""])
    for cid, n in summary.get("clusters", {}).items():
        lines.append(f"- `{cid}`: {n}")

    lines.extend(
        [
            "",
            "## Những chỗ vẫn thiếu source sạch",
            "",
        ]
    )
    for gap in summary.get("coverage_gaps", []):
        lines.append(f"- {gap}")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- **Workbook v2 gần 3-company review-ready?** {summary.get('review_ready_verdict', '')}",
            f"- **Thiếu ở đâu:** {summary.get('missing_where', '')}",
            "",
            "### Ba câu trả lời",
            "",
            f"1. Musinsa rebuild: **{summary.get('musinsa_new_seeds', 0)}** seed mới",
            f"2. Raysolution rebuild: **{summary.get('raysolution_new_seeds', 0)}** seed mới",
            f"3. Tổng coverage: **{summary.get('workbook_total', 0)}** seed / **{summary.get('unique_clusters', 0)}** fact cluster — {summary.get('coverage_verdict', '')}",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_rebuild_v2(
    *,
    corpus_path: Path,
    canonical_path: Path,
    musinsa_pool_path: Path,
    raysolution_pool_path: Path,
    candidates_path: Path,
    workbook_path: Path,
) -> Dict[str, Any]:
    corpus = read_jsonl(corpus_path)
    frozen = _frozen_hansem_anchors(canonical_path)

    musinsa_pool = audit_source_pool(corpus, "무신사")
    raysolution_pool = audit_source_pool(corpus, "레이시온")
    write_jsonl(musinsa_pool_path, musinsa_pool)
    write_jsonl(raysolution_pool_path, raysolution_pool)

    musinsa_text = _corpus_text_by_record(corpus, "무신사")
    rx_text = _corpus_text_by_record(corpus, "레이시온")
    musinsa_seeds = generate_seeds_from_pool(
        musinsa_pool, "무신사", corpus_text=musinsa_text, start_index=1
    )
    raysolution_seeds = generate_seeds_from_pool(
        raysolution_pool, "레이시온", corpus_text=rx_text, start_index=1
    )

    all_seeds = frozen + musinsa_seeds + raysolution_seeds
    write_jsonl(candidates_path, all_seeds)
    _write_workbook(all_seeds, workbook_path)

    musinsa_units = sum(1 for r in corpus if r.get("company") == "무신사")
    rx_units = sum(1 for r in corpus if r.get("company") == "레이시온")
    musinsa_eligible = sum(
        1 for p in musinsa_pool if p["cleaning_decision"] in {"keep_for_seed_generation", "conditional"}
    )
    rx_eligible = sum(
        1 for p in raysolution_pool if p["cleaning_decision"] in {"keep_for_seed_generation", "conditional"}
    )

    clusters = Counter(s.get("fact_cluster_id") for s in all_seeds)
    by_co = Counter(s["company"] for s in all_seeds)

    companies_with_seed = sum(1 for co in ("한샘", "무신사", "레이시온") if by_co.get(co, 0) > 0)
    min_non_hansem = min(by_co.get("무신사", 0), by_co.get("레이시온", 0))
    # Honest gate: need meaningful depth per company, not just presence
    review_ready = (
        companies_with_seed == 3
        and min_non_hansem >= 3
        and len(all_seeds) >= 10
    )

    summary = {
        "rebuild_version": REBUILD_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "hansem_frozen": len(frozen),
        "musinsa_new_seeds": len(musinsa_seeds),
        "raysolution_new_seeds": len(raysolution_seeds),
        "workbook_total": len(all_seeds),
        "unique_clusters": len(clusters),
        "seeds_by_company": dict(by_co),
        "clusters": dict(clusters),
        "musinsa_units_total": musinsa_units,
        "raysolution_units_total": rx_units,
        "musinsa_pool_eligible": musinsa_eligible,
        "raysolution_pool_eligible": rx_eligible,
        "musinsa_pool_drop": sum(1 for p in musinsa_pool if p["cleaning_decision"].startswith("drop")),
        "raysolution_pool_drop": sum(1 for p in raysolution_pool if p["cleaning_decision"].startswith("drop")),
        "coverage_gaps": [
            "무신사: thiếu report body narrative sạch — phụ thuộc press release / headline salvage",
            "레이시온: thiếu metric/governance narrative — chỉ có stakeholder disclosure sentence",
            "Cả hai: cần ingest PDF sustainability report thật (Impact Report / 2024 SR) để mở rộng cluster",
        ],
        "review_ready_verdict": (
            "**Chưa đủ** — có mặt 3 công ty nhưng Musinsa/Raysolution còn mỏng (press/headline salvage; RX chỉ 1 cluster)"
            if not review_ready
            else "**Gần đủ** cho review nội dung pilot"
        ),
        "missing_where": "무신사 (narrative report body), 레이시온 (ESG metric/governance body)",
        "coverage_verdict": (
            "partial 3-company workbook — đủ làm format anchor + pilot review, chưa đủ golden workbook đầy đủ"
        ),
        "seed_ids": [s["seed_id"] for s in all_seeds],
    }
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Rebuild reference seed workbook v2")
    parser.add_argument("--corpus", default="data/golden_set/v2/step1_corpus_units/corpus_units.jsonl")
    parser.add_argument(
        "--canonical",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_canonical_r2.jsonl",
    )
    parser.add_argument(
        "--musinsa-pool",
        default="data/golden_set/v2/reference_style/source_pool_musinsa_r2.jsonl",
    )
    parser.add_argument(
        "--raysolution-pool",
        default="data/golden_set/v2/reference_style/source_pool_raytheon_r2.jsonl",
    )
    parser.add_argument(
        "--candidates",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v2.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_v2.xlsx",
    )
    parser.add_argument("--report", default="reports/golden_set_reference_workbook_rebuild_v2.md")
    parser.add_argument("--summary-json", default="reports/_reference_workbook_rebuild_v2_summary.json")
    args = parser.parse_args(argv)

    summary = run_rebuild_v2(
        corpus_path=root / args.corpus,
        canonical_path=root / args.canonical,
        musinsa_pool_path=root / args.musinsa_pool,
        raysolution_pool_path=root / args.raysolution_pool,
        candidates_path=root / args.candidates,
        workbook_path=root / args.workbook,
    )
    write_rebuild_report(summary, root / args.report)
    (root / args.summary_json).write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(
        json.dumps(
            {
                "musinsa": summary["musinsa_new_seeds"],
                "raysolution": summary["raysolution_new_seeds"],
                "total": summary["workbook_total"],
                "clusters": summary["unique_clusters"],
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
