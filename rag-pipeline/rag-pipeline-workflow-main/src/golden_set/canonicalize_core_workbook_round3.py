"""Canonical Round 3 — core workbook for 한샘 + 무신사 only."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import METRIC_HINT_RE
from golden_set.canonicalize_reference_seed_workbook_r2 import (
    FACT_CLUSTER_RULES,
    TCFD_DEFINITION_ONLY,
    TRUNCATED_ENDINGS,
    _match_rules,
    infer_fact_target_id,
)
from golden_set.io_utils import read_jsonl, write_jsonl
from golden_set.prepare_manual_review_round2 import GENERIC_Q, _norm_ws

CANONICAL_VERSION = "ref_core_canonical_r3"
CORE_COMPANIES = frozenset({"한샘", "무신사"})

KEEP_DECISIONS = frozenset({"canonical_keep", "canonical_keep_after_merge"})
DROP_DUP = "canonical_drop_duplicate"
DROP_WEAK = "canonical_drop_weak"

META_HEAVY = [
    "sasb index",
    "tcfd index",
    "목차",
    "table of contents",
    "기자",
    "발행일",
    "모바일버전",
    "home 경제",
]

WORKBOOK_CANON_COLS = [
    "seed_id",
    "company",
    "question_type",
    "cluster_id",
    "manual_decision",
    "final_question",
    "final_disclosure",
    "canonical_decision",
    "canonical_reason",
    "canonical_anchor_flag",
    "canonical_merge_note",
    "canonical_fact_target",
    "canonical_prohibited_claims",
    "notes",
]


def _final_q(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("final_question") or row.get("question_draft") or "")


def _final_d(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("final_disclosure") or row.get("acceptable_disclosure") or "")


def _is_generic_question(q: str) -> bool:
    return bool(GENERIC_Q.search(q))


def _is_truncated(text: str) -> bool:
    t = _norm_ws(text)
    if len(t) < 50:
        return True
    if t.startswith(("완성하고", "또한")) and "한샘" not in t[:40] and "무신사" not in t[:40]:
        return True
    return any(t.endswith(end) for end in TRUNCATED_ENDINGS)


def _is_meta_heavy(disclosure: str) -> bool:
    lower = disclosure.lower()
    return sum(1 for m in META_HEAVY if m in lower) >= 1 and not METRIC_HINT_RE.search(disclosure)


def _specificity_score(q: str, disclosure: str) -> float:
    score = 0.0
    if not _is_generic_question(q):
        score += 5
    if len(disclosure) >= 80:
        score += 2
    if METRIC_HINT_RE.search(disclosure):
        score += 3
    if re.search(r"\d", disclosure):
        score += 2
    if _is_meta_heavy(disclosure):
        score -= 4
    if _is_truncated(disclosure):
        score -= 5
    if TCFD_DEFINITION_ONLY.search(disclosure) and "한샘" not in disclosure and "무신사" not in disclosure:
        score -= 6
    return score


def _canonical_question(company: str, fact_target: str, seed_q: str) -> str:
    templates = {
        "FC_NET_ZERO_2050": f"{company}는 2050년까지 어떤 탄소중립 목표를 공개했는가?",
        "FC_BOARD_2022": f"{company}는 2022년 이사회를 몇 회 개최하고 몇 건의 안건을 심의했는가?",
        "FC_ESG_GOVERNANCE": f"{company}의 ESG 거버넌스 체계는 어떻게 운영되는가?",
        "FC_TCFD": f"{company}는 기후변화 대응 공시에 어떤 프레임워크를 따르는가?",
        "FC_MATERIAL_8": f"{company}는 이중 중대성 평가를 통해 몇 개의 중대 이슈를 선정했는가?",
        "FC_KGCS_A": f"{company}는 KGCS ESG경영 평가에서 어떤 등급을 획득했는가?",
        "FC_HUMAN_RIGHTS": f"{company}는 인권경영 체계를 어떻게 강화했는가?",
    }
    if infer_fact_target_id(seed_q, "") == fact_target and not _is_generic_question(seed_q):
        return seed_q
    return templates.get(fact_target, seed_q)


def _trim_disclosure(disclosure: str, max_len: int = 420) -> str:
    text = _norm_ws(disclosure)
    if len(text) <= max_len:
        return text
    for sent in re.split(r"(?<=[.!?다])\s+", text):
        if len(sent) >= 50:
            return sent[:max_len]
    return text[:max_len]


def _intent_key(company: str, fact_target: str, question: str) -> str:
    q = _norm_ws(question).lower()
    return f"{company}::{fact_target}::{q[:80]}"


def _cluster_group_key(company: str, fact_target: str, disclosure: str) -> str:
    """Merge near-duplicates sharing fact target + disclosure fingerprint."""
    nums = "|".join(re.findall(r"\d+", disclosure)[:5])
    words = "|".join(re.findall(r"[가-힣]{4,}", disclosure)[:6])
    return f"{company}::{fact_target}::{words}::{nums}"


def _row_strength(row: Dict[str, Any], fact_target: str) -> float:
    q = _final_q(row)
    d = _final_d(row)
    base = float(row.get("rank") or 0)
    base += _specificity_score(q, d)
    if row.get("manual_decision") == "confirm":
        base += 3
    if row.get("canonical_cluster_action") == "anchor":
        base += 2
    return base


def _weak_drop_reason(row: Dict[str, Any], fact_target: str) -> Optional[str]:
    d = _final_d(row)
    q = _final_q(row)
    if not d or len(d) < 45:
        return "disclosure_too_short"
    if _is_truncated(d):
        return "truncated_disclosure"
    if fact_target == "FC_UNKNOWN" and _is_generic_question(q) and len(d) < 100:
        return "unknown_target_generic"
    if TCFD_DEFINITION_ONLY.search(d) and fact_target != "FC_TCFD":
        return "framework_definition_not_company_fact"
    if _is_meta_heavy(d) and not METRIC_HINT_RE.search(d):
        return "meta_heavy_disclosure"
    if fact_target == "FC_TCFD" and TCFD_DEFINITION_ONLY.search(d):
        return "tcfd_definition_only"
    return None


def preprocess_row(row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row)
    q = _final_q(row)
    d = _final_d(row)
    company = row.get("company") or ""
    fact_target = infer_fact_target_id(q, d)
    if fact_target == "FC_UNKNOWN":
        blob = f"{q} {d}"
        for cluster_id, q_patterns, _ in FACT_CLUSTER_RULES:
            if _match_rules(blob, q_patterns):
                fact_target = cluster_id
                break

    weak = _weak_drop_reason(row, fact_target)
    prohibited = row.get("prohibited_claims") or "원문에 없는 수치 추가 금지\n미공시 항목 단정 금지"

    if weak:
        out.update(
            {
                "canonical_version": CANONICAL_VERSION,
                "canonical_fact_target": fact_target,
                "canonical_decision": DROP_WEAK,
                "canonical_reason": weak,
                "canonical_anchor_flag": False,
                "canonical_merge_note": "",
                "final_question": q,
                "final_disclosure": d,
                "canonical_prohibited_claims": prohibited,
                "notes": f"Weak drop: {weak}",
                "_strength": -999,
            }
        )
        return out

    canon_q = _canonical_question(company, fact_target, q)
    canon_d = _trim_disclosure(d)
    out.update(
        {
            "canonical_version": CANONICAL_VERSION,
            "canonical_fact_target": fact_target,
            "cluster_id": row.get("cluster_id") or fact_target,
            "final_question": canon_q,
            "final_disclosure": canon_d,
            "canonical_prohibited_claims": prohibited,
            "_strength": _row_strength(row, fact_target),
            "_group_key": _cluster_group_key(company, fact_target, canon_d),
            "_intent_key": _intent_key(company, fact_target, canon_q),
        }
    )
    return out


def merge_clusters(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    candidates = [r for r in rows if r.get("canonical_decision") != DROP_WEAK]
    weak = [r for r in rows if r.get("canonical_decision") == DROP_WEAK]

    groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in candidates:
        groups[row["_group_key"]].append(row)

    canonical: List[Dict[str, Any]] = []
    dropped: List[Dict[str, Any]] = list(weak)

    for _gk, members in groups.items():
        ranked = sorted(members, key=lambda x: (-x["_strength"], x.get("seed_id", "")))
        anchor = ranked[0]
        merged_from = [m["seed_id"] for m in ranked[1:]]

        if len(ranked) > 1:
            decision = "canonical_keep_after_merge"
            merge_note = f"Anchor {anchor['seed_id']}; merged {len(merged_from)} dup(s): {', '.join(merged_from[:5])}"
            reason = "cluster_anchor_after_merge"
        elif anchor.get("manual_decision") == "revise":
            decision = "canonical_keep_after_merge"
            merge_note = "Single row after manual revise — normalized wording"
            reason = "normalized_from_revise"
        else:
            decision = "canonical_keep"
            merge_note = ""
            reason = "clean_core_anchor"

        anchor.update(
            {
                "canonical_decision": decision,
                "canonical_reason": reason,
                "canonical_anchor_flag": True,
                "canonical_merge_note": merge_note,
                "notes": merge_note or "Core canonical anchor.",
            }
        )
        canonical.append(anchor)

        for loser in ranked[1:]:
            dup = dict(loser)
            dup.update(
                {
                    "canonical_decision": DROP_DUP,
                    "canonical_reason": f"duplicate_of_{anchor['seed_id']}",
                    "canonical_anchor_flag": False,
                    "canonical_merge_note": f"Defer to anchor {anchor['seed_id']}",
                    "final_question": anchor.get("final_question"),
                    "final_disclosure": anchor.get("final_disclosure"),
                    "notes": f"Duplicate cluster — giữ {anchor['seed_id']}",
                }
            )
            dropped.append(dup)

    canonical.sort(key=lambda x: (x.get("company", ""), -x.get("_strength", 0)))
    return canonical, dropped


def _collect_rx_backlog(all_rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    backlog: List[Dict[str, Any]] = []
    for row in all_rows:
        co = row.get("company") or ""
        lane = row.get("manual_review_lane") or ""
        if co == "레이시온":
            backlog.append(row)
        elif lane in ("lane_c_rewrite_heavy", "reject_recommended"):
            backlog.append(row)
        elif row.get("manual_decision", "").startswith("backlog"):
            backlog.append(row)
    return backlog


def write_workbook(
    canonical: Sequence[Dict[str, Any]],
    dropped: Sequence[Dict[str, Any]],
    rx_backlog: Sequence[Dict[str, Any]],
    path: Path,
) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide.append(["Core Canonical Round 3 — 한샘 + 무신사"])
    guide.append(["RX", "Backlog sheet — không thuộc core canonical"])

    sheets = [
        ("Canonical_Core", canonical, "166534"),
        ("Dropped_Duplicates", [r for r in dropped if r.get("canonical_decision") == DROP_DUP], "991B1B"),
        ("Dropped_Weak", [r for r in dropped if r.get("canonical_decision") == DROP_WEAK], "7F1D1D"),
        ("RX_Backlog", rx_backlog, "B45309"),
    ]
    for name, rows, color in sheets:
        ws = wb.create_sheet(name)
        ws.append(WORKBOOK_CANON_COLS)
        fill = PatternFill("solid", fgColor=color)
        font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(WORKBOOK_CANON_COLS) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = fill
            c.font = font
        for r in rows:
            ws.append([r.get(c, "") for c in WORKBOOK_CANON_COLS])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _pick_examples(
    canonical: Sequence[Dict[str, Any]], dropped: Sequence[Dict[str, Any]]
) -> Dict[str, List[Dict[str, str]]]:
    ex: Dict[str, List[Dict[str, str]]] = {"merge": [], "specific": [], "weak": []}
    for r in canonical:
        if r.get("canonical_decision") == "canonical_keep_after_merge" and len(ex["merge"]) < 3:
            ex["merge"].append(
                {
                    "seed_id": r.get("seed_id", ""),
                    "note": r.get("canonical_merge_note", "")[:100],
                }
            )
    for r in dropped:
        if r.get("canonical_decision") == DROP_DUP and len(ex["specific"]) < 3:
            ex["specific"].append(
                {
                    "seed_id": r.get("seed_id", ""),
                    "anchor": r.get("canonical_merge_note", ""),
                }
            )
        if r.get("canonical_decision") == DROP_WEAK and len(ex["weak"]) < 3:
            ex["weak"].append(
                {
                    "seed_id": r.get("seed_id", ""),
                    "reason": r.get("canonical_reason", ""),
                }
            )
    return ex


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Core Canonical Round 3",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Canonicalize core workbook **한샘 + 무신사** từ manual review round 2:",
        "giữ fact target, merge duplicate cluster, chuẩn hóa Q/disclosure.",
        "",
        "## Vì sao chỉ canonical core 한샘 + 무신사",
        "",
        "- Core đủ mạnh: HS 49 + MS 11 canonical candidates sau manual review.",
        "- **레이시온** chỉ 1 survivor — source/data quality gap, tách backlog.",
        "- Không kéo RX vào core flow; không mở Lane C.",
        "",
        "## Input canonical candidates",
        "",
        f"**{summary.get('input_candidates', 0)}** row (`canonical_candidate_flag=true`, HS+MS).",
        "",
        "## Rule merge / dedupe / keep / weak drop",
        "",
        "| Decision | Điều kiện |",
        "|----------|-----------|",
        "| `canonical_keep` | Anchor sạch, confirm, không merge |",
        "| `canonical_keep_after_merge` | Anchor sau merge cluster hoặc normalize từ revise |",
        "| `canonical_drop_duplicate` | Trùng fact cluster/intent, giữ anchor mạnh hơn |",
        "| `canonical_drop_weak` | Meta-heavy, truncated, TCFD definition-only, unknown generic |",
        "",
        "## Kết quả",
        "",
        f"- Input: **{summary.get('input_candidates', 0)}**",
        f"- canonical_keep: **{summary.get('canonical_keep_count', 0)}**",
        f"- keep_after_merge: **{summary.get('canonical_keep_after_merge_count', 0)}**",
        f"- drop_duplicate: **{summary.get('canonical_drop_duplicate_count', 0)}**",
        f"- drop_weak: **{summary.get('canonical_drop_weak_count', 0)}**",
        f"- **Core canonical total:** **{summary.get('core_canonical_total', 0)}**",
        "",
        "### Breakdown theo công ty",
        "",
    ]
    for co, n in summary.get("by_company", {}).items():
        lines.append(f"- **{co}**: {n}")

    lines.extend(["", "### Breakdown theo question_type", ""])
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "### Breakdown theo fact cluster", ""])
    for cid, n in summary.get("by_cluster", {}).items():
        lines.append(f"- `{cid}`: {n}")

    lines.extend(["", "## Ví dụ", ""])
    lines.append("### Duplicate được merge đúng")
    for e in examples.get("merge", []):
        lines.append(f"- `{e['seed_id']}`: {e['note']}")
    lines.append("")
    lines.append("### Row generic thay bằng specific (drop dup)")
    for e in examples.get("specific", []):
        lines.append(f"- `{e['seed_id']}` → {e['anchor']}")
    lines.append("")
    lines.append("### Row weak bị drop")
    for e in examples.get("weak", []):
        lines.append(f"- `{e['seed_id']}`: `{e['reason']}`")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- Core canonical set: **{summary.get('core_canonical_total', 0)}** row",
            f"- Gold decision round cho core? **{summary.get('gold_decision_verdict', '')}**",
            f"- RX backlog: **{summary.get('rx_backlog_count', 0)}** row (source-acquisition dependent)",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_canonical_core(
    *,
    input_path: Path,
    canonical_jsonl: Path,
    dropped_jsonl: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    all_rows = read_jsonl(input_path)
    core_input = [
        r
        for r in all_rows
        if r.get("canonical_candidate_flag")
        and r.get("company") in CORE_COMPANIES
        and r.get("manual_decision") in ("confirm", "revise")
    ]

    preprocessed = [preprocess_row(r) for r in core_input]
    canonical, dropped = merge_clusters(preprocessed)

    rx_backlog = _collect_rx_backlog(all_rows)

    write_jsonl(canonical_jsonl, canonical)
    write_jsonl(dropped_jsonl, dropped)
    write_workbook(canonical, dropped, rx_backlog, workbook_path)

    canon_decisions = Counter(r.get("canonical_decision") for r in canonical)
    drop_decisions = Counter(r.get("canonical_decision") for r in dropped)

    by_co = dict(Counter(r.get("company") for r in canonical))
    by_qtype = dict(Counter(r.get("question_type") for r in canonical))
    by_cluster = dict(Counter(r.get("canonical_fact_target") for r in canonical))

    core_total = len(canonical)
    hs = by_co.get("한샘", 0)
    ms = by_co.get("무신사", 0)

    gold_ready = core_total >= 25 and hs >= 18 and ms >= 5

    summary = {
        "canonical_version": CANONICAL_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_candidates": len(core_input),
        "canonical_keep_count": canon_decisions.get("canonical_keep", 0),
        "canonical_keep_after_merge_count": canon_decisions.get("canonical_keep_after_merge", 0),
        "canonical_drop_duplicate_count": drop_decisions.get(DROP_DUP, 0),
        "canonical_drop_weak_count": drop_decisions.get(DROP_WEAK, 0),
        "core_canonical_total": core_total,
        "by_company": by_co,
        "by_question_type": by_qtype,
        "by_cluster": by_cluster,
        "rx_backlog_count": len(rx_backlog),
        "core_ready_for_gold_decision_flag": gold_ready,
        "gold_decision_verdict": (
            "Có — core đủ chặt để mở gold decision round (không promote trong task này)"
            if gold_ready
            else "Chưa — cần bổ sung fact hoặc human spot-check"
        ),
        "rx_status": "backlog_source_acquisition — 1 survivor + lane C/reject; không trong core canonical",
        "output_canonical_jsonl": str(canonical_jsonl),
        "output_dropped_jsonl": str(dropped_jsonl),
        "output_workbook": str(workbook_path),
    }

    examples = _pick_examples(canonical, dropped)
    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Canonicalize core workbook round 3 (HS+MS)")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_manual_round2_reviewed.jsonl",
    )
    parser.add_argument(
        "--canonical-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_core_canonical_round3.jsonl",
    )
    parser.add_argument(
        "--dropped-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_core_dropped_round3.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_core_canonical_round3.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_core_canonical_round3.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_core_canonical_round3_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_canonical_core(
        input_path=root / args.input,
        canonical_jsonl=root / args.canonical_jsonl,
        dropped_jsonl=root / args.dropped_jsonl,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "input_candidates",
                    "core_canonical_total",
                    "by_company",
                    "core_ready_for_gold_decision_flag",
                    "rx_backlog_count",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
