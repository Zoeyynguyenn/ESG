"""RTX v2.1 Workbook Review Round 1 — triage on fact-quality candidates."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import METRIC_HINT_RE, NUMBER_RE, YEAR_RE
from golden_set.io_utils import read_jsonl, write_jsonl
from golden_set.rtx_fact_quality import audit_candidate_row

REVIEW_VERSION = "ref_review_rtx_v21_r1"

TABLE_KINDS = {"appendix", "data_table", "questionnaire"}
AWKWARD_Q_PATTERNS = [
    re.compile(r"\(disclosed figure includes", re.I),
    re.compile(r"does RTX report for \d{4} \(disclosed", re.I),
]


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _passage(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("acceptable_disclosure") or row.get("source_excerpt") or "")


def _valid_years(text: str) -> List[str]:
    return sorted({y for y in YEAR_RE.findall(text) if 1995 <= int(y) <= 2035})


def _fact_cluster_id(row: Dict[str, Any]) -> str:
    ft = row.get("fact_target") or ""
    base = re.sub(r"\s*\(\d{4}\)\s*", "", ft)
    base = re.sub(r"\s*\[[^\]]+\]\s*", "", base).strip().lower()
    ftype = row.get("fact_target_type") or "unknown"
    reason = row.get("candidate_reason") or ""
    fid = reason.replace("canonical_fact:", "") if "canonical_fact:" in reason else ""
    return f"RTX::{ftype}::{fid or base[:40]}"


def _alignment_score(row: Dict[str, Any]) -> int:
    q = (row.get("question_draft") or "").lower()
    d = _passage(row).lower()
    ft = (row.get("fact_target") or "").lower()
    score = 0
    for tok in re.findall(r"[a-z]{4,}", ft):
        if tok in q:
            score += 1
        if tok in d:
            score += 1
    if any(y in q for y in _valid_years(q + " " + d)):
        score += 1
    return score


def _table_density(text: str) -> int:
    return text.count("|") + text.count("---")


def _reject_reason(row: Dict[str, Any]) -> Optional[Tuple[str, str]]:
    audit_errs = audit_candidate_row(row)
    if audit_errs:
        return audit_errs[0], f"Quality audit: {audit_errs[0]}"

    passage = _passage(row)
    if len(passage) < 30:
        return "weak_disclosure", "Excerpt quá ngắn sau v2.1."
    if _alignment_score(row) < 2:
        return "fact_mismatch", "Question/disclosure không align đủ với fact_target."
    if row.get("question_quality_status") == "dropped":
        return "low_review_value", "Row marked low quality."
    return None


def _trend_is_single_year(row: Dict[str, Any]) -> bool:
    if row.get("question_type") != "trend":
        return False
    years = _valid_years(_passage(row) + " " + (row.get("question_draft") or ""))
    return len(years) < 2


def _needs_rewrite(row: Dict[str, Any]) -> Tuple[bool, str]:
    reasons: List[str] = []
    q = row.get("question_draft") or ""
    passage = _passage(row)
    dk = row.get("document_kind") or ""

    if any(p.search(q) for p in AWKWARD_Q_PATTERNS):
        reasons.append("awkward_parenthetical_phrasing")
    if _trend_is_single_year(row):
        reasons.append("trend_should_be_quantitative")
    if dk in TABLE_KINDS and _table_density(passage) >= 3:
        reasons.append("semi_structured_disclosure_trim")
    if len(passage) > 380:
        reasons.append("disclosure_too_long")
    if q.count("?") > 1:
        reasons.append("question_format_cleanup")

    if not reasons:
        return False, ""
    return True, ";".join(reasons)


def _trim_disclosure(passage: str) -> str:
    if _table_density(passage) < 3:
        return passage[:380].strip()
    best = ""
    for part in re.split(r"\|", passage):
        part = _norm_ws(part)
        if len(part) < 25:
            continue
        score = len(NUMBER_RE.findall(part)) + (2 if METRIC_HINT_RE.search(part) else 0)
        if score > len(NUMBER_RE.findall(best)):
            best = part
    return (best or passage)[:380].strip()


def _suggest_rewrite(row: Dict[str, Any]) -> Tuple[str, str]:
    q = row.get("question_draft") or ""
    passage = _trim_disclosure(_passage(row))
    ft = row.get("fact_target") or ""
    years = _valid_years(passage + " " + q)

    new_q = q
    if _trend_is_single_year(row):
        base = re.sub(r"\s*\(\d{4}\)\s*", "", ft) or ft
        yr = f" for {years[-1]}" if years else ""
        new_q = f"What {base} does RTX report{yr}?"
    else:
        new_q = re.sub(r"\s*\(disclosed figure includes[^)]+\)", "", q).strip()
        if not new_q.endswith("?"):
            new_q += "?"

    new_q = _norm_ws(new_q)
    return new_q, passage


def _initial_decision(row: Dict[str, Any]) -> Tuple[str, str, str]:
    reject = _reject_reason(row)
    if reject:
        return "reject", reject[0], reject[1]

    rewrite, rreason = _needs_rewrite(row)
    if rewrite:
        return "rewrite", rreason, "Fact đúng; chỉnh nhẹ wording/disclosure cho review."

    q = row.get("question_draft") or ""
    if len(q) < 25:
        return "rewrite", "question_too_short", "Cần mở rộng question cho reviewer."

    return "keep", "v21_clean_grounded", "Question fact-specific, disclosure usable."


def _row_strength(row: Dict[str, Any]) -> float:
    rank = float(row.get("rank") or 0)
    bonus = {"keep": 10, "rewrite": 5, "reject": -100, "collapse_into_cluster": -40}
    decision = row.get("review_decision", "keep")
    rank += _alignment_score(row) * 2
    if row.get("document_kind") in ("appendix", "questionnaire"):
        rank += 1
    return rank + bonus.get(decision, 0)


def review_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    reviewed: List[Dict[str, Any]] = []
    for row in rows:
        out = dict(row)
        decision, reason, notes = _initial_decision(row)
        out.update(
            {
                "review_version": REVIEW_VERSION,
                "review_decision": decision,
                "review_reason": reason,
                "review_notes": notes,
                "cluster_id": _fact_cluster_id(row),
                "cluster_action": "rejected" if decision == "reject" else "pending",
                "rewritten_question_draft": "",
                "rewritten_disclosure_draft": "",
            }
        )
        if decision == "rewrite":
            rq, rd = _suggest_rewrite(row)
            out["rewritten_question_draft"] = rq
            out["rewritten_disclosure_draft"] = rd
        elif decision == "keep":
            out["rewritten_question_draft"] = row.get("question_draft") or ""
            out["rewritten_disclosure_draft"] = row.get("acceptable_disclosure") or ""
        reviewed.append(out)

    by_cluster: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in reviewed:
        if row["review_decision"] in ("keep", "rewrite"):
            by_cluster[row["cluster_id"]].append(row)

    for cluster_id, members in by_cluster.items():
        if len(members) <= 1:
            members[0]["cluster_action"] = "anchor"
            continue

        by_fact_base: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for m in members:
            ft = re.sub(r"\s*\(\d{4}\)\s*", "", m.get("fact_target") or "").lower().strip()
            by_fact_base[ft].append(m)

        for _base, group in by_fact_base.items():
            if len(group) <= 1:
                group[0]["cluster_action"] = "anchor"
                continue
            sorted_g = sorted(group, key=_row_strength, reverse=True)
            anchor = sorted_g[0]
            anchor["cluster_action"] = "anchor"
            for dup in sorted_g[1:]:
                dup_years = set(_valid_years(_passage(dup)))
                anchor_years = set(_valid_years(_passage(anchor)))
                if dup_years and anchor_years and dup_years == anchor_years:
                    dup["review_decision"] = "collapse_into_cluster"
                    dup["review_reason"] = "duplicate_year_cluster"
                    dup["review_notes"] = f"Trùng fact/year với anchor {anchor.get('seed_id')}"
                    dup["cluster_action"] = "collapsed_variant"
                    dup["rewritten_question_draft"] = ""
                    dup["rewritten_disclosure_draft"] = ""
                elif len(group) > 3:
                    dup["review_decision"] = "collapse_into_cluster"
                    dup["review_reason"] = "duplicate_fact_cluster"
                    dup["review_notes"] = f"Trùng cụm fact với anchor {anchor.get('seed_id')}"
                    dup["cluster_action"] = "collapsed_variant"
                    dup["rewritten_question_draft"] = ""
                    dup["rewritten_disclosure_draft"] = ""
                else:
                    dup["cluster_action"] = "anchor"

        for m in members:
            if m.get("cluster_action") == "pending":
                m["cluster_action"] = "anchor"

    return reviewed


WORKBOOK_COLUMNS = [
    "seed_id",
    "company",
    "question_type",
    "candidate_kind",
    "document_kind",
    "question_draft",
    "fact_target",
    "fact_target_type",
    "acceptable_disclosure",
    "review_decision",
    "review_reason",
    "cluster_id",
    "cluster_action",
    "rewritten_question_draft",
    "rewritten_disclosure_draft",
    "review_notes",
]


def _write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "Guide"
    guide.append(["RTX v2.1 Review Round 1"])
    guide.append(["Input: fact-quality candidates only — not v1/manual legacy"])

    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)

    for sheet_name, lane_key, color in (
        ("Working_Set", ("keep", "rewrite"), "166534"),
        ("Rejected", ("reject",), "991B1B"),
        ("Collapsed", ("collapse_into_cluster",), "B45309"),
    ):
        ws = wb.create_sheet(sheet_name)
        ws.append(WORKBOOK_COLUMNS)
        _style_header(ws, len(WORKBOOK_COLUMNS), color)
        if isinstance(lane_key, tuple) and lane_key[0] in ("keep", "rewrite"):
            subset = [r for r in rows if r.get("review_decision") in lane_key]
        else:
            subset = [r for r in rows if r.get("review_decision") == lane_key[0]]
        for row in sorted(subset, key=lambda x: (x.get("review_decision", ""), x.get("seed_id", ""))):
            ws.append([row.get(c, "") for c in WORKBOOK_COLUMNS])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _style_header(ws, ncol: int, color: str) -> None:
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncol + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    out: Dict[str, List[Dict[str, str]]] = {
        "keep": [],
        "rewrite": [],
        "collapse": [],
        "reject": [],
    }
    mapping = {
        "keep": "keep",
        "rewrite": "rewrite",
        "collapse_into_cluster": "collapse",
        "reject": "reject",
    }
    for row in rows:
        dec = row.get("review_decision", "")
        bucket = mapping.get(dec)
        if not bucket or len(out[bucket]) >= 3:
            continue
        out[bucket].append(
            {
                "seed_id": row.get("seed_id", ""),
                "question": (row.get("question_draft") or "")[:90],
                "reason": row.get("review_reason", ""),
                "notes": row.get("review_notes", ""),
            }
        )
    return out


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Workbook Review RTX V2.1 Round 1",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Mở lại Review Round 1 trên workbook **v2.1 fact-quality** (42 candidates).",
        "Không dùng v1 / manual_round2 legacy.",
        "",
        "## Vì sao v2.1 mới đủ mở review round 1",
        "",
        "- 0 exact duplicate question; mỗi row có `fact_target`",
        "- Post-audit quality errors = 0 trên input",
        "- Question layer đã fact-specific và aligned",
        "",
        "## Rule triage round 1",
        "",
        "| Decision | Điều kiện |",
        "|----------|-----------|",
        "| `keep` | Question tự nhiên, fact rõ, disclosure usable |",
        "| `rewrite` | Fact đúng; chỉnh wording/disclosure nhẹ (table/trend awkward) |",
        "| `reject` | Mismatch, residue, disclosure quá yếu |",
        "| `collapse_into_cluster` | Trùng fact cluster; year không thêm giá trị |",
        "",
        "## Kết quả tổng quan",
        "",
        f"- Input total: **{summary.get('input_total', 0)}**",
        f"- keep: **{summary.get('keep', 0)}**",
        f"- rewrite: **{summary.get('rewrite', 0)}**",
        f"- reject: **{summary.get('reject', 0)}**",
        f"- collapse_into_cluster: **{summary.get('collapse_into_cluster', 0)}**",
        f"- Reviewable (keep + rewrite): **{summary.get('reviewable_after_round1', 0)}**",
        "",
        "### Breakdown theo question_type (reviewable)",
        "",
    ]
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "### Breakdown theo document_kind (reviewable)", ""])
    for dk, n in summary.get("by_document_kind", {}).items():
        lines.append(f"- `{dk}`: {n}")

    lines.extend(["", "### Breakdown theo review_reason", ""])
    for reason, n in summary.get("by_reason", {}).items():
        lines.append(f"- `{reason}`: {n}")

    lines.extend(["", "## Ví dụ", ""])
    for title, key in (
        ("Keep tốt", "keep"),
        ("Rewrite hợp lý", "rewrite"),
        ("Collapse đúng", "collapse"),
        ("Reject", "reject"),
    ):
        lines.append(f"### {title}")
        for ex in examples.get(key, []):
            lines.append(f"- `{ex['seed_id']}`: {ex['question']} — {ex['reason']}")
        lines.append("")

    lines.extend(
        [
            "## Kết luận",
            "",
            f"- Reviewable rows: **{summary.get('reviewable_after_round1', 0)}**",
            f"- Đủ mở manual review prep: **{summary.get('manual_review_ready_verdict', '')}**",
            f"- `manual_review_ready_flag` = **{summary.get('manual_review_ready_flag', False)}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_review(
    *,
    input_path: Path,
    reviewed_jsonl: Path,
    rejected_jsonl: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    reviewed = review_rows(rows)

    rejected_out = [r for r in reviewed if r.get("review_decision") in ("reject", "collapse_into_cluster")]
    write_jsonl(reviewed_jsonl, reviewed)
    write_jsonl(rejected_jsonl, rejected_out)
    _write_workbook(reviewed, workbook_path)

    decisions = Counter(r.get("review_decision") for r in reviewed)
    active = [r for r in reviewed if r.get("review_decision") in ("keep", "rewrite")]

    by_reason = Counter(r.get("review_reason") for r in reviewed)
    reviewable = decisions.get("keep", 0) + decisions.get("rewrite", 0)

    manual_ready = (
        reviewable >= 12
        and decisions.get("reject", 0) <= 3
        and reviewable >= len(rows) * 0.5
    )

    examples = _pick_examples(reviewed)
    summary = {
        "review_version": REVIEW_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_total": len(rows),
        "keep": decisions.get("keep", 0),
        "rewrite": decisions.get("rewrite", 0),
        "reject": decisions.get("reject", 0),
        "collapse_into_cluster": decisions.get("collapse_into_cluster", 0),
        "reviewable_after_round1": reviewable,
        "by_question_type": dict(Counter(r.get("question_type") for r in active)),
        "by_document_kind": dict(Counter(r.get("document_kind") for r in active)),
        "by_reason": dict(by_reason.most_common(20)),
        "manual_review_ready_flag": manual_ready,
        "manual_review_ready_verdict": (
            "Có — đủ để mở manual review prep (lane split + polish)"
            if manual_ready
            else "Chưa — cần xem lại reject/collapse rate"
        ),
        "output_reviewed_jsonl": str(reviewed_jsonl),
        "output_rejected_jsonl": str(rejected_jsonl),
        "output_workbook": str(workbook_path),
    }

    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="RTX v2.1 workbook review round 1")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v2_1_fact_quality.jsonl",
    )
    parser.add_argument(
        "--reviewed-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v2_1_reviewed_round1.jsonl",
    )
    parser.add_argument(
        "--rejected-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v2_1_rejected_round1.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_rtx_v2_1_review_round1.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_workbook_review_rtx_v2_1_round1.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_workbook_review_rtx_v2_1_round1_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_review(
        input_path=root / args.input,
        reviewed_jsonl=root / args.reviewed_jsonl,
        rejected_jsonl=root / args.rejected_jsonl,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "input_total",
                    "keep",
                    "rewrite",
                    "reject",
                    "collapse_into_cluster",
                    "reviewable_after_round1",
                    "manual_review_ready_flag",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
