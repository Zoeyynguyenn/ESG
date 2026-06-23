"""Build a reference-style ESG seed workbook from corpus units.

Reset direction:
- multi-seed per strong passage
- softer dedupe (same fact + same intent only)
- workbook-first review flow similar to external v4 reference
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.io_utils import read_jsonl, write_jsonl

SEED_VERSION = "ref_seed_v1"

ESG_KEYWORDS = [
    "esg",
    "지속가능",
    "온실가스",
    "탄소",
    "기후",
    "인권",
    "안전보건",
    "공급망",
    "협력사",
    "이사회",
    "esg위원회",
    "tcfd",
    "kgcs",
    "materiality",
    "중대성",
    "net zero",
]

ESG_SENTENCE_KEYWORDS = [
    "esg",
    "지속가능",
    "온실가스",
    "탄소",
    "기후",
    "중대성",
    "중대 이슈",
    "인권",
    "안전보건",
    "공급망",
    "협력사",
    "책임 있는 조달",
    "이사회",
    "esg위원회",
    "tcfd",
    "kgcs",
    "cdp",
    "gri",
    "issb",
    "esrs",
]

NOISE_KEYWORDS = [
    "table of contents",
    "목차",
    "사이트맵",
    "정보공개",
    "민원",
    "다운로드",
    "국문 영문",
    "기사 공유",
    "주소복사",
    "네이버 채널",
    "다음 채널",
    "all rights reserved",
    "무단전재",
    "연합뉴스",
    "기자",
    "발행일",
    "댓글",
    "구독",
    "송고",
    "목록 글쓰기",
    "만족도 평가",
]

FINANCIAL_NOISE_PATTERNS = [
    "current price",
    "consensus",
    "eps",
    "ebitda",
    "revenue (wbn)",
    "영업이익",
    "당기순이익",
    "매출액",
    "매출 ",
    "오프라인 매장",
    "시장지위",
    "miraeasset",
    "@miraeasset.com",
]

OTHER_COMPANY_MARKERS = [
    "삼성전기",
    "삼성전자",
    "lg전자",
]

METRIC_HINT_RE = re.compile(
    r"(scope\s*[123]|tco2eq|kgcs|cdp|gr[i1]|issb|esrs|tcfd|%|명|회|개|건|억원|백만원)",
    re.IGNORECASE,
)
YEAR_RE = re.compile(r"20\d{2}")
NUMBER_RE = re.compile(r"\d[\d,\.]*")


@dataclass
class Candidate:
    company: str
    category: str
    question_type: str
    question_draft: str
    passage_text: str
    facts: List[str]
    acceptable_disclosure: str
    prohibited_claims: str
    source_unit_id: str
    source_record_id: str
    source_file: str
    source_type: str
    rank: float
    fingerprint: str
    notes: str


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _blob(row: Dict[str, Any]) -> str:
    return _norm_ws(
        " ".join(
            [
                str(row.get("company", "")),
                str(row.get("section_path", "")),
                str(row.get("source_type", "")),
                str(row.get("text", "")),
            ]
        )
    ).lower()


def _is_noise(row: Dict[str, Any]) -> bool:
    blob = _blob(row)
    text = row.get("text") or ""
    if len(_norm_ws(text)) < 120:
        return True
    if sum(1 for k in NOISE_KEYWORDS if k in blob) >= 2 and "중대" not in blob and "scope" not in blob:
        return True
    if blob.count("table of contents") >= 1 and blob.count("about this report") >= 1 and len(text) < 2000:
        return True
    if text.count("국문 영문") >= 3:
        return True
    if "e-mail" in blob and "tel." in blob and "지속가능경영보고서" in blob and len(text) < 2500:
        return True
    return False


def _has_esg_signal(row: Dict[str, Any]) -> bool:
    blob = _blob(row)
    if any(k in blob for k in ESG_KEYWORDS):
        return True
    return False


def _guess_category(row: Dict[str, Any]) -> str:
    blob = _blob(row)
    if "governance" in blob or "거버넌스" in blob or "이사회" in blob:
        return "G"
    if "social" in blob or "사회" in blob or "안전보건" in blob or "인권" in blob:
        return "S"
    return "E"


def _split_sentences(text: str) -> List[str]:
    raw = re.split(r"(?<=[\.\?\!다])\s+|\n+", text)
    out: List[str] = []
    for part in raw:
        s = _norm_ws(part)
        if len(s) >= 24:
            out.append(s)
    return out


def _sentence_score(sentence: str) -> int:
    lower = sentence.lower()
    if any(pat in lower for pat in FINANCIAL_NOISE_PATTERNS) and not any(k in lower for k in ESG_SENTENCE_KEYWORDS):
        return -5
    score = 0
    numbers = len(NUMBER_RE.findall(sentence))
    years = len(set(YEAR_RE.findall(sentence)))
    if METRIC_HINT_RE.search(sentence):
        score += 3
    if years >= 2:
        score += 4
    elif years == 1:
        score += 2
    if numbers >= 3:
        score += 3
    elif numbers >= 1:
        score += 1
    for kw in ("중대성", "중대 이슈", "net zero", "탄소중립", "tcfd", "kgcs", "esg위원회", "이사회", "협력사", "온실가스", "안전보건"):
        if kw.lower() in sentence.lower():
            score += 2
    return score


def _best_sentences(text: str, *, limit: int = 6) -> List[str]:
    scored = [(s, _sentence_score(s)) for s in _split_sentences(text)]
    scored.sort(key=lambda x: (x[1], len(x[0])), reverse=True)
    out: List[str] = []
    for s, score in scored:
        lower = s.lower()
        if score < 2:
            continue
        if not any(k in lower for k in ESG_SENTENCE_KEYWORDS) and not (
            METRIC_HINT_RE.search(s) and any(k in lower for k in ("scope", "온실가스", "kgcs", "tcfd", "esg", "안전보건"))
        ):
            continue
        out.append(s)
        if len(out) >= limit:
            break
    return out


def _cross_company_contaminated(company: str, text: str) -> bool:
    lower = text.lower()
    if company == "레이시온" and ("삼성전기" in text or "삼성" in text):
        return True
    if any(marker in text for marker in OTHER_COMPANY_MARKERS):
        return True
    if any(p in lower for p in FINANCIAL_NOISE_PATTERNS):
        return True
    return False


def _passage_has_esg_focus(passage: str) -> bool:
    lower = passage.lower()
    if any(p in lower for p in FINANCIAL_NOISE_PATTERNS):
        return False
    if "국문 영문" in passage or "e-mail" in lower or "tel." in lower:
        return False
    if any(k in passage for k in ("연합뉴스", "기자", "발행일", "댓글", "구독", "송고", "목록 글쓰기", "만족도 평가")):
        return False
    if any(k in lower for k in ESG_SENTENCE_KEYWORDS):
        return True
    return bool(METRIC_HINT_RE.search(passage) and any(k in lower for k in ("scope", "온실가스", "kgcs", "tcfd", "esg")))


def _make_question(company: str, passage: str, category: str, qtype: str) -> Optional[str]:
    lower = passage.lower()
    if "scope 1" in lower and len(set(YEAR_RE.findall(passage))) >= 2:
        return f"{company}의 온실가스 배출량 추이는 어떻게 나타나는가?"
    if "scope 1" in lower:
        return f"{company}의 온실가스 배출량은 어떻게 공시되어 있는가?"
    if "중대 이슈" in passage or "중대성" in passage:
        if re.search(r"\d+\s*개", passage):
            return f"{company}는 몇 개의 중대 이슈를 선정했는가?"
        return f"{company}가 선정한 중대 이슈는 무엇인가?"
    if "2050" in passage and ("net zero" in lower or "탄소중립" in passage):
        return f"{company}는 2050년까지 어떤 기후 목표를 추진하는가?"
    if "tcfd" in lower:
        return f"{company}의 기후변화 대응 공시는 무엇을 기반으로 하는가?"
    if "kgcs" in lower and "등급" in passage:
        return f"{company}는 ESG 평가에서 어떤 등급을 획득했는가?"
    if "esg위원회" in lower or "이사회" in passage:
        return f"{company}의 ESG 거버넌스 체계는 어떻게 운영되는가?"
    if "협력사" in passage and ("조달" in passage or "공급망" in passage):
        return f"{company}는 책임 있는 조달과 공급망 관리를 위해 어떤 원칙을 운영하는가?"
    if category == "S" and ("안전보건" in passage or "훈련" in passage or "교육" in passage):
        return f"{company}의 안전보건 운영 내용은 무엇인가?"
    if qtype == "trend":
        return f"{company}의 주요 ESG 지표 추이는 어떻게 변화했는가?"
    if qtype == "quantitative":
        return f"{company}의 주요 ESG 수치는 무엇인가?"
    if qtype == "qualitative":
        return f"{company}의 ESG 전략 또는 정책의 핵심 내용은 무엇인가?"
    return None


def _detect_qtypes(text: str) -> List[str]:
    qtypes: List[str] = []
    years = len(set(YEAR_RE.findall(text)))
    numbers = len(NUMBER_RE.findall(text))
    if years >= 2 and numbers >= 3:
        qtypes.append("trend")
    if numbers >= 1 and METRIC_HINT_RE.search(text):
        qtypes.append("quantitative")
    if any(k in text.lower() for k in ["전략", "정책", "위원회", "지배구조", "tcfd", "net zero", "조달", "공급망", "인권"]):
        qtypes.append("qualitative")
    if not qtypes:
        qtypes.append("qualitative")
    return qtypes[:3]


def _facts_tuple(company: str, facts: Sequence[str]) -> str:
    lines = []
    for i, fact in enumerate(facts[:4], start=1):
        lines.append(f"{company} | {i} | {fact}")
    return "\n".join(lines)


def _acceptable_disclosure(facts: Sequence[str]) -> str:
    text = " ".join(_norm_ws(f) for f in facts[:4])
    return text[:420].strip()


def _prohibited_claims(qtype: str) -> str:
    if qtype == "trend":
        return "원문에 없는 원인 추정 금지\n향후 개선 보장 주장 금지\n미공시 수치 보완 금지"
    if qtype == "quantitative":
        return "원문에 없는 수치 추가 금지\n단위 변경 추정 금지\n미공시 항목 단정 금지"
    return "완전 달성/전면 준수 단정 금지\n원문 밖 정책 확장 해석 금지\n미공시 성과 추정 금지"


def _fingerprint(company: str, qtype: str, passage: str) -> str:
    raw = f"{company}|{qtype}|{_norm_ws(passage).lower()}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def _candidate_rank(text: str, qtype: str, base_rank: float) -> float:
    score = base_rank
    if qtype == "trend":
        score += 8
    elif qtype == "quantitative":
        score += 6
    else:
        score += 4
    score += min(len(set(YEAR_RE.findall(text))) * 2, 6)
    score += min(len(NUMBER_RE.findall(text)), 5)
    return float(score)


def extract_candidates(row: Dict[str, Any]) -> List[Candidate]:
    if _is_noise(row) or not _has_esg_signal(row):
        return []
    if (row.get("source_type") or "").lower() not in {"sustainability_report", "official_sustainability_report"}:
        text_lower = (row.get("text") or "").lower()
        strong_hits = sum(1 for k in ESG_SENTENCE_KEYWORDS if k in text_lower)
        if strong_hits < 3:
            return []

    company = row.get("company", "")
    category = _guess_category(row)
    text = row.get("text", "")
    best = _best_sentences(text, limit=8)
    if not best:
        return []

    top_passage = " ".join(best[:3])
    qtypes = _detect_qtypes(top_passage)
    base_rank = float(row.get("substance_score", 0)) - float(row.get("noise_score", 0))
    out: List[Candidate] = []

    for qtype in qtypes:
        question = _make_question(company, top_passage, category, qtype)
        if not question:
            continue
        if qtype == "trend":
            facts = [s for s in best if len(set(YEAR_RE.findall(s))) >= 2][:3] or best[:2]
        elif qtype == "quantitative":
            facts = [s for s in best if NUMBER_RE.search(s)][:3] or best[:2]
        else:
            facts = best[:3]
        passage = " ".join(facts)
        if _cross_company_contaminated(company, passage):
            continue
        if not _passage_has_esg_focus(passage):
            continue
        fp = _fingerprint(company, qtype, passage)
        out.append(
            Candidate(
                company=company,
                category=category,
                question_type=qtype,
                question_draft=question,
                passage_text=passage,
                facts=list(facts),
                acceptable_disclosure=_acceptable_disclosure(facts),
                prohibited_claims=_prohibited_claims(qtype),
                source_unit_id=row.get("unit_id", ""),
                source_record_id=row.get("record_id", ""),
                source_file=row.get("source_file", ""),
                source_type=row.get("source_type", ""),
                rank=_candidate_rank(passage, qtype, base_rank),
                fingerprint=fp,
                notes=f"source_type={row.get('source_type','')}; section={row.get('section_path','')}",
            )
        )
    return out


def _seed_id(company: str, category: str, qtype: str, index: int) -> str:
    prefix = {
        "한샘": "HS",
        "무신사": "MS",
        "레이시온": "RX",
        "RTX": "RX",
    }.get(company, company[:2].upper())
    type_code = {"quantitative": "Q", "trend": "T", "qualitative": "L"}.get(qtype, "X")
    return f"{prefix}-{category}-{type_code}{index:02d}"


def _dedupe(candidates: Iterable[Candidate]) -> List[Candidate]:
    best_by_fp: Dict[str, Candidate] = {}
    for cand in candidates:
        prev = best_by_fp.get(cand.fingerprint)
        if prev is None or cand.rank > prev.rank:
            best_by_fp[cand.fingerprint] = cand
    return list(best_by_fp.values())


def _select_top(candidates: List[Candidate], target_total: int) -> List[Candidate]:
    by_company: Dict[str, List[Candidate]] = defaultdict(list)
    for cand in sorted(candidates, key=lambda x: x.rank, reverse=True):
        by_company[cand.company].append(cand)

    selected: List[Candidate] = []
    used_q: set[Tuple[str, str, str]] = set()

    # Round-robin by company for diversity.
    while len(selected) < target_total:
        progressed = False
        for company in sorted(by_company):
            pool = by_company[company]
            while pool:
                cand = pool.pop(0)
                key = (cand.company, cand.question_type, cand.question_draft)
                if key in used_q:
                    continue
                used_q.add(key)
                selected.append(cand)
                progressed = True
                break
            if len(selected) >= target_total:
                break
        if not progressed:
            break
    return selected


def _rows_for_jsonl(selected: Sequence[Candidate]) -> List[Dict[str, Any]]:
    counters: Counter[Tuple[str, str, str]] = Counter()
    rows: List[Dict[str, Any]] = []
    for cand in selected:
        counters[(cand.company, cand.category, cand.question_type)] += 1
        rows.append(
            {
                "seed_id": _seed_id(
                    cand.company,
                    cand.category,
                    cand.question_type,
                    counters[(cand.company, cand.category, cand.question_type)],
                ),
                "company": cand.company,
                "category": cand.category,
                "question_type": cand.question_type,
                "question_draft": cand.question_draft,
                "source_unit_id": cand.source_unit_id,
                "source_record_id": cand.source_record_id,
                "source_file": cand.source_file,
                "source_type": cand.source_type,
                "passage_text": cand.passage_text,
                "facts_tuple": _facts_tuple(cand.company, cand.facts),
                "acceptable_disclosure": cand.acceptable_disclosure,
                "prohibited_claims": cand.prohibited_claims,
                "status": "draft",
                "notes": cand.notes,
                "seed_version": SEED_VERSION,
                "rank": cand.rank,
            }
        )
    return rows


def _write_workbook(rows: Sequence[Dict[str, Any]], xlsx_path: Path) -> None:
    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "안내"
    guide_rows = [
        ["Reference-style Golden Seed Workbook"],
        ["Mục tiêu", "Khôi phục hướng workbook-first: một passage tốt có thể sinh nhiều seed ESG."],
        ["Nguyên tắc", "Không ép 1 unit = 1 QA; ưu tiên passage_text + facts_tuple + acceptable_disclosure."],
        ["Lưu ý", "Các dòng là candidate draft để reviewer/AI refine tiếp, không phải final gold."],
    ]
    for row in guide_rows:
        ws_guide.append(row)

    ws_main = wb.create_sheet("작성")
    header = [
        "seed_id",
        "company",
        "category",
        "question_type",
        "question_draft",
        "source_record_id",
        "source_unit_id",
        "source_type",
        "passage_text",
        "facts_tuple",
        "acceptable_disclosure",
        "prohibited_claims",
        "status",
        "notes",
    ]
    ws_main.append(header)
    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(color="FFFFFF", bold=True)
    for idx in range(1, len(header) + 1):
        cell = ws_main.cell(row=1, column=idx)
        cell.fill = fill
        cell.font = font
    for row in rows:
        ws_main.append([row.get(col, "") for col in header])

    ws_ref = wb.create_sheet("참조")
    ref_rows = [
        ["field", "description"],
        ["question_draft", "Câu hỏi nháp để query/eval, cần reviewer tinh chỉnh nếu cần."],
        ["passage_text", "Đoạn evidence gốc đã giữ ngữ cảnh ESG."],
        ["facts_tuple", "Danh sách atomic facts rút từ passage_text."],
        ["acceptable_disclosure", "Đáp án/đoạn công bố chấp nhận được ở mức seed."],
        ["prohibited_claims", "Các claim không được tự suy diễn."],
    ]
    for row in ref_rows:
        ws_ref.append(row)

    ws_sum = wb.create_sheet("요약")
    by_company = Counter(row["company"] for row in rows)
    by_type = Counter(row["question_type"] for row in rows)
    ws_sum.append(["metric", "value"])
    ws_sum.append(["total_rows", len(rows)])
    for company, count in sorted(by_company.items()):
        ws_sum.append([f"company:{company}", count])
    for qtype, count in sorted(by_type.items()):
        ws_sum.append([f"question_type:{qtype}", count])

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def run_reference_seed_builder(
    *,
    input_path: Path,
    output_jsonl: Path,
    output_xlsx: Path,
    target_total: int = 24,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    all_candidates: List[Candidate] = []
    for row in rows:
        all_candidates.extend(extract_candidates(row))

    deduped = _dedupe(all_candidates)
    selected = _select_top(deduped, target_total=target_total)
    out_rows = _rows_for_jsonl(selected)
    write_jsonl(output_jsonl, out_rows)
    _write_workbook(out_rows, output_xlsx)

    by_company = Counter(row["company"] for row in out_rows)
    by_type = Counter(row["question_type"] for row in out_rows)
    return {
        "input_units": len(rows),
        "raw_candidates": len(all_candidates),
        "deduped_candidates": len(deduped),
        "selected_rows": len(out_rows),
        "by_company": dict(by_company),
        "by_question_type": dict(by_type),
        "output_jsonl": str(output_jsonl),
        "output_xlsx": str(output_xlsx),
    }


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Build reference-style ESG seed workbook")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output-jsonl", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--target-total", type=int, default=24)
    args = parser.parse_args(argv)

    summary = run_reference_seed_builder(
        input_path=Path(args.input),
        output_jsonl=Path(args.output_jsonl),
        output_xlsx=Path(args.output_xlsx),
        target_total=args.target_total,
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
