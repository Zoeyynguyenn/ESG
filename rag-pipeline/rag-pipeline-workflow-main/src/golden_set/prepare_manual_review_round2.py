"""Manual Review Round 2 prep — lane split for reviewer (A/B/C + reject recommended)."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import ESG_SENTENCE_KEYWORDS, METRIC_HINT_RE
from golden_set.io_utils import read_jsonl, write_jsonl

PREP_VERSION = "ref_manual_round2_prep"

HEAVY_CHROME = [
    "advertisements",
    "advertisement",
    "모바일버전",
    "전체기사",
    "home 경제",
    "photo ",
    "phpto",
    "0000년 00월",
    "기자",
    "발행일",
    "댓글",
    "구독",
    "송고",
    "주소복사",
    "네이버 채널",
    "무단전재",
    "사진제공",
]

INDEX_META = [
    "sasb index",
    "tcfd index",
    "gri index",
    "esrs index",
    "content index",
    "disclosure index",
    "appendix",
    "table of contents",
    "목차",
]

JSON_BLOB = re.compile(r'[\{\[]\s*"[a-zA-Z_]+"\s*:')
TRUNCATED_START = ("완성하고", "또한", "먼저 오는", "대응전략")
TRUNCATED_END = ("BIS)과", "국제결제은행(BIS)과", "인증 제도인", "금융안정위원회(FSB)가")

GENERIC_Q = re.compile(
    r"(핵심 내용은 무엇인가|주요 ESG 수치는 무엇인가|주요 ESG 지표 추이는|ESG 전략 또는 정책)"
)
TCFD_DEF_ONLY = re.compile(r"TCFD는\s*2015년", re.IGNORECASE)


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _passage(row: Dict[str, Any]) -> str:
    return _norm_ws(
        row.get("rewritten_disclosure_draft")
        or row.get("acceptable_disclosure")
        or row.get("source_excerpt")
        or ""
    )


def _company_in_text(company: str, text: str) -> bool:
    if company in text:
        return True
    aliases = {"한샘": ["㈜한샘", "한샘이"], "무신사": ["musinsa"], "레이시온": ["raysolution"]}
    return any(a in text for a in aliases.get(company, []))


def _esg_fact_strength(passage: str, company: str) -> int:
    lower = passage.lower()
    score = sum(1 for kw in ESG_SENTENCE_KEYWORDS if kw.lower() in lower)
    if METRIC_HINT_RE.search(passage):
        score += 3
    if re.search(r"\d", passage):
        score += 2
    if _company_in_text(company, passage):
        score += 2
    if any(k in passage for k in ("중대 이슈", "탄소중립", "net zero", "kgcs", "이사회", "esg위원회")):
        score += 2
    return score


def _contamination_profile(row: Dict[str, Any]) -> Tuple[int, List[str]]:
    passage = _passage(row)
    blob = passage.lower()
    reasons: List[str] = []
    score = 0

    for m in HEAVY_CHROME:
        if m.lower() in blob:
            score += 2
            if "heavy_chrome" not in reasons:
                reasons.append("heavy_chrome")

    for m in INDEX_META:
        if m in blob:
            score += 3
            reasons.append("report_index_meta")

    if JSON_BLOB.search(passage) or ('"rec_' in passage or passage.count("{") >= 2):
        score += 5
        reasons.append("json_blob")

    if passage.startswith(TRUNCATED_START) or any(passage.rstrip().endswith(e) for e in TRUNCATED_END):
        score += 4
        reasons.append("truncated_excerpt")

    if len(passage) < 55:
        score += 3
        reasons.append("very_short_excerpt")

    if TCFD_DEF_ONLY.search(passage) and not _company_in_text(row.get("company", ""), passage):
        score += 4
        reasons.append("framework_definition_only")

    rr = row.get("review_reason") or ""
    if "light_news_chrome" in rr:
        score += 2
        reasons.append("news_chrome_flag")
    if "passage_too_broad" in rr:
        score += 2
        reasons.append("passage_too_broad")
    if "truncated_excerpt" in rr and "truncated_excerpt" not in reasons:
        score += 2
        reasons.append("truncated_flag")

    return score, list(dict.fromkeys(reasons))


def _disclosure_clean(passage: str, company: str) -> bool:
    score, reasons = 0, []
    # reuse contamination without row context
    blob = passage.lower()
    if any(m.lower() in blob for m in HEAVY_CHROME):
        return False
    if any(m in blob for m in INDEX_META):
        return False
    if JSON_BLOB.search(passage):
        return False
    if passage.startswith(TRUNCATED_START):
        return False
    if len(passage) < 55:
        return False
    if not _company_in_text(company, passage) and len(passage) < 120:
        return False
    return True


def _manual_reject_recommended(row: Dict[str, Any], contam: int, fact: int, reasons: List[str]) -> bool:
    passage = _passage(row)
    company = row.get("company") or ""

    if fact <= 2 and contam >= 5:
        return True
    if "framework_definition_only" in reasons and fact < 6:
        return True
    if len(passage) < 45 and fact < 5:
        return True
    if contam >= 7 and fact < 6:
        return True
    if TCFD_DEF_ONLY.search(passage) and fact < 7:
        return True
    if "json_blob" in reasons and fact < 6:
        return True
    if contam >= 10:
        return True
    # Truncated framework snippet without company-specific disclosure
    if "truncated_excerpt" in reasons and "report_index_meta" in reasons:
        return True
    if passage.rstrip().endswith(TRUNCATED_END) and not _company_in_text(company, passage[:80]):
        return True
    return False


def assign_lane(row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row)
    decision = row.get("review_decision") or ""
    contam, contam_reasons = _contamination_profile(row)
    fact = _esg_fact_strength(_passage(row), row.get("company") or "")
    passage = _passage(row)

    reject_rec = _manual_reject_recommended(row, contam, fact, contam_reasons)

    lane = "lane_c_rewrite_heavy"
    priority = "low"
    review_reason = ""
    rewrite_scope = "full_passage_cleanup"

    if reject_rec:
        lane = "reject_recommended"
        priority = "skip"
        review_reason = f"low_fact_high_noise; contam={contam}; fact={fact}; " + ",".join(contam_reasons[:4])
        rewrite_scope = "not_recommended"
    elif decision == "keep":
        if contam <= 1 and _disclosure_clean(passage, row.get("company", "")):
            lane = "lane_a_ready_keep"
            priority = "high"
            review_reason = "round1_keep_clean_grounded"
            rewrite_scope = "confirm_or_drop_only"
        elif contam <= 3:
            lane = "lane_b_rewrite_light"
            priority = "medium"
            review_reason = "keep_with_minor_cleanup"
            rewrite_scope = "wording_only"
        else:
            lane = "lane_c_rewrite_heavy"
            priority = "low"
            review_reason = "keep_but_disclosure_dirty; " + ",".join(contam_reasons[:3])
            rewrite_scope = "disclosure_extraction"
    elif decision == "rewrite":
        rr = row.get("review_reason") or ""
        generic_only = bool(GENERIC_Q.search(row.get("question_draft") or "")) and "generic" in rr
        clean = _disclosure_clean(passage, row.get("company", ""))

        if generic_only and clean and contam <= 2:
            lane = "lane_b_rewrite_light"
            priority = "medium"
            review_reason = "generic_question_clean_disclosure"
            rewrite_scope = "question_specificity"
        elif contam <= 2 and fact >= 5 and "truncated" not in rr:
            lane = "lane_b_rewrite_light"
            priority = "medium"
            review_reason = "salvageable_light_rewrite; " + rr
            rewrite_scope = "wording_and_specificity"
        elif contam >= 5 or any(
            r in contam_reasons
            for r in ("heavy_chrome", "json_blob", "report_index_meta", "truncated_excerpt")
        ):
            lane = "lane_c_rewrite_heavy"
            priority = "low"
            review_reason = "dirty_passage_salvage; " + ",".join(contam_reasons[:4])
            rewrite_scope = "extract_fact_from_dirty_passage"
        else:
            lane = "lane_b_rewrite_light" if contam <= 3 else "lane_c_rewrite_heavy"
            priority = "medium" if lane == "lane_b_rewrite_light" else "low"
            review_reason = f"rewrite_mixed; contam={contam}; {rr}"
            rewrite_scope = "wording_and_disclosure_trim" if lane == "lane_b_rewrite_light" else "disclosure_extraction"

    out.update(
        {
            "manual_prep_version": PREP_VERSION,
            "manual_review_lane": lane,
            "manual_priority": priority,
            "manual_review_reason": review_reason,
            "manual_reject_recommended": reject_rec,
            "manual_rewrite_scope": rewrite_scope,
            "contamination_score": contam,
            "fact_strength_score": fact,
            "contamination_flags": contam_reasons,
        }
    )
    return out


WORKBOOK_COLS = [
    "seed_id",
    "company",
    "question_type",
    "candidate_kind",
    "question_draft",
    "acceptable_disclosure",
    "review_decision",
    "review_reason",
    "rewritten_question_draft",
    "rewritten_disclosure_draft",
    "manual_review_lane",
    "manual_priority",
    "manual_review_reason",
    "manual_rewrite_scope",
    "manual_reject_recommended",
    "cluster_id",
]


def _style_header(ws, ncol: int, color: str = "1E3A5F") -> None:
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncol + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font


def write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide.append(["Manual Review Round 2 — Lane Split"])
    guide.append(["Thứ tự", "Lane A → Lane B → Lane C; Reject_Recommended = skip/audit"])

    lanes = [
        ("Lane_A_ReadyKeep", "lane_a_ready_keep", "166534"),
        ("Lane_B_RewriteLight", "lane_b_rewrite_light", "1D4ED8"),
        ("Lane_C_RewriteHeavy", "lane_c_rewrite_heavy", "B45309"),
        ("Reject_Recommended", "reject_recommended", "991B1B"),
    ]
    for sheet_name, lane_key, color in lanes:
        ws = wb.create_sheet(sheet_name)
        ws.append(WORKBOOK_COLS)
        _style_header(ws, len(WORKBOOK_COLS), color)
        subset = [r for r in rows if r.get("manual_review_lane") == lane_key]
        for r in sorted(subset, key=lambda x: (-float(x.get("rank") or 0), x.get("seed_id", ""))):
            ws.append([r.get(c, "") for c in WORKBOOK_COLS])

    ws_sum = wb.create_sheet("요약")
    ws_sum.append(["metric", "value"])
    counts = Counter(r.get("manual_review_lane") for r in rows)
    for k in (
        "lane_a_ready_keep",
        "lane_b_rewrite_light",
        "lane_c_rewrite_heavy",
        "reject_recommended",
    ):
        ws_sum.append([k, counts.get(k, 0)])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    out: Dict[str, List[Dict[str, str]]] = {
        "lane_a": [],
        "lane_b": [],
        "lane_c": [],
        "reject": [],
    }
    for lane_key, bucket in (
        ("lane_a_ready_keep", "lane_a"),
        ("lane_b_rewrite_light", "lane_b"),
        ("lane_c_rewrite_heavy", "lane_c"),
        ("reject_recommended", "reject"),
    ):
        for r in rows:
            if r.get("manual_review_lane") != lane_key:
                continue
            if len(out[bucket]) >= 3:
                break
            out[bucket].append(
                {
                    "seed_id": r.get("seed_id", ""),
                    "company": r.get("company", ""),
                    "question": (r.get("question_draft") or "")[:80],
                    "reason": r.get("manual_review_reason", ""),
                }
            )
    return out


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Manual Review Round 2 Prep",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Chuẩn bị workbook manual review round 2 với **lane split** để giảm review cost:",
        "reviewer không xử lý phẳng 107 row như nhau.",
        "",
        "## Vì sao cần lane split",
        "",
        "Round 1 `rewrite` trộn hai loại:",
        "- **rewrite_light**: fact sạch, chỉ cần chỉnh wording/specificity",
        "- **rewrite_heavy**: passage còn bẩn (ads, JSON, meta, truncated headline)",
        "",
        "Nếu review phẳng, reviewer tốn công vào row gần reject.",
        "",
        "## Rule chia lane",
        "",
        "| Lane | Điều kiện |",
        "|------|-----------|",
        "| `lane_a_ready_keep` | Round1 keep + disclosure sạch, confirm/drop nhanh |",
        "| `lane_b_rewrite_light` | Fact thật; chỉnh Q/disclosure gọn |",
        "| `lane_c_rewrite_heavy` | Fact salvageable nhưng passage bẩn |",
        "| `reject_recommended` | Fact yếu + noise cao; không đáng cứu |",
        "",
        "## Kết quả",
        "",
        f"- Lane A: **{summary.get('lane_a_count', 0)}**",
        f"- Lane B: **{summary.get('lane_b_count', 0)}**",
        f"- Lane C: **{summary.get('lane_c_count', 0)}**",
        f"- Reject recommended: **{summary.get('reject_recommended_count', 0)}**",
        "",
        "### Breakdown theo công ty",
        "",
    ]
    for co, stats in summary.get("by_company", {}).items():
        lines.append(
            f"- **{co}**: A={stats.get('lane_a_ready_keep', 0)}, "
            f"B={stats.get('lane_b_rewrite_light', 0)}, "
            f"C={stats.get('lane_c_rewrite_heavy', 0)}, "
            f"reject={stats.get('reject_recommended', 0)}"
        )

    lines.extend(["", "### Breakdown theo question_type", ""])
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "## Ví dụ mỗi lane", ""])
    for title, key in (
        ("Lane A", "lane_a"),
        ("Lane B", "lane_b"),
        ("Lane C", "lane_c"),
        ("Reject recommended", "reject"),
    ):
        lines.append(f"### {title}")
        for ex in examples.get(key, []):
            lines.append(f"- `{ex['seed_id']}` ({ex['company']}): {ex['question']} — {ex['reason']}")
        lines.append("")

    lines.extend(
        [
            "## Kết luận",
            "",
            f"- Reviewer bắt đầu từ: **{summary.get('reviewer_order_recommendation', '')}**",
            f"- Ước lượng row sống tới canonical round tiếp: **~{summary.get('likely_survivors_estimate', 0)}**",
            f"  (A ~90% + B ~75% + C ~45% của row không reject)",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_prep(
    *,
    input_path: Path,
    output_jsonl: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    active = [r for r in rows if r.get("review_decision") in ("keep", "rewrite")]
    prepared = [assign_lane(r) for r in active]

    write_jsonl(output_jsonl, prepared)
    write_workbook(prepared, workbook_path)

    lane_counts = Counter(r.get("manual_review_lane") for r in prepared)
    by_co: Dict[str, Dict[str, int]] = {}
    for co in sorted({r.get("company", "") for r in prepared}):
        sub = [r for r in prepared if r.get("company") == co]
        by_co[co] = dict(Counter(r.get("manual_review_lane") for r in sub))

    reviewable_lanes = [r for r in prepared if r.get("manual_review_lane") != "reject_recommended"]
    by_qtype = Counter(r.get("question_type") for r in reviewable_lanes)

    a = lane_counts.get("lane_a_ready_keep", 0)
    b = lane_counts.get("lane_b_rewrite_light", 0)
    c = lane_counts.get("lane_c_rewrite_heavy", 0)
    rej = lane_counts.get("reject_recommended", 0)

    survivors = int(round(a * 0.9 + b * 0.75 + c * 0.45))

    summary = {
        "prep_version": PREP_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_reviewable": len(active),
        "lane_a_count": a,
        "lane_b_count": b,
        "lane_c_count": c,
        "reject_recommended_count": rej,
        "by_company": by_co,
        "by_question_type": dict(by_qtype),
        "likely_survivors_estimate": survivors,
        "reviewer_order_recommendation": "Lane_A_ReadyKeep → Lane_B_RewriteLight → Lane_C_RewriteHeavy; bỏ qua Reject_Recommended trừ audit",
        "output_jsonl": str(output_jsonl),
        "output_workbook": str(workbook_path),
    }

    examples = _pick_examples(prepared)
    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Prepare manual review round 2 with lane split")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_reviewed_round1.jsonl",
    )
    parser.add_argument(
        "--output-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_manual_round2.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_v4_manual_round2.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_manual_review_round2_prep.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_manual_review_round2_prep_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_prep(
        input_path=root / args.input,
        output_jsonl=root / args.output_jsonl,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "lane_a_count",
                    "lane_b_count",
                    "lane_c_count",
                    "reject_recommended_count",
                    "likely_survivors_estimate",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
