"""Workbook-first seed curation Round 1 — classify reference seed candidates."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import (
    ESG_SENTENCE_KEYWORDS,
    FINANCIAL_NOISE_PATTERNS,
    METRIC_HINT_RE,
    NOISE_KEYWORDS,
    OTHER_COMPANY_MARKERS,
)
from golden_set.io_utils import read_jsonl, write_jsonl

CURATION_VERSION = "ref_curation_r1"

KEEP_DECISIONS = {"keep_strong", "keep_but_needs_rewrite"}
DROP_PREFIX = "drop_"

LISTING_MARKERS = [
    "table of contents",
    "목차",
    "게시판 목록",
    "조회 전체 건",
    "archive",
    "자료실",
    "국문 영문",
    "hanssem sustainability report 04",
    "appendix",
    "ceo message",
    "about this report",
    "보고서 개요",
]

NAV_CONTACT_MARKERS = [
    "정보공개",
    "민원",
    "사이트맵",
    "e-mail",
    "tel.",
    "만족도 평가",
    "faq english",
    "esg 소개 esg 기업정보",
    "기업 esg 조회",
    "참고사이트",
    "esg 강의실",
]

NEWS_CHROME_MARKERS = [
    "기자",
    "발행일",
    "기사 공유",
    "주소복사",
    "네이버 채널",
    "다음 채널",
    "무단전재",
    "advertisements",
    "댓글",
    "구독",
    "송고",
    "백세경제",
]

PRESS_RELEASE_MARKERS = [
    "발간했다고",
    "밝혔다",
    "(009240)",
    "지난 8일",
    "지난해",
]

GENERIC_QUESTION_PATTERNS = [
    re.compile(r"핵심 내용은 무엇인가"),
    re.compile(r"주요 ESG 수치는 무엇인가"),
    re.compile(r"주요 ESG 지표 추이는"),
    re.compile(r"ESG 전략 또는 정책"),
]


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _blob(seed: Dict[str, Any]) -> str:
    return _norm_ws(
        " ".join(
            [
                str(seed.get("passage_text", "")),
                str(seed.get("question_draft", "")),
                str(seed.get("facts_tuple", "")),
                str(seed.get("acceptable_disclosure", "")),
            ]
        )
    ).lower()


def _fact_lines(seed: Dict[str, Any]) -> List[str]:
    raw = seed.get("facts_tuple") or ""
    lines = []
    for line in str(raw).splitlines():
        parts = line.split("|")
        if len(parts) >= 3:
            lines.append(_norm_ws(parts[2]))
        elif line.strip():
            lines.append(_norm_ws(line))
    return [x for x in lines if len(x) >= 12]


def _esg_signal_score(blob: str, passage: str) -> int:
    score = 0
    for kw in ESG_SENTENCE_KEYWORDS:
        if kw.lower() in blob:
            score += 1
    if METRIC_HINT_RE.search(passage):
        score += 2
    if any(k in passage for k in ("중대 이슈", "탄소중립", "net zero", "tcfd", "kgcs", "esg위원회", "이사회")):
        score += 2
    return score


def _noise_hits(blob: str, markers: Sequence[str]) -> int:
    return sum(1 for m in markers if m.lower() in blob)


def _is_cross_company(company: str, blob: str, passage: str) -> bool:
    if company == "레이시온" and ("삼성전기" in passage or "삼성전자" in passage):
        return True
    if any(m in passage for m in OTHER_COMPANY_MARKERS):
        return True
    return False


def _is_financial_non_esg(blob: str, passage: str) -> bool:
    lower = blob
    fin_hits = sum(1 for p in FINANCIAL_NOISE_PATTERNS if p in lower)
    esg_hits = _esg_signal_score(blob, passage)
    return fin_hits >= 2 and esg_hits < 3


def _is_listing_archive(blob: str, passage: str) -> bool:
    if _noise_hits(blob, LISTING_MARKERS) >= 2:
        return True
    if "sustainability report" in blob and passage.count("·") >= 8 and len(passage) < 900:
        return True
    if passage.count("APPENDIX") >= 1 and "GOVERNANCE" in passage and len(_fact_lines({"facts_tuple": passage})) < 2:
        return True
    return False


def _is_nav_contact(blob: str) -> bool:
    if _noise_hits(blob, NAV_CONTACT_MARKERS) >= 3:
        return True
    if "esg 소개" in blob and "esg 통계" in blob and "faq" in blob:
        return True
    return False


def _is_press_release_noise(passage: str) -> bool:
    return sum(1 for m in PRESS_RELEASE_MARKERS if m in passage) >= 2


def _has_tcfd_definition_noise(passage: str) -> bool:
    return "tcfd는" in passage.lower() and "협의체" in passage


def _is_truncated_passage(passage: str) -> bool:
    if len(passage) < 160:
        return True
    if passage.startswith(("완성하고", "또한", "먼저 오는")) and "한샘" not in passage[:40]:
        return True
    return passage.rstrip().endswith(("BIS)과", "국제결제은행(BIS)과", "인증 제도인"))


def _is_news_chrome(blob: str, passage: str) -> bool:
    hits = _noise_hits(blob, NEWS_CHROME_MARKERS)
    if hits >= 2:
        return True
    if _is_press_release_noise(passage) and _esg_signal_score(blob, passage) < 6:
        return True
    return hits >= 1 and len(passage) < 400 and _esg_signal_score(blob, passage) < 4


def _is_too_generic(seed: Dict[str, Any]) -> bool:
    q = seed.get("question_draft") or ""
    passage = seed.get("passage_text") or ""
    if any(p.search(q) for p in GENERIC_QUESTION_PATTERNS):
        return True
    if len(passage) < 80:
        return True
    if passage.count("ESG") >= 6 and len(set(passage.split())) < 40:
        return True
    return False


def _needs_rewrite(seed: Dict[str, Any], blob: str, passage: str) -> Tuple[bool, str]:
    reasons: List[str] = []
    q = seed.get("question_draft") or ""
    if any(p.search(q) for p in GENERIC_QUESTION_PATTERNS):
        reasons.append("generic_question")
    if _noise_hits(blob, NEWS_CHROME_MARKERS) >= 1 or _is_press_release_noise(passage):
        reasons.append("light_news_chrome")
    if _has_tcfd_definition_noise(passage):
        reasons.append("mixed_definition_noise")
    if _is_truncated_passage(passage):
        reasons.append("truncated_passage")
    if len(passage) > 500 and passage.count(".") >= 4 and _is_press_release_noise(passage):
        reasons.append("passage_too_broad")
    if not reasons:
        return False, ""
    return True, ";".join(reasons)


def _is_clean_report_body(passage: str, blob: str) -> bool:
    if _is_press_release_noise(passage):
        return False
    if _has_tcfd_definition_noise(passage):
        return False
    if _is_truncated_passage(passage):
        return False
    if _noise_hits(blob, NEWS_CHROME_MARKERS) >= 1:
        return False
    return True


def _classify_seed(seed: Dict[str, Any]) -> Tuple[str, str, str]:
    company = seed.get("company") or ""
    passage = seed.get("passage_text") or ""
    blob = _blob(seed)
    esg_score = _esg_signal_score(blob, passage)

    if _is_cross_company(company, blob, passage):
        return "drop_cross_company_contamination", "other_company_or_analyst_marker", "Passage chứa marker công ty khác hoặc analyst/financial contamination."

    if _is_nav_contact(blob):
        return "drop_contact_or_navigation", "portal_nav_contact_page", "Passage là portal/nav/contact/FAQ listing, không phải ESG disclosure body."

    if _is_listing_archive(blob, passage):
        return "drop_listing_archive", "toc_archive_or_report_index", "Passage là TOC/archive/listing hoặc report chrome không có fact ESG cụ thể."

    if _is_financial_non_esg(blob, passage):
        return "drop_financial_non_esg", "financial_commentary_dominant", "Passage thiên về financial/analyst commentary, ESG signal yếu."

    if esg_score < 2:
        return "drop_not_esg_enough", "insufficient_esg_substance", "Không đủ tín hiệu ESG narrative/metric/governance trong passage."

    if _is_news_chrome(blob, passage) and esg_score < 5:
        return "drop_news_chrome", "news_page_chrome_heavy", "News chrome quá nặng, không đủ fact ESG salvageable."

    if _is_too_generic(seed) and esg_score < 4:
        return "drop_too_generic", "generic_site_or_question", "Câu hỏi/passage quá generic, không đủ specificity."

    rewrite, rewrite_reason = _needs_rewrite(seed, blob, passage)
    if rewrite:
        return "keep_but_needs_rewrite", rewrite_reason, "Fact ESG có thật nhưng cần rewrite question/disclosure."

    if not _is_clean_report_body(passage, blob):
        return "keep_but_needs_rewrite", "residual_noise", "Fact ESG có thật nhưng passage chưa đủ sạch."

    return "keep_strong", "clean_esg_passage", "ESG narrative/metric grounded, contamination thấp."


def _best_fact(facts: List[str], passage: str) -> str:
    if not facts:
        return _norm_ws(passage)[:280]
    scored = []
    for f in facts:
        s = 0
        if METRIC_HINT_RE.search(f):
            s += 3
        if any(k in f.lower() for k in ("중대", "탄소", "net zero", "tcfd", "kgcs", "이사회", "esg")):
            s += 2
        if "한샘" in f or "무신사" in f or "레이시온" in f:
            s += 1
        scored.append((s, f))
    scored.sort(key=lambda x: (-x[0], len(x[1])))
    return scored[0][1]


def _rewrite_drafts(seed: Dict[str, Any], fact: str) -> Dict[str, Optional[str]]:
    company = seed.get("company") or ""
    lower = fact.lower()
    q: Optional[str] = None

    if "8개 중대 이슈" in fact or "중대 이슈" in fact and re.search(r"\d+\s*개", fact):
        q = f"{company}는 이중 중대성 평가를 통해 몇 개의 중대 이슈를 선정했는가?"
    elif "2050" in fact and ("net zero" in lower or "탄소중립" in fact):
        q = f"{company}는 2050년까지 어떤 탄소중립 목표를 공개했는가?"
    elif "tcfd" in lower and company.lower() in fact.lower():
        q = f"{company}는 지속가능경영보고서에 어떤 기후 관련 공시 프레임워크를 수록했는가?"
    elif "kgcs" in lower and "등급" in fact:
        q = f"{company}는 KGCS ESG경영 평가에서 어떤 등급을 획득했는가?"
    elif "esg 위원회" in lower or "이사회" in fact:
        q = f"{company}의 ESG 거버넌스 체계는 어떻게 운영되는가?"
    elif "14회" in fact and "이사회" in fact:
        q = f"{company}는 2022년 이사회를 몇 회 개최했는가?"
    else:
        q = seed.get("question_draft")

    disclosure = fact[:420].strip()
    prohibited = seed.get("prohibited_claims") or (
        "원문에 없는 수치 추가 금지\n미공시 항목 단정 금지\n원문 밖 정책 확장 해석 금지"
    )
    return {
        "rewritten_question_draft": q,
        "rewritten_acceptable_disclosure": disclosure,
        "rewritten_prohibited_claims": prohibited,
    }


def _review_priority(decision: str, rank: float) -> str:
    if decision == "keep_strong":
        return "high" if rank >= 12 else "medium"
    if decision == "keep_but_needs_rewrite":
        return "medium" if rank >= 10 else "low"
    return "n/a"


def curate_seed(seed: Dict[str, Any]) -> Dict[str, Any]:
    decision, reason, notes = _classify_seed(seed)
    out = dict(seed)
    out.update(
        {
            "curation_version": CURATION_VERSION,
            "curation_decision": decision,
            "curation_reason": reason,
            "curation_notes": notes,
            "rewrite_needed": decision == "keep_but_needs_rewrite",
            "rewrite_reason": reason if decision == "keep_but_needs_rewrite" else "",
            "review_priority": _review_priority(decision, float(seed.get("rank") or 0)),
            "rewritten_question_draft": None,
            "rewritten_acceptable_disclosure": None,
            "rewritten_prohibited_claims": None,
        }
    )
    if decision == "keep_but_needs_rewrite":
        fact = _best_fact(_fact_lines(seed), seed.get("passage_text") or "")
        out.update(_rewrite_drafts(seed, fact))
    elif decision == "keep_strong":
        out["rewritten_question_draft"] = seed.get("question_draft")
        out["rewritten_acceptable_disclosure"] = seed.get("acceptable_disclosure")
        out["rewritten_prohibited_claims"] = seed.get("prohibited_claims")
    return out


def _write_workbook(rows: Sequence[Dict[str, Any]], xlsx_path: Path) -> None:
    wb = Workbook()
    ws_guide = wb.active
    ws_guide.title = "안내"
    for row in [
        ["Reference Seed Workbook — Curated R1"],
        ["Gate", "Workbook-first; không ép 1 unit = 1 QA"],
        ["Mục tiêu", "Seed workbook sạch hơn để reviewer refine tiếp"],
    ]:
        ws_guide.append(row)

    header = [
        "seed_id",
        "company",
        "category",
        "question_type",
        "curation_decision",
        "curation_reason",
        "review_priority",
        "question_draft",
        "rewritten_question_draft",
        "passage_text",
        "facts_tuple",
        "acceptable_disclosure",
        "rewritten_acceptable_disclosure",
        "prohibited_claims",
        "rewritten_prohibited_claims",
        "rewrite_needed",
        "curation_notes",
        "source_record_id",
        "status",
    ]
    ws_main = wb.create_sheet("작성")
    ws_main.append(header)
    fill = PatternFill("solid", fgColor="166534")
    font = Font(color="FFFFFF", bold=True)
    for idx in range(1, len(header) + 1):
        c = ws_main.cell(row=1, column=idx)
        c.fill = fill
        c.font = font
    for row in rows:
        ws_main.append([row.get(col, "") for col in header])

    ws_sum = wb.create_sheet("요약")
    ws_sum.append(["metric", "value"])
    ws_sum.append(["total_curated", len(rows)])
    for k, v in sorted(Counter(r["curation_decision"] for r in rows).items()):
        ws_sum.append([k, v])
    by_co = Counter(r["company"] for r in rows)
    for co, n in sorted(by_co.items()):
        ws_sum.append([f"company:{co}", n])

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def write_curation_report(summary: Dict[str, Any], examples: Dict[str, List[Dict]], report_path: Path) -> None:
    lines = [
        "# Golden Set — Reference Workbook Curation R1",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Làm sạch `reference_seed_workbook_v1` theo hướng **workbook-first**: phân loại contamination đúng lớp, giữ seed ESG thật, loại portal/news/financial noise — **không** quay lại gate `single-unit hard drop`.",
        "",
        "## Vì sao workflow cũ sai hướng",
        "",
        "1. Nhánh R2.1–R2.4 ép `1 unit → 1 QA → drop` → yield ~0 dù hệ thống vẫn tìm được ESG fact.",
        "2. Gate `>=8 usable` trên pilot Hansem 5–15 row không phản ánh khả năng sinh seed workbook.",
        "3. Workbook tham chiếu (`golden_set_3companies_v4.xlsx`) cho phép **multi-seed per passage** và review disclosure — không phải mini-pilot precision.",
        "",
        "## Quy tắc curation R1",
        "",
        "| Decision | Điều kiện |",
        "|----------|-----------|",
        "| `keep_strong` | ESG body sạch, metric/governance/materiality rõ, không chrome |",
        "| `keep_but_needs_rewrite` | Fact ESG thật, Q generic hoặc passage còn noise nhẹ |",
        "| `drop_news_chrome` | News page chrome nặng, fact không đủ salvage |",
        "| `drop_listing_archive` | TOC/archive/report index |",
        "| `drop_contact_or_navigation` | Portal/nav/contact/FAQ |",
        "| `drop_financial_non_esg` | Analyst/financial dominant |",
        "| `drop_cross_company_contamination` | Cross-company marker |",
        "| `drop_too_generic` | Site text/generic Q không đủ specificity |",
        "| `drop_not_esg_enough` | Thiếu tín hiệu ESG substance |",
        "",
        "## Kết quả tổng quan",
        "",
        f"| Chỉ số | Giá trị |",
        f"|--------|--------:|",
        f"| Tổng seed input | {summary.get('input_total', 0)} |",
        f"| keep_strong | {summary.get('keep_strong', 0)} |",
        f"| keep_but_needs_rewrite | {summary.get('keep_but_needs_rewrite', 0)} |",
        f"| curated total (keep) | {summary.get('curated_total', 0)} |",
        f"| rejected total | {summary.get('rejected_total', 0)} |",
        f"| usable thực sự (strong) | {summary.get('usable_strong', 0)} |",
        f"| salvageable (rewrite) | {summary.get('salvageable_rewrite', 0)} |",
        "",
        "### Rejected theo nhóm",
        "",
    ]
    for k, v in sorted(summary.get("rejected_by_decision", {}).items()):
        lines.append(f"- `{k}`: **{v}**")

    lines.extend(["", "## Ví dụ theo nhóm", ""])
    for group, items in examples.items():
        lines.append(f"### `{group}`")
        for ex in items[:2]:
            lines.append(f"- **{ex.get('seed_id')}** ({ex.get('company')}): {ex.get('curation_notes', '')[:120]}")
        lines.append("")

    lines.extend(
        [
            "## Đánh giá",
            "",
            f"- **Seed usable thực sự (keep_strong):** {summary.get('usable_strong', 0)}",
            f"- **Seed cần rewrite:** {summary.get('salvageable_rewrite', 0)}",
            f"- **Công ty thiếu narrative sạch:** {summary.get('companies_weak', [])}",
            "",
            "## Kết luận",
            "",
        ]
    )
    if summary.get("working_workbook_ready"):
        lines.extend(
            [
                f"- **Workbook curated R1 đủ cho review round tiếp theo?** **Có** — {summary.get('curated_total', 0)} seed keep (strong + rewrite).",
                "- **Bước kế tiếp:** `review + rewrite round 2` trên workbook curated — **không** benchmark, **không** mini-pilot distillation.",
                f"- **Cần build seed workbook v2 từ source sạch hơn?** {summary.get('rebuild_v2_recommendation', '')}",
            ]
        )
    else:
        lines.extend(
            [
                "- **Workbook curated R1 chưa đủ** — cần mở rộng corpus/source sạch trước review round 2.",
                "- **Bước kế tiếp:** build seed workbook v2 từ corpus đã lọc contamination.",
            ]
        )

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_curation_r1(
    *,
    input_jsonl: Path,
    curated_jsonl: Path,
    rejected_jsonl: Path,
    curated_xlsx: Path,
) -> Dict[str, Any]:
    seeds = read_jsonl(input_jsonl)
    curated: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []
    all_curated: List[Dict[str, Any]] = []
    examples: Dict[str, List[Dict]] = defaultdict(list)

    for seed in seeds:
        row = curate_seed(seed)
        all_curated.append(row)
        dec = row["curation_decision"]
        examples[dec].append(row)
        if dec in KEEP_DECISIONS:
            curated.append(row)
        else:
            rejected.append(row)

    curated.sort(key=lambda x: (-float(x.get("rank") or 0), x.get("seed_id", "")))
    write_jsonl(curated_jsonl, curated)
    write_jsonl(rejected_jsonl, rejected)
    _write_workbook(curated, curated_xlsx)

    by_decision = Counter(r["curation_decision"] for r in all_curated)
    curated_dec = Counter(r["curation_decision"] for r in curated)
    rejected_dec = Counter(r["curation_decision"] for r in rejected)
    by_company_keep = Counter(r["company"] for r in curated)
    by_company_all = Counter(s["company"] for s in seeds)

    weak_companies = [co for co in by_company_all if by_company_keep.get(co, 0) < 2]

    keep_strong = curated_dec.get("keep_strong", 0)
    keep_rewrite = curated_dec.get("keep_but_needs_rewrite", 0)
    curated_total = len(curated)

    summary = {
        "curation_version": CURATION_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_total": len(seeds),
        "keep_strong": keep_strong,
        "keep_but_needs_rewrite": keep_rewrite,
        "curated_total": curated_total,
        "rejected_total": len(rejected),
        "usable_strong": keep_strong,
        "salvageable_rewrite": keep_rewrite,
        "rejected_by_decision": dict(rejected_dec),
        "curated_by_company": dict(by_company_keep),
        "input_by_company": dict(by_company_all),
        "companies_weak": weak_companies,
        "working_workbook_ready": curated_total >= 8,
        "rebuild_v2_recommendation": (
            "Chưa bắt buộc — curated R1 đủ review round 2; v2 khi mở rộng corpus Hansem/무신사 narrative sạch."
            if curated_total >= 8
            else "Có — cần corpus/source sạch hơn trước khi review có ý nghĩa."
        ),
        "curated_jsonl": str(curated_jsonl),
        "rejected_jsonl": str(rejected_jsonl),
        "curated_xlsx": str(curated_xlsx),
        "promote_ready_seed_ids": [r["seed_id"] for r in curated if r["curation_decision"] == "keep_strong"],
    }
    return summary, dict(examples)


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Curate reference seed workbook R1")
    root = Path(__file__).resolve().parents[2]
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v1.jsonl",
    )
    parser.add_argument(
        "--curated-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_curated_r1.jsonl",
    )
    parser.add_argument(
        "--rejected-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rejected_r1.jsonl",
    )
    parser.add_argument(
        "--curated-xlsx",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_curated_r1.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_reference_workbook_curation_r1.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_reference_workbook_curation_r1_summary.json",
    )
    args = parser.parse_args(argv)

    summary, examples = run_curation_r1(
        input_jsonl=root / args.input,
        curated_jsonl=root / args.curated_jsonl,
        rejected_jsonl=root / args.rejected_jsonl,
        curated_xlsx=root / args.curated_xlsx,
    )
    write_curation_report(summary, examples, root / args.report)
    (root / args.summary_json).write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(
        {
            "input": summary["input_total"],
            "keep_strong": summary["keep_strong"],
            "keep_rewrite": summary["keep_but_needs_rewrite"],
            "rejected": summary["rejected_total"],
        },
        ensure_ascii=True,
    ))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
