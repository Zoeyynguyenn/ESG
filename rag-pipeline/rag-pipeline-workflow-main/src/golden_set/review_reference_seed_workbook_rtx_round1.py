"""RTX Workbook Review Round 1 — triage candidate inflation (keep/rewrite/reject/collapse)."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import METRIC_HINT_RE, NUMBER_RE, YEAR_RE
from golden_set.io_utils import read_jsonl, write_jsonl

REVIEW_VERSION = "ref_review_rtx_r1"

RTX_ESG_KEYWORDS = [
    "esg", "sustainability", "climate", "emission", "greenhouse", "carbon",
    "governance", "diversity", "stakeholder", "ethics", "compliance", "tcfd",
    "cdp", "materiality", "scope 1", "scope 2", "net zero", "renewable",
    "dei", "cybersecurity", "data privacy", "bribery", "human rights",
]

GENERIC_QUESTION_EXACT = {
    "What ESG-related policies or performance does RTX disclose?",
    "What quantitative ESG metrics does RTX disclose?",
    "How have RTX's key ESG metrics changed over time?",
    "How is ESG governance structured at RTX?",
    "What ethics and compliance practices does RTX disclose?",
    "How does RTX address data security and privacy?",
    "What greenhouse gas emissions does RTX disclose?",
    "What compliance resolutions has RTX disclosed related to government contracts?",
    "What diversity and inclusion commitments does RTX report?",
    "What material ESG topics does RTX identify?",
    "What sustainability reporting frameworks does RTX reference?",
}

GENERIC_QUESTION_PATTERNS = [
    re.compile(r"^What ESG-related policies or performance does RTX disclose\?$"),
    re.compile(r"^What quantitative ESG metrics does RTX disclose\?$"),
    re.compile(r"^How have RTX's key ESG metrics changed over time\?$"),
    re.compile(r"^How is ESG governance structured at RTX\?$"),
]

TABLE_HEAVY_KINDS = {"appendix", "data_table", "questionnaire"}
GOVERNANCE_INFLATE_KINDS = {"10k", "proxy_statement"}

FRAMEWORK_ONLY_MARKERS = [
    "table of contents",
    "form 10-k",
    "schedule 14a",
    "united states securities and exchange commission",
    "check the appropriate box",
    "select from:",
    "numeric input",
    "[fixed row]",
    "add row",
]

GOVERNANCE_BOILERPLATE = [
    "the board of directors",
    "our board",
    "corporate governance",
    "audit committee",
    "compensation committee",
    "nominating and corporate governance",
    "director independence",
    "code of ethics",
]


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _passage(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("source_excerpt") or row.get("acceptable_disclosure") or "")


def _blob(row: Dict[str, Any]) -> str:
    return _norm_ws(
        " ".join(
            str(row.get(k, ""))
            for k in ("question_draft", "acceptable_disclosure", "source_excerpt")
        )
    ).lower()


def _is_generic_question(q: str) -> bool:
    if q in GENERIC_QUESTION_EXACT:
        return True
    return any(p.search(q) for p in GENERIC_QUESTION_PATTERNS)


def _table_density(passage: str) -> int:
    return passage.count("|") + passage.count("---")


def _esg_signal(blob: str, passage: str) -> int:
    score = sum(1 for kw in RTX_ESG_KEYWORDS if kw in blob)
    if METRIC_HINT_RE.search(passage):
        score += 2
    if re.search(r"\b(scope|emission|intensity|governance|stakeholder)\b", blob):
        score += 1
    return score


def _specificity_score(passage: str, qtype: str) -> int:
    score = 0
    nums = NUMBER_RE.findall(passage)
    years = YEAR_RE.findall(passage)
    if len(nums) >= 2:
        score += 3
    elif len(nums) == 1:
        score += 1
    if len(set(years)) >= 2:
        score += 3
    elif years:
        score += 1
    if METRIC_HINT_RE.search(passage):
        score += 2
    if re.search(r"\b(scope\s*[12]|market-based|location-based|co2e|ghg)\b", passage, re.I):
        score += 2
    if re.search(r"\b(RTX|Raytheon|Collins|Pratt\s*&\s*Whitney)\b", passage, re.I):
        score += 1
    if _table_density(passage) >= 8 and len(passage) < 200:
        score -= 3
    if len(passage) < 60:
        score -= 2
    if qtype == "trend" and len(set(years)) >= 2 and len(nums) >= 2:
        score += 2
    return score


def _is_table_residue(passage: str) -> bool:
    if _table_density(passage) >= 10 and _esg_signal(passage.lower(), passage) < 3:
        return True
    if re.fullmatch(r"[\|\-\s\d\.\%\,\$\(\)N/A]+", passage[: min(180, len(passage))]):
        return True
    if passage.count("`") >= 2 and len(passage) < 120:
        return True
    return False


def _is_framework_meta_only(blob: str, passage: str) -> bool:
    hits = sum(1 for m in FRAMEWORK_ONLY_MARKERS if m in blob)
    if hits >= 2 and _specificity_score(passage, "qualitative") < 3:
        return True
    if hits >= 3:
        return True
    return False


def _is_governance_boilerplate(passage: str, document_kind: str) -> bool:
    if document_kind not in GOVERNANCE_INFLATE_KINDS:
        return False
    lower = passage.lower()
    gov_hits = sum(1 for g in GOVERNANCE_BOILERPLATE if g in lower)
    if gov_hits >= 2 and _specificity_score(passage, "qualitative") < 4:
        return True
    return False


def _reject_reason(row: Dict[str, Any]) -> Optional[Tuple[str, str]]:
    passage = _passage(row)
    blob = _blob(row)
    q = row.get("question_draft") or ""
    dk = row.get("document_kind") or ""

    if len(passage) < 35:
        return "weak_grounding", "Excerpt quá ngắn, grounding yếu."
    if _is_table_residue(passage):
        return "table_residue_only", "Table residue thuần, không đủ fact usable."
    if _is_framework_meta_only(blob, passage):
        return "framework_meta_only", "Framework/form meta hoặc CDP form chrome."
    esg = _esg_signal(blob, passage)
    has_nums = bool(NUMBER_RE.search(passage))
    spec = _specificity_score(passage, row.get("question_type", ""))
    if esg < 2 and not has_nums and spec < 2:
        return "insufficient_esg_substance", "Không đủ tín hiệu ESG substance."
    if _is_governance_boilerplate(passage, dk):
        return "governance_boilerplate", "Governance boilerplate 10-K/proxy chung chung."
    if _is_generic_question(q) and spec < 2 and not has_nums:
        return "generic_question_weak_grounding", "Câu hỏi generic + disclosure không đủ specificity."
    return None


def _metric_family(passage: str) -> str:
    lower = passage.lower()
    families = [
        ("scope_ghg", r"scope\s*[12]|greenhouse|ghg|co2e|emission"),
        ("energy_intensity", r"energy intensity|gj/\$m|renewable"),
        ("water", r"water withdrawal|water consumption|water stress"),
        ("waste", r"waste|hazardous waste|recycl"),
        ("diversity", r"diversity|inclusion|dei|female|underrepresented"),
        ("governance_board", r"board of directors|director|audit committee|esg committee"),
        ("ethics_compliance", r"ethics|compliance|bribery|deferred prosecution|fcpa"),
        ("cyber_data", r"cybersecurity|data privacy|data security"),
        ("climate_risk", r"climate risk|tcfd|physical risk|transition risk"),
        ("stakeholder", r"stakeholder|materiality|engagement"),
        ("safety", r"safety|injury|osha|lost.?time"),
        ("supply_chain", r"supplier|supply chain|ecovadis"),
    ]
    for name, pat in families:
        if re.search(pat, lower):
            return name
    words = re.findall(r"[a-z]{5,}", lower)[:4]
    return "|".join(words) if words else "general"


def _infer_cluster_id(row: Dict[str, Any]) -> str:
    hint = row.get("workbook_cluster_hint") or "FC_GENERAL"
    q = row.get("question_draft") or ""
    dk = row.get("document_kind") or "unknown"
    rec = row.get("source_record_id") or ""
    family = _metric_family(_passage(row))

    if dk in TABLE_HEAVY_KINDS:
        return f"RTX::{hint}::{family}::{rec}"
    if dk in GOVERNANCE_INFLATE_KINDS and hint in ("FC_ESG_GOVERNANCE", "FC_QUAL_POLICY"):
        return f"RTX::{hint}::{family}::{dk}"
    q_slug = hashlib.md5(q.encode("utf-8")).hexdigest()[:8]
    return f"RTX::{hint}::{q_slug}::{family}"


def _passage_fingerprint(passage: str) -> str:
    nums = re.findall(r"\d+(?:\.\d+)?%?", passage)[:6]
    words = re.findall(r"[a-zA-Z]{5,}", passage.lower())[:8]
    return "|".join(sorted(set(words[:5] + nums[:4])))


def _suggest_rewrite(row: Dict[str, Any], passage: str) -> Tuple[str, str]:
    lower = passage.lower()
    qtype = row.get("question_type") or "qualitative"
    family = _metric_family(passage)

    if "scope 1" in lower or "scope 2" in lower:
        q = "What Scope 1 and Scope 2 GHG emissions does RTX report?"
    elif "energy intensity" in lower:
        q = "What is RTX's disclosed energy intensity (GJ per revenue)?"
    elif "market-based" in lower and "scope 2" in lower:
        q = "What are RTX's market-based Scope 2 emissions?"
    elif re.search(r"deferred prosecution|fcpa|bribery", lower):
        q = "What compliance matter related to government contracts has RTX disclosed?"
    elif "materiality" in lower or "material topic" in lower:
        q = "What material ESG topics does RTX identify?"
    elif family == "diversity":
        q = "What workforce diversity metrics does RTX disclose?"
    elif family == "governance_board" and re.search(r"\d+", passage):
        q = "What board or committee governance facts does RTX disclose?"
    elif qtype == "trend" and len(set(YEAR_RE.findall(passage))) >= 2:
        metric = family.replace("_", " ")
        q = f"How has RTX's {metric} changed across reported years?"
    elif qtype == "quantitative" and NUMBER_RE.search(passage):
        q = f"What specific {family.replace('_', ' ')} metric does RTX disclose?"
    else:
        q = row.get("question_draft") or ""

    disclosure = passage[:420].strip()
    if _table_density(disclosure) >= 6:
        for part in re.split(r"\|", passage):
            part = _norm_ws(part)
            if len(part) >= 40 and NUMBER_RE.search(part):
                disclosure = part[:420]
                break
    return q, disclosure


def _needs_rewrite(row: Dict[str, Any], passage: str) -> Tuple[bool, str]:
    reasons: List[str] = []
    q = row.get("question_draft") or ""
    spec = _specificity_score(passage, row.get("question_type", ""))

    if _is_generic_question(q):
        reasons.append("generic_question_template")
    if _table_density(passage) >= 4 and spec >= 2:
        reasons.append("table_excerpt_needs_trim")
    if len(passage) > 500:
        reasons.append("passage_too_broad")
    if spec >= 2 and spec < 4 and not _is_generic_question(q):
        reasons.append("moderate_specificity")
    if not reasons:
        return False, ""
    if _is_generic_question(q) and spec < 2:
        return False, ""
    return True, ";".join(reasons)


def _initial_decision(row: Dict[str, Any]) -> Tuple[str, str, str]:
    reject = _reject_reason(row)
    if reject:
        return "reject", reject[0], reject[1]

    passage = _passage(row)
    q = row.get("question_draft") or ""
    spec = _specificity_score(passage, row.get("question_type", ""))

    rewrite, rreason = _needs_rewrite(row, passage)
    if rewrite:
        return "rewrite", rreason, "Fact có thật; cần chỉnh question/disclosure cho specificity."

    if _is_generic_question(q) and spec < 4:
        return "rewrite", "generic_question_needs_specificity", "Generic template — cần rewrite theo fact cụ thể."

    return "keep", "clean_grounded_fact", "Grounded rõ, specificity đủ, usable cho manual review."


def _row_strength(row: Dict[str, Any], decision: str) -> float:
    rank = float(row.get("rank") or 0)
    bonus = {"keep": 12, "rewrite": 6, "reject": -100, "collapse_into_cluster": -50}
    passage = _passage(row)
    spec = _specificity_score(passage, row.get("question_type", ""))
    rank += spec * 2
    if row.get("seed_origin_type") == "rtx_primary_candidate":
        rank += 3
    elif row.get("seed_origin_type") == "rtx_sec_filing_candidate":
        rank += 1
    if not _is_generic_question(row.get("question_draft") or ""):
        rank += 4
    return rank + bonus.get(decision, 0)


def review_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    reviewed: List[Dict[str, Any]] = []
    for row in rows:
        out = dict(row)
        decision, reason, notes = _initial_decision(row)
        cluster_id = _infer_cluster_id(row)
        out.update(
            {
                "review_version": REVIEW_VERSION,
                "review_decision": decision,
                "review_reason": reason,
                "review_notes": notes,
                "cluster_id": cluster_id,
                "cluster_action": "rejected" if decision == "reject" else "pending",
                "rewritten_question_draft": "",
                "rewritten_disclosure_draft": "",
            }
        )
        if decision == "rewrite":
            q, d = _suggest_rewrite(row, _passage(row))
            out["rewritten_question_draft"] = q
            out["rewritten_disclosure_draft"] = d
        elif decision == "keep":
            out["rewritten_question_draft"] = row.get("question_draft") or ""
            out["rewritten_disclosure_draft"] = row.get("acceptable_disclosure") or ""
        reviewed.append(out)

    by_cluster: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in reviewed:
        if row["review_decision"] in ("keep", "rewrite"):
            by_cluster[row["cluster_id"]].append(row)

    def _collapse_group(group: List[Dict[str, Any]], reason_suffix: str = "") -> None:
        if len(group) <= 1:
            group[0]["cluster_action"] = "anchor"
            return
        sorted_g = sorted(
            group,
            key=lambda x: _row_strength(x, x["review_decision"]),
            reverse=True,
        )
        anchor = sorted_g[0]
        anchor["cluster_action"] = "anchor"
        for dup in sorted_g[1:]:
            dup["review_decision"] = "collapse_into_cluster"
            dup["review_reason"] = "duplicate_cluster_variant"
            dup["review_notes"] = f"Trùng cụm với anchor {anchor.get('seed_id')}{reason_suffix}"
            dup["cluster_action"] = "collapsed_variant"
            dup["rewritten_question_draft"] = ""
            dup["rewritten_disclosure_draft"] = ""

    for _cluster_id, members in by_cluster.items():
        groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for m in members:
            fp = _passage_fingerprint(_passage(m))
            groups[fp].append(m)

        for _fp, group in groups.items():
            _collapse_group(group)

        remaining = [m for m in members if m.get("review_decision") in ("keep", "rewrite")]
        if len(remaining) <= 1:
            continue

        dk = remaining[0].get("document_kind") or ""
        if dk in TABLE_HEAVY_KINDS:
            by_rec: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
            for m in remaining:
                by_rec[m.get("source_record_id") or "unknown"].append(m)
            for rec_group in by_rec.values():
                _collapse_group(rec_group, " (table record cap)")

        remaining = [m for m in members if m.get("review_decision") in ("keep", "rewrite")]
        if dk in GOVERNANCE_INFLATE_KINDS and len(remaining) > 4:
            by_q: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
            for m in remaining:
                by_q[m.get("question_draft") or ""].append(m)
            for q_group in by_q.values():
                if len(q_group) > 3:
                    _collapse_group(q_group[:], " (governance cap)")

        for m in members:
            if m.get("cluster_action") == "pending":
                m["cluster_action"] = "anchor"

    # Global cap per generic question + metric family (anti template inflation)
    generic_active = [
        m for m in reviewed
        if m.get("review_decision") in ("keep", "rewrite") and _is_generic_question(m.get("question_draft") or "")
    ]
    by_q_family: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for m in generic_active:
        key = f"{m.get('question_draft')}::{_metric_family(_passage(m))}"
        by_q_family[key].append(m)

    family_caps = {
        "What ESG-related policies or performance does RTX disclose?": 4,
        "What quantitative ESG metrics does RTX disclose?": 5,
        "How have RTX's key ESG metrics changed over time?": 4,
        "How is ESG governance structured at RTX?": 4,
        "What ethics and compliance practices does RTX disclose?": 3,
    }
    default_family_cap = 3

    for key, group in by_q_family.items():
        q = group[0].get("question_draft") or ""
        cap = family_caps.get(q, default_family_cap)
        if len(group) <= cap:
            continue
        sorted_g = sorted(group, key=lambda x: _row_strength(x, x["review_decision"]), reverse=True)
        for dup in sorted_g[cap:]:
            dup["review_decision"] = "collapse_into_cluster"
            dup["review_reason"] = "generic_template_inflation_cap"
            dup["review_notes"] = f"Cap generic template — giữ top {cap} cho family {_metric_family(_passage(dup))}"
            dup["cluster_action"] = "collapsed_variant"
            dup["rewritten_question_draft"] = ""
            dup["rewritten_disclosure_draft"] = ""

    question_total_caps = {
        "What ESG-related policies or performance does RTX disclose?": 45,
        "What quantitative ESG metrics does RTX disclose?": 70,
        "How have RTX's key ESG metrics changed over time?": 35,
        "How is ESG governance structured at RTX?": 35,
        "What ethics and compliance practices does RTX disclose?": 25,
        "How does RTX address data security and privacy?": 15,
    }
    default_question_cap = 12

    by_question: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for m in reviewed:
        if m.get("review_decision") in ("keep", "rewrite"):
            by_question[m.get("question_draft") or ""].append(m)

    for q, group in by_question.items():
        cap = question_total_caps.get(q, default_question_cap)
        if len(group) <= cap:
            continue
        sorted_g = sorted(group, key=lambda x: _row_strength(x, x["review_decision"]), reverse=True)
        for dup in sorted_g[cap:]:
            dup["review_decision"] = "collapse_into_cluster"
            dup["review_reason"] = "generic_question_total_cap"
            dup["review_notes"] = f"Cap tổng {cap} row cho question template"
            dup["cluster_action"] = "collapsed_variant"
            dup["rewritten_question_draft"] = ""
            dup["rewritten_disclosure_draft"] = ""

    return reviewed


WORKBOOK_COLUMNS = [
    "seed_id",
    "company",
    "question_type",
    "candidate_kind",
    "document_kind",
    "question_draft",
    "acceptable_disclosure",
    "source_record_id",
    "review_decision",
    "review_reason",
    "cluster_id",
    "cluster_action",
    "rewritten_question_draft",
    "rewritten_disclosure_draft",
    "review_notes",
]


def _write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "Guide"
    guide.append(["RTX Reference Seed Workbook — Review Round 1"])
    guide.append(["Decisions: keep / rewrite / reject / collapse — not canonical final"])

    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)

    ws = wb.create_sheet("Working_Set")
    ws.append(WORKBOOK_COLUMNS)
    for col in range(1, len(WORKBOOK_COLUMNS) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    active = [r for r in rows if r.get("review_decision") in ("keep", "rewrite")]
    for row in sorted(active, key=lambda x: (x.get("document_kind", ""), x.get("review_decision", ""), x.get("seed_id", ""))):
        ws.append([row.get(h, "") for h in WORKBOOK_COLUMNS])

    ws_rej = wb.create_sheet("Rejected")
    ws_rej.append(WORKBOOK_COLUMNS)
    for col in range(1, len(WORKBOOK_COLUMNS) + 1):
        c = ws_rej.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    for row in sorted(
        [r for r in rows if r.get("review_decision") == "reject"],
        key=lambda x: (x.get("review_reason", ""), x.get("seed_id", "")),
    ):
        ws_rej.append([row.get(h, "") for h in WORKBOOK_COLUMNS])

    ws_col = wb.create_sheet("Collapsed")
    ws_col.append(WORKBOOK_COLUMNS)
    for col in range(1, len(WORKBOOK_COLUMNS) + 1):
        c = ws_col.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    for row in sorted(
        [r for r in rows if r.get("review_decision") == "collapse_into_cluster"],
        key=lambda x: (x.get("cluster_id", ""), x.get("seed_id", "")),
    ):
        ws_col.append([row.get(h, "") for h in WORKBOOK_COLUMNS])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    examples: Dict[str, List[Dict[str, str]]] = {
        "table_collapse": [],
        "generic_rewrite": [],
        "weak_reject": [],
        "strong_keep": [],
    }
    for r in rows:
        sid = r.get("seed_id", "")
        if r.get("review_decision") == "collapse_into_cluster" and r.get("document_kind") in TABLE_HEAVY_KINDS:
            if len(examples["table_collapse"]) < 3:
                examples["table_collapse"].append(
                    {"seed_id": sid, "cluster_id": r.get("cluster_id", ""), "notes": r.get("review_notes", "")}
                )
        if r.get("review_decision") == "rewrite" and "generic" in (r.get("review_reason") or ""):
            if len(examples["generic_rewrite"]) < 3:
                examples["generic_rewrite"].append(
                    {
                        "seed_id": sid,
                        "before_q": r.get("question_draft", ""),
                        "after_q": r.get("rewritten_question_draft", ""),
                    }
                )
        if r.get("review_decision") == "reject":
            if len(examples["weak_reject"]) < 3:
                examples["weak_reject"].append(
                    {"seed_id": sid, "reason": r.get("review_reason", ""), "excerpt": _passage(r)[:100]}
                )
        if r.get("review_decision") == "keep" and r.get("cluster_action") == "anchor":
            if len(examples["strong_keep"]) < 3 and not _is_generic_question(r.get("question_draft") or ""):
                examples["strong_keep"].append(
                    {"seed_id": sid, "question": r.get("question_draft", ""), "excerpt": _passage(r)[:100]}
                )
    return examples


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Workbook Review RTX Round 1",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Triage `reference_seed_candidates_rtx_v1.jsonl` (3170 rows) thành workbook reviewable hơn",
        "trước manual review — giảm candidate inflation, không benchmark/canonical.",
        "",
        "## Vì sao không review thẳng 3170 row",
        "",
        "- Chỉ **11** unique `question_draft` — generic template inflation",
        "- Nhiều row cùng cluster, khác một dòng table hoặc số liệu liền kề",
        "- 10-K/proxy governance sections inflate yield rất nhanh",
        "",
        "## Rule triage round 1",
        "",
        "| Decision | Điều kiện |",
        "|----------|-----------|",
        "| `keep` | Fact rõ, specificity đủ, không generic yếu, anchor trong cluster |",
        "| `rewrite` | Fact có thật; question generic hoặc table excerpt cần gọt |",
        "| `reject` | Table residue, framework meta, governance boilerplate, generic+weak |",
        "| `collapse_into_cluster` | Trùng fact cluster với anchor mạnh hơn |",
        "",
        "## Kết quả tổng quan",
        "",
        f"- Total input: **{summary.get('input_total', 0)}**",
        f"- keep: **{summary.get('keep', 0)}**",
        f"- rewrite: **{summary.get('rewrite', 0)}**",
        f"- reject: **{summary.get('reject', 0)}**",
        f"- collapse_into_cluster: **{summary.get('collapse_into_cluster', 0)}**",
        f"- **Reviewable sau round 1 (keep + rewrite):** **{summary.get('reviewable_after_round1', 0)}**",
        "",
        "### Breakdown theo question_type (reviewable)",
        "",
    ]
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "### Breakdown theo document_kind (reviewable)", ""])
    for dk, n in summary.get("by_document_kind", {}).items():
        lines.append(f"- `{dk}`: {n}")

    lines.extend(["", "### Breakdown theo rejection reason", ""])
    for reason, n in summary.get("by_reason", {}).items():
        lines.append(f"- `{reason}`: {n}")

    lines.extend(["", "### Breakdown theo cluster action", ""])
    for action, n in summary.get("by_cluster_action", {}).items():
        lines.append(f"- `{action}`: {n}")

    lines.extend(["", "## Ví dụ", ""])
    lines.append("### Table duplicate bị collapse")
    for ex in examples.get("table_collapse", []):
        lines.append(f"- `{ex['seed_id']}` cluster `{ex['cluster_id']}` — {ex['notes']}")

    lines.append("")
    lines.append("### Generic row được rewrite")
    for ex in examples.get("generic_rewrite", []):
        lines.append(f"- `{ex['seed_id']}`: `{ex['before_q']}` → `{ex['after_q']}`")

    lines.append("")
    lines.append("### Weak row bị reject")
    for ex in examples.get("weak_reject", []):
        lines.append(f"- `{ex['seed_id']}` (`{ex['reason']}`): {ex['excerpt']}…")

    lines.append("")
    lines.append("### Strong row được keep")
    for ex in examples.get("strong_keep", []):
        lines.append(f"- `{ex['seed_id']}`: `{ex['question']}` — {ex['excerpt']}…")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- Reviewable rows (keep + rewrite): **{summary.get('reviewable_after_round1', 0)}**",
            f"- Manual review round 2 ready? **{summary.get('manual_review_ready_verdict', '')}**",
            f"- Flag: `manual_review_ready_flag` = **{summary.get('manual_review_ready_flag', False)}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_review(
    *,
    input_path: Path,
    reviewed_jsonl: Path,
    rejected_jsonl: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    reviewed = review_rows(rows)

    rejected_out = [r for r in reviewed if r.get("review_decision") in ("reject", "collapse_into_cluster")]
    write_jsonl(reviewed_jsonl, reviewed)
    write_jsonl(rejected_jsonl, rejected_out)
    _write_workbook(reviewed, workbook_path)

    decisions = Counter(r.get("review_decision") for r in reviewed)
    active = [r for r in reviewed if r.get("review_decision") in ("keep", "rewrite")]

    by_reason = Counter(r.get("review_reason") for r in reviewed if r.get("review_decision") == "reject")
    by_qtype_active = Counter(r.get("question_type") for r in active)
    by_dk_active = Counter(r.get("document_kind") for r in active)
    by_cluster_action = Counter(r.get("cluster_action") for r in reviewed)

    reviewable = decisions.get("keep", 0) + decisions.get("rewrite", 0)
    dk_coverage = len([dk for dk, n in by_dk_active.items() if n >= 5])
    collapse_rate = decisions.get("collapse_into_cluster", 0) / max(len(rows), 1)

    reduction_rate = 1 - (reviewable / max(len(rows), 1))
    kinds_represented = len(by_dk_active)
    manual_ready = (
        150 <= reviewable <= 500
        and kinds_represented >= 5
        and reduction_rate >= 0.85
    )

    examples = _pick_examples(reviewed)
    summary = {
        "review_version": REVIEW_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_total": len(rows),
        "keep": decisions.get("keep", 0),
        "rewrite": decisions.get("rewrite", 0),
        "reject": decisions.get("reject", 0),
        "collapse_into_cluster": decisions.get("collapse_into_cluster", 0),
        "reviewable_after_round1": reviewable,
        "by_question_type": dict(by_qtype_active),
        "by_document_kind": dict(by_dk_active),
        "by_reason": dict(by_reason.most_common(20)),
        "by_cluster_action": dict(by_cluster_action),
        "manual_review_ready_flag": manual_ready,
        "manual_review_ready_verdict": (
            "Có — đủ sạch để mở manual review round 2 (keep+rewrite workbook)"
            if manual_ready
            else "Một phần — vẫn có thể review có chọn lọc; kiểm tra Working_Set sheet"
        ),
        "output_reviewed_jsonl": str(reviewed_jsonl),
        "output_rejected_jsonl": str(rejected_jsonl),
        "output_workbook": str(workbook_path),
    }

    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="RTX workbook review round 1")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v1.jsonl",
    )
    parser.add_argument(
        "--reviewed-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_reviewed_round1.jsonl",
    )
    parser.add_argument(
        "--rejected-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_rejected_round1.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_rtx_review_round1.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_workbook_review_rtx_round1.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_workbook_review_rtx_round1_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_review(
        input_path=root / args.input,
        reviewed_jsonl=root / args.reviewed_jsonl,
        rejected_jsonl=root / args.rejected_jsonl,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "input_total",
                    "keep",
                    "rewrite",
                    "reject",
                    "collapse_into_cluster",
                    "reviewable_after_round1",
                    "manual_review_ready_flag",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
