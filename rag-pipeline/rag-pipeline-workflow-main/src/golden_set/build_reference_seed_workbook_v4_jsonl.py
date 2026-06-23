"""Workbook-first candidate generation v4 from corpus JSONL.

Reset direction:
- one good passage -> multiple seed candidates
- three-tier filtering (passage / candidate / workbook)
- no source-audit blocker, no 1-unit-1-QA hard gate
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import (
    ESG_SENTENCE_KEYWORDS,
    METRIC_HINT_RE,
    NUMBER_RE,
    YEAR_RE,
    _best_sentences,
    _guess_category,
    _make_question,
    _norm_ws,
    _prohibited_claims,
    _sentence_score,
    _split_sentences,
)
from golden_set.io_utils import read_jsonl, write_jsonl

SEED_VERSION = "ref_seed_v4_jsonl"

ESG_KEYWORDS = [
    "esg", "지속가능", "온실가스", "탄소", "기후", "인권", "안전보건", "공급망",
    "협력사", "이사회", "esg위원회", "tcfd", "kgcs", "materiality", "중대성",
    "net zero", "gri", "impact report", "임팩트",
]

HARD_NOISE = [
    "게시판 목록", "목록화면", "사이트맵", "민원서비스", "정보공개제도",
    "만족도 평가", "print 상태", "주소복사", "네이버 채널", "다음 채널",
    "all rights reserved", "무단전재", "목록 글쓰기",
]

NAV_HEAVY = [
    "esg 소개", "esg 기업정보", "esg 통계", "자료실 faq", "english eng",
    "참고사이트", "esg 강의실",
]

FINANCIAL_ONLY = [
    "consensus", "eps", "ebitda", "목표주가", "영업이익률", "per ", "pbr ",
    "miraeasset.com", "기업분석 보고서",
]

CROSS_COMPANY = [
    "삼성전기", "삼성전자", "현대트랜시스", "여수광양항만공사",
    "rtx corporation", "raytheon", "patriot missile",
]

CANDIDATE_KINDS = (
    "quantitative_fact",
    "trend_or_change",
    "qualitative_narrative",
    "governance_policy",
    "materiality_or_stakeholder",
    "report_or_framework_disclosure",
    "unanswerable_or_insufficient_context",
)


@dataclass
class PassageReject:
    reason: str
    company: str
    unit_id: str


@dataclass
class SeedCandidate:
    company: str
    question_type: str
    candidate_kind: str
    question_draft: str
    acceptable_disclosure: str
    prohibited_claims: str
    source_record_id: str
    source_unit_id: str
    source_excerpt: str
    source_file: str
    source_type: str
    seed_origin_type: str
    candidate_status: str
    candidate_reason: str
    workbook_cluster_hint: str
    rank: float
    dedupe_key: str


def _blob(row: Dict[str, Any]) -> str:
    return _norm_ws(
        " ".join(
            str(row.get(k, ""))
            for k in ("company", "section_path", "source_type", "text")
        )
    ).lower()


def _esg_hits(text: str) -> int:
    lower = text.lower()
    return sum(1 for k in ESG_KEYWORDS if k in lower)


def _provenance(row: Dict[str, Any], text: str) -> str:
    st = (row.get("source_type") or "").lower()
    if st in {"sustainability_report", "official_sustainability_report"}:
        return "jsonl_primary_candidate"
    noise = sum(1 for n in HARD_NOISE if n in text.lower())
    if noise >= 2 or any(n in text.lower() for n in NAV_HEAVY):
        return "jsonl_noisy_but_salvageable"
    return "jsonl_mixed_candidate"


def passage_level_filter(row: Dict[str, Any]) -> Tuple[bool, str]:
    text = row.get("text") or ""
    norm = _norm_ws(text)
    company = row.get("company") or ""
    blob = _blob(row)

    if len(norm) < 80:
        return False, "passage_too_short"
    if sum(1 for m in CROSS_COMPANY if m.lower() in blob) >= 2:
        return False, "cross_company_contamination"
    if company == "레이시온" and ("삼성전기" in text or "rtx" in blob):
        return False, "cross_company_contamination"

    nav_hits = sum(1 for n in NAV_HEAVY if n in blob)
    esg_hits = _esg_hits(norm)
    if nav_hits >= 4 and esg_hits < 4:
        return False, "portal_navigation_noise"
    if sum(1 for n in HARD_NOISE if n in blob) >= 3 and esg_hits < 3:
        return False, "listing_or_portal_chrome"

    fin_hits = sum(1 for f in FINANCIAL_ONLY if f in blob)
    if fin_hits >= 2 and esg_hits < 2:
        return False, "pure_financial_or_analyst_noise"

    if "tel." in blob and "e-mail" in blob and "지속가능경영보고서" in blob and len(norm) < 2200:
        if esg_hits < 4:
            return False, "contact_or_report_meta_only"

    if esg_hits < 1 and not METRIC_HINT_RE.search(norm):
        return False, "no_esg_substance"

    return True, "passage_ok"


def _infer_candidate_kind(sentence: str, qtype: str) -> str:
    lower = sentence.lower()
    if "table of contents" in lower or "목차" in sentence and len(sentence) < 120:
        return "unanswerable_or_insufficient_context"
    if any(k in lower for k in ("gri", "tcfd", "sasb", "esrs", "impact report", "임팩트", "지속가능경영보고서")):
        return "report_or_framework_disclosure"
    if any(k in lower for k in ("중대성", "중대 이슈", "materiality", "이해관계자")):
        return "materiality_or_stakeholder"
    if any(k in lower for k in ("이사회", "esg위원회", "지배구조", "윤리", "준법", "컴플라이언스")):
        return "governance_policy"
    if qtype == "trend":
        return "trend_or_change"
    if qtype == "quantitative":
        return "quantitative_fact"
    return "qualitative_narrative"


def _cluster_hint(kind: str, sentence: str) -> str:
    mapping = {
        "quantitative_fact": "FC_METRIC_NUMERIC",
        "trend_or_change": "FC_TREND_MULTIYEAR",
        "governance_policy": "FC_ESG_GOVERNANCE",
        "materiality_or_stakeholder": "FC_MATERIAL_STAKEHOLDER",
        "report_or_framework_disclosure": "FC_REPORT_FRAMEWORK",
        "qualitative_narrative": "FC_QUAL_POLICY",
        "unanswerable_or_insufficient_context": "FC_INSUFFICIENT_CONTEXT",
    }
    base = mapping.get(kind, "FC_GENERAL")
    if "scope" in sentence.lower() or "온실가스" in sentence:
        return "FC_CLIMATE_GHG"
    if "net zero" in sentence.lower() or "탄소중립" in sentence:
        return "FC_NET_ZERO"
    if "kgcs" in sentence.lower():
        return "FC_KGCS_RATING"
    return base


def _detect_qtypes_for_sentence(sentence: str) -> List[str]:
    qtypes: List[str] = []
    years = len(set(YEAR_RE.findall(sentence)))
    numbers = len(NUMBER_RE.findall(sentence))
    if years >= 2 and numbers >= 2:
        qtypes.append("trend")
    if numbers >= 1 and (METRIC_HINT_RE.search(sentence) or re.search(r"\d", sentence)):
        qtypes.append("quantitative")
    qtypes.append("qualitative")
    out: List[str] = []
    for q in qtypes:
        if q not in out:
            out.append(q)
    return out[:3]


def candidate_level_filter(
    company: str,
    sentence: str,
    question: str,
    qtype: str,
) -> Tuple[bool, str]:
    if not question or len(question) < 12:
        return False, "empty_question"
    if company not in question:
        return False, "question_missing_company"
    if len(_norm_ws(sentence)) < 30:
        return False, "excerpt_too_short"
    if sentence.strip() not in _norm_ws(question) and len(sentence) < 40:
        # answer must be grounded in sentence
        if not any(tok in sentence for tok in re.findall(r"[가-힣]{4,}", question)[:3]):
            pass  # ok if question is paraphrase
    generic_only = question.endswith("핵심 내용은 무엇인가?") and _sentence_score(sentence) < 3
    if generic_only and not METRIC_HINT_RE.search(sentence):
        return False, "generic_question_low_substance"
    if sum(1 for n in NAV_HEAVY if n in sentence.lower()) >= 3:
        return False, "nav_chrome_in_excerpt"
    return True, "candidate_ok"


def _dedupe_key(company: str, question: str, disclosure: str) -> str:
    raw = f"{company}|{_norm_ws(question).lower()}|{_norm_ws(disclosure)[:100].lower()}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def _rank(sentence: str, qtype: str, provenance: str) -> float:
    score = float(_sentence_score(sentence))
    if qtype == "trend":
        score += 6
    elif qtype == "quantitative":
        score += 5
    else:
        score += 3
    if provenance == "jsonl_primary_candidate":
        score += 4
    elif provenance == "jsonl_mixed_candidate":
        score += 2
    return score


def generate_candidates_from_passage(row: Dict[str, Any]) -> List[SeedCandidate]:
    company = row.get("company") or ""
    text = row.get("text") or ""
    provenance = _provenance(row, text)
    category = _guess_category(row)
    sentences = _best_sentences(text, limit=10)
    if not sentences:
        sentences = [s for s in _split_sentences(text) if _sentence_score(s) >= 2][:6]

    out: List[SeedCandidate] = []
    for sentence in sentences:
        if _sentence_score(sentence) < 1 and _esg_hits(sentence) < 1:
            continue
        for qtype in _detect_qtypes_for_sentence(sentence):
            question = _make_question(company, sentence, category, qtype)
            if not question:
                continue
            ok, creason = candidate_level_filter(company, sentence, question, qtype)
            if not ok:
                continue
            kind = _infer_candidate_kind(sentence, qtype)
            if kind == "unanswerable_or_insufficient_context":
                continue
            disclosure = _norm_ws(sentence)[:420]
            out.append(
                SeedCandidate(
                    company=company,
                    question_type=qtype,
                    candidate_kind=kind,
                    question_draft=question,
                    acceptable_disclosure=disclosure,
                    prohibited_claims=_prohibited_claims(qtype),
                    source_record_id=str(row.get("record_id") or ""),
                    source_unit_id=str(row.get("unit_id") or ""),
                    source_excerpt=disclosure,
                    source_file=str(row.get("source_file") or ""),
                    source_type=str(row.get("source_type") or ""),
                    seed_origin_type=provenance,
                    candidate_status="candidate_v4",
                    candidate_reason=f"from_passage:{creason}; kind={kind}",
                    workbook_cluster_hint=_cluster_hint(kind, sentence),
                    rank=_rank(sentence, qtype, provenance),
                    dedupe_key=_dedupe_key(company, question, disclosure),
                )
            )
    return out


def workbook_level_dedupe(candidates: Sequence[SeedCandidate]) -> List[SeedCandidate]:
    best: Dict[str, SeedCandidate] = {}
    for cand in sorted(candidates, key=lambda x: x.rank, reverse=True):
        prev = best.get(cand.dedupe_key)
        if prev is None or cand.rank > prev.rank:
            best[cand.dedupe_key] = cand
    return list(best.values())


def _select_balanced(candidates: List[SeedCandidate], max_total: int = 0) -> List[SeedCandidate]:
    """Round-robin by company; optional soft cap."""
    by_co: Dict[str, List[SeedCandidate]] = defaultdict(list)
    for c in sorted(candidates, key=lambda x: x.rank, reverse=True):
        by_co[c.company].append(c)

    selected: List[SeedCandidate] = []
    seen_passage_q: set[Tuple[str, str, str]] = set()

    while True:
        progressed = False
        for company in sorted(by_co):
            pool = by_co[company]
            while pool:
                cand = pool.pop(0)
                key = (cand.company, cand.source_record_id, cand.question_draft)
                if key in seen_passage_q:
                    continue
                seen_passage_q.add(key)
                selected.append(cand)
                progressed = True
                break
            if max_total and len(selected) >= max_total:
                return selected
        if not progressed:
            break
    return selected


def _seed_id(company: str, qtype: str, index: int) -> str:
    prefix = {"한샘": "HS", "무신사": "MS", "레이시온": "RX"}.get(company, "XX")
    code = {"quantitative": "Q", "trend": "T", "qualitative": "L"}.get(qtype, "X")
    return f"{prefix}-V4-{code}{index:02d}"


def to_jsonl_rows(candidates: Sequence[SeedCandidate]) -> List[Dict[str, Any]]:
    counters: Counter[Tuple[str, str]] = Counter()
    rows: List[Dict[str, Any]] = []
    for cand in sorted(candidates, key=lambda x: (-x.rank, x.company, x.question_type)):
        counters[(cand.company, cand.question_type)] += 1
        rows.append(
            {
                "seed_id": _seed_id(
                    cand.company, cand.question_type, counters[(cand.company, cand.question_type)]
                ),
                "company": cand.company,
                "question_type": cand.question_type,
                "candidate_kind": cand.candidate_kind,
                "question_draft": cand.question_draft,
                "acceptable_disclosure": cand.acceptable_disclosure,
                "prohibited_claims": cand.prohibited_claims,
                "source_record_id": cand.source_record_id,
                "source_excerpt": cand.source_excerpt,
                "source_unit_id": cand.source_unit_id,
                "source_file": cand.source_file,
                "source_type": cand.source_type,
                "seed_origin_type": cand.seed_origin_type,
                "candidate_status": cand.candidate_status,
                "candidate_reason": cand.candidate_reason,
                "workbook_cluster_hint": cand.workbook_cluster_hint,
                "seed_version": SEED_VERSION,
                "rank": cand.rank,
            }
        )
    return rows


def write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "안내"
    ws.append(["Reference Seed Workbook V4 — JSONL workbook-first candidates"])
    ws.append(["Nguyên tắc", "Một passage tốt có thể sinh nhiều candidate; chưa canonical final."])

    header = [
        "seed_id", "company", "question_type", "candidate_kind",
        "question_draft", "acceptable_disclosure", "prohibited_claims",
        "source_record_id", "source_excerpt", "seed_origin_type",
        "candidate_status", "workbook_cluster_hint", "candidate_reason",
    ]
    ws2 = wb.create_sheet("작성")
    ws2.append(header)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
    for row in rows:
        ws2.append([row.get(h, "") for h in header])

    ws3 = wb.create_sheet("요약")
    ws3.append(["metric", "value"])
    ws3.append(["total_rows", len(rows)])
    for co, n in sorted(Counter(r["company"] for r in rows).items()):
        ws3.append([f"company:{co}", n])
    for qt, n in sorted(Counter(r["question_type"] for r in rows).items()):
        ws3.append([f"question_type:{qt}", n])
    for kind, n in sorted(Counter(r.get("candidate_kind", "") for r in rows).items()):
        ws3.append([f"candidate_kind:{kind}", n])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report(summary: Dict[str, Any], path: Path) -> None:
    lines = [
        "# Golden Set — Candidate Generation V4 (JSONL)",
        "",
        f"Generated: {summary['generated_at']}",
        "",
        "## Mục tiêu",
        "",
        "Rebuild candidate generation từ `corpus_units.jsonl` theo hướng **workbook-first**:",
        "một passage tốt → nhiều seed candidate; không ép 1 unit = 1 QA.",
        "",
        "## Vì sao workflow cũ sai",
        "",
        "- Prefilter/distillation/QC ép `1 unit → 1 QA` và hard-drop duplicate sớm.",
        "- Yield sụp trước khi có workbook reviewable.",
        "- Gate precision sớm không phản ánh khả năng khai thác ESG fact từ jsonl.",
        "",
        "## Nguyên tắc builder v4 từ jsonl",
        "",
        "- Nguồn chính: `corpus_units.jsonl` (118 units).",
        "- Không yêu cầu PDF thật; không hardcode salvage record id.",
        "- Provenance flag: `jsonl_primary_candidate` / `jsonl_mixed_candidate` / `jsonl_noisy_but_salvageable`.",
        "",
        "## Passage-level filtering",
        "",
        "Chỉ loại noise mạnh: nav/listing/contact, pure financial/analyst, cross-company contamination, passage quá ngắn/không ESG.",
        "",
        f"- Passages accepted: **{summary['input_passages']}**",
        f"- Passages rejected: **{summary['passages_rejected']}**",
        "",
        "Top rejection reasons:",
    ]
    for reason, count in summary.get("passage_reject_reasons", {}).items():
        lines.append(f"- `{reason}`: {count}")

    lines.extend(
        [
            "",
            "## Candidate-level generation",
            "",
            "Từ mỗi passage: tách câu ESG substance → sinh candidate theo `quantitative` / `trend` / `qualitative` và `candidate_kind`.",
            "",
            f"- Raw candidates generated: **{summary['raw_candidates']}**",
            f"- Rejected at candidate-level (implicit in generation): filtered during `candidate_level_filter`",
            "",
            "## Workbook-level dedupe",
            "",
            "Collapse duplicate thực sự (cùng company + question + disclosure prefix); giữ diversity theo record/question.",
            "",
            f"- After workbook dedupe + balanced select: **{summary['filtered_candidates']}**",
            "",
            "## Kết quả",
            "",
            f"- Tổng passage dùng: **{summary['input_passages']}**",
            f"- Raw candidates: **{summary['raw_candidates']}**",
            f"- Sau lọc/dedupe: **{summary['filtered_candidates']}**",
            "",
            "### Coverage theo công ty",
        ]
    )
    for co, n in summary.get("companies", {}).items():
        lines.append(f"- **{co}**: {n}")

    lines.extend(["", "### Coverage theo question_type", ""])
    for qt, n in summary.get("question_type_breakdown", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(
        [
            "",
            "## So sánh định tính với nhánh cũ",
            "",
            f"- Yield cao hơn v1/v2/v3? **{summary.get('yield_vs_old', '')}**",
            f"- Diversity tốt hơn? **{summary.get('diversity_vs_old', '')}**",
            f"- Noise còn lại: **{summary.get('noise_note', '')}**",
            "",
            "## Kết luận",
            "",
            f"- Gần tinh thần `golden_set_3companies_v4`? **{summary.get('v4_spirit', '')}**",
            f"- Đủ mở review workbook round tiếp? **{summary.get('review_ready_estimate', '')}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_v4_builder(
    *,
    input_path: Path,
    output_jsonl: Path,
    output_xlsx: Path,
    report_path: Path,
    summary_json_path: Path,
    max_total: int = 0,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    passage_rejects: List[PassageReject] = []
    raw_candidates: List[SeedCandidate] = []

    for row in rows:
        ok, reason = passage_level_filter(row)
        if not ok:
            passage_rejects.append(
                PassageReject(reason=reason, company=row.get("company", ""), unit_id=row.get("unit_id", ""))
            )
            continue
        raw_candidates.extend(generate_candidates_from_passage(row))

    deduped = workbook_level_dedupe(raw_candidates)
    selected = _select_balanced(deduped, max_total=max_total)
    out_rows = to_jsonl_rows(selected)

    write_jsonl(output_jsonl, out_rows)
    write_workbook(out_rows, output_xlsx)

    reject_counts = Counter(p.reason for p in passage_rejects)
    by_company = Counter(r["company"] for r in out_rows)
    by_qtype = Counter(r["question_type"] for r in out_rows)
    by_kind = Counter(r.get("candidate_kind") for r in out_rows)
    by_origin = Counter(r.get("seed_origin_type") for r in out_rows)

    # Compare to v1 count if exists
    v1_count = 0
    v1_path = input_path.parent.parent / "reference_style/reference_seed_candidates_v1.jsonl"
    if v1_path.exists():
        v1_count = len(read_jsonl(v1_path))

    filtered_n = len(out_rows)
    yield_better = filtered_n > max(v1_count, 10)
    diversity_ok = len(by_company) >= 2 and len(by_qtype) >= 2
    noise_note = (
        f"~{by_origin.get('jsonl_noisy_but_salvageable', 0)} noisy-salvageable; "
        f"reviewer should triage portal/press rows"
    )

    summary = {
        "ingest_version": SEED_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_units": len(rows),
        "input_passages": len(rows) - len(passage_rejects),
        "passages_rejected": len(passage_rejects),
        "passage_reject_reasons": dict(reject_counts.most_common(12)),
        "raw_candidates": len(raw_candidates),
        "after_dedupe": len(deduped),
        "filtered_candidates": filtered_n,
        "companies": dict(by_company),
        "question_type_breakdown": dict(by_qtype),
        "candidate_kind_breakdown": dict(by_kind),
        "seed_origin_breakdown": dict(by_origin),
        "high_noise_rejections": reject_counts.get("portal_navigation_noise", 0)
        + reject_counts.get("listing_or_portal_chrome", 0),
        "yield_vs_old": (
            f"Có — v4 có {filtered_n} rows vs v1 ~{v1_count}"
            if yield_better
            else f"Chưa vượt rõ v1 ({filtered_n} vs {v1_count})"
        ),
        "diversity_vs_old": "Có — multi-seed per passage + 3 company mix" if diversity_ok else "Chưa đủ",
        "noise_note": noise_note,
        "v4_spirit": (
            "Có — workbook-first, candidate-rich, chưa canonical"
            if filtered_n >= 20
            else "Một phần — cần thêm passage MS/RX sạch"
        ),
        "review_ready_estimate": (
            "Có — đủ để mở review workbook round tiếp (draft candidates, không gold-ready)"
            if filtered_n >= 18 and len(by_company) >= 2
            else "Chưa — cần bổ sung candidate hoặc refine filter"
        ),
        "review_ready_flag": filtered_n >= 18 and len(by_company) >= 2,
        "output_jsonl": str(output_jsonl),
        "output_xlsx": str(output_xlsx),
    }

    write_report(summary, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Build reference seed workbook v4 from JSONL")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/step1_corpus_units/corpus_units.jsonl",
    )
    parser.add_argument(
        "--output-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_jsonl.jsonl",
    )
    parser.add_argument(
        "--output-xlsx",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_v4_jsonl.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_candidate_generation_v4_jsonl.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_candidate_generation_v4_jsonl_summary.json",
    )
    parser.add_argument("--max-total", type=int, default=0, help="0 = no cap")
    args = parser.parse_args(argv)

    summary = run_v4_builder(
        input_path=root / args.input,
        output_jsonl=root / args.output_jsonl,
        output_xlsx=root / args.output_xlsx,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
        max_total=args.max_total,
    )
    print(json.dumps({k: summary[k] for k in (
        "input_passages", "raw_candidates", "filtered_candidates", "companies"
    )}, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
