"""Freeze Gold Core V1 from round 4 approved rows (한샘 + 무신사)."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.io_utils import read_jsonl, write_jsonl
from golden_set.prepare_manual_review_round2 import _norm_ws

GOLD_CORE_VERSION = "gold_core_v1"
GOLD_STATUS = "frozen_approved"

GOLD_FIELDS = [
    "gold_id",
    "company",
    "question_type",
    "question",
    "ground_truth_context",
    "ground_truth_answer",
    "facts_tuple",
    "prohibited_claims",
    "source_record_id",
    "source_file",
    "fact_cluster_id",
    "gold_version",
    "gold_status",
    "notes",
]

WORKBOOK_GOLD_COLS = GOLD_FIELDS + ["seed_id", "resolved_cluster_hint"]


def _normalize_prohibited(text: str, qtype: str) -> str:
    t = _norm_ws(text).replace("; ", "\n")
    if t and "금지" in t:
        return t
    if qtype == "quantitative":
        return "원문에 없는 수치 추가 금지\n단위 변경 추정 금지\n미공시 항목 단정 금지"
    if qtype == "trend":
        return "원문에 없는 원인 추정 금지\n향후 개선 보장 금지\n미공시 수치 보완 금지"
    return "완전 달성/전면 준수 단정 금지\n원문 밖 정책 확장 해석 금지\n미공시 성과 추정 금지"


def _split_atomic_facts(disclosure: str, company: str, limit: int = 4) -> List[str]:
    text = _norm_ws(disclosure)
    if not text:
        return []
    parts = re.split(r"(?<=[.!?다])\s+|[;；]\s*|\s+▲", text)
    facts: List[str] = []
    for part in parts:
        p = _norm_ws(part)
        if len(p) < 20:
            continue
        if company not in p and len(p) < 60:
            continue
        facts.append(p[:220])
        if len(facts) >= limit:
            break
    if not facts and text:
        facts = [text[:220]]
    return facts[:limit]


def _facts_tuple(company: str, facts: Sequence[str]) -> str:
    lines = []
    for i, fact in enumerate(facts[:4], start=1):
        lines.append(f"{company} | {i} | {fact}")
    return "\n".join(lines)


def _ground_truth_answer(disclosure: str, facts: Sequence[str]) -> str:
    if facts:
        return facts[0][:420]
    return _norm_ws(disclosure)[:420]


def _fact_cluster_id(row: Dict[str, Any]) -> str:
    return (
        row.get("resolved_cluster_hint")
        or row.get("canonical_fact_target")
        or row.get("cluster_id")
        or "FC_UNKNOWN"
    )


def _gold_id(company: str, index: int) -> str:
    prefix = {"한샘": "HS", "무신사": "MS"}.get(company, "XX")
    return f"GC1-{prefix}-{index:03d}"


def freeze_row(row: Dict[str, Any], gold_index: int) -> Dict[str, Any]:
    company = row.get("company") or ""
    qtype = row.get("question_type") or "qualitative"
    question = _norm_ws(row.get("final_question") or "")
    context = _norm_ws(row.get("final_disclosure") or "")
    facts = _split_atomic_facts(context, company)
    prohib_raw = row.get("canonical_prohibited_claims") or row.get("prohibited_claims") or ""

    notes_parts = [
        f"seed_id={row.get('seed_id', '')}",
        f"gold_decision={row.get('gold_decision', '')}",
    ]
    if row.get("needs_cluster_rename"):
        notes_parts.append(f"cluster_hint={row.get('resolved_cluster_hint', '')}")
    if row.get("canonical_merge_note"):
        notes_parts.append(str(row.get("canonical_merge_note"))[:120])

    return {
        "gold_id": _gold_id(company, gold_index),
        "company": company,
        "question_type": qtype,
        "question": question,
        "ground_truth_context": context,
        "ground_truth_answer": _ground_truth_answer(context, facts),
        "facts_tuple": _facts_tuple(company, facts),
        "prohibited_claims": _normalize_prohibited(prohib_raw, qtype),
        "source_record_id": row.get("source_record_id") or "",
        "source_file": row.get("source_file") or "records/company_evidence.jsonl",
        "fact_cluster_id": _fact_cluster_id(row),
        "gold_version": GOLD_CORE_VERSION,
        "gold_status": GOLD_STATUS,
        "notes": "; ".join(notes_parts),
        "seed_id": row.get("seed_id", ""),
        "resolved_cluster_hint": row.get("resolved_cluster_hint", ""),
    }


def write_workbook(
    gold_rows: Sequence[Dict[str, Any]],
    hold_rows: Sequence[Dict[str, Any]],
    path: Path,
) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide.append(["Gold Core V1 — frozen from round4 approved"])
    guide.append(["Hold", "Expansion lane — not in v1 freeze"])

    ws = wb.create_sheet("Gold_Core_V1")
    ws.append(WORKBOOK_GOLD_COLS)
    fill = PatternFill("solid", fgColor="166534")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(WORKBOOK_GOLD_COLS) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    for r in gold_rows:
        ws.append([r.get(c, "") for c in WORKBOOK_GOLD_COLS])

    ws_h = wb.create_sheet("Hold_Backlog")
    hold_cols = ["seed_id", "company", "question_type", "final_question", "gold_reason", "resolved_cluster_hint"]
    ws_h.append(hold_cols)
    for r in hold_rows:
        ws_h.append([r.get(c, "") for c in hold_cols])

    ws_rx = wb.create_sheet("RX_Backlog_Status")
    ws_rx.append(["item", "status"])
    ws_rx.append(["rx_in_core", "excluded"])
    ws_rx.append(["rx_status", "source-acquisition dependent"])
    ws_rx.append(["hold_expansion", f"{len(hold_rows)} rows — refine round 5"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_eval_md(gold_rows: Sequence[Dict[str, Any]], path: Path) -> None:
    lines = [
        "# Golden Set — Gold Core V1 (KO eval mapping)",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "| ID | Company | Question | Expected Source | Ground Truth Answer | Record ID | Question Type | Cluster |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for r in gold_rows:
        ans = (r.get("ground_truth_answer") or "").replace("|", "\\|")[:120]
        q = (r.get("question") or "").replace("|", "\\|")[:100]
        lines.append(
            f"| {r.get('gold_id', '')} | {r.get('company', '')} | {q} | "
            f"{r.get('source_file', '')} | {ans} | {r.get('source_record_id', '')} | "
            f"{r.get('question_type', '')} | {r.get('fact_cluster_id', '')} |"
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_report(summary: Dict[str, Any], path: Path) -> None:
    lines = [
        "# Golden Set — Freeze Gold Core V1",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Freeze **gold_core_v1** từ 26 row `gold_approve` — artifact gold chính thức nhỏ nhưng usable cho 한샘 + 무신사.",
        "",
        "## Vì sao freeze 26 approved ngay lúc này",
        "",
        "- Core round 4 đã tách rõ approve/hold/reject.",
        "- 26 approved đủ làm baseline gold core; 17 hold là lane mở rộng sau.",
        "- Không chặn freeze vì hold chưa refine.",
        "",
        "## Input approved set",
        "",
        f"**{summary.get('input_approved_count', 0)}** rows từ `golden_set_core_round4_approved.jsonl`.",
        "",
        "## Schema final của gold core",
        "",
        "Fields: `gold_id`, `company`, `question_type`, `question`, `ground_truth_context`,",
        "`ground_truth_answer`, `facts_tuple`, `prohibited_claims`, `source_record_id`,",
        "`source_file`, `fact_cluster_id`, `gold_version`, `gold_status`, `notes`.",
        "",
        "## Kết quả",
        "",
        f"- **gold_core_v1_count:** {summary.get('gold_core_v1_count', 0)}",
        "",
        "### Breakdown theo công ty",
        "",
    ]
    for co, n in summary.get("by_company", {}).items():
        lines.append(f"- **{co}**: {n}")

    lines.extend(["", "### Breakdown theo question_type", ""])
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "### Breakdown theo fact cluster", ""])
    for cid, n in summary.get("by_cluster", {}).items():
        lines.append(f"- `{cid}`: {n}")

    lines.extend(
        [
            "",
            "## Đánh giá",
            "",
            "- **Core gold nhỏ nhưng usable:** 26 row frozen, không inflate từ hold.",
            f"- **Hold backlog:** {summary.get('hold_backlog_count', 0)} row — refine round 5.",
            "- **RX:** ngoài core — source-acquisition dependent.",
            "",
            "## Kết luận",
            "",
            f"- gold_core_v1 sẵn sàng artifact chính thức? **{summary.get('artifact_ready_verdict', '')}**",
            f"- Bước tiếp theo: **{summary.get('next_step_recommendation', '')}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_freeze(
    *,
    approved_path: Path,
    hold_path: Path,
    output_jsonl: Path,
    workbook_path: Path,
    eval_md_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    approved = read_jsonl(approved_path)
    hold: List[Dict[str, Any]] = []
    if hold_path.exists():
        hold = read_jsonl(hold_path)

    counters: Dict[str, int] = defaultdict(int)
    gold_rows: List[Dict[str, Any]] = []
    for row in sorted(approved, key=lambda x: (x.get("company", ""), x.get("seed_id", ""))):
        co = row.get("company") or ""
        counters[co] += 1
        gold_rows.append(freeze_row(row, counters[co]))

    write_jsonl(output_jsonl, [{k: r[k] for k in GOLD_FIELDS} for r in gold_rows])
    write_workbook(gold_rows, hold, workbook_path)
    write_eval_md(gold_rows, eval_md_path)

    summary = {
        "gold_version": GOLD_CORE_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_approved_count": len(approved),
        "gold_core_v1_count": len(gold_rows),
        "by_company": dict(Counter(r["company"] for r in gold_rows)),
        "by_question_type": dict(Counter(r["question_type"] for r in gold_rows)),
        "by_cluster": dict(Counter(r["fact_cluster_id"] for r in gold_rows)),
        "hold_backlog_count": len(hold),
        "rx_backlog_status": "source-acquisition dependent — excluded from gold_core_v1",
        "artifact_ready_verdict": "Có — gold_core_v1 frozen, usable làm artifact chính thức core",
        "next_step_recommendation": "Refine hold round 5 (17 rows) song song eval mapping từ eval_gold_core_v1_ko.md — không benchmark trong round freeze",
        "output_jsonl": str(output_jsonl),
        "output_workbook": str(workbook_path),
        "output_eval_md": str(eval_md_path),
    }

    write_report(summary, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Freeze gold core v1 from approved rows")
    parser.add_argument(
        "--approved",
        default="data/golden_set/v2/step6_gold/golden_set_core_round4_approved.jsonl",
    )
    parser.add_argument(
        "--hold",
        default="data/golden_set/v2/step6_gold/golden_set_core_round4_hold.jsonl",
    )
    parser.add_argument(
        "--output-jsonl",
        default="data/golden_set/v2/step6_gold/golden_set_core_v1.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/step6_gold/golden_set_core_v1.xlsx",
    )
    parser.add_argument(
        "--eval-md",
        default="data/golden_set/v2/step6_gold/eval_gold_core_v1_ko.md",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_freeze_gold_core_v1.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_freeze_gold_core_v1_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_freeze(
        approved_path=root / args.approved,
        hold_path=root / args.hold,
        output_jsonl=root / args.output_jsonl,
        workbook_path=root / args.workbook,
        eval_md_path=root / args.eval_md,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(json.dumps({k: summary[k] for k in ("gold_core_v1_count", "by_company", "hold_backlog_count")}, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
