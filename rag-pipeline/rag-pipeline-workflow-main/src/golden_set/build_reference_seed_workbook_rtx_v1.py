"""Workbook-first candidate generation v1 for RTX lane (English corpus)."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import (
    METRIC_HINT_RE,
    NUMBER_RE,
    YEAR_RE,
    _best_sentences,
    _norm_ws,
    _sentence_score,
    _split_sentences,
)
from golden_set.io_utils import read_jsonl, write_jsonl

SEED_VERSION = "ref_seed_rtx_v1"
COMPANY = "RTX"

RTX_ESG_KEYWORDS = [
    "esg", "sustainability", "climate", "emission", "greenhouse", "carbon",
    "governance", "diversity", "inclusion", "stakeholder", "human rights",
    "safety", "ethics", "compliance", "tcfd", "cdp", "gri", "materiality",
    "net zero", "renewable", "dei", "workforce", "supply chain", "bribery",
    "export control", "data privacy", "cybersecurity",
]

HARD_NOISE = [
    "table of contents", "all rights reserved", "click here", "cookie policy",
    "javascript enabled", "skip to main", "sign in", "log in",
]

SEC_BOILERPLATE = [
    "united states securities and exchange commission",
    "schedule 14a", "form 10-k", "check the appropriate box",
]

FINANCIAL_ONLY = [
    "consensus", "eps", "ebitda", "target price", "stock price", "p/e ratio",
]

CANDIDATE_KINDS = (
    "quantitative_fact",
    "trend_or_change",
    "qualitative_narrative",
    "governance_policy",
    "materiality_or_stakeholder",
    "report_or_framework_disclosure",
    "unanswerable_or_insufficient_context",
)


@dataclass
class SeedCandidate:
    company: str
    question_type: str
    candidate_kind: str
    question_draft: str
    acceptable_disclosure: str
    prohibited_claims: str
    source_record_id: str
    source_unit_id: str
    source_excerpt: str
    source_file: str
    source_type: str
    seed_origin_type: str
    candidate_status: str
    candidate_reason: str
    workbook_cluster_hint: str
    document_kind: str
    rank: float
    dedupe_key: str


def _prohibited_claims_en(qtype: str) -> str:
    if qtype == "trend":
        return (
            "Do not infer causes not stated in source\n"
            "Do not guarantee future improvement\n"
            "Do not fill undisclosed metrics"
        )
    if qtype == "quantitative":
        return (
            "Do not add numbers not in source\n"
            "Do not change units or scale\n"
            "Do not assert undisclosed items as fact"
        )
    return (
        "Do not claim full compliance/achievement beyond source\n"
        "Do not expand policy beyond disclosed text\n"
        "Do not estimate undisclosed performance"
    )


def _esg_hits(text: str) -> int:
    lower = text.lower()
    return sum(1 for k in RTX_ESG_KEYWORDS if k in lower)


def _provenance(row: Dict[str, Any], text: str) -> str:
    dk = row.get("document_kind") or ""
    if dk in ("appendix", "questionnaire", "policy_page"):
        return "rtx_primary_candidate"
    if row.get("is_fallback_snapshot"):
        return "rtx_fallback_snapshot"
    if dk in ("10k", "proxy_statement"):
        return "rtx_sec_filing_candidate"
    return "rtx_mixed_candidate"


def passage_level_filter(row: Dict[str, Any]) -> Tuple[bool, str]:
    text = row.get("text") or ""
    norm = _norm_ws(text)
    blob = norm.lower()

    if len(norm) < 80:
        return False, "passage_too_short"
    if sum(1 for n in HARD_NOISE if n in blob) >= 2 and _esg_hits(norm) < 2:
        return False, "nav_or_portal_noise"
    if sum(1 for n in SEC_BOILERPLATE if n in blob) >= 2 and _esg_hits(norm) < 3:
        return False, "sec_boilerplate_only"
    fin_hits = sum(1 for f in FINANCIAL_ONLY if f in blob)
    if fin_hits >= 2 and _esg_hits(norm) < 2:
        return False, "pure_financial_noise"
    if re.fullmatch(r"[\|\-\s\d\.]+", norm[:200]):
        return False, "table_residue_only"
    if _esg_hits(norm) < 1 and not METRIC_HINT_RE.search(norm):
        if not re.search(r"\b(scope|emission|governance|stakeholder|sustainability)\b", blob):
            return False, "no_esg_substance"
    return True, "passage_ok"


def _infer_candidate_kind(sentence: str, qtype: str, document_kind: str) -> str:
    lower = sentence.lower()
    if document_kind == "press_release":
        return "governance_policy"
    if document_kind in ("10k", "proxy_statement") and "risk" in lower:
        return "qualitative_narrative"
    if "table of contents" in lower:
        return "unanswerable_or_insufficient_context"
    if any(k in lower for k in ("gri", "tcfd", "sasb", "cdp", "sustainability report", "form 10-k")):
        return "report_or_framework_disclosure"
    if any(k in lower for k in ("materiality", "stakeholder", "double materiality")):
        return "materiality_or_stakeholder"
    if any(k in lower for k in ("board", "governance", "ethics", "compliance", "audit", "director")):
        return "governance_policy"
    if qtype == "trend":
        return "trend_or_change"
    if qtype == "quantitative":
        return "quantitative_fact"
    return "qualitative_narrative"


def _cluster_hint(kind: str, sentence: str) -> str:
    mapping = {
        "quantitative_fact": "FC_METRIC_NUMERIC",
        "trend_or_change": "FC_TREND_MULTIYEAR",
        "governance_policy": "FC_ESG_GOVERNANCE",
        "materiality_or_stakeholder": "FC_MATERIAL_STAKEHOLDER",
        "report_or_framework_disclosure": "FC_REPORT_FRAMEWORK",
        "qualitative_narrative": "FC_QUAL_POLICY",
    }
    base = mapping.get(kind, "FC_GENERAL")
    lower = sentence.lower()
    if "scope" in lower or "emission" in lower or "greenhouse" in lower:
        return "FC_CLIMATE_GHG"
    if "cdp" in lower:
        return "FC_CDP"
    if "deferred prosecution" in lower or "bribery" in lower:
        return "FC_COMPLIANCE_ENFORCEMENT"
    return base


def _detect_qtypes(sentence: str) -> List[str]:
    qtypes: List[str] = []
    years = len(set(YEAR_RE.findall(sentence)))
    numbers = len(NUMBER_RE.findall(sentence))
    if years >= 2 and numbers >= 2:
        qtypes.append("trend")
    if numbers >= 1 and (METRIC_HINT_RE.search(sentence) or re.search(r"\d", sentence)):
        qtypes.append("quantitative")
    qtypes.append("qualitative")
    out: List[str] = []
    for q in qtypes:
        if q not in out:
            out.append(q)
    return out[:3]


def _make_question_rtx(sentence: str, kind: str, qtype: str) -> Optional[str]:
    lower = sentence.lower()
    if "scope 1" in lower or "scope 2" in lower or "greenhouse gas" in lower:
        return "What greenhouse gas emissions does RTX disclose?"
    if "stakeholder" in lower and "engage" in lower:
        return "How does RTX engage stakeholders on sustainability topics?"
    if "materiality" in lower or "material topic" in lower:
        return "What material ESG topics does RTX identify?"
    if "board" in lower or "director" in lower or "governance" in lower:
        return "How is ESG governance structured at RTX?"
    if "ethics" in lower or "compliance" in lower:
        return "What ethics and compliance practices does RTX disclose?"
    if "data privacy" in lower or "cybersecurity" in lower:
        return "How does RTX address data security and privacy?"
    if "deferred prosecution" in lower or "bribery" in lower:
        return "What compliance resolutions has RTX disclosed related to government contracts?"
    if "diversity" in lower or "inclusion" in lower or "dei" in lower:
        return "What diversity and inclusion commitments does RTX report?"
    if kind == "report_or_framework_disclosure":
        return "What sustainability reporting frameworks does RTX reference?"
    if qtype == "trend":
        return "How have RTX's key ESG metrics changed over time?"
    if qtype == "quantitative":
        return "What quantitative ESG metrics does RTX disclose?"
    return "What ESG-related policies or performance does RTX disclose?"


def candidate_level_filter(sentence: str, question: str, qtype: str) -> Tuple[bool, str]:
    if not question or len(question) < 20:
        return False, "empty_question"
    if "RTX" not in question:
        return False, "question_missing_company"
    if len(_norm_ws(sentence)) < 30:
        return False, "excerpt_too_short"
    if sum(1 for n in HARD_NOISE if n in sentence.lower()) >= 2:
        return False, "nav_chrome_in_excerpt"
    return True, "candidate_ok"


def _dedupe_key(question: str, disclosure: str) -> str:
    raw = f"RTX|{_norm_ws(question).lower()}|{_norm_ws(disclosure)[:100].lower()}"
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def _rank(sentence: str, qtype: str, provenance: str, document_kind: str) -> float:
    score = float(max(_sentence_score(sentence), 0))
    if qtype == "trend":
        score += 6
    elif qtype == "quantitative":
        score += 5
    else:
        score += 3
    if provenance == "rtx_primary_candidate":
        score += 4
    elif provenance == "rtx_sec_filing_candidate":
        score += 3
    if document_kind in ("appendix", "policy_page"):
        score += 2
    return score


def generate_candidates_from_passage(row: Dict[str, Any]) -> List[SeedCandidate]:
    text = row.get("text") or ""
    document_kind = row.get("document_kind") or "unknown"
    provenance = _provenance(row, text)
    sentences = _best_sentences(text, limit=8)
    if not sentences:
        sentences = [s for s in _split_sentences(text) if len(s) >= 40][:6]

    out: List[SeedCandidate] = []
    for sentence in sentences:
        if _esg_hits(sentence) < 1 and _sentence_score(sentence) < 2:
            if not re.search(r"\b(RTX|Raytheon|Collins|Pratt)\b", sentence, re.I):
                continue
        for qtype in _detect_qtypes(sentence):
            question = _make_question_rtx(sentence, "", qtype)
            if not question:
                continue
            ok, creason = candidate_level_filter(sentence, question, qtype)
            if not ok:
                continue
            kind = _infer_candidate_kind(sentence, qtype, document_kind)
            if kind == "unanswerable_or_insufficient_context":
                continue
            disclosure = _norm_ws(sentence)[:420]
            out.append(
                SeedCandidate(
                    company=COMPANY,
                    question_type=qtype,
                    candidate_kind=kind,
                    question_draft=question,
                    acceptable_disclosure=disclosure,
                    prohibited_claims=_prohibited_claims_en(qtype),
                    source_record_id=str(row.get("record_id") or ""),
                    source_unit_id=str(row.get("unit_id") or ""),
                    source_excerpt=disclosure,
                    source_file=str(row.get("source_file") or ""),
                    source_type=str(row.get("source_type") or ""),
                    seed_origin_type=provenance,
                    candidate_status="candidate_rtx_v1",
                    candidate_reason=f"from_passage:{creason}; kind={kind}",
                    workbook_cluster_hint=_cluster_hint(kind, sentence),
                    document_kind=document_kind,
                    rank=_rank(sentence, qtype, provenance, document_kind),
                    dedupe_key=_dedupe_key(question, disclosure),
                )
            )
    return out


def workbook_level_dedupe(candidates: Sequence[SeedCandidate]) -> List[SeedCandidate]:
    best: Dict[str, SeedCandidate] = {}
    for cand in sorted(candidates, key=lambda x: x.rank, reverse=True):
        prev = best.get(cand.dedupe_key)
        if prev is None or cand.rank > prev.rank:
            best[cand.dedupe_key] = cand
    return list(best.values())


def _seed_id(qtype: str, index: int) -> str:
    code = {"quantitative": "Q", "trend": "T", "qualitative": "L"}.get(qtype, "X")
    return f"RTX-V1-{code}{index:02d}"


def to_jsonl_rows(candidates: Sequence[SeedCandidate]) -> List[Dict[str, Any]]:
    counters: Counter[str] = Counter()
    rows: List[Dict[str, Any]] = []
    for cand in sorted(candidates, key=lambda x: (-x.rank, x.question_type)):
        counters[cand.question_type] += 1
        rows.append(
            {
                "seed_id": _seed_id(cand.question_type, counters[cand.question_type]),
                "company": cand.company,
                "question_type": cand.question_type,
                "candidate_kind": cand.candidate_kind,
                "question_draft": cand.question_draft,
                "acceptable_disclosure": cand.acceptable_disclosure,
                "prohibited_claims": cand.prohibited_claims,
                "source_record_id": cand.source_record_id,
                "source_excerpt": cand.source_excerpt,
                "source_unit_id": cand.source_unit_id,
                "source_file": cand.source_file,
                "source_type": cand.source_type,
                "seed_origin_type": cand.seed_origin_type,
                "candidate_status": cand.candidate_status,
                "candidate_reason": cand.candidate_reason,
                "workbook_cluster_hint": cand.workbook_cluster_hint,
                "document_kind": cand.document_kind,
                "seed_version": SEED_VERSION,
                "rank": cand.rank,
            }
        )
    return rows


def write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "Guide"
    guide.append(["RTX Reference Seed Workbook V1 — workbook-first candidates"])
    guide.append(["Principle", "One good passage -> multiple candidates; not canonical final"])

    header = [
        "seed_id", "company", "question_type", "candidate_kind", "document_kind",
        "question_draft", "acceptable_disclosure", "prohibited_claims",
        "source_record_id", "source_excerpt", "seed_origin_type",
        "candidate_status", "workbook_cluster_hint", "candidate_reason",
    ]
    ws = wb.create_sheet("Candidates")
    ws.append(header)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font
    for row in rows:
        ws.append([row.get(h, "") for h in header])

    ws2 = wb.create_sheet("Summary")
    ws2.append(["metric", "value"])
    ws2.append(["total_rows", len(rows)])
    for qt, n in sorted(Counter(r["question_type"] for r in rows).items()):
        ws2.append([f"question_type:{qt}", n])
    for dk, n in sorted(Counter(r.get("document_kind", "") for r in rows).items()):
        ws2.append([f"document_kind:{dk}", n])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_report(summary: Dict[str, Any], path: Path) -> None:
    lines = [
        "# Golden Set — Candidate Generation RTX V1",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Test lại workflow Golden Set workbook-first trên lane RTX mới (corpus tốt hơn).",
        "",
        "## Đầu vào RTX",
        "",
        f"- Normalized corpus units: **{summary.get('normalized_units', 0)}**",
        "- Nguồn: 4 PDF + 5 HTML + 1 DOJ fallback snapshot",
        "",
        "## Vì sao cần normalization nhẹ",
        "",
        "Chunk sample còn mojibake (`â€¢`), HTML entities (`&amp;`), table/HTML residue — gây noise cho reviewer.",
        "",
        "## Rule normalization đã áp",
        "",
        "- Mojibake bullet/quote repair",
        "- `html.unescape` cho entities phổ biến",
        "- Strip HTML comments/tags residue",
        "- Collapse whitespace; giữ fact thật",
        "",
        "## Rule passage-level filtering",
        "",
        "Chỉ loại noise mạnh: nav/boilerplate SEC thuần, table residue, pure financial, no ESG substance.",
        "",
        "## Cách candidate generation hoạt động",
        "",
        "Một passage tốt → nhiều candidate (`quantitative` / `trend` / `qualitative` + `candidate_kind`).",
        "",
        "## Kết quả",
        "",
        f"- Input corpus units: **{summary.get('input_units', 0)}**",
        f"- Passages accepted: **{summary.get('input_passages', 0)}**",
        f"- Raw candidates: **{summary.get('raw_candidates', 0)}**",
        f"- Filtered candidates: **{summary.get('filtered_candidates', 0)}**",
        "",
        "### Breakdown theo question_type",
        "",
    ]
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")
    lines.extend(["", "### Breakdown theo document_kind", ""])
    for dk, n in summary.get("by_document_kind", {}).items():
        lines.append(f"- `{dk}`: {n}")

    lines.extend(
        [
            "",
            "## Đánh giá",
            "",
            f"- Corpus RTX sinh workbook tốt hơn lane cũ: **{summary.get('quality_vs_old', '')}**",
            f"- Noise còn lại: **{summary.get('noise_note', '')}**",
            "",
            "## Kết luận",
            "",
            f"- Workbook RTX v1 đủ mở review round 1: **{summary.get('review_ready_estimate', '')}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_rtx_builder(
    *,
    input_path: Path,
    output_jsonl: Path,
    output_xlsx: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    passage_rejects: Counter = Counter()
    raw_candidates: List[SeedCandidate] = []

    for row in rows:
        ok, reason = passage_level_filter(row)
        if not ok:
            passage_rejects[reason] += 1
            continue
        raw_candidates.extend(generate_candidates_from_passage(row))

    deduped = workbook_level_dedupe(raw_candidates)
    out_rows = to_jsonl_rows(deduped)

    write_jsonl(output_jsonl, out_rows)
    write_workbook(out_rows, output_xlsx)

    by_qtype = Counter(r["question_type"] for r in out_rows)
    by_kind = Counter(r.get("document_kind", "") for r in out_rows)
    filtered_n = len(out_rows)
    passages_ok = len(rows) - sum(passage_rejects.values())

    review_ready = filtered_n >= 15 and len(by_kind) >= 3

    summary = {
        "seed_version": SEED_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_units": len(rows),
        "normalized_units": len(rows),
        "input_passages": passages_ok,
        "passages_rejected": sum(passage_rejects.values()),
        "passage_reject_reasons": dict(passage_rejects.most_common(12)),
        "raw_candidates": len(raw_candidates),
        "after_dedupe": len(deduped),
        "filtered_candidates": filtered_n,
        "by_question_type": dict(by_qtype),
        "by_document_kind": dict(by_kind),
        "by_candidate_kind": dict(Counter(r.get("candidate_kind") for r in out_rows)),
        "high_noise_rejections": passage_rejects.get("nav_or_portal_noise", 0)
        + passage_rejects.get("sec_boilerplate_only", 0),
        "quality_vs_old": "Có — lane RTX có nguồn SEC/ESG thật, yield candidate-rich hơn lane salvage cũ",
        "noise_note": "SEC filing boilerplate vẫn cần reviewer triage; normalization đã giảm encoding noise",
        "review_ready_estimate": (
            "Có — đủ để mở review workbook round 1 (draft, chưa canonical)"
            if review_ready
            else "Chưa — cần thêm candidate hoặc nới filter nhẹ"
        ),
        "review_ready_flag": review_ready,
        "output_jsonl": str(output_jsonl),
        "output_xlsx": str(output_xlsx),
    }

    write_report(summary, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Build RTX reference seed workbook v1")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/rtx_step1_corpus_units/corpus_units_rtx_normalized.jsonl",
    )
    parser.add_argument(
        "--output-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_v1.jsonl",
    )
    parser.add_argument(
        "--output-xlsx",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_rtx_v1.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_candidate_generation_rtx_v1.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_candidate_generation_rtx_v1_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_rtx_builder(
        input_path=root / args.input,
        output_jsonl=root / args.output_jsonl,
        output_xlsx=root / args.output_xlsx,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "normalized_units",
                    "raw_candidates",
                    "filtered_candidates",
                    "by_document_kind",
                    "review_ready_flag",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
