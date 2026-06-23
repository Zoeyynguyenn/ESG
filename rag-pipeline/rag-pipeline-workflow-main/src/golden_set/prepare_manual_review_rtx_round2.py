"""RTX Manual Review Round 2 prep — rewrite polish + lane split."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import METRIC_HINT_RE, NUMBER_RE, YEAR_RE
from golden_set.io_utils import read_jsonl, write_jsonl

PREP_VERSION = "ref_manual_rtx_round2_prep"

TABLE_HEAVY_KINDS = {"appendix", "data_table", "questionnaire"}
GOVERNANCE_KINDS = {"10k", "proxy_statement"}

SEC_META = [
    "select from:",
    "table of contents",
    "form 10-k",
    "schedule 14a",
    "check the appropriate box",
    "numeric input",
    "[fixed row]",
    "<!-- image -->",
]

PIPE_TOKEN_RE = re.compile(r"\b[a-z]{3,}\|[a-z]{3,}")

FAMILY_LABELS = {
    "scope_ghg": "greenhouse gas emissions",
    "energy_intensity": "energy intensity",
    "water": "water use",
    "waste": "waste generation",
    "diversity": "workforce diversity",
    "governance_board": "board governance",
    "ethics_compliance": "ethics and compliance",
    "cyber_data": "data security and privacy",
    "climate_risk": "climate risk",
    "stakeholder": "stakeholder engagement",
    "safety": "workplace safety",
    "supply_chain": "supply chain sustainability",
    "general": "ESG performance",
}


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _base_passage(row: Dict[str, Any]) -> str:
    return _norm_ws(
        row.get("rewritten_disclosure_draft")
        or row.get("acceptable_disclosure")
        or row.get("source_excerpt")
        or ""
    )


def _base_question(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("rewritten_question_draft") or row.get("question_draft") or "")


def _table_density(text: str) -> int:
    return text.count("|") + text.count("---")


def _infer_metric_label(passage: str) -> str:
    lower = passage.lower()
    rules = [
        ("energy consumption reduction since 2019", r"energy consumption since 2019|reduction in energy consumption"),
        ("energy intensity (GJ per revenue)", r"energy intensity|gj/\$m"),
        ("ergonomic risk reduction", r"ergonomic risk"),
        ("Scope 1 and Scope 2 GHG emissions", r"scope\s*[12]|greenhouse gas|ghg|co2e"),
        ("market-based Scope 2 emissions", r"market-based.*scope 2"),
        ("water withdrawal or consumption", r"water withdrawal|water consumption"),
        ("workforce diversity", r"diversity|inclusion|underrepresented|female"),
        ("board or committee governance", r"board of directors|audit committee|esg committee"),
        ("ethics and compliance program", r"ethics|compliance|code of conduct"),
        ("data security and privacy", r"cybersecurity|data privacy|data security"),
        ("climate risk management", r"climate risk|tcfd|physical risk|transition risk"),
        ("stakeholder engagement", r"stakeholder|materiality|engagement"),
        ("safety performance", r"safety|injury|osha|lost.?time"),
        ("supplier sustainability", r"supplier|supply chain|ecovadis"),
        ("renewable electricity use", r"renewable electricity|renewable energy"),
        ("R&D investment for sustainability", r"r&d|research and development"),
    ]
    for label, pat in rules:
        if re.search(pat, lower):
            return label
    return "ESG metric or policy"


def _extract_disclosure_snippet(passage: str) -> str:
    if _table_density(passage) < 4:
        return passage[:420].strip()

    best = ""
    for part in re.split(r"\|", passage):
        part = _norm_ws(part)
        if len(part) < 25:
            continue
        score = len(NUMBER_RE.findall(part)) * 2 + (2 if METRIC_HINT_RE.search(part) else 0)
        if score > len(NUMBER_RE.findall(best)) * 2:
            best = part
    if best:
        return best[:420]

    for sent in re.split(r"(?<=[.!?])\s+", passage):
        sent = _norm_ws(sent)
        if len(sent) >= 40 and NUMBER_RE.search(sent):
            return sent[:420]
    return passage[:420].strip()


def _is_mechanical_question(q: str) -> bool:
    if "|" in q:
        return True
    if PIPE_TOKEN_RE.search(q):
        return True
    if re.search(r"How has RTX's [a-z]+\|[a-z]+", q):
        return True
    generic = (
        "What ESG-related policies or performance does RTX disclose?",
        "What quantitative ESG metrics does RTX disclose?",
        "How have RTX's key ESG metrics changed over time?",
        "How is ESG governance structured at RTX?",
    )
    return q in generic


def polish_rewrite(row: Dict[str, Any]) -> Tuple[str, str, bool]:
    """Return polished_question, polished_disclosure, was_mechanical_before."""
    passage = _base_passage(row)
    q_in = _base_question(row)
    qtype = row.get("question_type") or "qualitative"
    mechanical_before = _is_mechanical_question(q_in)
    label = _infer_metric_label(passage)
    disclosure = _extract_disclosure_snippet(passage)

    if row.get("review_decision") == "keep" and not mechanical_before:
        return q_in, disclosure, mechanical_before

    if "scope 1" in passage.lower() or "scope 2" in passage.lower():
        question = "What Scope 1 and Scope 2 GHG emissions does RTX report?"
    elif "energy intensity" in passage.lower():
        question = "What is RTX's disclosed energy intensity (GJ per revenue)?"
    elif "energy consumption" in passage.lower() and "2019" in passage:
        question = "What reduction in energy consumption since 2019 does RTX report?"
    elif "ergonomic" in passage.lower():
        question = "How much has RTX reduced high and elevated ergonomic risks since 2015?"
    elif "market-based" in passage.lower() and "scope 2" in passage.lower():
        question = "What are RTX's market-based Scope 2 emissions?"
    elif re.search(r"deferred prosecution|fcpa|bribery", passage, re.I):
        question = "What compliance resolution related to government contracts has RTX disclosed?"
    elif qtype == "trend" and (len(set(YEAR_RE.findall(passage))) >= 2 or "%" in passage):
        question = f"How has RTX's {label} changed over reported years?"
    elif qtype == "quantitative" and NUMBER_RE.search(passage):
        question = f"What {label} does RTX disclose?"
    elif mechanical_before or _is_mechanical_question(q_in):
        question = f"What does RTX disclose about {label}?"
    else:
        question = q_in

    question = re.sub(r"\s+", " ", question).strip()
    question = question.replace("|", " ").strip()
    return question, disclosure, mechanical_before


def _esg_fact_strength(passage: str) -> int:
    lower = passage.lower()
    score = 0
    if NUMBER_RE.search(passage):
        score += 3
    if METRIC_HINT_RE.search(passage):
        score += 2
    if len(set(YEAR_RE.findall(passage))) >= 2:
        score += 2
    if re.search(r"\b(scope|emission|governance|stakeholder|sustainability|climate)\b", lower):
        score += 2
    if re.search(r"\b(RTX|Raytheon|Collins|Pratt)\b", passage, re.I):
        score += 1
    return score


def _disclosure_noise(passage: str, document_kind: str) -> Tuple[int, List[str]]:
    blob = passage.lower()
    flags: List[str] = []
    score = 0
    if _table_density(passage) >= 8:
        score += 3
        flags.append("table_heavy")
    elif _table_density(passage) >= 4:
        score += 1
        flags.append("table_moderate")
    for m in SEC_META:
        if m in blob:
            score += 2
            flags.append("sec_form_meta")
    if len(passage) < 50:
        score += 3
        flags.append("very_short")
    if len(passage) > 500:
        score += 1
        flags.append("passage_broad")
    if passage.count("|") >= 6 and len(re.findall(r"[a-zA-Z]{4,}", passage)) < 8:
        score += 3
        flags.append("pipe_fragment")
    if document_kind in TABLE_HEAVY_KINDS and _table_density(passage) >= 5:
        score += 1
    return score, flags


def _manual_reject_recommended(row: Dict[str, Any], noise: int, fact: int, flags: List[str]) -> bool:
    if fact <= 2 and noise >= 5:
        return True
    if "pipe_fragment" in flags and fact < 4:
        return True
    if "very_short" in flags and fact < 3:
        return True
    if noise >= 8 and fact < 5:
        return True
    if row.get("review_reason") == "generic_question_weak_grounding" and fact < 3:
        return True
    return False


def assign_lane(row: Dict[str, Any], polished_q: str, polished_d: str, was_mechanical: bool) -> Dict[str, Any]:
    out = dict(row)
    decision = row.get("review_decision") or ""
    raw_passage = _base_passage(row)
    passage = polished_d or raw_passage
    dk = row.get("document_kind") or ""
    noise, flags = _disclosure_noise(raw_passage, dk)
    polished_noise, _ = _disclosure_noise(passage, dk)
    fact = _esg_fact_strength(passage)
    reject_rec = _manual_reject_recommended(row, noise, fact, flags)

    lane = "lane_c_rewrite_heavy"
    priority = "low"
    review_reason = ""
    rewrite_scope = "full_passage_cleanup"

    clean_disclosure = (
        noise <= 2
        and polished_noise <= 2
        and _table_density(raw_passage) < 4
        and len(passage) >= 55
    )
    polished_ok = not _is_mechanical_question(polished_q)
    table_salvage = noise >= 4 or _table_density(raw_passage) >= 6 or dk in TABLE_HEAVY_KINDS and noise >= 2

    if reject_rec:
        lane = "reject_recommended"
        priority = "skip"
        review_reason = f"low_fact_high_noise; noise={noise}; fact={fact}; " + ",".join(flags[:4])
        rewrite_scope = "not_recommended"
    elif decision == "keep":
        if clean_disclosure and polished_ok:
            lane = "lane_a_ready_keep"
            priority = "high"
            review_reason = "round1_keep_clean_grounded"
            rewrite_scope = "confirm_or_drop_only"
        elif noise <= 3:
            lane = "lane_b_rewrite_light"
            priority = "medium"
            review_reason = "keep_with_minor_cleanup"
            rewrite_scope = "wording_only"
        else:
            lane = "lane_c_rewrite_heavy"
            priority = "low"
            review_reason = "keep_but_disclosure_dirty; " + ",".join(flags[:3])
            rewrite_scope = "disclosure_extraction"
    elif decision == "rewrite":
        if was_mechanical and polished_ok and clean_disclosure:
            lane = "lane_b_rewrite_light"
            priority = "medium"
            review_reason = "mechanical_draft_polished_to_usable"
            rewrite_scope = "confirm_polished_draft"
        elif table_salvage and polished_ok:
            lane = "lane_c_rewrite_heavy"
            priority = "low"
            review_reason = "table_salvage_after_polish; " + ",".join(flags[:4])
            rewrite_scope = "extract_fact_from_table"
        elif was_mechanical and polished_ok and noise <= 4:
            lane = "lane_b_rewrite_light"
            priority = "medium"
            review_reason = "polished_question_ok_disclosure_trim"
            rewrite_scope = "disclosure_trim"
        elif clean_disclosure and polished_ok and fact >= 5:
            lane = "lane_b_rewrite_light"
            priority = "medium"
            review_reason = "salvageable_light_rewrite"
            rewrite_scope = "wording_and_specificity"
        elif noise >= 5 or "table_heavy" in flags or "pipe_fragment" in flags:
            lane = "lane_c_rewrite_heavy"
            priority = "low"
            review_reason = "table_or_semi_structured_salvage; " + ",".join(flags[:4])
            rewrite_scope = "extract_fact_from_table"
        elif dk in GOVERNANCE_KINDS and noise >= 3:
            lane = "lane_c_rewrite_heavy"
            priority = "low"
            review_reason = "10k_proxy_governance_extraction; " + ",".join(flags[:3])
            rewrite_scope = "governance_fact_extraction"
        else:
            lane = "lane_b_rewrite_light" if noise <= 3 else "lane_c_rewrite_heavy"
            priority = "medium" if lane == "lane_b_rewrite_light" else "low"
            review_reason = f"rewrite_mixed; noise={noise}; fact={fact}"
            rewrite_scope = "wording_and_disclosure_trim" if lane == "lane_b_rewrite_light" else "disclosure_extraction"

    out.update(
        {
            "manual_prep_version": PREP_VERSION,
            "polished_question_draft": polished_q,
            "polished_disclosure_draft": polished_d,
            "polish_was_mechanical": was_mechanical,
            "manual_review_lane": lane,
            "manual_priority": priority,
            "manual_review_reason": review_reason,
            "manual_reject_recommended": reject_rec,
            "manual_rewrite_scope": rewrite_scope,
            "disclosure_noise_score": noise,
            "fact_strength_score": fact,
            "disclosure_noise_flags": flags,
        }
    )
    return out


WORKBOOK_COLS = [
    "seed_id",
    "company",
    "question_type",
    "candidate_kind",
    "document_kind",
    "question_draft",
    "acceptable_disclosure",
    "review_decision",
    "review_reason",
    "rewritten_question_draft",
    "rewritten_disclosure_draft",
    "polished_question_draft",
    "polished_disclosure_draft",
    "manual_review_lane",
    "manual_priority",
    "manual_review_reason",
    "manual_rewrite_scope",
    "manual_reject_recommended",
    "cluster_id",
]


def _style_header(ws, ncol: int, color: str = "1E3A5F") -> None:
    fill = PatternFill("solid", fgColor=color)
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, ncol + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = font


def write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "Guide"
    guide.append(["RTX Manual Review Round 2 — Lane Split + Polish"])
    guide.append(["Order", "Lane_A → Lane_B → Lane_C; skip Reject_Recommended unless audit"])

    lanes = [
        ("Lane_A_ReadyKeep", "lane_a_ready_keep", "166534"),
        ("Lane_B_RewriteLight", "lane_b_rewrite_light", "1D4ED8"),
        ("Lane_C_RewriteHeavy", "lane_c_rewrite_heavy", "B45309"),
        ("Reject_Recommended", "reject_recommended", "991B1B"),
    ]
    for sheet_name, lane_key, color in lanes:
        ws = wb.create_sheet(sheet_name)
        ws.append(WORKBOOK_COLS)
        _style_header(ws, len(WORKBOOK_COLS), color)
        subset = [r for r in rows if r.get("manual_review_lane") == lane_key]
        for r in sorted(subset, key=lambda x: (-float(x.get("rank") or 0), x.get("seed_id", ""))):
            ws.append([r.get(c, "") for c in WORKBOOK_COLS])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    out: Dict[str, List[Dict[str, str]]] = {
        "lane_a": [],
        "lane_b_polish": [],
        "lane_c": [],
        "reject": [],
    }
    for lane_key, bucket in (
        ("lane_a_ready_keep", "lane_a"),
        ("lane_b_rewrite_light", "lane_b_polish"),
        ("lane_c_rewrite_heavy", "lane_c"),
        ("reject_recommended", "reject"),
    ):
        for r in rows:
            if r.get("manual_review_lane") != lane_key:
                continue
            if len(out[bucket]) >= 3:
                break
            ex: Dict[str, str] = {
                "seed_id": r.get("seed_id", ""),
                "polished_q": (r.get("polished_question_draft") or "")[:90],
                "reason": r.get("manual_review_reason", ""),
            }
            if bucket == "lane_b_polish" and r.get("polish_was_mechanical"):
                ex["before_q"] = (r.get("rewritten_question_draft") or "")[:90]
            out[bucket].append(ex)
    return out


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Manual Review RTX Round 2 Prep",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Chuẩn bị workbook RTX cho manual review round 2: lane split + rewrite polish,",
        "không review phẳng 221 row.",
        "",
        "## Vì sao cần lane split cho RTX",
        "",
        "- Lượng row từ 10-K/proxy lớn; nhiều row định lượng/xu hướng từ bảng",
        "- Round 1 `rewrite` trộn fact sạch và passage semi-structured",
        "- Reviewer cần ưu tiên confirm trước, salvage sau",
        "",
        "## Vì sao cần rewrite polish trước manual review",
        "",
        f"- **{summary.get('mechanical_before_count', 0)}** row có `rewritten_question_draft` dạng pipe-token",
        f"- Sau polish còn **{summary.get('mechanical_after_count', 0)}** row mechanical",
        "- Giảm việc reviewer sửa câu hỏi kiểu `foo|bar|baz`",
        "",
        "## Rule chia lane",
        "",
        "| Lane | Điều kiện |",
        "|------|-----------|",
        "| `lane_a_ready_keep` | Keep + disclosure sạch, confirm/drop nhanh |",
        "| `lane_b_rewrite_light` | Fact thật; polish đủ tốt, chỉnh nhẹ |",
        "| `lane_c_rewrite_heavy` | Table-heavy / 10-K governance extraction |",
        "| `reject_recommended` | Fact yếu + noise cao; skip sớm |",
        "",
        "## Kết quả",
        "",
        f"- Lane A: **{summary.get('lane_a_count', 0)}**",
        f"- Lane B: **{summary.get('lane_b_count', 0)}**",
        f"- Lane C: **{summary.get('lane_c_count', 0)}**",
        f"- Reject recommended: **{summary.get('reject_recommended_count', 0)}**",
        "",
        "### Breakdown theo question_type (lanes A+B+C)",
        "",
    ]
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "### Breakdown theo document_kind (lanes A+B+C)", ""])
    for dk, n in summary.get("by_document_kind", {}).items():
        lines.append(f"- `{dk}`: {n}")

    lines.extend(["", "## Ví dụ", ""])
    lines.append("### Lane A — keep-ready")
    for ex in examples.get("lane_a", []):
        lines.append(f"- `{ex['seed_id']}`: {ex['polished_q']} — {ex['reason']}")

    lines.append("")
    lines.append("### Lane B — polish cải thiện draft máy móc")
    for ex in examples.get("lane_b_polish", []):
        before = ex.get("before_q", "")
        if before:
            lines.append(f"- `{ex['seed_id']}`: `{before}` → `{ex['polished_q']}`")
        else:
            lines.append(f"- `{ex['seed_id']}`: {ex['polished_q']}")

    lines.append("")
    lines.append("### Lane C — rewrite heavy")
    for ex in examples.get("lane_c", []):
        lines.append(f"- `{ex['seed_id']}`: {ex['polished_q']} — {ex['reason']}")

    lines.append("")
    lines.append("### Reject recommended")
    for ex in examples.get("reject", []):
        lines.append(f"- `{ex['seed_id']}`: {ex['reason']}")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- Reviewer bắt đầu từ: **{summary.get('reviewer_order_recommendation', '')}**",
            f"- Ước lượng row sống tới canonical round: **~{summary.get('likely_survivors_estimate', 0)}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_prep(
    *,
    input_path: Path,
    output_jsonl: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    active = [r for r in rows if r.get("review_decision") in ("keep", "rewrite")]

    mechanical_before = 0
    mechanical_after = 0
    prepared: List[Dict[str, Any]] = []

    for row in active:
        pq, pd, was_mech = polish_rewrite(row)
        if was_mech:
            mechanical_before += 1
        if _is_mechanical_question(pq):
            mechanical_after += 1
        prepared.append(assign_lane(row, pq, pd, was_mech))

    write_jsonl(output_jsonl, prepared)
    write_workbook(prepared, workbook_path)

    lane_counts = Counter(r.get("manual_review_lane") for r in prepared)
    review_lanes = [r for r in prepared if r.get("manual_review_lane") != "reject_recommended"]

    a = lane_counts.get("lane_a_ready_keep", 0)
    b = lane_counts.get("lane_b_rewrite_light", 0)
    c = lane_counts.get("lane_c_rewrite_heavy", 0)
    rej = lane_counts.get("reject_recommended", 0)

    survivors = int(round(a * 0.9 + b * 0.75 + c * 0.45))

    summary = {
        "prep_version": PREP_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_reviewable": len(active),
        "lane_a_count": a,
        "lane_b_count": b,
        "lane_c_count": c,
        "reject_recommended_count": rej,
        "mechanical_before_count": mechanical_before,
        "mechanical_after_count": mechanical_after,
        "mechanical_reduced_by": mechanical_before - mechanical_after,
        "by_question_type": dict(Counter(r.get("question_type") for r in review_lanes)),
        "by_document_kind": dict(Counter(r.get("document_kind") for r in review_lanes)),
        "likely_survivors_estimate": survivors,
        "reviewer_order_recommendation": (
            "Lane_A_ReadyKeep → Lane_B_RewriteLight → Lane_C_RewriteHeavy; "
            "bỏ qua Reject_Recommended trừ audit"
        ),
        "output_jsonl": str(output_jsonl),
        "output_workbook": str(workbook_path),
    }

    examples = _pick_examples(prepared)
    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="RTX manual review round 2 prep")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_reviewed_round1.jsonl",
    )
    parser.add_argument(
        "--output-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_rtx_manual_round2.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_rtx_manual_round2.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_manual_review_rtx_round2_prep.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_manual_review_rtx_round2_prep_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_prep(
        input_path=root / args.input,
        output_jsonl=root / args.output_jsonl,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "lane_a_count",
                    "lane_b_count",
                    "lane_c_count",
                    "reject_recommended_count",
                    "mechanical_before_count",
                    "mechanical_after_count",
                    "likely_survivors_estimate",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
