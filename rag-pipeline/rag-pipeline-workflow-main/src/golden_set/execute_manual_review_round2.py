"""Manual Review Round 2 execution — Lane A + Lane B only (confirm/revise/drop)."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import METRIC_HINT_RE
from golden_set.io_utils import read_jsonl, write_jsonl
from golden_set.prepare_manual_review_round2 import (
    GENERIC_Q,
    _company_in_text,
    _norm_ws,
    _passage,
)

PORTAL_NOISE = [
    "개인정보처리방침",
    "홈페이지 저작권",
    "cctv 운영관리",
    "청렴마일리지",
    "자료실 공지사항",
]

OTHER_COMPANY_IN_RX = ["기아는", "현대트랜시스", "삼성전자", "삼성전기"]

EXEC_VERSION = "ref_manual_round2_exec"

LANE_AB = frozenset({"lane_a_ready_keep", "lane_b_rewrite_light"})
LANE_BACKLOG = frozenset({"lane_c_rewrite_heavy", "reject_recommended"})

WORKBOOK_COLS = [
    "seed_id",
    "company",
    "question_type",
    "candidate_kind",
    "question_draft",
    "acceptable_disclosure",
    "manual_review_lane",
    "manual_decision",
    "manual_reason",
    "final_question",
    "final_disclosure",
    "final_notes",
    "canonical_candidate_flag",
    "canonical_cluster_action",
    "cluster_id",
]


def _disclosure(row: Dict[str, Any]) -> str:
    return _norm_ws(
        row.get("rewritten_disclosure_draft")
        or row.get("acceptable_disclosure")
        or row.get("source_excerpt")
        or ""
    )


def _question(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("rewritten_question_draft") or row.get("question_draft") or "")


def _is_generic_question(q: str) -> bool:
    return bool(GENERIC_Q.search(q))


def _build_specific_question(company: str, disclosure: str, qtype: str) -> str:
    lower = disclosure.lower()
    if "8개 중대 이슈" in disclosure or ("중대 이슈" in disclosure and re.search(r"\d+\s*개", disclosure)):
        return f"{company}는 이중 중대성 평가를 통해 몇 개의 중대 이슈를 선정했는가?"
    if "2050" in disclosure and ("탄소중립" in disclosure or "넷제로" in disclosure or "net zero" in lower):
        return f"{company}는 2050년까지 어떤 탄소중립 목표를 공개했는가?"
    if "kgcs" in lower and "등급" in disclosure:
        return f"{company}는 KGCS ESG경영 평가에서 어떤 등급을 획득했는가?"
    if "14회" in disclosure and "이사회" in disclosure:
        return f"{company}는 2022년 이사회를 몇 회 개최했는가?"
    if "esg위원회" in lower or ("이사회" in disclosure and "기후" in disclosure):
        return f"{company}의 ESG 거버넌스 체계는 어떻게 운영되는가?"
    if "임팩트" in disclosure or "impact report" in lower:
        return f"{company}는 어떤 임팩트/지속가능경영 보고서를 발간했는가?"
    if "이해관계자" in disclosure or "stakeholder" in lower:
        return f"{company}는 이해관계자 소통·참여를 어떻게 운영하는가?"
    if "인권" in disclosure:
        return f"{company}의 인권경영 정책 또는 실사 현황은 무엇인가?"
    if "온실가스" in disclosure or "scope" in lower:
        return f"{company}의 온실가스 배출 또는 감축 목표는 무엇인가?"
    if qtype == "quantitative" and METRIC_HINT_RE.search(disclosure):
        return f"{company}가 공시한 주요 ESG 수치는 무엇인가?"
    if qtype == "trend":
        return f"{company}의 주요 ESG 지표 또는 목표는 어떻게 변화했는가?"
    return ""


def _trim_disclosure(disclosure: str, company: str, max_len: int = 420) -> str:
    text = _norm_ws(disclosure)
    if len(text) <= max_len:
        return text
    for sent in re.split(r"(?<=[.!?다])\s+", text):
        if len(sent) >= 40 and _company_in_text(company, sent):
            return sent[:max_len]
    return text[:max_len]


def _fact_fingerprint(disclosure: str) -> str:
    nums = re.findall(r"\d+", disclosure)[:6]
    words = re.findall(r"[가-힣]{4,}", disclosure)[:8]
    raw = "|".join(words + nums)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:12]


def _initial_review(row: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    """Return decision, reason, final_q, final_d, notes."""
    lane = row.get("manual_review_lane") or ""
    company = row.get("company") or ""
    qtype = row.get("question_type") or "qualitative"
    disc = _disclosure(row)
    q = _question(row)
    contam = int(row.get("contamination_score") or 0)
    fact = int(row.get("fact_strength_score") or 0)

    if not disc or len(disc) < 40:
        return "drop", "insufficient_disclosure", q, disc, "Excerpt quá ngắn sau lane prep."

    blob = disc.lower()
    if sum(1 for m in PORTAL_NOISE if m.lower() in blob) >= 2:
        return "drop", "portal_nav_noise", q, disc, "Portal/nav noise — không phải ESG fact body."

    if company == "레이시온" and any(m in disc for m in OTHER_COMPANY_IN_RX):
        return "drop", "cross_company_in_disclosure", q, disc, "Disclosure thuộc công ty khác (RX package contamination)."

    if not _company_in_text(company, disc) and fact < 5:
        # RX portal excerpts often omit company name but may still carry ESG menu signal
        if company == "레이시온" and fact < 6:
            return "drop", "company_not_in_disclosure", q, disc, "Disclosure không anchor company."
        if company != "레이시온":
            return "drop", "company_not_in_disclosure", q, disc, "Disclosure không anchor company."

    specific_q = _build_specific_question(company, disc, qtype)
    final_d = _trim_disclosure(disc, company)

    if lane == "lane_a_ready_keep":
        if contam <= 1 and not _is_generic_question(q) and len(final_d) >= 55:
            return (
                "confirm",
                "lane_a_clean_grounded",
                q,
                final_d,
                "Lane A: fact grounded, confirm cho canonical round.",
            )
        final_q = specific_q or q
        return (
            "revise",
            "lane_a_minor_wording",
            final_q,
            final_d,
            "Lane A: chỉnh nhẹ wording trước canonical.",
        )

    # lane_b_rewrite_light
    if _is_generic_question(q):
        if specific_q:
            return (
                "revise",
                "generic_question_to_specific",
                specific_q,
                final_d,
                "Lane B: đổi câu hỏi generic sang fact-specific.",
            )
        if fact >= 5:
            return (
                "revise",
                "generic_question_manual_spec",
                f"{company}가 공시한 ESG 관련 내용은 무엇인가?",
                final_d,
                "Lane B: generic Q — cần reviewer tinh chỉnh thêm nếu cần.",
            )
        return "drop", "generic_unsalvageable", q, final_d, "Generic Q + fact yếu."

    if contam >= 3:
        return "drop", "residual_noise_in_lane_b", q, final_d, "Lane B nhưng disclosure còn noise."

    if specific_q and specific_q != q and _is_generic_question(q) is False:
        # question ok but could align better
        if len(q) < 18:
            return "revise", "question_too_short", specific_q, final_d, "Lane B: mở rộng câu hỏi."
        return "confirm", "lane_b_clean_after_prep", q, final_d, "Lane B: đủ rõ, confirm."

    if row.get("manual_rewrite_scope") == "question_specificity":
        final_q = specific_q or q
        return (
            "revise",
            "question_specificity",
            final_q,
            final_d,
            "Lane B: chỉnh specificity câu hỏi.",
        )

    return "confirm", "lane_b_usable", q, final_d, "Lane B: confirm sau prep."


def _apply_cluster_dedupe(reviewed_ab: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Drop weaker duplicates within same company+cluster+fingerprint among confirm/revise."""
    groups: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
    for row in reviewed_ab:
        if row.get("manual_decision") not in ("confirm", "revise"):
            continue
        key = (
            row.get("company", ""),
            row.get("cluster_id", ""),
            _fact_fingerprint(row.get("final_disclosure") or ""),
        )
        groups[key].append(row)

    drop_ids: set[str] = set()
    anchor_ids: Dict[Tuple[str, str, str], str] = {}

    for key, members in groups.items():
        if len(members) <= 1:
            anchor_ids[key] = members[0]["seed_id"]
            continue
        ranked = sorted(members, key=lambda x: (-float(x.get("rank") or 0), x.get("seed_id", "")))
        anchor_ids[key] = ranked[0]["seed_id"]
        for dup in ranked[1:]:
            drop_ids.add(dup["seed_id"])

    for row in reviewed_ab:
        key = (
            row.get("company", ""),
            row.get("cluster_id", ""),
            _fact_fingerprint(row.get("final_disclosure") or ""),
        )
        if row.get("seed_id") in drop_ids:
            row["manual_decision"] = "drop"
            row["manual_reason"] = "duplicate_cluster_lane_ab"
            row["final_notes"] = f"Trùng cụm với anchor {anchor_ids.get(key, '')}"
            row["canonical_candidate_flag"] = False
            row["canonical_cluster_action"] = "defer_to_anchor"
        elif row.get("manual_decision") in ("confirm", "revise"):
            if anchor_ids.get(key) == row.get("seed_id"):
                row["canonical_cluster_action"] = "anchor" if key[1] else "standalone"
            else:
                row["canonical_cluster_action"] = "standalone"
    return reviewed_ab


def _finalize_canonical_flags(row: Dict[str, Any]) -> None:
    dec = row.get("manual_decision")
    if dec == "confirm":
        row["canonical_candidate_flag"] = True
        if not row.get("canonical_cluster_action"):
            row["canonical_cluster_action"] = "anchor"
    elif dec == "revise":
        row["canonical_candidate_flag"] = True
        if not row.get("canonical_cluster_action"):
            row["canonical_cluster_action"] = "revise_before_canonical"
    else:
        row["canonical_candidate_flag"] = False
        if not row.get("canonical_cluster_action"):
            row["canonical_cluster_action"] = "dropped"


def review_lane_ab(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ab = [r for r in rows if r.get("manual_review_lane") in LANE_AB]
    out: List[Dict[str, Any]] = []
    for row in ab:
        rec = dict(row)
        decision, reason, fq, fd, notes = _initial_review(row)
        rec.update(
            {
                "exec_version": EXEC_VERSION,
                "manual_decision": decision,
                "manual_reason": reason,
                "final_question": fq,
                "final_disclosure": fd,
                "final_notes": notes,
                "canonical_candidate_flag": False,
                "canonical_cluster_action": "",
            }
        )
        _finalize_canonical_flags(rec)
        out.append(rec)

    out = _apply_cluster_dedupe(out)
    for rec in out:
        _finalize_canonical_flags(rec)
    return out


def _passthrough_backlog(row: Dict[str, Any]) -> Dict[str, Any]:
    rec = dict(row)
    lane = rec.get("manual_review_lane") or ""
    rec.update(
        {
            "exec_version": EXEC_VERSION,
            "manual_decision": "backlog_lane_c" if lane == "lane_c_rewrite_heavy" else "backlog_reject",
            "manual_reason": "not_processed_round2_exec",
            "final_question": "",
            "final_disclosure": "",
            "final_notes": "Giữ backlog — xử lý sau nếu còn bandwidth.",
            "canonical_candidate_flag": False,
            "canonical_cluster_action": "backlog",
        }
    )
    return rec


def write_workbook(all_rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    guide.append(["Manual Review Round 2 — Execution (Lane A+B)"])
    guide.append(["Lane C", "Backlog — chưa xử lý trong round này"])

    sheets = [
        ("Confirmed", "confirm", "166534"),
        ("Revised", "revise", "1D4ED8"),
        ("Dropped", "drop", "991B1B"),
        ("LaneC_Backlog", None, "B45309"),
    ]
    for sheet_name, decision_key, color in sheets:
        ws = wb.create_sheet(sheet_name)
        ws.append(WORKBOOK_COLS)
        fill = PatternFill("solid", fgColor=color)
        font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(WORKBOOK_COLS) + 1):
            c = ws.cell(row=1, column=col)
            c.fill = fill
            c.font = font

        if sheet_name == "LaneC_Backlog":
            subset = [
                r
                for r in all_rows
                if r.get("manual_decision") in ("backlog_lane_c", "backlog_reject")
                or r.get("manual_review_lane") in LANE_BACKLOG
            ]
        else:
            subset = [r for r in all_rows if r.get("manual_decision") == decision_key]

        for r in sorted(subset, key=lambda x: (x.get("company", ""), x.get("seed_id", ""))):
            ws.append([r.get(c, "") for c in WORKBOOK_COLS])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    ex: Dict[str, List[Dict[str, str]]] = {"confirm": [], "revise": [], "drop": []}
    for dec in ("confirm", "revise", "drop"):
        for r in rows:
            if r.get("manual_decision") != dec:
                continue
            if r.get("manual_review_lane") not in LANE_AB:
                continue
            if len(ex[dec]) >= 3:
                break
            ex[dec].append(
                {
                    "seed_id": r.get("seed_id", ""),
                    "company": r.get("company", ""),
                    "final_q": (r.get("final_question") or "")[:90],
                    "reason": r.get("manual_reason", ""),
                }
            )
    return ex


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Manual Review Round 2 Execution",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Thực thi manual review trên **Lane A + Lane B** (confirm/revise/drop);",
        "Lane C giữ backlog. Chuẩn bị workbook cho canonical round kế tiếp.",
        "",
        "## Vì sao chỉ xử lý Lane A + B",
        "",
        "- Lane A/B = phần tạo giá trị chính (confirm nhanh + rewrite nhẹ).",
        "- Lane C = passage bẩn — chỉ xử lý nếu còn bandwidth.",
        "- Không review phẳng 107 row.",
        "",
        "## Kết quả tổng quan",
        "",
        f"- Input Lane A+B: **{summary.get('input_lane_ab', 0)}**",
        f"- confirm: **{summary.get('confirm_count', 0)}**",
        f"- revise: **{summary.get('revise_count', 0)}**",
        f"- drop: **{summary.get('drop_count', 0)}**",
        "",
        "### Breakdown theo công ty",
        "",
    ]
    for co, stats in summary.get("by_company", {}).items():
        lines.append(
            f"- **{co}**: confirm={stats.get('confirm', 0)}, "
            f"revise={stats.get('revise', 0)}, drop={stats.get('drop', 0)}, "
            f"canonical_candidates={stats.get('canonical_candidate', 0)}"
        )

    lines.extend(["", "### Breakdown theo question_type", ""])
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "## Ví dụ", ""])
    for title, key in (("Row confirm tốt", "confirm"), ("Row revise tốt", "revise"), ("Row drop (Lane B)", "drop")):
        lines.append(f"### {title}")
        for e in examples.get(key, []):
            lines.append(f"- `{e['seed_id']}` ({e['company']}): {e['final_q']} — `{e['reason']}`")
        lines.append("")

    lines.extend(
        [
            "## Đánh giá",
            "",
            f"- Canonical candidate estimate: **{summary.get('canonical_candidate_estimate', 0)}**",
            f"- 레이시온 survivors (confirm+revise): **{summary.get('raysolution_survivors', 0)}**",
            f"- Lane C backlog: **{summary.get('lane_c_backlog_count', 0)}**",
            "",
            "## Kết luận",
            "",
            f"- Mở canonical round kế tiếp? **{summary.get('ready_for_canonical_verdict', '')}**",
            f"- Lane C tiếp tục? **{summary.get('lane_c_recommendation', '')}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_execution(
    *,
    input_path: Path,
    output_jsonl: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    reviewed_ab = review_lane_ab(rows)
    backlog = [_passthrough_backlog(r) for r in rows if r.get("manual_review_lane") in LANE_BACKLOG]
    all_out = reviewed_ab + backlog

    write_jsonl(output_jsonl, all_out)
    write_workbook(all_out, workbook_path)

    ab_decisions = Counter(r.get("manual_decision") for r in reviewed_ab)
    canonical_candidates = [r for r in reviewed_ab if r.get("canonical_candidate_flag")]
    by_co: Dict[str, Dict[str, int]] = {}
    for co in sorted({r.get("company", "") for r in reviewed_ab}):
        sub = [r for r in reviewed_ab if r.get("company") == co]
        by_co[co] = {
            "confirm": sum(1 for r in sub if r.get("manual_decision") == "confirm"),
            "revise": sum(1 for r in sub if r.get("manual_decision") == "revise"),
            "drop": sum(1 for r in sub if r.get("manual_decision") == "drop"),
            "canonical_candidate": sum(1 for r in sub if r.get("canonical_candidate_flag")),
        }

    rx_survivors = sum(
        1
        for r in reviewed_ab
        if r.get("company") == "레이시온" and r.get("manual_decision") in ("confirm", "revise")
    )
    lane_c_count = sum(1 for r in rows if r.get("manual_review_lane") == "lane_c_rewrite_heavy")
    canon_est = len(canonical_candidates)

    hs_canon = by_co.get("한샘", {}).get("canonical_candidate", 0)
    # Canonical round có thể mở trên core Hansem/MS; RX mỏng là expected
    ready = canon_est >= 45 and hs_canon >= 22

    summary = {
        "exec_version": EXEC_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_lane_ab": len(reviewed_ab),
        "confirm_count": ab_decisions.get("confirm", 0),
        "revise_count": ab_decisions.get("revise", 0),
        "drop_count": ab_decisions.get("drop", 0),
        "canonical_candidate_estimate": canon_est,
        "raysolution_survivors": rx_survivors,
        "by_company": by_co,
        "by_question_type": dict(
            Counter(r.get("question_type") for r in canonical_candidates)
        ),
        "lane_c_backlog_count": lane_c_count,
        "ready_for_canonical_round_flag": ready,
        "ready_for_canonical_verdict": (
            "Có — đủ anchor để mở canonical round (Lane A/B); Lane C giữ backlog"
            if ready
            else "Một phần — cần human spot-check 레이시온 hoặc bổ sung fact trước canonical full"
        ),
        "lane_c_recommendation": "Giữ backlog — chỉ mở Lane C nếu cần mở rộng coverage RX/MS sau canonical core",
        "output_jsonl": str(output_jsonl),
        "output_workbook": str(workbook_path),
    }

    examples = _pick_examples(reviewed_ab)
    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Execute manual review round 2 on Lane A+B")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_manual_round2.jsonl",
    )
    parser.add_argument(
        "--output-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_manual_round2_reviewed.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_v4_manual_round2_reviewed.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_manual_review_round2_execution.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_manual_review_round2_execution_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_execution(
        input_path=root / args.input,
        output_jsonl=root / args.output_jsonl,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {
                k: summary[k]
                for k in (
                    "input_lane_ab",
                    "confirm_count",
                    "revise_count",
                    "drop_count",
                    "canonical_candidate_estimate",
                    "raysolution_survivors",
                    "ready_for_canonical_round_flag",
                )
            },
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
