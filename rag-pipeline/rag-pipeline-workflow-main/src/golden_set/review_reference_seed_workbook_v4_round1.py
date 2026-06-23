"""Workbook Review Round 1 — triage v4 JSONL candidates (keep/rewrite/reject/collapse)."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from golden_set.build_reference_seed_workbook import (
    ESG_SENTENCE_KEYWORDS,
    FINANCIAL_NOISE_PATTERNS,
    METRIC_HINT_RE,
    OTHER_COMPANY_MARKERS,
)
from golden_set.io_utils import read_jsonl, write_jsonl

REVIEW_VERSION = "ref_review_v4_r1"

LISTING_MARKERS = [
    "목록 글쓰기",
    "게시판 목록",
    "조회수",
    "노출 글주소",
    "정렬하기",
    "회원명",
    "회원아이디",
    "이전글",
    "다음글",
    "스크랩 신고",
    "mb)",
    "share.google",
    "table of contents",
    "목차",
    "archive",
    "자료실",
    "appendix",
    "ceo message",
    "about this report",
]

INDEX_META_MARKERS = [
    "sasb index",
    "tcfd index",
    "gri index",
    "esrs index",
    "content index",
    "disclosure index",
    "index tcfd",
    "index sasb",
]

NAV_CONTACT_MARKERS = [
    "정보공개",
    "민원",
    "사이트맵",
    "e-mail",
    "tel.",
    "만족도 평가",
    "faq english",
    "esg 소개 esg 기업정보",
    "기업 esg 조회",
    "참고사이트",
    "esg 강의실",
]

NEWS_CHROME_MARKERS = [
    "기자",
    "발행일",
    "기사 공유",
    "주소복사",
    "네이버 채널",
    "다음 채널",
    "무단전재",
    "댓글",
    "구독",
    "송고",
    "모바일버전",
    "전체기사",
    "home 경제",
]

CROSS_COMPANY_NAMES = [
    "현대트랜시스",
    "삼성전자",
    "삼성전기",
    "신세계",
    "cj제일제당",
    "db손해보험",
    "한국타이어",
    "gs칼텍스",
    "hd현대오일뱅크",
    "hj중공업",
    "여수광양항만",
    "rtx corporation",
    "raytheon",
]

GENERIC_QUESTION_PATTERNS = [
    re.compile(r"핵심 내용은 무엇인가"),
    re.compile(r"주요 ESG 수치는 무엇인가"),
    re.compile(r"주요 ESG 지표 추이는"),
    re.compile(r"ESG 전략 또는 정책"),
]

TCFD_DEFINITION_ONLY = re.compile(r"TCFD는\s*2015년.*협의체", re.IGNORECASE)
PAGE_NUMBER_BLOB = re.compile(r"(?:\b\d{2,3}\b\s+){4,}\d{2,3}")


def _norm_ws(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip())


def _blob(row: Dict[str, Any]) -> str:
    return _norm_ws(
        " ".join(
            str(row.get(k, ""))
            for k in ("question_draft", "acceptable_disclosure", "source_excerpt")
        )
    ).lower()


def _passage(row: Dict[str, Any]) -> str:
    return _norm_ws(row.get("source_excerpt") or row.get("acceptable_disclosure") or "")


def _noise_hits(blob: str, markers: Sequence[str]) -> int:
    return sum(1 for m in markers if m.lower() in blob)


def _esg_signal(blob: str, passage: str) -> int:
    score = sum(1 for kw in ESG_SENTENCE_KEYWORDS if kw.lower() in blob)
    if METRIC_HINT_RE.search(passage):
        score += 2
    if any(k in passage for k in ("중대 이슈", "탄소중립", "net zero", "tcfd", "kgcs", "esg위원회", "이사회")):
        score += 2
    return score


def _company_in_passage(company: str, passage: str) -> bool:
    if company in passage:
        return True
    aliases = {
        "한샘": ["㈜한샘", "한샘이"],
        "무신사": ["musinsa"],
        "레이시온": ["raysolution", "레이 시 온"],
    }
    lower = passage.lower()
    return any(a.lower() in lower for a in aliases.get(company, []))


def _is_cross_company(company: str, passage: str, blob: str) -> bool:
    other_hits = [n for n in CROSS_COMPANY_NAMES if n.lower() in blob]
    if not other_hits:
        return any(m in passage for m in OTHER_COMPANY_MARKERS)
    if company == "레이시온" and any(n in passage for n in ("현대트랜시스", "삼성전기", "삼성전자")):
        return True
    if company == "무신사" and len(other_hits) >= 2 and not _company_in_passage(company, passage[:200]):
        return True
    if not _company_in_passage(company, passage) and len(other_hits) >= 1:
        return True
    return False


def _is_listing_index_noise(blob: str, passage: str) -> bool:
    if _noise_hits(blob, LISTING_MARKERS) >= 2:
        return True
    if "조회수" in blob and "글쓰기" in blob:
        return True
    if _noise_hits(blob, INDEX_META_MARKERS) >= 1 and _esg_signal(blob, passage) < 5:
        return True
    if PAGE_NUMBER_BLOB.search(passage) and _esg_signal(blob, passage) < 4:
        return True
    if passage.count("지속가능경영보고서") >= 3 and len(passage) > 400 and _noise_hits(blob, LISTING_MARKERS) >= 1:
        return True
    return False


def _is_nav_contact(blob: str) -> bool:
    return _noise_hits(blob, NAV_CONTACT_MARKERS) >= 3


def _is_framework_only(passage: str, company: str) -> bool:
    if TCFD_DEFINITION_ONLY.search(passage) and not _company_in_passage(company, passage):
        return True
    lower = passage.lower()
    if any(idx in lower for idx in ("sasb index", "tcfd index", "gri index")):
        if not _company_in_passage(company, passage) or len(passage) < 120:
            return True
    return False


def _is_financial_noise(blob: str, passage: str) -> bool:
    fin = sum(1 for p in FINANCIAL_NOISE_PATTERNS if p in blob)
    return fin >= 2 and _esg_signal(blob, passage) < 3


def _is_truncated(passage: str) -> bool:
    if len(passage) < 50:
        return True
    if passage.startswith(("완성하고", "또한", "먼저 오는")) and "한샘" not in passage[:40]:
        return True
    return passage.rstrip().endswith(("BIS)과", "국제결제은행(BIS)과", "인증 제도인"))


def _needs_rewrite(row: Dict[str, Any], blob: str, passage: str) -> Tuple[bool, str]:
    reasons: List[str] = []
    q = row.get("question_draft") or ""
    if any(p.search(q) for p in GENERIC_QUESTION_PATTERNS):
        reasons.append("generic_question")
    if _noise_hits(blob, NEWS_CHROME_MARKERS) >= 1:
        reasons.append("light_news_chrome")
    if TCFD_DEFINITION_ONLY.search(passage) and _company_in_passage(row.get("company", ""), passage):
        reasons.append("tcfd_definition_mixed")
    if _is_truncated(passage):
        reasons.append("truncated_excerpt")
    if len(passage) > 500 and ("발간했다" in passage or "밝혔다" in passage):
        reasons.append("passage_too_broad")
    if not reasons:
        return False, ""
    return True, ";".join(reasons)


def _reject_reason(row: Dict[str, Any]) -> Optional[Tuple[str, str]]:
    company = row.get("company") or ""
    passage = _passage(row)
    blob = _blob(row)

    if _is_cross_company(company, passage, blob):
        return "cross_company_contamination", "Passage chứa công ty khác hoặc listing cross-company."
    if _is_nav_contact(blob):
        return "portal_nav_contact", "Portal/nav/contact page — không phải ESG disclosure body."
    if _is_listing_index_noise(blob, passage):
        return "listing_index_meta", "Listing/index/report-meta thuần (목록, 조회수, SASB/TCFD Index)."
    if _is_financial_noise(blob, passage):
        return "financial_non_esg", "Financial/analyst dominant, ESG signal yếu."
    if _is_framework_only(passage, company):
        return "framework_only_no_company_fact", "Framework/index definition không có company fact cụ thể."
    if _esg_signal(blob, passage) < 2:
        return "insufficient_esg_substance", "Không đủ tín hiệu ESG substance trong excerpt."
    if len(passage) < 30:
        return "weak_grounding", "Excerpt quá ngắn, grounding yếu."
    return None


def _infer_cluster_id(row: Dict[str, Any]) -> str:
    hint = row.get("workbook_cluster_hint") or "FC_GENERAL"
    passage = _passage(row)
    company = row.get("company") or ""
    q = row.get("question_draft") or ""

    if "FC_NET_ZERO" in hint or ("2050" in passage and "탄소" in passage):
        return f"{company}::FC_NET_ZERO"
    if "FC_MATERIAL" in hint or "8개 중대" in passage:
        return f"{company}::FC_MATERIAL_8"
    if "FC_KGCS" in hint or "kgcs" in passage.lower():
        return f"{company}::FC_KGCS"
    if "FC_ESG_GOVERNANCE" in hint or "esg위원회" in passage.lower() or ("이사회" in passage and "14회" in passage):
        if "14회" in passage and "44건" in passage:
            return f"{company}::FC_BOARD_2022"
        return f"{company}::FC_ESG_GOVERNANCE"
    if "FC_REPORT_FRAMEWORK" in hint or "tcfd" in passage.lower():
        if TCFD_DEFINITION_ONLY.search(passage):
            return f"{company}::FC_TCFD_DEF"
        return f"{company}::FC_REPORT_FRAMEWORK"
    if "FC_CLIMATE" in hint or "온실가스" in passage:
        return f"{company}::FC_CLIMATE_GHG"

    key = _norm_ws(f"{q}|{passage[:80]}").lower()
    digest = hashlib.md5(key.encode("utf-8")).hexdigest()[:10]
    return f"{company}::{hint}::{digest}"


def _cluster_fingerprint(row: Dict[str, Any]) -> str:
    passage = _passage(row)
    nums = re.findall(r"\d+", passage)
    words = re.findall(r"[가-힣]{4,}", passage)[:6]
    return "|".join(words + nums[:4])


def _suggest_rewrite(row: Dict[str, Any], passage: str) -> Tuple[str, str]:
    company = row.get("company") or ""
    lower = passage.lower()
    qtype = row.get("question_type") or "qualitative"

    if "8개 중대 이슈" in passage or ("중대 이슈" in passage and re.search(r"\d+\s*개", passage)):
        q = f"{company}는 이중 중대성 평가를 통해 몇 개의 중대 이슈를 선정했는가?"
    elif "2050" in passage and ("net zero" in lower or "탄소중립" in passage or "넷제로" in passage):
        q = f"{company}는 2050년까지 어떤 탄소중립 목표를 공개했는가?"
    elif "kgcs" in lower and "등급" in passage:
        q = f"{company}는 KGCS ESG경영 평가에서 어떤 등급을 획득했는가?"
    elif "14회" in passage and "이사회" in passage:
        q = f"{company}는 2022년 이사회를 몇 회 개최했는가?"
    elif "esg위원회" in lower or ("이사회" in passage and "기후" in passage):
        q = f"{company}의 ESG 거버넌스 체계는 어떻게 운영되는가?"
    elif "임팩트" in passage or "impact report" in lower:
        q = f"{company}는 어떤 임팩트/지속가능경영 보고서를 발간했는가?"
    else:
        q = row.get("question_draft") or ""

    disclosure = passage[:420].strip()
    if _is_truncated(disclosure) and len(passage) > 80:
        for sent in re.split(r"(?<=[.!?다])\s+", passage):
            if len(sent) >= 40 and _company_in_passage(company, sent):
                disclosure = sent[:420]
                break

    return q, disclosure


def _initial_decision(row: Dict[str, Any]) -> Tuple[str, str, str]:
    reject = _reject_reason(row)
    if reject:
        return "reject", reject[0], reject[1]

    blob = _blob(row)
    passage = _passage(row)
    rewrite, rreason = _needs_rewrite(row, blob, passage)
    if rewrite:
        return "rewrite", rreason, "Fact ESG có thật; cần chỉnh question/disclosure cho workbook."
    return "keep", "clean_grounded_fact", "Grounded rõ, company đúng, usable cho review."


def _row_strength(row: Dict[str, Any], decision: str) -> float:
    rank = float(row.get("rank") or 0)
    bonus = {"keep": 10, "rewrite": 5, "reject": -100, "collapse_into_cluster": -50}
    passage = _passage(row)
    if _company_in_passage(row.get("company", ""), passage):
        rank += 3
    if row.get("seed_origin_type") == "jsonl_primary_candidate":
        rank += 2
    return rank + bonus.get(decision, 0)


def review_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    reviewed: List[Dict[str, Any]] = []
    for row in rows:
        out = dict(row)
        decision, reason, notes = _initial_decision(row)
        cluster_id = _infer_cluster_id(row)
        out.update(
            {
                "review_version": REVIEW_VERSION,
                "review_decision": decision,
                "review_reason": reason,
                "review_notes": notes,
                "cluster_id": cluster_id,
                "cluster_action": "rejected" if decision == "reject" else "pending",
                "rewritten_question_draft": "",
                "rewritten_disclosure_draft": "",
            }
        )
        if decision == "rewrite":
            q, d = _suggest_rewrite(row, _passage(row))
            out["rewritten_question_draft"] = q
            out["rewritten_disclosure_draft"] = d
        elif decision == "keep":
            out["rewritten_question_draft"] = row.get("question_draft") or ""
            out["rewritten_disclosure_draft"] = row.get("acceptable_disclosure") or ""
        reviewed.append(out)

    # Cluster collapse among keep/rewrite
    by_cluster: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in reviewed:
        if row["review_decision"] in ("keep", "rewrite"):
            by_cluster[row["cluster_id"]].append(row)

    for cluster_id, members in by_cluster.items():
        if len(members) <= 1:
            members[0]["cluster_action"] = "anchor"
            continue

        # Sub-group by fingerprint similarity within cluster
        groups: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for m in members:
            fp = _cluster_fingerprint(m)
            groups[fp].append(m)

        for _fp, group in groups.items():
            if len(group) <= 1:
                group[0]["cluster_action"] = "anchor"
                continue
            sorted_g = sorted(
                group,
                key=lambda x: _row_strength(x, x["review_decision"]),
                reverse=True,
            )
            anchor = sorted_g[0]
            anchor["cluster_action"] = "anchor"
            for dup in sorted_g[1:]:
                dup["review_decision"] = "collapse_into_cluster"
                dup["review_reason"] = "duplicate_cluster_variant"
                dup["review_notes"] = f"Trùng cụm với anchor {anchor.get('seed_id')}"
                dup["cluster_action"] = "collapsed_variant"
                dup["rewritten_question_draft"] = ""
                dup["rewritten_disclosure_draft"] = ""

    return reviewed


def _write_workbook(rows: Sequence[Dict[str, Any]], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "안내"
    ws.append(["Reference Seed Workbook V4 — Review Round 1"])
    ws.append(["Gate", "keep / rewrite / reject / collapse — chưa canonical final"])

    header = [
        "seed_id",
        "company",
        "question_type",
        "candidate_kind",
        "question_draft",
        "acceptable_disclosure",
        "source_record_id",
        "seed_origin_type",
        "review_decision",
        "review_reason",
        "cluster_id",
        "cluster_action",
        "rewritten_question_draft",
        "rewritten_disclosure_draft",
        "review_notes",
    ]
    ws2 = wb.create_sheet("작성")
    ws2.append(header)
    fill = PatternFill("solid", fgColor="1E3A5F")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(header) + 1):
        c = ws2.cell(row=1, column=col)
        c.fill = fill
        c.font = font

    active = [r for r in rows if r.get("review_decision") in ("keep", "rewrite")]
    for row in sorted(active, key=lambda x: (x.get("company", ""), x.get("review_decision", ""), x.get("seed_id", ""))):
        ws2.append([row.get(h, "") for h in header])

    ws3 = wb.create_sheet("Rejected")
    rej_header = header + ["workbook_cluster_hint"]
    ws3.append(rej_header)
    for row in sorted(
        [r for r in rows if r.get("review_decision") in ("reject", "collapse_into_cluster")],
        key=lambda x: (x.get("review_decision", ""), x.get("company", "")),
    ):
        ws3.append([row.get(h, "") for h in rej_header])

    ws4 = wb.create_sheet("요약")
    ws4.append(["metric", "value"])
    counts = _summary_counts(rows)
    for k in ("input_total", "keep", "rewrite", "reject", "collapse_into_cluster"):
        ws4.append([k, counts.get(k, 0)])
    for co, stats in counts.get("by_company", {}).items():
        for decision, n in sorted(stats.items()):
            ws4.append([f"{co}:{decision}", n])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _summary_counts(rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    decisions = Counter(r.get("review_decision") for r in rows)
    by_co: Dict[str, Counter] = defaultdict(Counter)
    for r in rows:
        by_co[r.get("company", "")][r.get("review_decision", "")] += 1
    return {
        "input_total": len(rows),
        "keep": decisions.get("keep", 0),
        "rewrite": decisions.get("rewrite", 0),
        "reject": decisions.get("reject", 0),
        "collapse_into_cluster": decisions.get("collapse_into_cluster", 0),
        "by_company": {co: dict(cnt) for co, cnt in sorted(by_co.items())},
    }


def _pick_examples(rows: Sequence[Dict[str, Any]]) -> Dict[str, List[Dict[str, str]]]:
    examples: Dict[str, List[Dict[str, str]]] = {
        "cross_company_reject": [],
        "listing_reject": [],
        "rewrite_salvage": [],
        "collapse_dup": [],
    }
    for r in rows:
        sid = r.get("seed_id", "")
        if r.get("review_decision") == "reject" and r.get("review_reason") == "cross_company_contamination":
            if len(examples["cross_company_reject"]) < 3:
                examples["cross_company_reject"].append(
                    {"seed_id": sid, "company": r.get("company", ""), "excerpt": _passage(r)[:120]}
                )
        if r.get("review_decision") == "reject" and r.get("review_reason") == "listing_index_meta":
            if len(examples["listing_reject"]) < 3:
                examples["listing_reject"].append(
                    {"seed_id": sid, "company": r.get("company", ""), "excerpt": _passage(r)[:120]}
                )
        if r.get("review_decision") == "rewrite" and r.get("cluster_action") == "anchor":
            if len(examples["rewrite_salvage"]) < 3:
                examples["rewrite_salvage"].append(
                    {
                        "seed_id": sid,
                        "before_q": r.get("question_draft", ""),
                        "after_q": r.get("rewritten_question_draft", ""),
                    }
                )
        if r.get("review_decision") == "collapse_into_cluster":
            if len(examples["collapse_dup"]) < 3:
                examples["collapse_dup"].append(
                    {"seed_id": sid, "cluster_id": r.get("cluster_id", ""), "notes": r.get("review_notes", "")}
                )
    return examples


def write_report(summary: Dict[str, Any], examples: Dict[str, List[Dict[str, str]]], path: Path) -> None:
    lines = [
        "# Golden Set — Workbook Review Round 1",
        "",
        f"Generated: {summary.get('generated_at', '')}",
        "",
        "## Mục tiêu",
        "",
        "Triage `reference_seed_candidates_v4_jsonl.jsonl` thành workbook **reviewable**:",
        "loại noise mạnh (cross-company, listing/index, framework-only), collapse duplicate cluster,",
        "giữ fact thật qua `keep`/`rewrite` — **không** canonical final, **không** gold promotion.",
        "",
        "## Tình trạng v4 trước review",
        "",
        f"- Input: **{summary.get('input_total', 0)}** candidate rows (v4 JSONL)",
        "- Vấn đề: cross-company leakage (레이시온/현대트랜시스), listing/index (무신사), report meta/index (한샘)",
        "- Yield đã về nhưng review cost cao — thiếu tầng triage workbook",
        "",
        "## Rule triage round 1",
        "",
        "| Decision | Điều kiện |",
        "|----------|-----------|",
        "| `keep` | Grounded rõ, company đúng, Q/A usable, không listing/nav/meta |",
        "| `rewrite` | Fact có thật; Q generic hoặc excerpt còn noise nhẹ / truncated |",
        "| `reject` | Cross-company, listing/index/nav, framework-only, grounding yếu |",
        "| `collapse_into_cluster` | Trùng cụm fact với anchor mạnh hơn trong cùng `cluster_id` |",
        "",
        "## Kết quả tổng quan",
        "",
        f"- Total input: **{summary.get('input_total', 0)}**",
        f"- keep: **{summary.get('keep', 0)}**",
        f"- rewrite: **{summary.get('rewrite', 0)}**",
        f"- reject: **{summary.get('reject', 0)}**",
        f"- collapse_into_cluster: **{summary.get('collapse_into_cluster', 0)}**",
        f"- **Reviewable sau round 1 (keep + rewrite):** **{summary.get('reviewable_after_round1', 0)}**",
        "",
        "### Breakdown theo công ty",
        "",
    ]
    for co, stats in summary.get("by_company", {}).items():
        lines.append(f"- **{co}**: keep {stats.get('keep', 0)}, rewrite {stats.get('rewrite', 0)}, "
                     f"reject {stats.get('reject', 0)}, collapse {stats.get('collapse_into_cluster', 0)}")

    lines.extend(["", "### Breakdown theo question_type", ""])
    for qt, n in summary.get("by_question_type", {}).items():
        lines.append(f"- `{qt}`: {n}")

    lines.extend(["", "### Breakdown theo rejection reason", ""])
    for reason, n in summary.get("by_reason", {}).items():
        lines.append(f"- `{reason}`: {n}")

    lines.extend(["", "### Breakdown theo cluster action", ""])
    for action, n in summary.get("by_cluster_action", {}).items():
        lines.append(f"- `{action}`: {n}")

    lines.extend(["", "## Ví dụ cụ thể", ""])
    lines.append("### Cross-company bị reject")
    for ex in examples.get("cross_company_reject", []):
        lines.append(f"- `{ex['seed_id']}` ({ex['company']}): {ex['excerpt']}…")

    lines.append("")
    lines.append("### Listing/index bị reject")
    for ex in examples.get("listing_reject", []):
        lines.append(f"- `{ex['seed_id']}` ({ex['company']}): {ex['excerpt']}…")

    lines.append("")
    lines.append("### Row salvageable được rewrite")
    for ex in examples.get("rewrite_salvage", []):
        lines.append(f"- `{ex['seed_id']}`: `{ex['before_q']}` → `{ex['after_q']}`")

    lines.append("")
    lines.append("### Row duplicate bị collapse")
    for ex in examples.get("collapse_dup", []):
        lines.append(f"- `{ex['seed_id']}` cluster `{ex['cluster_id']}` — {ex['notes']}")

    lines.extend(
        [
            "",
            "## Kết luận",
            "",
            f"- Reviewable rows (keep + rewrite): **{summary.get('reviewable_after_round1', 0)}**",
            f"- Manual review round 2 ready? **{summary.get('manual_review_ready_verdict', '')}**",
            f"- Flag: `manual_review_ready_flag` = **{summary.get('manual_review_ready_flag', False)}**",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_review(
    *,
    input_path: Path,
    reviewed_jsonl: Path,
    rejected_jsonl: Path,
    workbook_path: Path,
    report_path: Path,
    summary_json_path: Path,
) -> Dict[str, Any]:
    rows = read_jsonl(input_path)
    reviewed = review_rows(rows)

    active = [r for r in reviewed if r.get("review_decision") in ("keep", "rewrite")]
    rejected = [r for r in reviewed if r.get("review_decision") in ("reject", "collapse_into_cluster")]

    write_jsonl(reviewed_jsonl, reviewed)
    write_jsonl(rejected_jsonl, rejected)
    _write_workbook(reviewed, workbook_path)

    decisions = Counter(r.get("review_decision") for r in reviewed)
    by_co: Dict[str, Dict[str, int]] = {}
    for co in sorted({r.get("company", "") for r in reviewed}):
        sub = [r for r in reviewed if r.get("company") == co]
        by_co[co] = dict(Counter(r.get("review_decision") for r in sub))

    by_reason = Counter(
        r.get("review_reason") for r in reviewed if r.get("review_decision") == "reject"
    )
    by_qtype_active = Counter(r.get("question_type") for r in active)
    by_cluster_action = Counter(r.get("cluster_action") for r in reviewed)

    reviewable = decisions.get("keep", 0) + decisions.get("rewrite", 0)
    companies_with_active = len({r.get("company") for r in active})
    min_per_co = min(
        (sum(1 for r in active if r.get("company") == co) for co in ("한샘", "무신사", "레이시온")),
        default=0,
    )
    reject_rate = decisions.get("reject", 0) / max(len(rows), 1)
    manual_ready = (
        reviewable >= 40
        and companies_with_active >= 3
        and min_per_co >= 8
        and reject_rate <= 0.55
    )

    examples = _pick_examples(reviewed)
    summary = {
        "review_version": REVIEW_VERSION,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "input_total": len(rows),
        "keep": decisions.get("keep", 0),
        "rewrite": decisions.get("rewrite", 0),
        "reject": decisions.get("reject", 0),
        "collapse_into_cluster": decisions.get("collapse_into_cluster", 0),
        "reviewable_after_round1": reviewable,
        "by_company": by_co,
        "by_reason": dict(by_reason.most_common(20)),
        "by_question_type": dict(by_qtype_active),
        "by_cluster_action": dict(by_cluster_action),
        "manual_review_ready_flag": manual_ready,
        "manual_review_ready_verdict": (
            "Có — đủ sạch để mở manual review round 2 (keep+rewrite workbook)"
            if manual_ready
            else "Một phần — còn noise hoặc yield thấp; vẫn có thể review có chọn lọc"
        ),
        "output_reviewed_jsonl": str(reviewed_jsonl),
        "output_rejected_jsonl": str(rejected_jsonl),
        "output_workbook": str(workbook_path),
    }

    write_report(summary, examples, report_path)
    summary_json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main(argv: Optional[Sequence[str]] = None) -> int:
    root = Path(__file__).resolve().parents[2]
    parser = argparse.ArgumentParser(description="Review reference seed workbook v4 round 1")
    parser.add_argument(
        "--input",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_jsonl.jsonl",
    )
    parser.add_argument(
        "--reviewed-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_reviewed_round1.jsonl",
    )
    parser.add_argument(
        "--rejected-jsonl",
        default="data/golden_set/v2/reference_style/reference_seed_candidates_v4_rejected_round1.jsonl",
    )
    parser.add_argument(
        "--workbook",
        default="data/golden_set/v2/reference_style/reference_seed_workbook_v4_review_round1.xlsx",
    )
    parser.add_argument(
        "--report",
        default="reports/golden_set_workbook_review_round1.md",
    )
    parser.add_argument(
        "--summary-json",
        default="reports/_workbook_review_round1_summary.json",
    )
    args = parser.parse_args(argv)

    summary = run_review(
        input_path=root / args.input,
        reviewed_jsonl=root / args.reviewed_jsonl,
        rejected_jsonl=root / args.rejected_jsonl,
        workbook_path=root / args.workbook,
        report_path=root / args.report,
        summary_json_path=root / args.summary_json,
    )
    print(
        json.dumps(
            {k: summary[k] for k in ("input_total", "keep", "rewrite", "reject", "collapse_into_cluster", "reviewable_after_round1")},
            ensure_ascii=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
